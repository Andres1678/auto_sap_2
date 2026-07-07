from flask import request, jsonify, Blueprint, current_app as app, g
from backend.models import (
    db, Modulo, Consultor, Registro, BaseRegistro, BaseRegistroInfoCoeSapFuncional, Login,
    Rol, Equipo, Horario, Oportunidad, Cliente,
    Permiso, RolPermiso, EquipoPermiso, ConsultorPermiso,
    Ocupacion, Tarea, TareaAlias, Ocupacion, RegistroExcel,
    ConsultorPresupuesto, Proyecto, ProyectoFase, ProyectoModulo, ProyectoFaseProyecto,
    ProyectoMapeo,
    ProyectoPresupuestoMensual, ProyectoPerfilPlan, ProyectoCostoAdicional,
    Perfil, ModuloPerfil, ConsultorPerfil, ProyectoModulo, ProyectoPerfil,
    ProyectoPerfilPlan, ProyectoCostoAdicional, ProyectoMapeo, ProyectoPerfilConsultor, CoeSapFuncionalCalificacion,
    CoeSapFuncionalCalificacionHora
)
from datetime import datetime, timedelta, time, date
from functools import wraps
from sqlalchemy import or_, text, func, extract, and_, cast, Integer, literal
from sqlalchemy.orm import relationship, backref, joinedload, aliased, selectinload
import unicodedata, re
from collections import defaultdict
import pandas as pd
from io import BytesIO
from sqlalchemy.exc import SQLAlchemyError, IntegrityError
import logging
import traceback
import math
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from openpyxl import load_workbook
from zoneinfo import ZoneInfo
import bcrypt
import holidays
import secrets



bp = Blueprint('routes', __name__, url_prefix="/api")

_HORARIO_RE = re.compile(r"^\s*\d{2}:\d{2}\s*-\s*\d{2}:\d{2}\s*$", re.I)

FUNCIONAL_HORARIOS_PERMITIDOS = [
    "08:30 - 18:00",
    "07:30 - 16:00",
    "07:00 - 16:00",
]

FUNCIONAL_HORARIO_DEFAULT = "08:30 - 18:00"

FUNCIONAL_HORARIOS_ANTERIORES = {
    "08:00 - 18:00": "08:30 - 18:00",
    "07:00 - 17:00": "08:30 - 18:00",
}


def _norm_equipo_horario(value):
    s = str(value or "").strip().upper()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return s


def _normalizar_horario_text(value):
    s = str(value or "").strip()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"\s*-\s*", " - ", s)
    return s


def _normalizar_horario_trabajo_por_equipo(horario_trabajo, equipo):
    equipo_norm = _norm_equipo_horario(equipo)
    horario = _normalizar_horario_text(horario_trabajo)

    if equipo_norm == "FUNCIONAL":
        if horario in FUNCIONAL_HORARIOS_PERMITIDOS:
            return horario

        if horario in FUNCIONAL_HORARIOS_ANTERIORES:
            return FUNCIONAL_HORARIOS_ANTERIORES[horario]

        return FUNCIONAL_HORARIO_DEFAULT

    return horario or None

ROLE_TEAM_MAP = {
    "ADMIN_BASIS": "BASIS",
    "ADMIN_FUNCIONAL": "FUNCIONAL",
    "ADMIN_IMPLEMENTACION": "IMPLEMENTACION",
    "ADMIN_ARQUITECTURA": "ARQUITECTURA",
    "ADMIN_CONSULTORIA": "CONSULTORIA",
    "ADMIN_GESTION_DE_PROYECTOS": "GESTION_DE_PROYECTOS",
}

ROLE_POOL_ROLES = {
    "ADMIN_GESTION_PREVENTA",
    "ADMIN_OPORTUNIDADES"
}

GRAFICOS_ALL_ROLES = {
    "ADMIN",
    "ADMIN_GERENTES",
    "ADMIN_GESTION_PREVENTA",
}

def _consultor_role_id(consultor_login):
    rid = getattr(consultor_login, "rol_id", None)
    if rid:
        return int(rid)

    rol_obj = getattr(consultor_login, "rol_obj", None)
    rid = getattr(rol_obj, "id", None)
    return int(rid or 0)

def apply_scope(query, rol, usuario_login):
    if rol == "ADMIN":
        return query  # sin filtro

    if rol == "CONSULTOR":
        return query.filter(Registro.usuario == usuario_login)

    if rol == "ADMIN_OPORTUNIDADES":
        # excepción: o devuelves vacío, o filtras distinto según tu caso
        return query.filter(text("1=0"))  # ejemplo: no aplica a horas

    # admins por equipo
    team = ROLE_TEAM_MAP.get(rol)
    if team:
        return query.filter(Registro.equipo == team)

    # fallback seguro: si no conoces el rol, no muestres nada
    return query.filter(text("1=0"))


def permission_required(codigo_permiso):
    def decorator(fn):
        @auth_required
        @wraps(fn)
        def wrapper(*args, **kwargs):
            consultor = g.current_user
            permisos = obtener_permisos_finales(consultor)

            if codigo_permiso not in permisos:
                return jsonify({"mensaje": f"Permiso '{codigo_permiso}' requerido"}), 403

            return fn(*args, **kwargs)
        return wrapper
    return decorator


def norm_fecha(v):
    if pd.isna(v):
        return None
    if hasattr(v, 'date'):
        return v.date()
    return v

def norm_hora(v):
    if pd.isna(v):
        return None
    if hasattr(v, 'strftime'):
        return v.strftime('%H:%M')
    return str(v)

def safe_float(value):
    try:
        if value is None:
            return 0.0
        if isinstance(value, float) and math.isnan(value):
            return 0.0
        if isinstance(value, str):
            value = value.strip().replace(",", ".")
        return float(value)
    except Exception:
        return 0.0

def normalizar_ocupaciones(registros):

    cambios = False

    for r in registros:
        raw = str(r.ocupacion_azure or "").strip()
        tarea_nombre = (r.tipo_tarea or "").strip()
        
        if " - " in raw:
            continue

        
        if raw.isdigit():
            occ = Ocupacion.query.get(int(raw))
            if occ:
                r.ocupacion_azure = f"{occ.codigo} - {occ.nombre}"
                cambios = True
                continue

        
        tarea = None

       
        if tarea_nombre:
            tarea = Tarea.query.filter(func.lower(Tarea.nombre) == tarea_nombre.lower()).first()

        
        if not tarea:
            alias = TareaAlias.query.filter(
                func.lower(TareaAlias.alias) == tarea_nombre.lower()
            ).first()
            if alias:
                tarea = alias.tarea

        
        if not tarea:
            tarea = Tarea.query.filter(Tarea.nombre.ilike(f"%{tarea_nombre}%")).first()

        
        occ = None
        if tarea:
            if hasattr(tarea, "ocupaciones") and tarea.ocupaciones:
                occ = tarea.ocupaciones[0]
            elif hasattr(tarea, "ocupacion") and tarea.ocupacion:
                occ = tarea.ocupacion
        if occ:
            r.ocupacion_azure = f"{occ.codigo} - {occ.nombre}"
        else:
            r.ocupacion_azure = "00 - SIN CLASIFICAR"

    if cambios:
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()

    return registros


def obtener_permisos_finales(consultor):
    permisos = set()

    # -----------------------------
    # Permisos del Rol
    # -----------------------------
    if consultor.rol_obj:
        for rp in consultor.rol_obj.permisos_asignados:
            permisos.add(rp.permiso.codigo)

   
    if consultor.equipo_obj:
        for ep in consultor.equipo_obj.permisos_asignados:
            permisos.add(ep.permiso.codigo)

    
    for cp in consultor.permisos_especiales:
        permisos.add(cp.permiso.codigo)

    
    
    for p in getattr(consultor, "permisos", []) or []:
        if hasattr(p, "codigo"):       
            permisos.add(p.codigo)
        else:                          
            permisos.add(str(p).strip().upper())

    return permisos

def _get_perfiles_permitidos_proyecto(proyecto_id):
    modulo_ids = [
        int(x.modulo_id)
        for x in (
            ProyectoModulo.query
            .filter(ProyectoModulo.proyecto_id == proyecto_id)
            .filter(ProyectoModulo.activo == True)
            .all()
        )
        if x.modulo_id
    ]

    if not modulo_ids:
        return []

    rows = (
        Perfil.query
        .join(ModuloPerfil, ModuloPerfil.perfil_id == Perfil.id)
        .filter(ModuloPerfil.modulo_id.in_(modulo_ids))
        .filter(ModuloPerfil.activo == True)
        .filter(Perfil.activo == True)
        .distinct()
        .order_by(Perfil.orden.asc(), Perfil.nombre.asc())
        .all()
    )

    return rows


def _is_admin_role(rol: str) -> bool:
    r = str(rol or "").strip().upper()
    return r == "ADMIN" or r.startswith("ADMIN_")


def _is_admin_request(rol_header: str, consultor) -> bool:
    # 1) rol enviado por header/query
    if _is_admin_role(rol_header):
        return True

    # 2) rol real del consultor en BD (rol_obj)
    if consultor and getattr(consultor, "rol_obj", None) and getattr(consultor.rol_obj, "nombre", None):
        return _is_admin_role(consultor.rol_obj.nombre)

    # 3) fallback por campo plano consultor.rol si lo usas en algún lado
    if consultor and getattr(consultor, "rol", None):
        return _is_admin_role(consultor.rol)

    return False


def _validar_horario(horario: str):
    if not horario or not isinstance(horario, str):
        return False, "Horario requerido."
    s = horario.strip().upper()
    if s == "DISPONIBLE":
        return True, None
    if _HORARIO_RE.match(horario):
        try:
            ini_str, fin_str = [p.strip() for p in horario.split("-")]
            h1 = datetime.strptime(ini_str, "%H:%M")
            h2 = datetime.strptime(fin_str, "%H:%M")
            if h1 == h2:
                return False, "El horario no puede tener la misma hora de inicio y fin."
            return True, None
        except Exception:
            return False, "Horario inválido."
    return False, "Formato de horario inválido. Usa HH:MM-HH:MM o 'DISPONIBLE'."

def _client_ip():
    fwd = request.headers.get("X-Forwarded-For")
    if fwd:
        return fwd.split(",")[0].strip()
    return request.remote_addr

def _extract_bearer_token():
    auth = request.headers.get("Authorization", "")
    if not auth or not auth.startswith("Bearer "):
        return None
    return auth.replace("Bearer ", "", 1).strip() or None


def _get_session_from_token():
    token = _extract_bearer_token()
    if not token:
        return None

    sesion = Login.query.filter_by(token=token, activo=True).first()
    return sesion


def _get_consultor_from_token():
    sesion = _get_session_from_token()
    if not sesion:
        return None, None

    consultor = (
        Consultor.query.options(
            joinedload(Consultor.rol_obj)
                .joinedload(Rol.permisos_asignados)
                .joinedload(RolPermiso.permiso),
            joinedload(Consultor.equipo_obj)
                .joinedload(Equipo.permisos_asignados)
                .joinedload(EquipoPermiso.permiso),
            joinedload(Consultor.permisos_especiales)
                .joinedload(ConsultorPermiso.permiso),
            joinedload(Consultor.modulos),
            joinedload(Consultor.horario_obj),
        )
        .filter(Consultor.id == sesion.consultor_id)
        .first()
    )

    return sesion, consultor


def auth_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if request.method == "OPTIONS":
            return ("", 204)

        sesion, consultor = _get_consultor_from_token()
        if not sesion or not consultor:
            return jsonify({"mensaje": "Sesión no válida o expirada"}), 401

        if not bool(getattr(consultor, "activo", True)):
            sesion.activo = False
            sesion.fecha_logout = datetime.utcnow()
            db.session.commit()
            return jsonify({"mensaje": "Usuario inactivo. Contacte al administrador."}), 403

        g.current_session = sesion
        g.current_user = consultor
        return fn(*args, **kwargs)
    return wrapper

def pick(d: dict, *keys, default=None):
    """Devuelve el primer valor no vacío encontrado en las claves dadas."""
    for k in keys:
        if k in d and d[k] not in (None, "", "null", "None"):
            return d[k]
    return default

def _rol_from_request():
    data = request.get_json(silent=True) or {}
    return (request.headers.get('X-User-Rol')
            or request.args.get('rol')
            or data.get('rol')
            or '').strip().upper()

def admin_required(fn):
    @auth_required
    @wraps(fn)
    def wrapper(*args, **kwargs):
        consultor = g.current_user

        rol_real = (
            consultor.rol_obj.nombre
            if getattr(consultor, "rol_obj", None) and getattr(consultor.rol_obj, "nombre", None)
            else (getattr(consultor, "rol", "") or "")
        )

        if not _is_admin_request(rol_real, consultor):
            return jsonify({"mensaje": "Solo ADMIN"}), 403

        return fn(*args, **kwargs)
    return wrapper




def _norm_default(value, default=None):
    """Normaliza valores vacíos a un valor por defecto."""
    if value is None:
        return default
    s = str(value).strip()
    return s if s not in ("", "null", "None") else default

def _ensure_rol(nombre: str):
    if not nombre:
        return None
    r = Rol.query.filter(func.lower(Rol.nombre) == nombre.lower()).first()
    if not r:
        r = Rol(nombre=nombre)
        db.session.add(r)
        db.session.flush()
    return r

def _ensure_equipo(nombre: str):
    if not nombre:
        return None
    e = Equipo.query.filter(func.lower(Equipo.nombre) == nombre.lower()).first()
    if not e:
        e = Equipo(nombre=nombre)
        db.session.add(e)
        db.session.flush()
    return e

def _ensure_horario(rango: str):
    if not rango:
        return None
    h = Horario.query.filter(func.lower(Horario.rango) == rango.lower()).first()
    if not h:
        h = Horario(rango=rango)
        db.session.add(h)
        db.session.flush()
    return h

def get_val(d, *keys):
    """Busca claves posibles sin error de KeyError."""
    for k in keys:
        if isinstance(d, dict) and k in d:
            return d[k]
    return None

def clip(value, field=None):
    """Limpia valores de texto para evitar errores."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return str(value)
    s = str(value).strip()
    return s if s else None

def excel_date_to_iso(value):
    """Convierte fechas Excel (número o string) a formato ISO."""
    try:
        if isinstance(value, (int, float)):
            base = datetime(1899, 12, 30)
            date_val = base + timedelta(days=int(value))
            return date_val.strftime("%Y-%m-%d")
        s = str(value).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except Exception:
                continue
    except Exception:
        pass
    return None

def to_sql_time_str(val):
    """Convierte un string tipo '8:30' a formato HH:MM:SS compatible SQL."""
    if not val:
        return None
    try:
        s = str(val).strip()
        if len(s.split(":")) == 2:
            s += ":00"
        datetime.strptime(s, "%H:%M:%S")
        return s
    except Exception:
        return None

def norm(s: str) -> str:
    return (s or "").strip().upper()

def role_scope(rol_nombre: str, user_equipo: str = ""):
    """
    Retorna un dict con el scope efectivo:
    - {"mode": "ALL"}            => ve todo
    - {"mode": "USER"}           => solo su usuario
    - {"mode": "TEAM", "team":X} => solo equipo X
    """
    r = norm(rol_nombre)

    # ADMIN global
    if r == "ADMIN":
        return {"mode": "ALL"}

    # Cualquier ADMIN_* => se restringe por equipo.
    # Ej: ADMIN_BASIS -> BASIS
    if r.startswith("ADMIN_"):
        team = r.replace("ADMIN_", "").replace("_", " ")
        # Si en DB el equipo está como "GERENCIA DE PROYECTOS" pero el rol dice GESTION_DE_PROYECTOS
        # puedes mapear acá si lo necesitas:
        map_roles = {
            "GESTION DE PROYECTOS": "GERENCIA DE PROYECTOS",
        }
        team = map_roles.get(team, team)
        return {"mode": "TEAM", "team": team}

    # CONSULTOR o cualquier otro => usuario
    return {"mode": "USER"}

def _perfil_build_code(nombre):
    s = str(nombre or "").strip()
    s = unicodedata.normalize("NFD", s)
    s = s.encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^A-Za-z0-9]+", "_", s.upper()).strip("_")
    return s[:50] or "PERFIL"


def perfil_to_dict(x: Perfil, include_modulos=False):
    out = {
        "id": x.id,
        "codigo": x.codigo,
        "nombre": x.nombre,
        "descripcion": x.descripcion,
        "activo": bool(x.activo),
        "orden": int(x.orden or 0),
    }

    if include_modulos:
        out["modulos"] = [
            {
                "id": mp.modulo.id,
                "nombre": mp.modulo.nombre,
                "activo": bool(mp.activo),
                "modulo_perfil_id": mp.id,
            }
            for mp in (x.modulos or [])
            if mp.modulo
        ]

    return out


def _perfil_to_dict(x: Perfil):
    return {
        "id": x.id,
        "codigo": x.codigo,
        "nombre": x.nombre,
        "descripcion": x.descripcion,
        "activo": bool(x.activo),
        "orden": int(x.orden or 0),
    }


def modulo_perfil_to_dict(x: ModuloPerfil):
    return {
        "id": x.id,
        "modulo_id": x.modulo_id,
        "perfil_id": x.perfil_id,
        "activo": bool(x.activo),
        "modulo": {
            "id": x.modulo.id,
            "nombre": x.modulo.nombre,
        } if x.modulo else None,
        "perfil": perfil_to_dict(x.perfil) if x.perfil else None,
    }

def _parse_date_safe(v):
    if v in (None, "", "null", "None"):
        return None
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(str(v).strip()[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def consultor_perfil_to_dict(x: ConsultorPerfil):
    return {
        "id": x.id,
        "consultor_id": x.consultor_id,
        "perfil_id": x.perfil_id,
        "activo": bool(x.activo),
        "fecha_inicio": x.fecha_inicio.isoformat() if x.fecha_inicio else None,
        "fecha_fin": x.fecha_fin.isoformat() if x.fecha_fin else None,
        "consultor": {
            "id": x.consultor.id,
            "nombre": x.consultor.nombre,
            "usuario": x.consultor.usuario,
        } if x.consultor else None,
        "perfil": perfil_to_dict(x.perfil) if x.perfil else None,
    }

def _consultor_modulos_payload(consultor):
    """
    Devuelve los módulos del consultor desde:
    1. Relación muchos a muchos consultor.modulos
    2. Campo directo consultor.modulo_id
    """
    modulos_map = {}

    # Relación muchos a muchos
    for m in getattr(consultor, "modulos", []) or []:
        if m and getattr(m, "id", None):
            modulos_map[int(m.id)] = {
                "id": int(m.id),
                "nombre": m.nombre,
            }

    # Campo directo modulo_id
    modulo_id = getattr(consultor, "modulo_id", None)

    if modulo_id and int(modulo_id) not in modulos_map:
        modulo = Modulo.query.get(int(modulo_id))

        if modulo:
            modulos_map[int(modulo.id)] = {
                "id": int(modulo.id),
                "nombre": modulo.nombre,
            }

    return list(modulos_map.values())
# ===============================
# Catálogos
# ===============================
@bp.route('/logout', methods=['POST'])
@auth_required
def logout():
    try:
        sesion = g.current_session
        sesion.activo = False
        sesion.fecha_logout = datetime.utcnow()
        db.session.commit()

        return jsonify({"mensaje": "Sesión cerrada correctamente"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"mensaje": f"Error cerrando sesión: {str(e)}"}), 500

@bp.route('/roles', methods=['GET'])
@permission_required("ROLES_ADMIN")  
def listar_roles():
    roles = Rol.query.order_by(Rol.nombre.asc()).all()
    return jsonify([{"id": r.id, "nombre": r.nombre} for r in roles]), 200


@bp.route('/equipos', methods=['GET'])
def listar_equipos():
    equipos = Equipo.query.order_by(Equipo.nombre).all()
    return jsonify([{"id": e.id, "nombre": e.nombre} for e in equipos]), 200

@bp.route('/horarios', methods=['GET'])
def listar_horarios():
    horarios = Horario.query.order_by(Horario.rango).all()
    return jsonify([{"id": h.id, "rango": h.rango} for h in horarios]), 200

# ===============================
# LOGIN
# ===============================

@bp.route('/login', methods=['POST'])
def login():
    data = request.get_json(silent=True) or {}

    usuario = (data.get("usuario") or "").strip().lower()
    password = data.get("password") or ""
    horario = data.get("horario")

    if not usuario or not password:
        return jsonify({"mensaje": "Usuario y password son obligatorios"}), 400

    consultor = (
        Consultor.query.options(
            joinedload(Consultor.rol_obj)
                .joinedload(Rol.permisos_asignados)
                .joinedload(RolPermiso.permiso),
            joinedload(Consultor.equipo_obj)
                .joinedload(Equipo.permisos_asignados)
                .joinedload(EquipoPermiso.permiso),
            joinedload(Consultor.permisos_especiales)
                .joinedload(ConsultorPermiso.permiso),
            joinedload(Consultor.modulos),
            joinedload(Consultor.horario_obj),
        )
        .filter(func.lower(Consultor.usuario) == usuario)
        .first()
    )

    if not consultor:
        return jsonify({"mensaje": "Credenciales incorrectas"}), 401

    if not bool(getattr(consultor, "activo", True)):
        return jsonify({"mensaje": "Usuario inactivo. Contacte al administrador."}), 403

    stored = (consultor.password or "")

    if stored.startswith("$2"):
        ok_pass = bcrypt.checkpw(password.encode("utf-8"), stored.encode("utf-8"))
    else:
        ok_pass = (stored == password)

    if not ok_pass:
        return jsonify({"mensaje": "Credenciales incorrectas"}), 401

    if horario:
        ok, msg = _validar_horario(horario)
        if not ok:
            return jsonify({"mensaje": msg or "Horario inválido"}), 400

        rango = horario.strip()
        h = Horario.query.filter_by(rango=rango).first()
        if not h:
            h = Horario(rango=rango)
            db.session.add(h)
            db.session.flush()

        consultor.horario_id = h.id

    # Cerrar sesiones activas anteriores del mismo consultor
    Login.query.filter_by(consultor_id=consultor.id, activo=True).update({
        "activo": False,
        "fecha_logout": datetime.utcnow()
    })

    token = secrets.token_urlsafe(48)

    login_log = Login(
        consultor_id=consultor.id,
        usuario=consultor.usuario,
        horario_asignado=(consultor.horario_obj.rango if consultor.horario_obj else (horario or "N/D")),
        ip_address=_client_ip(),
        user_agent=request.headers.get("User-Agent", ""),
        fecha_login=datetime.utcnow(),
        token=token,
        activo=True
    )

    db.session.add(login_log)
    db.session.commit()

    permisos_list = sorted(list(obtener_permisos_finales(consultor)))

    modulos_payload = _consultor_modulos_payload(consultor)

    user_payload = {
        "id": consultor.id,
        "usuario": consultor.usuario,
        "nombre": consultor.nombre,
        "rol": consultor.rol_obj.nombre.upper() if consultor.rol_obj else "CONSULTOR",
        "equipo": consultor.equipo_obj.nombre.upper() if consultor.equipo_obj else "SIN EQUIPO",
        "horario": consultor.horario_obj.rango if consultor.horario_obj else (horario or "N/D"),
        "consultor_id": consultor.id,
        "modulos": modulos_payload,
        "modulo": modulos_payload[0]["nombre"] if modulos_payload else "SIN MODULO",
        "permisos": permisos_list
    }

    return jsonify({
        "token": token,
        "user": user_payload
    }), 200

def _to_bool(v, default=True):
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v == 1
    s = str(v).strip().lower()
    return s in ("1", "true", "si", "sí", "yes", "y")

@bp.route('/consultores', methods=['POST'])
@permission_required("CONSULTORES_CREAR")
def crear_consultor():
    data = request.get_json() or {}

    c = Consultor(
        usuario=data.get('usuario'),
        nombre=data.get('nombre'),
        password=data.get('password'),
        activo=_to_bool(data.get("activo"), default=True),  # ✅ NUEVO
    )
    _apply_catalog_fields_to_consultor(c, data)

    db.session.add(c)
    db.session.flush()

    modulos_ids = data.get('modulos', [])
    if modulos_ids:
        mods = Modulo.query.filter(Modulo.id.in_(modulos_ids)).all()
        c.modulos = mods

    try:
        db.session.commit()
        return jsonify({"mensaje": "Consultor creado correctamente"}), 201
    except IntegrityError:
        db.session.rollback()
        return jsonify({"mensaje": "Error: usuario duplicado o datos inválidos"}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({"mensaje": f"Error interno: {e}"}), 500
    
@bp.route("/consultores/<int:consultor_id>/activo", methods=["PUT"])
@permission_required("CONSULTORES_EDITAR")
def set_activo_consultor(consultor_id):
    data = request.get_json() or {}
    c = Consultor.query.get_or_404(consultor_id)
    c.activo = _to_bool(data.get("activo"), default=True)
    db.session.commit()
    return jsonify({"mensaje": "Estado actualizado", "activo": bool(c.activo)}), 200


# ===============================
# Helper: asignar catálogos y módulos
# ===============================
def _apply_catalog_fields_to_consultor(consultor, data):
    """Actualiza relaciones de catálogo (Rol, Equipo, Horario, Módulos)."""
    try:
        # --- Rol ---
        rol_name = data.get("rol")
        if rol_name:
            rol = Rol.query.filter_by(nombre=rol_name).first()
            consultor.rol_id = rol.id if rol else None

        # --- Equipo ---
        equipo_name = data.get("equipo")
        if equipo_name:
            equipo = Equipo.query.filter_by(nombre=equipo_name).first()
            consultor.equipo_id = equipo.id if equipo else None

        # --- Horario ---
        horario_name = data.get("horario")
        if horario_name:
            horario = Horario.query.filter_by(rango=horario_name).first()
            consultor.horario_id = horario.id if horario else None

        # --- Módulos ---
        modulos_ids = data.get("modulos")
        if isinstance(modulos_ids, list):
            mods = Modulo.query.filter(Modulo.id.in_(modulos_ids)).all()
            consultor.modulos = mods

    except Exception as e:
        logging.error(f"Error aplicando catálogos a consultor: {e}")
        raise

# ===============================
# PUT /api/consultores/<id> — Editar consultor
# ===============================
@bp.route('/consultores/<int:id>', methods=['PUT'])
@permission_required("CONSULTORES_EDITAR")
def editar_consultor(id):
    data = request.get_json() or {}
    c = Consultor.query.get_or_404(id)

    try:
        # --- Campos básicos ---
        c.usuario = data.get('usuario', c.usuario)
        c.nombre = data.get('nombre', c.nombre)
        if data.get('password'):
            c.password = data['password']

        # --- Campos relacionales ---
        _apply_catalog_fields_to_consultor(c, data)

        db.session.commit()
        return jsonify({"mensaje": "Consultor actualizado correctamente"}), 200

    except SQLAlchemyError as e:
        db.session.rollback()
        logging.error(f"Error SQLAlchemy al actualizar consultor {id}: {e}")
        return jsonify({"mensaje": f"Error al actualizar consultor: {str(e)}"}), 500

    except Exception as e:
        db.session.rollback()
        logging.error(f"Error general al actualizar consultor {id}: {e}")
        return jsonify({"mensaje": f"Error interno: {str(e)}"}), 500

# ===============================
# DELETE /api/consultores/<id> — Eliminar consultor
# ===============================
@bp.route('/consultores/<int:id>', methods=['DELETE'])
@permission_required("CONSULTORES_ELIMINAR")
def eliminar_consultor(id):
    c = Consultor.query.get_or_404(id)
    try:
        db.session.delete(c)
        db.session.commit()
        return jsonify({"mensaje": "Consultor eliminado correctamente"}), 200
    except SQLAlchemyError as e:
        db.session.rollback()
        logging.error(f"Error SQLAlchemy al eliminar consultor {id}: {e}")
        return jsonify({"mensaje": f"Error al eliminar consultor: {str(e)}"}), 500
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error general al eliminar consultor {id}: {e}")
        return jsonify({"mensaje": f"Error interno: {str(e)}"}), 500

# ===============================
# Registros (listar / crear / editar / eliminar / resumen)
# ===============================

def registro_to_dict(r: Registro):
    consul = r.consultor
    modulo_name = r.modulo or (consul.modulo.nombre if (consul and consul.modulo) else None)
    equipo_name = r.equipo
    if not equipo_name and consul and consul.equipo_id:
        eq = Equipo.query.get(consul.equipo_id)
        equipo_name = eq.nombre if eq else None
    return {
        'id': r.id,
        'fecha': r.fecha,
        'cliente': r.cliente,
        'nroCasoCliente': r.nro_caso_cliente,
        'nroCasoInterno': r.nro_caso_interno,
        'nroCasoEscaladoSap': r.nro_caso_escalado,
        'tipoTarea': r.tipo_tarea,
        'horaInicio': r.hora_inicio,
        'horaFin': r.hora_fin,
        'tiempoInvertido': r.tiempo_invertido,
        'actividadMalla': r.actividad_malla,
        'oncall': r.oncall,
        'desborde': r.desborde,
        'tiempoFacturable': r.tiempo_facturable,
        'horasAdicionales': _calcular_horas_adicionales_por_horario(
            r.hora_inicio,
            r.hora_fin,
            getattr(r, "horario_trabajo", None),
            equipo_name,
            r.horas_adicionales,
        ),
        'horarioTrabajo': _normalizar_horario_trabajo_por_equipo(
            getattr(r, "horario_trabajo", None),
            equipo_name,
        ),
        'horario_trabajo': _normalizar_horario_trabajo_por_equipo(
            getattr(r, "horario_trabajo", None),
            equipo_name,
        ),
        'descripcion': r.descripcion,
        'totalHoras': r.total_horas,
        'consultor': consul.nombre if consul else None,
        'modulo': modulo_name,
        'equipo': equipo_name,
        'bloqueado': r.bloqueado
    }

def _basis_defaults_from_payload(data: dict):
    return {
        "actividad_malla": _norm_default(data.get("actividadMalla") or data.get("actividad_malla"), "N/APLICA"),
        "oncall":          _norm_default(data.get("oncall"), "N/A"),
        "desborde":        _norm_default(data.get("desborde"), "N/A"),
        "nro_escalado": (
            data.get("nroCasoEscaladoSap")
            if data.get("nroCasoEscaladoSap") not in (None, "", "null", "None")
            else data.get("nro_caso_escalado")
),

    }

@bp.route('/consultores/horario', methods=['GET'])
def horario_consultor():
    usuario = request.args.get('usuario')
    if not usuario:
        return jsonify({'mensaje': 'Parámetro "usuario" requerido'}), 400
    consultor = Consultor.query.filter_by(usuario=usuario).first()
    if not consultor:
        return jsonify({'mensaje': 'Usuario no encontrado'}), 404

    opciones = [h.rango for h in Horario.query.order_by(Horario.rango).all()]
    h_actual = None
    if consultor.horario_id:
        hh = Horario.query.get(consultor.horario_id)
        h_actual = hh.rango if hh else None
    return jsonify({'horario': h_actual, 'opciones': opciones})

def _parse_hhmm_to_minutes(value):
    s = str(value or "").strip()
    if not re.match(r"^\d{2}:\d{2}$", s):
        return None
    try:
        h, m = s.split(":")
        h = int(h)
        m = int(m)
        if h < 0 or h > 23 or m < 0 or m > 59:
            return None
        return h * 60 + m
    except Exception:
        return None


def _ranges_overlap(start_a, end_a, start_b, end_b):
    return max(start_a, end_a) != min(start_a, end_a) and max(start_a, start_b) < min(end_a, end_b)


def _calcular_tiempo_horas(hora_inicio, hora_fin):
    ini = _parse_hhmm_to_minutes(hora_inicio)
    fin = _parse_hhmm_to_minutes(hora_fin)
    if ini is None or fin is None or fin <= ini:
        return 0
    mins = fin - ini
    return round(mins / 60.0, 2)


def _buscar_registro_traslapado(fecha, usuario_consultor, hora_inicio, hora_fin, exclude_id=None):
    nuevo_ini = _parse_hhmm_to_minutes(hora_inicio)
    nuevo_fin = _parse_hhmm_to_minutes(hora_fin)

    if nuevo_ini is None or nuevo_fin is None or nuevo_fin <= nuevo_ini:
        return None

    q = Registro.query.filter(
        Registro.fecha == fecha,
        func.lower(Registro.usuario_consultor) == func.lower(str(usuario_consultor).strip())
    )

    if exclude_id:
        q = q.filter(Registro.id != int(exclude_id))

    existentes = q.all()

    for r in existentes:
        old_ini = _parse_hhmm_to_minutes(r.hora_inicio)
        old_fin = _parse_hhmm_to_minutes(r.hora_fin)

        if old_ini is None or old_fin is None or old_fin <= old_ini:
            continue

        if max(nuevo_ini, old_ini) < min(nuevo_fin, old_fin):
            return r

    return None

def _minutes_to_hhmm(value: int) -> str:
    h = value // 60
    m = value % 60
    return f"{h:02d}:{m:02d}"


def _parse_horario_range_to_minutes(horario_trabajo: str):
    if not horario_trabajo:
        return None

    s = str(horario_trabajo).strip()
    if not re.match(r"^\d{2}:\d{2}\s*-\s*\d{2}:\d{2}$", s):
        return None

    ini_txt, fin_txt = [x.strip() for x in s.split("-")]
    ini = _parse_hhmm_to_minutes(ini_txt)
    fin = _parse_hhmm_to_minutes(fin_txt)

    if ini is None or fin is None:
        return None

    if fin <= ini:
        return None

    return {"start": ini, "end": fin}


def _dividir_registro_por_horario(hora_inicio: str, hora_fin: str, horario_trabajo: str):
    """
    Retorna una lista de fragmentos:
    [
      {"hora_inicio": "13:00", "hora_fin": "18:00", "horas_adicionales": "No"},
      {"hora_inicio": "18:00", "hora_fin": "19:00", "horas_adicionales": "Sí"},
    ]
    """
    ini = _parse_hhmm_to_minutes(hora_inicio)
    fin = _parse_hhmm_to_minutes(hora_fin)

    if ini is None or fin is None or fin <= ini:
        return []

    rango = _parse_horario_range_to_minutes(horario_trabajo)

    # si no hay horario válido, deja un solo registro
    if not rango:
        return [{
            "hora_inicio": hora_inicio,
            "hora_fin": hora_fin,
            "horas_adicionales": "N/D",
        }]

    work_start = rango["start"]
    work_end = rango["end"]

    fragmentos = []

    # Parte antes del horario laboral => extra
    if ini < work_start:
        extra_fin = min(fin, work_start)
        if extra_fin > ini:
            fragmentos.append({
                "hora_inicio": _minutes_to_hhmm(ini),
                "hora_fin": _minutes_to_hhmm(extra_fin),
                "horas_adicionales": "Sí",
            })

    # Parte dentro del horario laboral => normal
    normal_ini = max(ini, work_start)
    normal_fin = min(fin, work_end)
    if normal_fin > normal_ini:
        fragmentos.append({
            "hora_inicio": _minutes_to_hhmm(normal_ini),
            "hora_fin": _minutes_to_hhmm(normal_fin),
            "horas_adicionales": "No",
        })

    # Parte después del horario laboral => extra
    if fin > work_end:
        extra_ini = max(ini, work_end)
        if fin > extra_ini:
            fragmentos.append({
                "hora_inicio": _minutes_to_hhmm(extra_ini),
                "hora_fin": _minutes_to_hhmm(fin),
                "horas_adicionales": "Sí",
            })

    # fallback
    if not fragmentos:
        fragmentos.append({
            "hora_inicio": hora_inicio,
            "hora_fin": hora_fin,
            "horas_adicionales": "No",
        })

    return fragmentos


def _calcular_horas_adicionales_por_horario(hora_inicio, hora_fin, horario_trabajo, equipo=None, fallback="N/D"):
    horario_normalizado = _normalizar_horario_trabajo_por_equipo(horario_trabajo, equipo)
    fragmentos = _dividir_registro_por_horario(hora_inicio, hora_fin, horario_normalizado)

    if not fragmentos:
        return fallback or "N/D"

    valores = {str(f.get("horas_adicionales") or "").strip().upper() for f in fragmentos}

    if "SÍ" in valores or "SI" in valores:
        return "Sí"

    if valores == {"NO"}:
        return "No"

    return fallback or "N/D"

def _norm_text_basic(value):
    s = str(value or "").strip().upper()
    s = unicodedata.normalize("NFD", s)
    return "".join(ch for ch in s if unicodedata.category(ch) != "Mn")


def _is_vacaciones_payload(data: dict) -> bool:
    if not bool(data.get("generarRangoVacaciones")):
        return False

    tipo = pick(data, "tipoTarea", "tipo_tarea", default="")
    tipo_norm = _norm_text_basic(tipo)

    # En tu catálogo se ve como "15 - Vacaciones / Incapacidades"
    return tipo_norm.startswith("15") or "VACACIONES" in tipo_norm


def _parse_iso_date_for_range(value):
    s = str(value or "").strip()
    if not s:
        return None

    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def _build_iso_dates_range(start_value, end_value):
    start = _parse_iso_date_for_range(start_value)
    end = _parse_iso_date_for_range(end_value)

    if not start or not end or end < start:
        return []

    dates = []
    current = start

    while current <= end:
        dates.append(current.isoformat())
        current = current + timedelta(days=1)

    return dates

@bp.route('/registrar-hora', methods=['POST'])
def registrar_hora():
    data = request.get_json(force=True, silent=True) or {}

    # ------------------------------------------------------------------
    # 1) BUSCAR CONSULTOR
    # ------------------------------------------------------------------
    usuario = pick(data, 'usuario', 'username', 'user') or request.headers.get('X-User-Usuario')
    consultor = None

    if usuario:
        consultor = Consultor.query.filter(
            func.lower(Consultor.usuario) == str(usuario).strip().lower()
        ).first()

    cid = pick(data, 'consultor_id', 'consultorId', 'id')
    if not consultor and cid:
        try:
            consultor = Consultor.query.get(int(cid))
        except Exception:
            consultor = None

    if not consultor:
        cname = pick(data, 'consultor', 'nombre')
        if cname:
            consultor = Consultor.query.filter(
                func.lower(Consultor.nombre) == str(cname).strip().lower()
            ).first()

    if not consultor:
        return jsonify({'mensaje': 'Consultor no encontrado'}), 404

    # ------------------------------------------------------------------
    # 2) CAMPOS REQUERIDOS
    # ------------------------------------------------------------------
    fecha = pick(data, 'fecha')
    cliente = pick(data, 'cliente')
    hora_inicio = pick(data, 'horaInicio', 'hora_inicio')
    hora_fin = pick(data, 'horaFin', 'hora_fin')

    es_rango_vacaciones = _is_vacaciones_payload(data)

    if not cliente or not hora_inicio or not hora_fin:
        return jsonify({'mensaje': 'Campos obligatorios faltantes'}), 400

    if es_rango_vacaciones:
        fecha_inicio_vacaciones = pick(
            data,
            'fechaInicioVacaciones',
            'fecha_inicio_vacaciones'
        )
        fecha_fin_vacaciones = pick(
            data,
            'fechaFinVacaciones',
            'fecha_fin_vacaciones'
        )

        fechas_a_crear = _build_iso_dates_range(
            fecha_inicio_vacaciones,
            fecha_fin_vacaciones
        )

        if not fechas_a_crear:
            return jsonify({
                'mensaje': 'Rango de vacaciones inválido. Verifica fecha de inicio y fecha de fin.'
            }), 400

        fecha = fechas_a_crear[0]
    else:
        if not fecha:
            return jsonify({'mensaje': 'Campos obligatorios faltantes'}), 400

        fechas_a_crear = [fecha]

    tiempo_calculado = _calcular_tiempo_horas(hora_inicio, hora_fin)
    if tiempo_calculado <= 0:
        return jsonify({'mensaje': 'Hora fin debe ser mayor a hora inicio'}), 400

    # ------------------------------------------------------------------
    # 3) VALIDAR TRASLAPE EN BACKEND
    # ------------------------------------------------------------------
    for fecha_item in fechas_a_crear:
        conflicto = _buscar_registro_traslapado(
            fecha=fecha_item,
            usuario_consultor=consultor.usuario,
            hora_inicio=hora_inicio,
            hora_fin=hora_fin,
            exclude_id=None
        )

        if conflicto:
            return jsonify({
                'mensaje': (
                    f'Ya existe un registro que se cruza con este rango el día {fecha_item}: '
                    f'{conflicto.hora_inicio} - {conflicto.hora_fin} (ID: {conflicto.id})'
                )
            }), 409

    # ------------------------------------------------------------------
    # 4) DETERMINAR TAREA
    # ------------------------------------------------------------------
    tarea_id = pick(data, "tarea_id")
    tarea_obj = None

    if tarea_id:
        try:
            tarea_id = int(tarea_id)
            tarea_obj = Tarea.query.get(tarea_id)
            if not tarea_obj:
                return jsonify({'mensaje': 'Tarea inválida'}), 400
        except Exception:
            return jsonify({'mensaje': 'Tarea inválida'}), 400
    else:
        tipoTareaRaw = pick(data, "tipoTarea", "tipo_tarea")
        if tipoTareaRaw:
            codigo = str(tipoTareaRaw).split("-", 1)[0].strip()
            tarea_obj = Tarea.query.filter(Tarea.codigo == codigo).first()
            if tarea_obj:
                tarea_id = tarea_obj.id

    # ------------------------------------------------------------------
    # 5) VALORES NUMÉRICOS
    # ------------------------------------------------------------------
    tiempo_invertido = float(
        pick(data, 'tiempoInvertido', 'tiempo_invertido', default=tiempo_calculado) or tiempo_calculado
    )
    tiempo_facturable = float(
        pick(data, 'tiempoFacturable', 'tiempo_facturable', default=0) or 0
    )
    total_horas = float(
        pick(data, 'totalHoras', 'total_horas', default=tiempo_calculado) or tiempo_calculado
    )

    # ------------------------------------------------------------------
    # 6) MÓDULO
    # ------------------------------------------------------------------
    modulo_final = pick(data, "modulo")
    
    if modulo_final:
        modulo_final = str(modulo_final).strip()
    else:
        modulos_payload = _consultor_modulos_payload(consultor)
        modulo_final = modulos_payload[0]["nombre"] if modulos_payload else "SIN MODULO"

    # ------------------------------------------------------------------
    # 7) HORARIO DE TRABAJO
    # ------------------------------------------------------------------
    horario_trabajo_raw = (
        pick(data, 'horario_trabajo', 'horarioTrabajo')
        or (Horario.query.get(consultor.horario_id).rango if consultor.horario_id else None)
    )

    equipo_para_horario = pick(data, 'equipo')
    if not equipo_para_horario and consultor.equipo_id:
        eq_horario = Equipo.query.get(consultor.equipo_id)
        equipo_para_horario = eq_horario.nombre if eq_horario else None

    horario_trabajo = _normalizar_horario_trabajo_por_equipo(
        horario_trabajo_raw,
        equipo_para_horario
    )

    # ------------------------------------------------------------------
    # 7.1) DIVIDIR EL REGISTRO SEGÚN EL HORARIO
    # ------------------------------------------------------------------
    if es_rango_vacaciones:
        fragmentos = [{
            "hora_inicio": hora_inicio,
            "hora_fin": hora_fin,
            "horas_adicionales": pick(
                data,
                "horasAdicionales",
                "horas_adicionales",
                default="No"
            ),
        }]
    else:
        fragmentos = _dividir_registro_por_horario(
            hora_inicio,
            hora_fin,
            horario_trabajo
        )

        if not fragmentos:
            return jsonify({'mensaje': 'No se pudo dividir el rango horario'}), 400

    # validar traslape por cada fragmento
    for fecha_item in fechas_a_crear:
        for frag in fragmentos:
            conflicto = _buscar_registro_traslapado(
                fecha=fecha_item,
                usuario_consultor=consultor.usuario,
                hora_inicio=frag["hora_inicio"],
                hora_fin=frag["hora_fin"],
                exclude_id=None
            )

            if conflicto:
                return jsonify({
                    'mensaje': (
                        f'Ya existe un registro que se cruza con este rango el día {fecha_item}: '
                        f'{conflicto.hora_inicio} - {conflicto.hora_fin} (ID: {conflicto.id})'
                    )
                }), 409

    # ------------------------------------------------------------------
    # 8) EQUIPO
    # ------------------------------------------------------------------
    equipo_final = pick(data, 'equipo')
    if not equipo_final and consultor.equipo_id:
        eq = Equipo.query.get(consultor.equipo_id)
        equipo_final = eq.nombre if eq else None

    if isinstance(equipo_final, str):
        equipo_final = equipo_final.strip().upper()

    # ------------------------------------------------------------------
    # 9) CAMPOS BASIS
    # ------------------------------------------------------------------
    es_basis = (equipo_final == "BASIS")

    bd = _basis_defaults_from_payload(data) if es_basis else {
        "actividad_malla": "N/APLICA",
        "oncall": "N/A",
        "desborde": "N/A",
        "nro_escalado": pick(data, 'nroCasoEscaladoSap', 'nro_caso_escalado')
    }

    # ------------------------------------------------------------------
    # 10) OCUPACIÓN
    # ------------------------------------------------------------------
    ocupacion_id = pick(data, "ocupacion_id")

    if ocupacion_id:
        try:
            ocupacion_id = int(ocupacion_id)
        except Exception:
            return jsonify({'mensaje': 'Ocupación inválida'}), 400

        occ = Ocupacion.query.get(ocupacion_id)
        if not occ:
            return jsonify({'mensaje': 'Ocupación inválida'}), 400
    else:
        if tarea_id:
            t = Tarea.query.options(db.joinedload(Tarea.ocupaciones)).get(tarea_id)
            if t and t.ocupaciones:
                ocupacion_id = t.ocupaciones[0].id
            else:
                ocupacion_id = None

    # --------------------------------------------------
    # 10.1) Validación cliente restringido por ocupación
    # --------------------------------------------------
    cliente_upper = str(cliente or "").strip().upper()
    occ_codigo = ""

    if ocupacion_id:
        occ_obj = Ocupacion.query.get(ocupacion_id)
        occ_codigo = str(getattr(occ_obj, "codigo", "") or "").strip()

        # Ocupaciones que NO pueden usar HITSS/CLARO
        if occ_codigo in {"01", "02"} and cliente_upper == "HITSS/CLARO":
            return jsonify({
                "mensaje": "Las ocupaciones 01 y 02 no pueden registrarse para el cliente HITSS/CLARO"
            }), 400

        # Ocupación 03 SOLO puede usar HITSS/CLARO
        if occ_codigo == "03" and cliente_upper != "HITSS/CLARO":
            return jsonify({
                "mensaje": "La ocupación 03 solo puede registrarse para el cliente HITSS/CLARO"
            }), 400

    # ------------------------------------------------------------------
    # 11) PROYECTOS
    # ------------------------------------------------------------------
    proyecto_id = pick(data, "proyecto_id")
    fase_proyecto_id = pick(data, "fase_proyecto_id", "faseProyectoId")

    try:
        proyecto_id = int(proyecto_id) if proyecto_id not in (None, "", "null", "None") else None
    except Exception:
        return jsonify({'mensaje': 'proyecto_id inválido'}), 400

    try:
        fase_proyecto_id = int(fase_proyecto_id) if fase_proyecto_id not in (None, "", "null", "None") else None
    except Exception:
        return jsonify({'mensaje': 'fase_proyecto_id inválido'}), 400

    if proyecto_id:
        if not Proyecto.query.get(proyecto_id):
            return jsonify({'mensaje': 'Proyecto no existe'}), 400

    if fase_proyecto_id:
        if not ProyectoFase.query.get(fase_proyecto_id):
            return jsonify({'mensaje': 'Fase de proyecto no existe'}), 400

    if proyecto_id and not fase_proyecto_id:
        p = Proyecto.query.get(proyecto_id)
        fase_proyecto_id = getattr(p, "fase_id", None) if p else None

    # ------------------------------------------------------------------
    # 12) CREAR UNO O VARIOS REGISTROS
    # ------------------------------------------------------------------
    try:
        nuevos = []

        for fecha_item in fechas_a_crear:
            for frag in fragmentos:
                frag_hora_inicio = frag["hora_inicio"]
                frag_hora_fin = frag["hora_fin"]
                frag_horas_adic = frag["horas_adicionales"]

                frag_tiempo = _calcular_tiempo_horas(frag_hora_inicio, frag_hora_fin)

                nuevo = Registro(
                    fecha=fecha_item,
                    cliente=cliente,
                    nro_caso_cliente=_norm_default(pick(data, 'nroCasoCliente', 'nro_caso_cliente'), '0'),
                    nro_caso_interno=_norm_default(pick(data, 'nroCasoInterno', 'nro_caso_interno'), '0'),
                    nro_caso_escalado=bd["nro_escalado"],

                    tarea_id=tarea_id,
                    ocupacion_id=ocupacion_id,

                    proyecto_id=proyecto_id,
                    fase_proyecto_id=fase_proyecto_id,

                    hora_inicio=frag_hora_inicio,
                    hora_fin=frag_hora_fin,
                    tiempo_invertido=frag_tiempo,
                    tiempo_facturable=tiempo_facturable,

                    actividad_malla=bd["actividad_malla"],
                    oncall=bd["oncall"],
                    desborde=bd["desborde"],

                    horas_adicionales=frag_horas_adic,
                    descripcion=_norm_default(pick(data, 'descripcion'), ''),

                    total_horas=frag_tiempo,
                    modulo=modulo_final,
                    horario_trabajo=horario_trabajo,
                    usuario_consultor=consultor.usuario,
                    equipo=equipo_final,
                )

                db.session.add(nuevo)
                nuevos.append(nuevo)

        db.session.commit()

        return jsonify({
            'mensaje': 'Registro guardado correctamente',
            'cantidad_registros': len(nuevos),
            'es_rango_vacaciones': bool(es_rango_vacaciones),
            'registros': [
                {
                    'id': r.id,
                    'fecha': r.fecha,
                    'horaInicio': r.hora_inicio,
                    'horaFin': r.hora_fin,
                    'horasAdicionales': r.horas_adicionales,
                    'totalHoras': r.total_horas,
                }
                for r in nuevos
            ]
        }), 201

    except IntegrityError as e:
        db.session.rollback()
        return jsonify({'mensaje': f'No se pudo guardar el registro (integridad): {e}'}), 400

    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'No se pudo guardar el registro: {e}'}), 500
    
# ============================================================
#  ✅ HELPERS 
# ============================================================

def _norm_user(u: str) -> str:
    return (u or "").strip().lower()
    

def _norm_role(r: str) -> str:
    return (r or "").strip().upper()

def _get_usuario_from_request() -> str:
    try:
        sesion, consultor = _get_consultor_from_token()
        if consultor:
            return (consultor.usuario or "").strip().lower()
    except Exception:
        pass

    return (request.headers.get("X-User-Usuario") or "").strip().lower()


def _get_rol_from_request() -> str:
    try:
        sesion, consultor = _get_consultor_from_token()
        if consultor:
            rol = (
                consultor.rol_obj.nombre
                if getattr(consultor, "rol_obj", None) and getattr(consultor.rol_obj, "nombre", None)
                else (getattr(consultor, "rol", "") or "")
            )
            return str(rol).strip().upper()
    except Exception:
        pass

    return (request.headers.get("X-User-Rol") or "").strip().upper()


def _is_admin_total(role: str) -> bool:
    r = _norm_role(role)
    # SOLO admin global
    return r in {"ADMIN", "SUPERADMIN", "ADMIN_TOTAL"}

def _is_admin_equipo(role: str) -> bool:
    return _norm_role(role) in {
        "ADMIN_FUNCIONAL",
        "ADMIN_BASIS",
        "ADMIN_IMPLEMENTACION",
        "ADMIN_GESTION_DE_PROYECTOS",
        "ADMIN_CONSULTORIA",
        "ADMIN_ARQUITECTURA",
    }

def scope_for(consultor_login, rol_req: str):
    """
    ADMIN                       -> ("ALL", None)
    ADMIN_GESTION_PREVENTA      -> ("ROLE_POOL", rol_id)
    ADMIN_*                     -> ("TEAM", equipo_id)
    CONSULTOR / otros           -> ("SELF", usuario_norm)
    """
    rol = (rol_req or (consultor_login.rol_obj.nombre if consultor_login.rol_obj else "") or "").strip().upper()
    usuario_norm = (consultor_login.usuario or "").strip().lower()

    if rol == "ADMIN":
        return "ALL", None

    if rol in ROLE_POOL_ROLES:
        role_id = _consultor_role_id(consultor_login)
        return "ROLE_POOL", role_id

    if rol.startswith("ADMIN_"):
        return "TEAM", int(consultor_login.equipo_id) if consultor_login.equipo_id else 0

    return "SELF", usuario_norm



# ============================================================
# ✅ ENDPOINT MODIFICADO (COMPLETO)
# ============================================================
def _scope_for_graficos(consultor_login, rol_req: str):
    """
    Reglas:
      - ADMIN, ADMIN_GERENTES, ADMIN_GESTION_PREVENTA -> ALL
      - ADMIN_* (BASIS/FUNCIONAL/...)                 -> TEAM
      - resto                                         -> scope_for normal
    """
    rol = (rol_req or "").strip().upper()

    if rol in GRAFICOS_ALL_ROLES:
        return "ALL", None

    if rol.startswith("ADMIN_"):
        equipo_id = consultor_login.equipo_id if consultor_login else None
        if equipo_id:
            return "TEAM", int(equipo_id)
        return "SELF", None

    return scope_for(consultor_login, rol_req)

# ============================================================
#  ENDPOINT: SOLO PARA GRAFICOS
# ============================================================
def _safe_fecha_iso(v):
    if v is None:
        return None

    if hasattr(v, "isoformat"):
        try:
            return v.isoformat()
        except Exception:
            pass

    s = str(v).strip()
    return s[:10] if s else None

def _apply_project_filter_graficos(q, proyecto_id: int):
    proyecto = (
        Proyecto.query.options(joinedload(Proyecto.mapeos))
        .filter(Proyecto.id == proyecto_id)
        .first()
    )

    if not proyecto:
        return q.filter(text("1=0"))

    clauses = [Registro.proyecto_id == proyecto.id]

    codigo = (proyecto.codigo or "").strip().upper()

    campo_nro = func.upper(func.coalesce(Registro.nro_caso_cliente, ""))
    campo_desc = func.upper(func.coalesce(Registro.descripcion, ""))

    if codigo:
        clauses.append(campo_nro == codigo)
        clauses.append(campo_desc == codigo)
        clauses.append(campo_nro.like(f"%{codigo}%"))
        clauses.append(campo_desc.like(f"%{codigo}%"))

    for mp in (proyecto.mapeos or []):
        if not bool(mp.activo):
            continue

        valor = (mp.valor_origen or "").strip().upper()
        tipo = (mp.tipo_match or "EXACT").strip().upper()

        if not valor:
            continue

        if tipo == "EXACT":
            clauses.append(campo_nro == valor)
            clauses.append(campo_desc == valor)

        elif tipo == "CONTAINS":
            clauses.append(campo_nro.like(f"%{valor}%"))
            clauses.append(campo_desc.like(f"%{valor}%"))

        elif tipo == "REGEX":
            try:
                clauses.append(campo_nro.op("REGEXP")(valor))
                clauses.append(campo_desc.op("REGEXP")(valor))
            except Exception:
                pass

    return q.filter(or_(*clauses))

def _apply_project_filter_shared(q, proyecto_id: int):
    proyecto = (
        Proyecto.query.options(joinedload(Proyecto.mapeos))
        .filter(Proyecto.id == proyecto_id)
        .first()
    )

    if not proyecto:
        return q.filter(text("1=0"))

    clauses = [Registro.proyecto_id == proyecto.id]

    codigo = (proyecto.codigo or "").strip().upper()
    campo_nro = func.upper(func.coalesce(Registro.nro_caso_cliente, ""))
    campo_desc = func.upper(func.coalesce(Registro.descripcion, ""))

    if codigo:
        clauses.append(campo_nro == codigo)
        clauses.append(campo_desc == codigo)
        clauses.append(campo_nro.like(f"%{codigo}%"))
        clauses.append(campo_desc.like(f"%{codigo}%"))

    for mp in (proyecto.mapeos or []):
        if not bool(mp.activo):
            continue

        valor = (mp.valor_origen or "").strip().upper()
        tipo = (mp.tipo_match or "EXACT").strip().upper()

        if not valor:
            continue

        if tipo == "EXACT":
            clauses.append(campo_nro == valor)
            clauses.append(campo_desc == valor)
        elif tipo == "CONTAINS":
            clauses.append(campo_nro.like(f"%{valor}%"))
            clauses.append(campo_desc.like(f"%{valor}%"))
        elif tipo == "REGEX":
            try:
                clauses.append(campo_nro.op("REGEXP")(valor))
                clauses.append(campo_desc.op("REGEXP")(valor))
            except Exception:
                pass

    return q.filter(or_(*clauses))

@bp.route('/registros/graficos', methods=['GET'])
@permission_required("GRAFICOS_VER")
def obtener_registros_graficos():
    try:
        usuario = _get_usuario_from_request()
        rol_req = _get_rol_from_request()

        if not usuario:
            return jsonify({'error': 'Usuario no enviado'}), 400

        usuario_norm = (usuario or "").strip().lower()

        consultor_login = (
            Consultor.query.options(
                joinedload(Consultor.rol_obj),
                joinedload(Consultor.equipo_obj),
            )
            .filter(func.lower(Consultor.usuario) == usuario_norm)
            .first()
        )

        if not consultor_login:
            return jsonify({'error': 'Consultor no encontrado'}), 404

        scope, val = _scope_for_graficos(consultor_login, rol_req)

        C = aliased(Consultor)
        E = aliased(Equipo)

        q = (
            Registro.query
            .options(
                joinedload(Registro.consultor).joinedload(Consultor.equipo_obj),
                joinedload(Registro.tarea),
                joinedload(Registro.ocupacion),
                joinedload(Registro.proyecto),
                joinedload(Registro.fase_proyecto),
            )
            .outerjoin(C, func.lower(Registro.usuario_consultor) == func.lower(C.usuario))
            .outerjoin(E, C.equipo_id == E.id)
        )

        # ----------------------------------------------------------
        # Scope
        # ----------------------------------------------------------
        if scope == "SELF":
            q = q.filter(func.lower(Registro.usuario_consultor) == usuario_norm)

        elif scope == "TEAM":
            if not int(val or 0):
                return jsonify({'error': 'Consultor sin equipo asignado'}), 403
            q = q.filter(C.equipo_id == int(val))

        elif scope == "ALL":
            pass

        # ----------------------------------------------------------
        # Filtro opcional por equipo
        # ----------------------------------------------------------
        equipo_filter = (request.args.get("equipo") or "").strip().upper()

        if equipo_filter:
            if scope in ("TEAM", "SELF"):
                eq_login = ""
                if consultor_login.equipo_obj:
                    eq_login = (consultor_login.equipo_obj.nombre or "").strip().upper()

                if equipo_filter != eq_login:
                    return jsonify({'error': 'No autorizado para consultar otro equipo'}), 403

            q = q.filter(func.upper(E.nombre) == equipo_filter)

        # ----------------------------------------------------------
        # Filtros opcionales backend
        # ----------------------------------------------------------
        filtro_mes = (request.args.get("mes") or "").strip()
        filtro_desde = (request.args.get("desde") or "").strip()
        filtro_hasta = (request.args.get("hasta") or "").strip()
        filtro_modulo = (request.args.get("modulo") or "").strip()
        filtro_cliente = (request.args.get("cliente") or "").strip()
        filtro_consultor = (request.args.get("consultor") or "").strip()
        filtro_proyecto_id = (request.args.get("proyecto_id") or "").strip()

        # ----------------------------------------------------------
        # Filtro por mes o rango
        # ----------------------------------------------------------
        if filtro_mes:
            partes = filtro_mes.split("-")
            if len(partes) != 2:
                return jsonify({"error": "mes inválido, usa YYYY-MM"}), 400

            try:
                y = int(partes[0])
                m = int(partes[1])
            except ValueError:
                return jsonify({"error": "mes inválido, usa YYYY-MM"}), 400

            if m < 1 or m > 12:
                return jsonify({"error": "mes inválido, usa YYYY-MM"}), 400

            prefijo_mes = f"{y:04d}-{m:02d}"

            q = q.filter(
                func.substr(func.cast(Registro.fecha, db.String), 1, 7) == prefijo_mes
            )
        else:
            if filtro_desde:
                q = q.filter(func.cast(Registro.fecha, db.String) >= filtro_desde)
            if filtro_hasta:
                q = q.filter(func.cast(Registro.fecha, db.String) <= filtro_hasta)

        if filtro_modulo:
            q = q.filter(func.upper(Registro.modulo) == filtro_modulo.upper())

        if filtro_cliente:
            q = q.filter(Registro.cliente.ilike(f"%{filtro_cliente}%"))

        if filtro_consultor:
            q = q.filter(C.nombre.ilike(f"%{filtro_consultor}%"))

        if filtro_proyecto_id:
            try:
                q = _apply_project_filter_graficos(q, int(filtro_proyecto_id))
            except Exception:
                return jsonify({"error": "proyecto_id inválido"}), 400

        # ----------------------------------------------------------
        # Orden
        # ----------------------------------------------------------
        q = q.order_by(Registro.fecha.desc(), Registro.id.desc())

        # ----------------------------------------------------------
        # IMPORTANTE:
        # No limitar cuando se consulta un mes o rango, porque
        # las gráficas deben sumar TODO el periodo visible.
        # El limit solo se usa como protección cuando NO hay
        # filtro temporal.
        # ----------------------------------------------------------
        tiene_filtro_temporal = bool(filtro_mes or filtro_desde or filtro_hasta)

        if tiene_filtro_temporal:
            registros = q.all()
        else:
            max_rows = request.args.get("max_rows", type=int) or 5000
            max_rows = min(max(max_rows, 1), 10000)
            registros = q.limit(max_rows).all()

        # ----------------------------------------------------------
        # Serialización
        # ----------------------------------------------------------
        data = []

        for r in registros:
            tarea = getattr(r, "tarea", None)
            ocup = getattr(r, "ocupacion", None)

            if tarea and getattr(tarea, "codigo", None) and getattr(tarea, "nombre", None):
                tipo_tarea_str = f"{tarea.codigo} - {tarea.nombre}"
            else:
                tipo_tarea_str = (getattr(r, "tipo_tarea", "") or "").strip() or None

            equipo_nombre = None
            if r.consultor and getattr(r.consultor, "equipo_obj", None):
                equipo_nombre = (r.consultor.equipo_obj.nombre or "").strip().upper()

            equipo_raw = getattr(r, "equipo", None)

            proyecto = getattr(r, "proyecto", None)
            fase_proyecto = getattr(r, "fase_proyecto", None)

            data.append({
                "id": r.id,
                "fecha": _safe_fecha_iso(r.fecha),
                "modulo": r.modulo,
                "cliente": r.cliente,
                "equipo": equipo_nombre or (str(equipo_raw).strip().upper() if equipo_raw else "SIN EQUIPO"),
                "nroCasoCliente": r.nro_caso_cliente,
                "nroCasoInterno": r.nro_caso_interno,
                "nroCasoEscaladoSap": r.nro_caso_escalado,

                "ocupacion_id": r.ocupacion_id,
                "ocupacion_codigo": ocup.codigo if ocup else None,
                "ocupacion_nombre": ocup.nombre if ocup else None,

                "tarea_id": r.tarea_id,
                "tipoTarea": tipo_tarea_str,
                "tarea": {
                    "id": tarea.id,
                    "codigo": getattr(tarea, "codigo", None),
                    "nombre": getattr(tarea, "nombre", None),
                } if tarea else None,

                "consultor": r.consultor.nombre if r.consultor else None,
                "usuario_consultor": (r.usuario_consultor or "").strip().lower(),
                "horaInicio": r.hora_inicio,
                "horaFin": r.hora_fin,
                "tiempoInvertido": r.tiempo_invertido,
                "tiempoFacturable": r.tiempo_facturable,
                "horasAdicionales": _calcular_horas_adicionales_por_horario(
                    r.hora_inicio,
                    r.hora_fin,
                    getattr(r, "horario_trabajo", None),
                    equipo_nombre or getattr(r, "equipo", None),
                    r.horas_adicionales,
                ),
                "horarioTrabajo": _normalizar_horario_trabajo_por_equipo(
                    getattr(r, "horario_trabajo", None),
                    equipo_nombre or getattr(r, "equipo", None),
                ),
                "horario_trabajo": _normalizar_horario_trabajo_por_equipo(
                    getattr(r, "horario_trabajo", None),
                    equipo_nombre or getattr(r, "equipo", None),
                ),
                "descripcion": r.descripcion,
                "totalHoras": r.total_horas,
                "bloqueado": bool(r.bloqueado),
                "oncall": r.oncall,
                "desborde": r.desborde,
                "actividadMalla": r.actividad_malla,

                "proyecto_id": r.proyecto_id,
                "fase_proyecto_id": r.fase_proyecto_id,
                "proyecto": {
                    "id": proyecto.id,
                    "codigo": proyecto.codigo,
                    "nombre": proyecto.nombre,
                    "activo": bool(getattr(proyecto, "activo", True)),
                } if proyecto else None,
                "fase_proyecto": {
                    "id": fase_proyecto.id,
                    "nombre": fase_proyecto.nombre,
                } if fase_proyecto else None,
                "proyecto_codigo": proyecto.codigo if proyecto else None,
                "proyecto_nombre": proyecto.nombre if proyecto else None,
                "proyecto_fase": fase_proyecto.nombre if fase_proyecto else None,
            })

        return jsonify(data), 200

    except Exception as e:
        err = traceback.format_exc()
        app.logger.error(f"❌ Error en /registros/graficos: {e}\n{err}")
        return jsonify({
            "error": "Error interno del servidor",
            "detalle": str(e)
        }), 500
    
USUARIOS_PUEDE_SEMANAS_ANTERIORES = {
}

@bp.route('/registros', methods=['GET'])
def obtener_registros():
    try:
        usuario = _get_usuario_from_request()
        rol_req = _get_rol_from_request()

        if not usuario:
            return jsonify({'error': 'Usuario no enviado'}), 400

        usuario_norm = (usuario or "").strip().lower()

        consultor_login = (
            Consultor.query.options(
                joinedload(Consultor.rol_obj),
                joinedload(Consultor.equipo_obj),
            )
            .filter(func.lower(Consultor.usuario) == usuario_norm)
            .first()
        )
        if not consultor_login:
            return jsonify({'error': 'Consultor no encontrado'}), 404

        scope, val = scope_for(consultor_login, rol_req)

        C = aliased(Consultor)
        E = aliased(Equipo)

        q = (
            Registro.query
            .options(
                joinedload(Registro.consultor).joinedload(Consultor.equipo_obj),
                joinedload(Registro.tarea),
                joinedload(Registro.ocupacion),
                joinedload(Registro.proyecto),
                joinedload(Registro.fase_proyecto),
            )
            .outerjoin(C, func.lower(Registro.usuario_consultor) == func.lower(C.usuario))
            .outerjoin(E, C.equipo_id == E.id)
        )

        # -----------------------------
        # Scope
        # -----------------------------
        if scope == "SELF":
            q = q.filter(func.lower(Registro.usuario_consultor) == usuario_norm)

        elif scope == "TEAM":
            if not int(val or 0):
                return jsonify({'error': 'Consultor sin equipo asignado'}), 403
            q = q.filter(C.equipo_id == int(val))

        elif scope == "ROLE_POOL":
            if not int(val or 0):
                return jsonify({'error': 'Consultor sin rol asignado'}), 403
            q = q.filter(C.rol_id == int(val))

        # -----------------------------
        # Filtros backend
        # -----------------------------
        filtro_id = (request.args.get("id") or "").strip()
        filtro_fecha = (request.args.get("fecha") or "").strip()
        filtro_equipo = (request.args.get("equipo") or "").strip().upper()
        filtro_mes = (request.args.get("mes") or "").strip()
        filtro_anio = (request.args.get("anio") or "").strip()
        filtro_nro_caso = (request.args.get("nroCasoCliente") or "").strip()

        filtro_clientes = [v.strip() for v in _get_list_arg("cliente") if str(v).strip()]
        filtro_consultores = [v.strip() for v in _get_list_arg("consultor") if str(v).strip()]
        filtro_horas_adic = [str(v).strip().upper() for v in _get_list_arg("horasAdicionales") if str(v).strip()]
        filtro_tarea_ids = []
        filtro_ocupacion_ids = []

        for v in _get_list_arg("tarea_id"):
            try:
                filtro_tarea_ids.append(int(v))
            except Exception:
                pass

        for v in _get_list_arg("ocupacion_id"):
            try:
                filtro_ocupacion_ids.append(int(v))
            except Exception:
                pass

        filtro_clientes = list(dict.fromkeys(filtro_clientes))
        filtro_consultores = list(dict.fromkeys(filtro_consultores))
        filtro_horas_adic = list(dict.fromkeys(filtro_horas_adic))
        filtro_tarea_ids = list(dict.fromkeys(filtro_tarea_ids))
        filtro_ocupacion_ids = list(dict.fromkeys(filtro_ocupacion_ids))

        if filtro_id and filtro_id.isdigit():
            q = q.filter(Registro.id == int(filtro_id))

        if filtro_fecha:
            q = q.filter(Registro.fecha == filtro_fecha)

        if filtro_clientes:
            q = q.filter(Registro.cliente.in_(filtro_clientes))

        if filtro_consultores:
            q = q.filter(C.nombre.in_(filtro_consultores))

        if filtro_equipo:
            if scope == "TEAM":
                eq_login = (consultor_login.equipo_obj.nombre or "").strip().upper() if consultor_login.equipo_obj else ""
                if filtro_equipo != eq_login:
                    return jsonify({'error': 'No autorizado para consultar otro equipo'}), 403
            q = q.filter(func.upper(E.nombre) == filtro_equipo)

        if filtro_mes:
            try:
                q = q.filter(extract("month", cast(Registro.fecha, db.Date)) == int(filtro_mes))
            except Exception:
                pass

        if filtro_anio:
            try:
                q = q.filter(extract("year", cast(Registro.fecha, db.Date)) == int(filtro_anio))
            except Exception:
                pass

        if filtro_nro_caso:
            q = q.filter(Registro.nro_caso_cliente.ilike(f"%{filtro_nro_caso}%"))

        if filtro_horas_adic:
            q = q.filter(func.upper(Registro.horas_adicionales).in_(filtro_horas_adic))

        if filtro_tarea_ids:
            q = q.filter(Registro.tarea_id.in_(filtro_tarea_ids))

        if filtro_ocupacion_ids:
            q = q.filter(Registro.ocupacion_id.in_(filtro_ocupacion_ids))

        # -----------------------------
        # Paginación
        # -----------------------------
        page = max(int(request.args.get("page", 1)), 1)
        per_page = min(max(int(request.args.get("per_page", 50)), 1), 200)

        total = q.count()

        registros = (
            q.order_by(Registro.fecha.desc(), Registro.id.desc())
             .offset((page - 1) * per_page)
             .limit(per_page)
             .all()
        )

        data = []
        for r in registros:
            tarea = r.tarea
            ocup = r.ocupacion

            if tarea and getattr(tarea, "codigo", None) and getattr(tarea, "nombre", None):
                tipo_tarea_str = f"{tarea.codigo} - {tarea.nombre}"
            else:
                tipo_tarea_str = (r.tipo_tarea or "").strip() or None

            equipo_nombre = None
            if r.consultor and r.consultor.equipo_obj:
                equipo_nombre = (r.consultor.equipo_obj.nombre or "").strip().upper()

            proyecto = getattr(r, "proyecto", None)
            fase_proyecto = getattr(r, "fase_proyecto", None)

            data.append({
                "id": r.id,
                "fecha": r.fecha,
                "modulo": r.modulo,
                "cliente": r.cliente,
                "equipo": equipo_nombre or (r.equipo or "").strip().upper() or "SIN EQUIPO",
                "nroCasoCliente": r.nro_caso_cliente,
                "nroCasoInterno": r.nro_caso_interno,
                "nroCasoEscaladoSap": r.nro_caso_escalado,

                "ocupacion_id": r.ocupacion_id,
                "ocupacion_codigo": ocup.codigo if ocup else None,
                "ocupacion_nombre": ocup.nombre if ocup else None,

                "tarea_id": r.tarea_id,
                "tipoTarea": tipo_tarea_str,
                "tarea": {
                    "id": tarea.id,
                    "codigo": getattr(tarea, "codigo", None),
                    "nombre": getattr(tarea, "nombre", None),
                } if tarea else None,

                "consultor": r.consultor.nombre if r.consultor else None,
                "usuario_consultor": (r.usuario_consultor or "").strip().lower(),

                "horaInicio": r.hora_inicio,
                "horaFin": r.hora_fin,
                "tiempoInvertido": r.tiempo_invertido,
                "tiempoFacturable": r.tiempo_facturable,
                "horasAdicionales": _calcular_horas_adicionales_por_horario(
                    r.hora_inicio,
                    r.hora_fin,
                    getattr(r, "horario_trabajo", None),
                    equipo_nombre or getattr(r, "equipo", None),
                    r.horas_adicionales,
                ),
                "horarioTrabajo": _normalizar_horario_trabajo_por_equipo(
                    getattr(r, "horario_trabajo", None),
                    equipo_nombre or getattr(r, "equipo", None),
                ),
                "horario_trabajo": _normalizar_horario_trabajo_por_equipo(
                    getattr(r, "horario_trabajo", None),
                    equipo_nombre or getattr(r, "equipo", None),
                ),
                "descripcion": r.descripcion,
                "totalHoras": r.total_horas,

                "bloqueado": bool(r.bloqueado),
                "oncall": r.oncall,
                "desborde": r.desborde,
                "actividadMalla": r.actividad_malla,

                "proyecto_id": r.proyecto_id,
                "fase_proyecto_id": r.fase_proyecto_id,

                "proyecto": {
                    "id": proyecto.id,
                    "codigo": proyecto.codigo,
                    "nombre": proyecto.nombre,
                    "activo": bool(getattr(proyecto, "activo", True)),
                } if proyecto else None,

                "fase_proyecto": {
                    "id": fase_proyecto.id,
                    "nombre": fase_proyecto.nombre,
                } if fase_proyecto else None,

                "proyecto_codigo": proyecto.codigo if proyecto else None,
                "proyecto_nombre": proyecto.nombre if proyecto else None,
                "proyecto_fase": fase_proyecto.nombre if fase_proyecto else None,
            })

        return jsonify({
            "data": data,
            "total": total,
            "page": page,
            "per_page": per_page,
            "total_pages": math.ceil(total / per_page) if per_page else 1
        }), 200

    except Exception as e:
        app.logger.exception("❌ Error en obtener_registros (/registros)")
        return jsonify({'error': str(e)}), 500

@bp.route("/resumen-horas", methods=["GET"])
def resumen_horas():
    try:
        # ----------------------------------------------------------
        # 1) Usuario / rol desde request
        # ----------------------------------------------------------
        usuario = _get_usuario_from_request()
        rol_req = _get_rol_from_request()

        if not usuario:
            return jsonify({"error": "Usuario no enviado"}), 400

        usuario_norm = (usuario or "").strip().lower()

        # ----------------------------------------------------------
        # 2) Consultor login
        # ----------------------------------------------------------
        consultor_login = (
            Consultor.query
            .options(joinedload(Consultor.rol_obj), joinedload(Consultor.equipo_obj))
            .filter(func.lower(Consultor.usuario) == usuario_norm)
            .first()
        )
        if not consultor_login:
            return jsonify({"error": "Consultor no encontrado"}), 404

        # ----------------------------------------------------------
        # 3) Scope (SELF / TEAM / ALL)
        # ----------------------------------------------------------
        scope, val = scope_for(consultor_login, rol_req)

        # ----------------------------------------------------------
        # 4) Filtros opcionales por fecha (ISO recomendado: YYYY-MM-DD)
        #    soporta: ?desde=2026-01-01&hasta=2026-01-31
        # ----------------------------------------------------------
        desde = (request.args.get("desde") or "").strip()
        hasta = (request.args.get("hasta") or "").strip()

        # ----------------------------------------------------------
        # 5) Query base (contadores por equipo)
        #    OJO: tu Registro NO tiene equipo_id, por eso usamos Consultor->Equipo
        # ----------------------------------------------------------
        q = (
            db.session.query(
                Equipo.nombre.label("equipo_nombre"),
                func.count(Registro.id).label("total_registros"),
                func.coalesce(func.sum(Registro.total_horas), 0).label("total_horas"),
            )
            .select_from(Registro)
            .join(Consultor, func.lower(Registro.usuario_consultor) == func.lower(Consultor.usuario))
            .outerjoin(Equipo, Consultor.equipo_id == Equipo.id)
        )

        # ----------------------------------------------------------
        # 6) Aplicar scope
        # ----------------------------------------------------------
        if scope == "SELF":
            q = q.filter(func.lower(Registro.usuario_consultor) == usuario_norm)
        elif scope == "TEAM":
            q = q.filter(Consultor.equipo_id == int(val))
        elif scope == "ROLE_POOL":
            q = q.filter(Consultor.rol_id == int(val))

        # ----------------------------------------------------------
        # 7) Filtro por fecha (si viene)
        #    (si tu fecha es ISO YYYY-MM-DD, BETWEEN funciona perfecto)
        # ----------------------------------------------------------
        if desde and hasta:
            q = q.filter(Registro.fecha.between(desde, hasta))
        elif desde:
            q = q.filter(Registro.fecha >= desde)
        elif hasta:
            q = q.filter(Registro.fecha <= hasta)

        # ----------------------------------------------------------
        # 8) Agrupar
        # ----------------------------------------------------------
        q = q.group_by(Equipo.nombre)

        rows = q.all()

        # ----------------------------------------------------------
        # 9) Total general (para "Todos")
        # ----------------------------------------------------------
        qt = (
            db.session.query(
                func.count(Registro.id).label("total_registros"),
                func.coalesce(func.sum(Registro.total_horas), 0).label("total_horas"),
            )
            .select_from(Registro)
            .join(Consultor, func.lower(Registro.usuario_consultor) == func.lower(Consultor.usuario))
        )

        if scope == "SELF":
            qt = qt.filter(func.lower(Registro.usuario_consultor) == usuario_norm)
        elif scope == "TEAM":
            qt = qt.filter(Consultor.equipo_id == int(val))
        elif scope == "ROLE_POOL":
            qt = qt.filter(Consultor.rol_id == int(val))

        if desde and hasta:
            qt = qt.filter(Registro.fecha.between(desde, hasta))
        elif desde:
            qt = qt.filter(Registro.fecha >= desde)
        elif hasta:
            qt = qt.filter(Registro.fecha <= hasta)

        total_row = qt.first()

        total_registros = int(total_row.total_registros or 0)
        total_horas = float(total_row.total_horas or 0)

        # ----------------------------------------------------------
        # 10) Formato para el FRONT (chips por equipo)
        # ----------------------------------------------------------
        equipos = []
        for r in rows:
            nombre = (r.equipo_nombre or "SIN EQUIPO").strip().upper()
            equipos.append({
                "equipo": nombre,
                "total": int(r.total_registros or 0),
                "totalHoras": float(r.total_horas or 0),
            })

        # ordenar por total desc
        equipos.sort(key=lambda x: x["total"], reverse=True)

        return jsonify({
            "total": total_registros,
            "totalHoras": total_horas,
            "equipos": equipos,
        }), 200

    except Exception as e:
        app.logger.exception("❌ Error en /resumen-horas")
        return jsonify({"error": str(e)}), 500



@bp.route('/eliminar-registro/<int:id>', methods=['DELETE'])
@permission_required("REGISTROS_ELIMINAR")
def eliminar_registro(id):
    data = request.json or {}
    rol = (data.get('rol') or '').strip().upper()
    nombre = data.get('nombre')

    registro = Registro.query.get(id)
    if not registro:
        return jsonify({'mensaje': 'Registro no encontrado'}), 404

    if rol != 'ADMIN' and (not registro.consultor or registro.consultor.nombre != (nombre or "")):
        return jsonify({'mensaje': 'No autorizado'}), 403

    db.session.delete(registro)
    db.session.commit()
    return jsonify({'mensaje': 'Registro eliminado'}), 200

def _week_bounds_bogota(ref_dt=None):
    tz = ZoneInfo("America/Bogota")
    now = ref_dt.astimezone(tz) if ref_dt else datetime.now(tz)

    today = now.date()
    # lunes = 0, domingo = 6
    start = today - timedelta(days=today.weekday())
    end = start + timedelta(days=6)

    return start, end


def _parse_iso_date_safe(value):
    s = str(value or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def _fecha_en_semana_vigente(fecha_value):
    fecha = _parse_iso_date_safe(fecha_value)
    if not fecha:
        return False

    start, end = _week_bounds_bogota()
    return start <= fecha <= end


def _fecha_no_futura(fecha_value):
    fecha = _parse_iso_date_safe(fecha_value)
    if not fecha:
        return False

    tz = ZoneInfo("America/Bogota")
    hoy = datetime.now(tz).date()
    return fecha <= hoy


@bp.route('/editar-registro/<int:id>', methods=['PUT'])
@permission_required("REGISTROS_EDITAR")
def editar_registro(id):
    data = request.get_json(silent=True) or {}

    # ----------------------------------------------------------
    # 1) Usuario autenticado por header
    # ----------------------------------------------------------
    usuario_header = (request.headers.get("X-User-Usuario") or "").strip().lower()
    if not usuario_header:
        return jsonify({'mensaje': 'Usuario no enviado'}), 401

    consultor_login = Consultor.query.filter(
        func.lower(Consultor.usuario) == usuario_header
    ).first()
    if not consultor_login:
        return jsonify({'mensaje': 'Usuario no encontrado'}), 404

    rol_real = ""
    if getattr(consultor_login, "rol_obj", None) and getattr(consultor_login.rol_obj, "nombre", None):
        rol_real = consultor_login.rol_obj.nombre
    else:
        rol_real = getattr(consultor_login, "rol", "") or ""

    es_admin = _is_admin_request(rol_real, consultor_login)

    # ----------------------------------------------------------
    # 2) Registro + autorización
    # ----------------------------------------------------------
    registro = Registro.query.get(id)
    if not registro:
        return jsonify({'mensaje': 'Registro no encontrado'}), 404

    dueno = (registro.usuario_consultor or "").strip().lower()
    if not es_admin and dueno and dueno != usuario_header:
        return jsonify({'mensaje': 'No autorizado'}), 403

    try:
        # ----------------------------------------------------------
        # 3) Tomar valores finales antes de guardar
        # ----------------------------------------------------------
        nueva_fecha = pick(data, 'fecha', default=registro.fecha)
        nuevo_hora_inicio = pick(data, 'horaInicio', 'hora_inicio', default=registro.hora_inicio)
        nuevo_hora_fin = pick(data, 'horaFin', 'hora_fin', default=registro.hora_fin)

        tiempo_calculado = _calcular_tiempo_horas(nuevo_hora_inicio, nuevo_hora_fin)
        if tiempo_calculado <= 0:
            return jsonify({'mensaje': 'Hora fin debe ser mayor a hora inicio'}), 400

        equipo_para_horario = pick(data, 'equipo', default=registro.equipo)
        if not equipo_para_horario and getattr(consultor_login, "equipo_obj", None):
            equipo_para_horario = consultor_login.equipo_obj.nombre

        horario_trabajo = _normalizar_horario_trabajo_por_equipo(
            pick(data, 'horarioTrabajo', 'horario_trabajo', default=getattr(registro, "horario_trabajo", None)),
            equipo_para_horario,
        )

        horas_adicionales_calculadas = _calcular_horas_adicionales_por_horario(
            nuevo_hora_inicio,
            nuevo_hora_fin,
            horario_trabajo,
            equipo_para_horario,
            registro.horas_adicionales,
        )

        # ----------------------------------------------------------
        # 3.1) No permitir fechas futuras
        # ----------------------------------------------------------
        if not _fecha_no_futura(nueva_fecha):
            return jsonify({
                'mensaje': 'No puedes actualizar el registro con una fecha futura.'
            }), 403

        # ----------------------------------------------------------
        # 4) Validar traslape en backend
        # ----------------------------------------------------------
        conflicto = _buscar_registro_traslapado(
            fecha=nueva_fecha,
            usuario_consultor=registro.usuario_consultor,
            hora_inicio=nuevo_hora_inicio,
            hora_fin=nuevo_hora_fin,
            exclude_id=registro.id
        )
        if conflicto:
            return jsonify({
                'mensaje': f'Ya existe un registro que se cruza con este rango: {conflicto.hora_inicio} - {conflicto.hora_fin} (ID: {conflicto.id})'
            }), 409

        # ----------------------------------------------------------
        # 5) Campos básicos
        # ----------------------------------------------------------
        registro.fecha = nueva_fecha
        registro.cliente = pick(data, 'cliente', default=registro.cliente)

        registro.nro_caso_cliente = pick(
            data,
            'nroCasoCliente',
            'nro_caso_cliente',
            default=registro.nro_caso_cliente
        )
        registro.nro_caso_interno = pick(
            data,
            'nroCasoInterno',
            'nro_caso_interno',
            default=registro.nro_caso_interno
        )

        nro_escalado = pick(data, 'nroCasoEscaladoSap', 'nro_caso_escalado')
        if nro_escalado is not None:
            registro.nro_caso_escalado = nro_escalado

        # ----------------------------------------------------------
        # 6) Tarea
        # ----------------------------------------------------------
        tarea_id = pick(data, "tarea_id")
        tipoTareaTexto = pick(data, "tipoTarea", "tipo_tarea")

        if tarea_id not in (None, "", "null", "None"):
            try:
                tarea_id_int = int(tarea_id)
            except Exception:
                return jsonify({'mensaje': 'Tarea inválida'}), 400

            tarea_obj = Tarea.query.get(tarea_id_int)
            if not tarea_obj:
                return jsonify({'mensaje': 'Tarea inválida'}), 400

            registro.tarea_id = tarea_obj.id

            if not tipoTareaTexto:
                registro.tipo_tarea = f"{tarea_obj.codigo} - {tarea_obj.nombre}"

        if tipoTareaTexto:
            registro.tipo_tarea = str(tipoTareaTexto).strip()

        # ----------------------------------------------------------
        # 7) Ocupación
        # ----------------------------------------------------------
        ocupacion_id = pick(data, "ocupacion_id")
        if ocupacion_id not in (None, "", "null", "None"):
            try:
                ocupacion_id_int = int(ocupacion_id)
            except Exception:
                return jsonify({'mensaje': 'Ocupación inválida'}), 400

            ocup_obj = Ocupacion.query.get(ocupacion_id_int)
            if not ocup_obj:
                return jsonify({'mensaje': 'Ocupación inválida'}), 400

            registro.ocupacion_id = ocupacion_id_int

        if (ocupacion_id in (None, "", "null", "None")) and registro.tarea_id:
            tarea_db = Tarea.query.options(db.joinedload(Tarea.ocupaciones)).get(registro.tarea_id)
            if tarea_db and getattr(tarea_db, "ocupaciones", None) and tarea_db.ocupaciones:
                registro.ocupacion_id = tarea_db.ocupaciones[0].id

        # ----------------------------------------------------------
        # 7.1) Validación cliente restringido por ocupación
        # ----------------------------------------------------------
        cliente_validar = pick(data, 'cliente', default=registro.cliente)
        cliente_upper = str(cliente_validar or "").strip().upper()

        occ_obj = Ocupacion.query.get(registro.ocupacion_id) if registro.ocupacion_id else None
        occ_codigo = str(getattr(occ_obj, "codigo", "") or "").strip()

        # Ocupaciones que NO pueden usar HITSS/CLARO
        if occ_codigo in {"01", "02"} and cliente_upper == "HITSS/CLARO":
            return jsonify({
                'mensaje': 'Las ocupaciones 01 y 02 no pueden registrarse para el cliente HITSS/CLARO'
            }), 400

        # Ocupación 03 SOLO puede usar HITSS/CLARO
        if occ_codigo == "03" and cliente_upper != "HITSS/CLARO":
            return jsonify({
                "mensaje": "La ocupación 03 solo puede registrarse para el cliente HITSS/CLARO"
            }), 400

        # ----------------------------------------------------------
        # 8) Fechas/horas y valores numéricos
        # ----------------------------------------------------------
        registro.hora_inicio = nuevo_hora_inicio
        registro.hora_fin = nuevo_hora_fin

        registro.tiempo_invertido = pick(
            data,
            'tiempoInvertido',
            'tiempo_invertido',
            default=tiempo_calculado
        )
        registro.tiempo_facturable = pick(
            data,
            'tiempoFacturable',
            'tiempo_facturable',
            default=registro.tiempo_facturable
        )
        registro.horas_adicionales = horas_adicionales_calculadas
        registro.horario_trabajo = horario_trabajo
        registro.descripcion = pick(data, 'descripcion', default=registro.descripcion)
        registro.total_horas = pick(
            data,
            'totalHoras',
            'total_horas',
            default=tiempo_calculado
        )

        # ----------------------------------------------------------
        # 9) Módulo / Equipo
        # ----------------------------------------------------------
        modulo_in = pick(data, 'modulo')
        if modulo_in is not None:
            registro.modulo = modulo_in

        equipo_in = pick(data, 'equipo')
        if equipo_in is not None:
            registro.equipo = str(equipo_in).strip().upper() if isinstance(equipo_in, str) else equipo_in

        # ----------------------------------------------------------
        # 10) Proyectos
        # ----------------------------------------------------------
        proyecto_id = pick(data, "proyecto_id")
        fase_proyecto_id = pick(data, "fase_proyecto_id", "faseProyectoId")

        if proyecto_id in ("", "null", "None"):
            proyecto_id = None
        if fase_proyecto_id in ("", "null", "None"):
            fase_proyecto_id = None

        if proyecto_id is not None:
            try:
                proyecto_id = int(proyecto_id) if proyecto_id is not None else None
            except Exception:
                return jsonify({'mensaje': 'proyecto_id inválido'}), 400

            if proyecto_id and not Proyecto.query.get(proyecto_id):
                return jsonify({'mensaje': 'Proyecto no existe'}), 400

            registro.proyecto_id = proyecto_id

        if fase_proyecto_id is not None:
            try:
                fase_proyecto_id = int(fase_proyecto_id) if fase_proyecto_id is not None else None
            except Exception:
                return jsonify({'mensaje': 'fase_proyecto_id inválido'}), 400

            if fase_proyecto_id and not ProyectoFase.query.get(fase_proyecto_id):
                return jsonify({'mensaje': 'Fase de proyecto no existe'}), 400

            registro.fase_proyecto_id = fase_proyecto_id

        if registro.proyecto_id and not registro.fase_proyecto_id:
            p = Proyecto.query.get(registro.proyecto_id)
            if p and getattr(p, "fase_id", None):
                registro.fase_proyecto_id = p.fase_id

        # ----------------------------------------------------------
        # 11) Campos BASIS
        # ----------------------------------------------------------
        bd = _basis_defaults_from_payload(data)

        if 'actividadMalla' in data or 'actividad_malla' in data:
            registro.actividad_malla = bd["actividad_malla"]
        if 'oncall' in data:
            registro.oncall = bd["oncall"]
        if 'desborde' in data:
            registro.desborde = bd["desborde"]

        if not registro.actividad_malla:
            registro.actividad_malla = "N/APLICA"
        if not registro.oncall:
            registro.oncall = "N/A"
        if not registro.desborde:
            registro.desborde = "N/A"

        db.session.commit()
        return jsonify({
            'mensaje': 'Registro actualizado',
            'registro': {
                'id': registro.id,
                'fecha': registro.fecha,
                'horaInicio': registro.hora_inicio,
                'horaFin': registro.hora_fin
            }
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'mensaje': f'No se pudo actualizar el registro: {e}'}), 500

@bp.route('/toggle-bloqueado/<int:id>', methods=['PUT'])
def toggle_bloqueado(id):
    data = request.json or {}
    if (data.get('rol') or '').strip().upper() != 'ADMIN':
        return jsonify({'mensaje': 'No autorizado'}), 403

    registro = Registro.query.get(id)
    if not registro:
        return jsonify({'mensaje': 'Registro no encontrado'}), 404

    registro.bloqueado = not registro.bloqueado
    db.session.commit()
    return jsonify({'bloqueado': registro.bloqueado}), 200

# ===============================
# BaseRegistro (carga masiva / listado)
# ===============================

def to_float(valor):
    try:
        if isinstance(valor, str) and "," in valor and "." not in valor:
            valor = valor.replace(",", ".")
        return float(valor)
    except (TypeError, ValueError):
        return 0.0

def parse_horas_adicionales(valor):
    if valor is None or valor == "":
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip().lower()
    if s in ("si", "sí", "s", "yes", "y", "true", "1"): return 1.0
    if s in ("no", "n", "false", "0"): return 0.0
    try:
        if "," in s and "." not in s:
            s = s.replace(",", ".")
        return float(s)
    except ValueError:
        return 0.0

def inferir_modulo_por_consultor(nombre_consultor):
    if not nombre_consultor:
        return None
    c = Consultor.query.filter_by(nombre=str(nombre_consultor).strip()).first()
    if c and c.modulo:
        return c.modulo.nombre
    return None

def partir_fecha_mas_campos(fecha_val, dia_val=None, mes_val=None, anio_val=None):
    try:
        d = int(dia_val) if dia_val not in (None, "") else 0
        m = int(mes_val) if mes_val not in (None, "") else 0
        y = int(float(anio_val)) if anio_val not in (None, "") else 0
        if 1 <= d <= 31 and 1 <= m <= 12 and 1900 <= y <= 2100:
            return d, m, y
    except Exception:
        pass
    iso = excel_date_to_iso(fecha_val)
    try:
        dt = datetime.strptime(iso, "%Y-%m-%d")
        return dt.day, dt.month, dt.year
    except Exception:
        return 0, 0, 0

@bp.route('/cargar-registros-excel', methods=['POST'])
@permission_required("BASE_REGISTRO_IMPORTAR")
def cargar_registros_excel():
    data = request.get_json() or {}
    registros = data.get('registros', [])
    if not registros:
        return jsonify({'mensaje': 'No se recibieron registros'}), 400

    replace_all = bool(data.get('replace_all')) or (request.args.get('replace') in ('1', 'true', 'yes'))
    if replace_all:
        try:
            db.session.execute(text("TRUNCATE TABLE base_registro"))
            db.session.commit()
            app.logger.info("base_registro: TRUNCATE OK")
        except Exception:
            db.session.rollback()
            db.session.execute(text("DELETE FROM base_registro"))
            db.session.execute(text("ALTER TABLE base_registro AUTO_INCREMENT = 1"))
            db.session.commit()
            app.logger.info("base_registro: DELETE + reset AUTO_INCREMENT OK")

    BATCH_SIZE = 1000
    DEFAULT_MODULO = "SIN MODULO"
    DEFAULT_CLIENTE = "SIN CLIENTE"
    rellenados = defaultdict(int)

    ALLOWED_INSERT = {
        'fecha','modulo','cliente',
        'nro_caso_cliente','nro_caso_interno',
        'tipo_tarea','consultor',
        'hora_inicio','hora_fin','tiempo_invertido',
        'tiempo_facturable','horas_adicionales',
        'horas_convertidas','promedio',
        'descripcion','consolidado_cliente',
        'ocupacion_azure','tarea_azure',
        'extemporaneo','equipo'
    }

    def preparar_mapping(reg):
        fecha_raw = get_val(reg, "FECHA", "fecha")
        dia_raw   = get_val(reg, "DIA", "día", "dia")
        mes_raw   = get_val(reg, "MES", "mes", "mes1", "mes_1")
        anio_raw  = get_val(reg, "ANIO", "ano", "año", "ANIO")

        d, m, y = partir_fecha_mas_campos(fecha_raw, dia_raw, mes_raw, anio_raw)
        if 1 <= d <= 31 and 1 <= m <= 12 and 1900 <= y <= 2100:
            fecha_iso = f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
        else:
            fecha_iso = excel_date_to_iso(fecha_raw)

        consultor_val = clip(get_val(reg, "CONSULTOR", "consultor"), 'consultor')

        modulo_val = clip(get_val(reg, "MODULO", "modulo"), 'modulo')
        if not modulo_val:
            modulo_val = inferir_modulo_por_consultor(consultor_val) or DEFAULT_MODULO
            rellenados['modulo'] += 1

        cliente_val = clip(get_val(reg, "CLIENTE", "cliente"), 'cliente')
        if not cliente_val:
            cliente_val = DEFAULT_CLIENTE
            rellenados['cliente'] += 1

        mapp = {
            'fecha': clip(fecha_iso, 'fecha'),
            'modulo': modulo_val,
            'cliente': cliente_val,
            'nro_caso_cliente': clip(get_val(reg, "NRO CASO CLIENTE", "nro_caso_cliente"), 'nro_caso_cliente'),
            'nro_caso_interno': clip(get_val(reg, "NRO CASO INTERNO", "nro_caso_interno", "nro_caso_cliente_interno"), 'nro_caso_interno'),
            'tipo_tarea': clip(get_val(reg, "Tipo Tarea Azure", "tipo_tarea", "tipo_tarea_azure", "tarea"), 'tipo_tarea'),
            'consultor': consultor_val,
            'hora_inicio': to_sql_time_str(get_val(reg, "Hora Inicio", "hora_inicio", "inicio")),
            'hora_fin': to_sql_time_str(get_val(reg, "Hora Fin", "hora_fin", "fin")),
            'tiempo_invertido': to_sql_time_str(get_val(reg, "TIEMPO INVERTIDO", "tiempo_invertido", "duracion")),
            'tiempo_facturable': to_float(get_val(reg, "TIEMPO FACTURABLE A CLIENTE", "tiempo_facturable", "facturable")),
            'horas_adicionales': parse_horas_adicionales(get_val(reg, "HORAS_ADICIONALES", "horas_adicionales", "hora_adicional", "horas adicionales", "extra")),
            'horas_convertidas': to_float(get_val(reg, "Horas Convertidas", "horas_convertidas")),
            'promedio': to_float(get_val(reg, "PROMEDIO", "promedio")),
            'descripcion': get_val(reg, "Descripción", "descripcion", "detalle", "observaciones"),
            'consolidado_cliente': clip(get_val(reg, "CONCILIADO CON EL CLIENTE", "consolidado_cliente", "consolidado"), 'consolidado_cliente'),
            'ocupacion_azure': clip(get_val(reg, "Ocupacion Azure", "ocupacion_azure", "ocupacion azure"), 'ocupacion_azure'),
            'tarea_azure': clip(get_val(reg, "Tarea Azure", "tarea_azure", "tarea azure"), 'tarea_azure'),
            'extemporaneo': clip(get_val(reg, "EXTEMPORANEO", "extemporaneo"), 'extemporaneo'),
            'equipo': clip(get_val(reg, "Equipo", "equipo"), 'equipo'),
        }
        return {k: v for k, v in mapp.items() if k in ALLOWED_INSERT}

    def flush_batch(batch_maps):
        try:
            safe_batch = [{k: v for k, v in m.items() if k in ALLOWED_INSERT} for m in batch_maps]
            db.session.bulk_insert_mappings(BaseRegistro, safe_batch)
            db.session.commit()
            return len(safe_batch), 0
        except Exception as e:
            app.logger.warning("Fallo lote de %s filas, intentando fila a fila: %s", len(batch_maps), e)
            db.session.rollback()
            ok = 0; fail = 0
            for m in batch_maps:
                try:
                    safe = {k: v for k, v in m.items() if k in ALLOWED_INSERT}
                    db.session.add(BaseRegistro(**safe))
                    db.session.commit()
                    ok += 1
                except Exception as ex:
                    db.session.rollback()
                    fail += 1
                    if fail <= 5:
                        app.logger.error("Fila fallida: %r. Error: %s", m, ex)
            return ok, fail

    total = len(registros)
    inserted = 0; failed = 0; batch = []

    try:
        for idx, reg in enumerate(registros, start=1):
            if idx <= 3:
                app.logger.info("Fila %s (preview): %r", idx, reg)
            batch.append(preparar_mapping(reg))
            if len(batch) >= BATCH_SIZE:
                ok, ko = flush_batch(batch); inserted += ok; failed += ko; batch = []

        if batch:
            ok, ko = flush_batch(batch); inserted += ok; failed += ko

        return jsonify({
            'mensaje': 'Proceso finalizado',
            'replace_all': bool(replace_all),
            'total_recibidos': total,
            'insertados': inserted,
            'fallidos': failed,
        }), 200

    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error cargando registros desde Excel")
        return jsonify({'error': str(e)}), 500

def base_registro_to_dict(r : BaseRegistro):
    return {
        'id': r.id,
        'fecha': r.fecha,
        'cliente': r.cliente,
        'nroCasoCliente': r.nro_caso_cliente,
        'nroCasoInterno': r.nro_caso_interno,
        'tipoTarea': r.tipo_tarea,
        'horaInicio': r.hora_inicio,
        'horaFin': r.hora_fin,
        'tiempoInvertido': r.tiempo_invertido,
        'tiempoFacturable': r.tiempo_facturable,
        'horasAdicionales': r.horas_adicionales,
        'descripcion': r.descripcion,
        'modulo': r.modulo,
        'consultor': r.consultor,
        'equipo': getattr(r, 'equipo', None),
    }

@bp.route('/base-registros', methods=['GET'])
@admin_required
def listar_base_registros():
    page = int(request.args.get('page', 1))
    page_size = min(int(request.args.get('page_size', 50)), 1000)
    q = (request.args.get('q') or '').strip()
    modulo = (request.args.get('modulo') or '').strip()
    cliente = (request.args.get('cliente') or '').strip()
    consultor = (request.args.get('consultor') or '').strip()
    fdesde = (request.args.get('fecha_desde') or '').strip()
    fhasta = (request.args.get('fecha_hasta') or '').strip()

    qry = BaseRegistro.query

    # Intento de parse de fecha robusto (MySQL)
    f10 = func.left(BaseRegistro.fecha, 10)
    fecha_eff = func.coalesce(
        func.str_to_date(f10, '%Y-%m-%d'),
        func.str_to_date(f10, '%d/%m/%Y'),
        func.str_to_date(f10, '%d-%m-%Y')
    )

    if fdesde:
        qry = qry.filter(fecha_eff >= fdesde)
    if fhasta:
        qry = qry.filter(fecha_eff <= fhasta)
    if modulo:
        qry = qry.filter(BaseRegistro.modulo.ilike(f"%{modulo}%"))
    if cliente:
        qry = qry.filter(BaseRegistro.cliente.ilike(f"%{cliente}%"))
    if consultor:
        qry = qry.filter(BaseRegistro.consultor.ilike(f"%{consultor}%"))
    if q:
        like = f"%{q}%"
        qry = qry.filter(or_(
            BaseRegistro.tipo_tarea.ilike(like),
            BaseRegistro.consolidado_cliente.ilike(like),
            BaseRegistro.ocupacion_azure.ilike(like),
            BaseRegistro.tarea_azure.ilike(like),
            BaseRegistro.nro_caso_cliente.ilike(like),
            BaseRegistro.nro_caso_interno.ilike(like),
            BaseRegistro.modulo.ilike(like),
            BaseRegistro.cliente.ilike(like),
            BaseRegistro.consultor.ilike(like),
        ))

    total = qry.count()
    rows = (qry
            .order_by(fecha_eff.desc(), BaseRegistro.consultor.asc(), BaseRegistro.id.asc())
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all())

    return jsonify({
        'data': [base_registro_to_dict(r) for r in rows],
        'total': total,
        'page': page,
        'page_size': page_size,
    })

@bp.route('/consultores/modulos', methods=['GET'])
def get_consultor_modulos():
    try:
        usuario = request.args.get('usuario', '').strip().lower()

        if not usuario:
            return jsonify({'error': 'Parámetro "usuario" requerido'}), 400

        consultor = (
            Consultor.query
            .options(joinedload(Consultor.modulos))
            .filter(func.lower(Consultor.usuario) == usuario)
            .first()
        )

        if not consultor:
            return jsonify({'error': f'Consultor "{usuario}" no encontrado'}), 404

        modulos_payload = _consultor_modulos_payload(consultor)

        modulos = [m["nombre"] for m in modulos_payload]

        if not modulos:
            modulos = ["SIN MODULO"]

        return jsonify({
            "modulos": modulos,
            "modulos_detalle": modulos_payload,
            "modulo": modulos[0] if modulos else "SIN MODULO",
        }), 200

    except Exception as e:
        app.logger.exception("Error en get_consultor_modulos")
        return jsonify({'error': str(e)}), 500

@bp.route('/consultores', methods=['GET'])
@permission_required("CONSULTORES_VER")
def listar_consultores():
    try:
        nombre = (request.args.get('nombre') or '').strip()
        equipo = (request.args.get('equipo') or '').strip()

        query = (
            Consultor.query
            .options(
                joinedload(Consultor.rol_obj),
                joinedload(Consultor.equipo_obj),
                joinedload(Consultor.horario_obj),
                joinedload(Consultor.modulos)
            )
            .order_by(Consultor.nombre.asc())
        )

        if nombre:
            query = query.filter(Consultor.nombre.ilike(f"%{nombre}%"))
        if equipo:
            query = query.join(Equipo).filter(Equipo.nombre.ilike(f"%{equipo}%"))

        consultores = query.all()

        data = []
        for c in consultores:
            data.append({
                'id': c.id,
                'usuario': c.usuario,
                'nombre': c.nombre,
                'rol': c.rol_obj.nombre if c.rol_obj else None,
                'equipo': c.equipo_obj.nombre if c.equipo_obj else None,
                'horario': c.horario_obj.rango if c.horario_obj else None,
                'modulos': [{'id': m.id, 'nombre': m.nombre} for m in c.modulos]
            })

        return jsonify(data), 200

    except Exception as e:
        app.logger.exception("Error al listar consultores")
        return jsonify({'error': str(e)}), 500

@bp.route('/modulos', methods=['GET'])
@permission_required("ADMIN_MODULOS_GESTION")
def listar_modulos():
    try:
        modulos = Modulo.query.order_by(Modulo.nombre.asc()).all()
        data = [{'id': m.id, 'nombre': m.nombre} for m in modulos]
        return jsonify(data), 200
    except Exception as e:
        app.logger.exception("Error al listar módulos")
        return jsonify({'error': str(e)}), 500
    
@bp.route('/consultores/datos', methods=['GET'])
@permission_required("CONSULTORES_VER")
def get_datos_consultor():
    usuario = request.args.get('usuario')
    if not usuario:
        return jsonify({"mensaje": "Debe enviar el parámetro 'usuario'"}), 400

    consultor = (
        Consultor.query
        .options(
            joinedload(Consultor.rol_obj),
            joinedload(Consultor.equipo_obj),
            joinedload(Consultor.horario_obj),
            joinedload(Consultor.modulos)
        )
        .filter_by(usuario=usuario)
        .first()
    )

    if not consultor:
        return jsonify({"mensaje": "Consultor no encontrado"}), 404

    rol = consultor.rol_obj.nombre if consultor.rol_obj else None
    equipo = consultor.equipo_obj.nombre if consultor.equipo_obj else None
    horario = consultor.horario_obj.rango if consultor.horario_obj else None
    modulos_payload = _consultor_modulos_payload(consultor)
    modulos = [m["nombre"] for m in modulos_payload]

    return jsonify({
        "usuario": consultor.usuario,
        "nombre": consultor.nombre,
        "rol": rol,
        "equipo": equipo,
        "horario": horario,
        "modulos": modulos,
        "modulos_detalle": modulos_payload,
        "modulo": modulos[0] if modulos else "SIN MODULO",
        "activo": bool(consultor.activo),
    }), 200

# ===============================
# Oportunidades   

EXCLUDE_LIST = [
    "OTP",
    "OTE",
    "OTL",
    "PROSPECCION",
    "REGISTRO",
    "PENDIENTE APROBACION SAP",
    "0TP",
    "0TE",
    "0TL",
    "OT",
    "EJECUCION CONTRACTUAL",
    "N/A",
]


def _norm_key_for_match(v):
    s = str(v or "").replace("\u00A0", " ").strip().upper()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"\b0TP\b", "OTP", s)
    s = re.sub(r"\b0TE\b", "OTE", s)
    s = re.sub(r"\b0TL\b", "OTL", s)
    return s


EXCLUDE_NORM_SET = {_norm_key_for_match(x) for x in EXCLUDE_LIST if x}


def _get_list_arg(key: str):
    vals = request.args.getlist(key)
    if not vals:
        vals = request.args.getlist(f"{key}[]")
    return [str(v).strip() for v in vals if v is not None and str(v).strip() != ""]


def _sql_norm_estado(col):
    c = func.upper(func.trim(func.replace(col, "\u00A0", " ")))
    for a, b in [("Á", "A"), ("É", "E"), ("Í", "I"), ("Ó", "O"), ("Ú", "U"), ("Ü", "U"), ("Ñ", "N")]:
        c = func.replace(c, a, b)
    c = func.replace(c, "0TP", "OTP")
    c = func.replace(c, "0TE", "OTE")
    c = func.replace(c, "0TL", "OTL")
    c = func.replace(c, "  ", " ")
    c = func.replace(c, "  ", " ")
    return c


def _apply_excluded_states(query):
    col = _sql_norm_estado(Oportunidad.estado_oferta)
    query = query.filter(~col.in_(list(EXCLUDE_NORM_SET)))
    for x in EXCLUDE_NORM_SET:
        safe = x.replace("%", "").replace("_", "")
        query = query.filter(~col.like(f"%{safe}%"))
    return query


def _apply_oportunidades_filters(query, apply_exclusion: bool = True):
    if apply_exclusion:
        query = _apply_excluded_states(query)

    q = (request.args.get("q") or "").strip()

    anios = _get_list_arg("anio")
    meses = _get_list_arg("mes")
    tipos = _get_list_arg("tipo")

    direccion = _get_list_arg("direccion_comercial")
    gerencia = _get_list_arg("gerencia_comercial")
    cliente = _get_list_arg("nombre_cliente")

    estado_oferta = _get_list_arg("estado_oferta")
    resultado = _get_list_arg("resultado_oferta")

    estado_ot = _get_list_arg("estado_ot")
    ultimo_mes = _get_list_arg("ultimo_mes")
    calif = _get_list_arg("calificacion_oportunidad")

    fecha_acta_cierre_ot = _get_list_arg("fecha_acta_cierre_ot")
    fecha_cierre_oportunidad = _get_list_arg("fecha_cierre_oportunidad")

    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Oportunidad.nombre_cliente.ilike(like),
                Oportunidad.servicio.ilike(like),
                Oportunidad.estado_oferta.ilike(like),
                Oportunidad.resultado_oferta.ilike(like),
                Oportunidad.pais.ilike(like),
                Oportunidad.direccion_comercial.ilike(like),
                Oportunidad.gerencia_comercial.ilike(like),
            )
        )

    if anios:
        try:
            anios_int = [int(a) for a in anios]
            query = query.filter(extract("year", Oportunidad.fecha_creacion).in_(anios_int))
        except Exception:
            pass

    if meses:
        try:
            meses_int = [int(m) for m in meses]
            query = query.filter(extract("month", Oportunidad.fecha_creacion).in_(meses_int))
        except Exception:
            pass

    if direccion:
        query = query.filter(Oportunidad.direccion_comercial.in_(direccion))
    if gerencia:
        query = query.filter(Oportunidad.gerencia_comercial.in_(gerencia))
    if cliente:
        query = query.filter(Oportunidad.nombre_cliente.in_(cliente))

    if estado_oferta:
        estado_norm = [_norm_key_for_match(x) for x in estado_oferta]
        query = query.filter(_sql_norm_estado(Oportunidad.estado_oferta).in_(estado_norm))

    if resultado:
        resultado_norm = [_norm_key_for_match(x) for x in resultado]
        query = query.filter(_sql_norm_estado(Oportunidad.resultado_oferta).in_(resultado_norm))

    if estado_ot:
        query = query.filter(Oportunidad.estado_ot.in_(estado_ot))
    if ultimo_mes:
        query = query.filter(Oportunidad.ultimo_mes.in_(ultimo_mes))
    if calif:
        query = query.filter(Oportunidad.calificacion_oportunidad.in_(calif))

    if fecha_acta_cierre_ot:
        query = query.filter(func.date(Oportunidad.fecha_acta_cierre_ot).in_(fecha_acta_cierre_ot))
    if fecha_cierre_oportunidad:
        query = query.filter(func.date(Oportunidad.fecha_cierre_oportunidad).in_(fecha_cierre_oportunidad))

    if tipos:
        ESTADOS_ACTIVOS = {
            "EN PROCESO",
            "DIAGNOSTICO - LEVANTAMIENTO DE INFORMACION",
            "DIAGNOSTICO - LEVANTAMIENTO DE INFORMACIÓN",
            "EN ELABORACION",
            "ENTREGA COMERCIAL",
            "EN ESPERA DEL RFI / RFP",
            "RFI PRESENTADO",
            "SUSPENDIDA",
        }
        ESTADOS_CERRADOS = {
            "CERRADO",
            "CERRADA",
            "CERRADOS",
            "PERDIDA",
            "PERDIDO",
            "DECLINADA",
            "DECLINADO",
            "SUSPENDIDO",
            "PERDIDA - SIN FEEDBACK",
            "RFP PRESENTADO",
        }

        tipos_up = {t.upper().strip() for t in tipos}
        conds = []
        colN = _sql_norm_estado(Oportunidad.estado_oferta)

        if "GANADA" in tipos_up:
            conds.append(colN == "GANADA")

        if "ACTIVA" in tipos_up:
            activosN = [_norm_key_for_match(s) for s in ESTADOS_ACTIVOS]
            conds.append(colN.in_(activosN))

        if "CERRADA" in tipos_up or "CERRADO" in tipos_up:
            cerradosN = [_norm_key_for_match(s) for s in ESTADOS_CERRADOS]
            conds.append(colN.in_(cerradosN))

        if conds:
            query = query.filter(or_(*conds))

    return query


ESTADO_RESULTADO = {
    "REGISTRO": {"OPORTUNIDAD EN PROCESO"},
    "PROSPECCION": {"OPORTUNIDAD EN PROCESO"},
    "DIAGNOSTICO - LEVANTAMIENTO DE INFORMACIÓN": {"OPORTUNIDAD EN PROCESO"},
    "DIAGNOSTICO - LEVANTAMIENTO DE INFORMACION": {"OPORTUNIDAD EN PROCESO"},
    "PENDIENTE APROBACION SAP": {"PENDIENTE APROBACION SAP"},
    "EN ELABORACION": {"OPORTUNIDAD EN PROCESO"},
    "EN ESPERA DEL RFI / RFP": {"EN ESPERA DEL CLIENTE"},
    "RFI PRESENTADO": {"EN ESPERA DEL CLIENTE"},
    "ENTREGA COMERCIAL": {"OPORTUNIDAD EN PROCESO"},
    "GANADA": {
        "BOLSA DE HORAS / CONTINUIDAD DE LA OPERACIÓN",
        "EVOLUTIVO",
        "PROYECTO",
        "VAR",
        "VALORES AGREGADOS",
        "LICENCIAMIENTO",
    },
    "PERDIDA": {"OPORTUNIDAD PERDIDA"},
    "PERDIDA - SIN FEEDBACK": {"OPORTUNIDAD CERRADA"},
    "DECLINADA": {"OPORTUNIDAD CERRADA"},
    "SUSPENDIDA": {"EN ESPERA DEL CLIENTE"},
    "0TL": {"0TL"},
    "0TP": {"0TP"},
    "0TE": {"0TE"},
    "N/A": {"N/A"},
}

CATEGORIA_SUBCATEGORIA = {
    "CLIENTE": {"PRESUPUESTO NO ASIGNADO", "SUSPENDE POR DIRECTRIZ INTERNA"},
    "COMPETENCIA": {"MEJOR POSICIONAMIENTO", "CONDICIONES CONTRACTUALES", "PRESENCIA LOCAL"},
    "PRECIO": {"TARIFA NO COMPETITIVA", "NO CUMPLE PRESUPUESTO"},
    "PRODUCTO": {
        "OTRO PORTAFOLIO DE SOLUCION",
        "PRODUCTO NO SATISFACE LAS NECESIDADES",
        "SOLUCION PROPUESTA NO CUMPLIO",
        "TARIFA NO COMPETITIVA",
    },
    "REASIGNADO": {"SERVICIO DE CLARO EXISTENTE", "REASIGNADO"},
    "SEGUIMIENTO": {"COMERCIAL", "CLIENTE NO RESPONDE"},
}

DATE_FIELDS_API = {
    "fecha_creacion",
    "fecha_cierre_sm",
    "fecha_entrega_oferta_final",
    "vigencia_propuesta",
    "fecha_aceptacion_oferta",
    "fecha_cierre_oportunidad",
    "fecha_firma_aos",
    "fecha_compromiso",
    "fecha_cierre",
    "fecha_acta_cierre_ot",
    "proyeccion_ingreso",
}

INT_FIELDS_API = {
    "otc",
    "mrc",
    "mrc_normalizado",
    "valor_oferta_claro",
}

def _upper(v):
    return str(v).strip().upper() if v is not None else None

def _parse_iso_date(v):
    if v is None:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    s = str(v).strip()
    if s == "" or s.lower() in ("nan", "none", "null"):
        return None
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except Exception:
            return None
    return None

def _parse_decimal(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            return Decimal(str(v))
        except Exception:
            return None

    s = str(v).strip()
    if s == "" or s.lower() in ("nan", "none", "null"):
        return None

    s = re.sub(r"\s+", "", s)
    s = re.sub(r"(?i)COP|USD", "", s)
    s = re.sub(r"[$€£%]", "", s)

    if re.match(r"^[+-]?\d+(\.\d+)?([eE][+-]?\d+)?$", s):
        try:
            return Decimal(s)
        except Exception:
            return None

    last_comma = s.rfind(",")
    last_dot = s.rfind(".")

    if last_comma != -1 and last_dot != -1:
        if last_comma > last_dot:
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif last_comma != -1 and last_dot == -1:
        parts = s.split(",")
        if len(parts) == 2 and len(parts[1]) <= 6:
            s = parts[0].replace(".", "") + "." + parts[1]
        else:
            s = s.replace(",", "").replace(".", "")
    elif last_dot != -1 and last_comma == -1:
        parts = s.split(".")
        if len(parts) == 2 and len(parts[1]) != 3:
            pass
        else:
            s = s.replace(".", "")

    s = re.sub(r"[^\d\.-]", "", s)
    if s == "" or s in ("-", ".", "-."):
        return None

    try:
        return Decimal(s)
    except Exception:
        return None

def _to_int_round(v_dec):
    if v_dec is None:
        return None
    try:
        return int(v_dec.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    except Exception:
        return None

def _normalizar_mostrar_dashboard(v):
    """
    Regla:
    - Si viene NO / N / FALSE / 0 => guarda 'NO'
    - Si viene SI / SÍ / S / TRUE / 1 / YES => guarda 'SI'
    - Si viene vacío, None, null, nan => deja None
      En el dashboard None o vacío se interpreta como visible.
    - Cualquier otro valor no vacío se toma como 'SI'
    """
    if v is None:
        return None

    s = str(v).replace("\u00A0", " ").strip().upper()

    if s in ("", "NAN", "NONE", "NULL"):
        return None

    if s in ("NO", "N", "FALSE", "0"):
        return "NO"

    if s in ("SI", "SÍ", "S", "TRUE", "1", "YES", "Y"):
        return "SI"

    return "SI"


def clean_payload(data: dict) -> dict:
    out = {}
    data = data or {}

    for k, v in data.items():
        if isinstance(v, str):
            v = v.strip()
            if v == "":
                v = None

        if k in DATE_FIELDS_API:
            v = _parse_iso_date(v)

        if k in INT_FIELDS_API and k != "mrc_normalizado":
            v = _to_int_round(_parse_decimal(v))

        if k == "mrc_normalizado":
            v = None

        if k == "mostrar_dashboard":
            v = _normalizar_mostrar_dashboard(v)

        out[k] = v

    moneda = out.get("tipo_moneda")
    if moneda is not None:
        m = _upper(moneda)
        out["tipo_moneda"] = m if m in ("COP", "USD") else None

    estado = out.get("estado_oferta")
    resultado = out.get("resultado_oferta")

    if estado:
        estado_n = _upper(estado)
        allowed = ESTADO_RESULTADO.get(estado_n)

        if allowed:
            if resultado:
                res_n = _upper(resultado)
                allowed_up = {a.upper() for a in allowed}

                if res_n not in allowed_up:
                    out["resultado_oferta"] = None

            if not out.get("resultado_oferta") and len(allowed) == 1:
                out["resultado_oferta"] = next(iter(allowed))

    cat = out.get("categoria_perdida")
    sub = out.get("subcategoria_perdida")

    if cat:
        cat_n = _upper(cat)
        allowed = CATEGORIA_SUBCATEGORIA.get(cat_n)

        if allowed and sub:
            sub_n = _upper(sub)
            allowed_up = {a.upper() for a in allowed}

            if sub_n not in allowed_up:
                out["subcategoria_perdida"] = None

    otc = out.get("otc")
    mrc = out.get("mrc")

    mrc_norm = None

    try:
        if otc is not None and mrc is not None:
            mrc_norm = (Decimal(str(otc)) / Decimal("12")) + Decimal(str(mrc))
        elif otc is not None:
            mrc_norm = Decimal(str(otc)) / Decimal("12")
        elif mrc is not None:
            mrc_norm = Decimal(str(mrc))
    except Exception:
        mrc_norm = None

    out["mrc_normalizado"] = _to_int_round(mrc_norm)

    return out

ESTADO_RESULTADO_FORZADO = {
    "EN ESPERA DEL RFI / RFP": "EN ESPERA DEL CLIENTE",
    "RFI PRESENTADO": "EN ESPERA DEL CLIENTE",
    "SUSPENDIDA": "EN ESPERA DEL CLIENTE",
}

def normalize_oportunidad_dict(row: dict) -> dict:
    row = dict(row or {})

    estado = _upper(row.get("estado_oferta"))
    resultado = _upper(row.get("resultado_oferta"))

    if estado in ESTADO_RESULTADO_FORZADO:
        row["resultado_oferta"] = ESTADO_RESULTADO_FORZADO[estado]

    if estado:
        row["estado_oferta"] = estado

    if resultado and estado not in ESTADO_RESULTADO_FORZADO:
        row["resultado_oferta"] = resultado

    return row


@bp.route("/oportunidades/import", methods=["POST"])
@permission_required("OPORTUNIDADES_CREAR")
def importar_oportunidades():
    file = request.files.get("file")
    if not file:
        return jsonify({"mensaje": "Archivo no recibido"}), 400

    if Oportunidad.query.count() > 0:
        return jsonify({"mensaje": "La carga inicial ya fue realizada"}), 400

    df = pd.read_excel(BytesIO(file.read()), dtype=str)

    def norm_col(c):
        c = str(c).replace("\u00A0", " ").strip().upper()
        c = re.sub(r"\s+", " ", c)
        return c

    df.columns = [norm_col(c) for c in df.columns]

    colmap = {
        "NOMBRE CLIENTE": "nombre_cliente",
        "SERVICIO": "servicio",
        "FECHA DE ASIGNACION": "fecha_creacion",
        "FECHA CREACIÓN": "fecha_creacion",
        "TIPO CLIENTE": "tipo_cliente",
        "TIPO DE SOLICITUD": "tipo_solicitud",
        "CASO SM": "caso_sm",
        "FECHA CIERRE SM": "fecha_cierre_sm",
        "SALESFORCE": "salesforce",
        "ULTIMOS 6 MESES": "ultimos_6_meses",
        "ULTIMO MES": "ultimo_mes",
        "RETRASO": "retraso",
        "ESTADO OFERTA": "estado_oferta",
        "RESULTADO OFERTA": "resultado_oferta",
        "CALIFICACION OPORTUNIDAD": "calificacion_oportunidad",
        "ORIGEN DE LA OPORTUNIDAD": "origen_oportunidad",
        "DIRECCION COMERCIAL": "direccion_comercial",
        "GERENCIA COMERCIAL": "gerencia_comercial",
        "COMERCIAL ASIGNADO": "comercial_asignado",
        "CONSULTOR COMERCIAL": "consultor_comercial",
        "COMERCIAL ASIGNADO HITSS": "comercial_asignado_hitss",
        "OBSERVACIONES": "observaciones",
        "CATEGORIA PERDIDA": "categoria_perdida",
        "SUBCATEGORIA PERDIDA": "subcategoria_perdida",
        "FECHA ENTREGA OFERTA FINAL AL CLIENTE": "fecha_entrega_oferta_final",
        "VIGENCIA DE LA PROPUESTA": "vigencia_propuesta",
        "FECHA ACEPTACIÓN DE LA OFERTA": "fecha_aceptacion_oferta",
        "TIPO DE MONEDA": "tipo_moneda",
        "OTC": "otc",
        "OTR": "otc",
        "MRC": "mrc",
        "MRC NORMALIZADO": "mrc_normalizado",
        "VALOR OFERTA CLARO": "valor_oferta_claro",
        "DURACION": "duracion",
        "PAIS": "pais",
        "FECHA DE CIERRE OPORTUNIDAD": "fecha_cierre_oportunidad",
        "CODIGO PROYECTO (PRC)": "codigo_prc",
        "FECHA FIRMA AOS": "fecha_firma_aos",
        "PM ASIGNADO CLARO": "pm_asignado_claro",
        "PM ASIGNADO GLOBAL HITSS": "pm_asignado_hitss",
        "DESCRIPCION OT": "descripcion_ot",
        "NO. ENLACE": "num_enlace",
        "NO. INCIDENTE": "num_incidente",
        "NO OT": "num_ot",
        "ESTADO OT": "estado_ot",
        "PROYECCION DEL INGRESO PROYECTO / EVOLUTIVO": "proyeccion_ingreso",
        "FECHA COMPROMISO": "fecha_compromiso",
        "FECHA DE CIERRE": "fecha_cierre",
        "ESTADO PROYECTO / EVOLUTIVO": "estado_proyecto",
        "AÑO CREACIÓN OT": "anio_creacion_ot",
        "FECHA ACTA DE CIERRE Y/O OT": "fecha_acta_cierre_ot",
        "SEGUIMIENTO ORDENES DE TRABAJO": "seguimiento_ot",
        "TIPO DE SERVICIO": "tipo_servicio",
        "SEMESTRE DE EJECUCIÓN": "semestre_ejecucion",
        "PUBLICACIÓN SHAREPOINT": "publicacion_sharepoint",
    }

    DATE_FIELDS = {
        "fecha_creacion",
        "fecha_cierre_sm",
        "fecha_entrega_oferta_final",
        "vigencia_propuesta",
        "fecha_aceptacion_oferta",
        "fecha_cierre_oportunidad",
        "fecha_firma_aos",
        "fecha_compromiso",
        "fecha_cierre",
        "fecha_acta_cierre_ot",
        "proyeccion_ingreso",
    }

    MONEY_FIELDS = {"otc", "mrc", "mrc_normalizado", "valor_oferta_claro"}

    def parse_date(val):
        if val is None:
            return None
        s = str(val).strip()
        if s == "" or s.lower() in ("nan", "none", "null"):
            return None
        try:
            d = pd.to_datetime(s, errors="coerce", dayfirst=True)
            return None if pd.isna(d) else d.date()
        except Exception:
            return None

    def parse_str(val):
        if val is None:
            return None
        s = str(val).replace("\u00A0", " ").strip()
        if s == "" or s.lower() in ("nan", "none", "null"):
            return None
        return s

    def parse_money_int(val):
        if val is None:
            return None

        s = str(val).strip()
        if s == "" or s.lower() in ("nan", "none", "null"):
            return None

        s = re.sub(r"\s+", "", s)
        s = re.sub(r"(?i)COP|USD", "", s)
        s = re.sub(r"[$€£%]", "", s)

        if re.match(r"^[+-]?\d+(\.\d+)?([eE][+-]?\d+)?$", s):
            try:
                d = Decimal(s)
                return int(d.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
            except InvalidOperation:
                return None

        last_comma = s.rfind(",")
        last_dot = s.rfind(".")

        if last_comma != -1 and last_dot != -1:
            if last_comma > last_dot:
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        elif last_comma != -1 and last_dot == -1:
            parts = s.split(",")
            if len(parts) == 2 and len(parts[1]) <= 6:
                s = parts[0].replace(".", "") + "." + parts[1]
            else:
                s = s.replace(",", "").replace(".", "")
        elif last_dot != -1 and last_comma == -1:
            parts = s.split(".")
            if len(parts) == 2 and len(parts[1]) != 3:
                pass
            else:
                s = s.replace(".", "")

        s = re.sub(r"[^\d\.-]", "", s)
        if s == "" or s in ("-", ".", "-."):
            return None

        try:
            d = Decimal(s)
            return int(d.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        except InvalidOperation:
            return None

    data_list = []

    for _, row in df.iterrows():
        obj = {}

        for col_excel, field in colmap.items():
            if col_excel in df.columns:
                raw = row.get(col_excel)

                if field in DATE_FIELDS:
                    obj[field] = parse_date(raw)
                elif field in MONEY_FIELDS:
                    obj[field] = parse_money_int(raw)
                else:
                    obj[field] = parse_str(raw)

        fecha = obj.get("fecha_creacion")
        if fecha:
            mes = fecha.month
            anio = fecha.year
            obj["semestre"] = f"{'1ER' if mes <= 6 else '2DO'} SEMESTRE {anio}"
        else:
            obj["semestre"] = None

        otc = obj.get("otc")
        mrc = obj.get("mrc")

        try:
            if otc is not None and mrc is not None:
                obj["mrc_normalizado"] = int(
                    (Decimal(otc) / Decimal("12") + Decimal(mrc)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
                )
            elif otc is not None:
                obj["mrc_normalizado"] = int(
                    (Decimal(otc) / Decimal("12")).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
                )
            elif mrc is not None:
                obj["mrc_normalizado"] = int(Decimal(mrc).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
            else:
                obj["mrc_normalizado"] = None
        except Exception:
            obj["mrc_normalizado"] = None

        obj = clean_payload(obj)

        data_list.append(Oportunidad(**obj))

    try:
        db.session.bulk_save_objects(data_list)
        db.session.commit()
        return jsonify({"mensaje": f"Carga inicial exitosa ({len(data_list)} registros)"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"mensaje": f"Error al guardar: {str(e)}"}), 500

def _merge_unique_sorted(values, extra_values=None):
    vals = set()

    for v in values or []:
        s = str(v).strip()
        if s:
            vals.add(s)

    for v in extra_values or []:
        s = str(v).strip()
        if s:
            vals.add(s)

    return sorted(vals, key=lambda x: x.upper())

def distinct_clientes_model():
    rows = (
        Cliente.query.with_entities(Cliente.nombre_cliente)
        .filter(Cliente.nombre_cliente.isnot(None))
        .filter(func.trim(Cliente.nombre_cliente) != "")
        .order_by(Cliente.nombre_cliente.asc())
        .all()
    )
    return [r[0] for r in rows]


@bp.route("/oportunidades/filters", methods=["GET"])
@permission_required("OPORTUNIDADES_VER")
def oportunidades_filters():
    try:
        # TABLA PRINCIPAL: sin exclusión
        base = Oportunidad.query

        anios = (
            base.with_entities(extract("year", Oportunidad.fecha_creacion).label("y"))
            .filter(Oportunidad.fecha_creacion.isnot(None))
            .distinct()
            .order_by("y")
            .all()
        )

        meses = (
            base.with_entities(extract("month", Oportunidad.fecha_creacion).label("m"))
            .filter(Oportunidad.fecha_creacion.isnot(None))
            .distinct()
            .order_by("m")
            .all()
        )

        def distinct_col(col):
            rows = (
                base.with_entities(col)
                .filter(col.isnot(None))
                .filter(func.trim(col) != "")
                .distinct()
                .order_by(col.asc())
                .all()
            )
            return [r[0] for r in rows]

        def distinct_date(col):
            rows = (
                base.with_entities(func.date(col))
                .filter(col.isnot(None))
                .distinct()
                .order_by(func.date(col).asc())
                .all()
            )
            out = []
            for r in rows:
                v = r[0]
                if v is None:
                    continue
                if hasattr(v, "strftime"):
                    out.append(v.strftime("%Y-%m-%d"))
                else:
                    out.append(str(v))
            return out

        return jsonify(
            {
                "anios": [int(r.y) for r in anios if r.y is not None],
                "meses": [int(r.m) for r in meses if r.m is not None],
                "direccion_comercial": distinct_col(Oportunidad.direccion_comercial),
                "gerencia_comercial": distinct_col(Oportunidad.gerencia_comercial),
                "nombre_cliente": _merge_unique_sorted(
                    distinct_clientes_model(),
                    distinct_col(Oportunidad.nombre_cliente)
                ),
                "servicio": distinct_col(Oportunidad.servicio),

                "estado_oferta": _merge_unique_sorted(
                    distinct_col(Oportunidad.estado_oferta),
                    [
                        "EN ESPERA DEL RFI / RFP",
                        "RFI PRESENTADO",
                        "SUSPENDIDA",
                    ],
                ),

                "resultado_oferta": _merge_unique_sorted(
                    distinct_col(Oportunidad.resultado_oferta),
                    [
                        "EN ESPERA DEL CLIENTE",
                    ],
                ),

                "estado_ot": distinct_col(Oportunidad.estado_ot),
                "ultimo_mes": distinct_col(Oportunidad.ultimo_mes),
                "calificacion_oportunidad": distinct_col(Oportunidad.calificacion_oportunidad),
                "fecha_acta_cierre_ot": distinct_date(Oportunidad.fecha_acta_cierre_ot),
                "fecha_cierre_oportunidad": distinct_date(Oportunidad.fecha_cierre_oportunidad),
                "tipos": ["GANADA", "ACTIVA", "CERRADA"],
            }
        ), 200

    except Exception:
        return jsonify({
            "mensaje": "Error interno en /oportunidades/filters",
            "trace": traceback.format_exc()
        }), 500


@bp.route("/oportunidades", methods=["GET"])
@permission_required("OPORTUNIDADES_VER")
def listar_oportunidades():
    try:
        query = Oportunidad.query
        query = _apply_oportunidades_filters(query, apply_exclusion=False)

        query = query.order_by(Oportunidad.id.desc())
        data = [normalize_oportunidad_dict(o.to_dict()) for o in query.limit(5000).all()]
        return jsonify(data), 200

    except Exception:
        return jsonify({"mensaje": "Error interno en /oportunidades", "trace": traceback.format_exc()}), 500

@bp.route("/oportunidades", methods=["POST"])
@permission_required("OPORTUNIDADES_CREAR")
def crear_oportunidad():
    try:
        data = clean_payload(request.get_json() or {})
        o = Oportunidad(**data)
        db.session.add(o)
        db.session.commit()
        return jsonify(o.to_dict()), 201
    except Exception:
        db.session.rollback()
        return jsonify({"mensaje": "Error creando oportunidad", "trace": traceback.format_exc()}), 500


@bp.route("/oportunidades/<int:id>", methods=["PUT"])
@permission_required("OPORTUNIDADES_EDITAR")
def editar_oportunidad(id):
    try:
        data = clean_payload(request.get_json() or {})
        o = Oportunidad.query.get_or_404(id)
        for k, v in data.items():
            if hasattr(o, k):
                setattr(o, k, v)
        db.session.commit()
        return jsonify({"mensaje": "Actualizado correctamente"}), 200
    except Exception:
        db.session.rollback()
        return jsonify({"mensaje": "Error editando oportunidad", "trace": traceback.format_exc()}), 500


@bp.route("/oportunidades/<int:id>", methods=["DELETE"])
@permission_required("OPORTUNIDADES_ELIMINAR")
def eliminar_oportunidad(id):
    try:
        o = Oportunidad.query.get_or_404(id)
        db.session.delete(o)
        db.session.commit()
        return jsonify({"mensaje": "Eliminado correctamente"}), 200
    except Exception:
        db.session.rollback()
        return jsonify({"mensaje": "Error eliminando oportunidad", "trace": traceback.format_exc()}), 500

# ========== CLIENTES ==========

@bp.route('/clientes', methods=['GET'])
@permission_required("CLIENTES_VER")
def listar_clientes():
    q = (request.args.get('q') or "").strip()
    query = Cliente.query

    if q:
        like = f"%{q}%"
        query = query.filter(Cliente.nombre_cliente.ilike(like))

    data = [c.to_dict() for c in query.order_by(Cliente.nombre_cliente).all()]
    return jsonify(data), 200

@bp.route('/clientes', methods=['POST'])
@permission_required("CLIENTES_CREAR")
def crear_cliente():
    data = request.get_json() or {}

    c = Cliente(nombre_cliente=data.get("nombre_cliente").strip())

    db.session.add(c)
    try:
        db.session.commit()
        return jsonify({"mensaje": "Cliente creado correctamente"}), 201
    except IntegrityError:
        db.session.rollback()
        return jsonify({"mensaje": "Cliente duplicado"}), 400


@bp.route('/clientes/<int:id>', methods=['PUT'])
@permission_required("CLIENTES_EDITAR")
def editar_cliente(id):
    c = Cliente.query.get_or_404(id)
    data = request.get_json() or {}

    c.nombre_cliente = data.get("nombre_cliente", c.nombre_cliente).strip()

    try:
        db.session.commit()
        return jsonify({"mensaje": "Cliente actualizado correctamente"}), 200
    except IntegrityError:
        db.session.rollback()
        return jsonify({"mensaje": "Cliente ya existe"}), 400


@bp.route('/clientes/<int:id>', methods=['DELETE'])
@permission_required("CLIENTES_ELIMINAR")
def eliminar_cliente(id):
    c = Cliente.query.get_or_404(id)
    db.session.delete(c)
    db.session.commit()
    return jsonify({"mensaje": "Cliente eliminado correctamente"}), 200

# ========== PERMISOS ==========

@bp.route('/permisos', methods=['GET'])
@permission_required("PERMISOS_VER")
def listar_permisos():
    permisos = Permiso.query.order_by(Permiso.codigo).all()
    return jsonify([p.to_dict() for p in permisos]), 200

@bp.route('/permisos', methods=['POST'])
@permission_required("PERMISOS_CREAR")
def crear_permiso():
    data = request.get_json() or {}
    codigo = data.get("codigo", "").strip().upper()

    if not codigo:
        return jsonify({"mensaje": "El código es obligatorio"}), 400

    if Permiso.query.filter_by(codigo=codigo).first():
        return jsonify({"mensaje": "El permiso ya existe"}), 400

    p = Permiso(codigo=codigo, descripcion=data.get("descripcion"))
    db.session.add(p)
    db.session.commit()

    return jsonify({"mensaje": "Permiso creado", "permiso": p.to_dict()}), 201

@bp.route('/permisos/<int:id>', methods=['DELETE'])
@permission_required("PERMISOS_ELIMINAR")
def eliminar_permiso(id):
    p = Permiso.query.get_or_404(id)
    db.session.delete(p)
    db.session.commit()
    return jsonify({"mensaje": "Permiso eliminado"}), 200

@bp.route('/roles/<int:rol_id>/permisos', methods=['GET'])
@permission_required("ROLES_VER")
def permisos_por_rol(rol_id):
    perms = (
        db.session.query(Permiso)
        .join(RolPermiso, RolPermiso.permiso_id == Permiso.id)
        .filter(RolPermiso.rol_id == rol_id)
        .order_by(Permiso.codigo.asc())
        .all()
    )
    return jsonify([p.to_dict() for p in perms]), 200

@bp.route('/roles/<int:rol_id>/permisos', methods=['POST'])
@permission_required("ROLES_EDITAR")
def asignar_permiso_rol(rol_id):
    data = request.get_json() or {}
    permiso_id = data.get("permiso_id")

    if not permiso_id:
        return jsonify({"mensaje": "permiso_id requerido"}), 400

    if RolPermiso.query.filter_by(rol_id=rol_id, permiso_id=permiso_id).first():
        return jsonify({"mensaje": "Permiso ya asignado"}), 400

    rp = RolPermiso(rol_id=rol_id, permiso_id=permiso_id)
    db.session.add(rp)
    db.session.commit()

    return jsonify({"mensaje": "Permiso asignado al rol"}), 201

@bp.route('/roles/<int:rol_id>/permisos/<int:permiso_id>', methods=['DELETE'])
@permission_required("ROLES_EDITAR")
def quitar_permiso_rol(rol_id, permiso_id):
    rp = RolPermiso.query.filter_by(rol_id=rol_id, permiso_id=permiso_id).first()

    if not rp:
        return jsonify({"mensaje": "No estaba asignado"}), 404

    db.session.delete(rp)
    db.session.commit()

    return jsonify({"mensaje": "Permiso removido del rol"}), 200

@bp.route('/equipos/<int:equipo_id>/permisos', methods=['GET'])
@permission_required("EQUIPOS_VER")
def permisos_por_equipo(equipo_id):
    perms = (
        db.session.query(Permiso)
        .join(EquipoPermiso, EquipoPermiso.permiso_id == Permiso.id)
        .filter(EquipoPermiso.equipo_id == equipo_id)
        .order_by(Permiso.codigo.asc())
        .all()
    )
    return jsonify([p.to_dict() for p in perms]), 200

@bp.route('/equipos/<int:equipo_id>/permisos', methods=['POST'])
def asignar_permiso_equipo(equipo_id):
    data = request.get_json() or {}
    permiso_id = data.get("permiso_id")

    if not permiso_id:
        return jsonify({"mensaje": "permiso_id requerido"}), 400

    if EquipoPermiso.query.filter_by(equipo_id=equipo_id, permiso_id=permiso_id).first():
        return jsonify({"mensaje": "Ya asignado"}), 400

    ep = EquipoPermiso(equipo_id=equipo_id, permiso_id=permiso_id)
    db.session.add(ep)
    db.session.commit()

    return jsonify({"mensaje": "Permiso asignado al equipo"}), 201

@bp.route('/equipos/<int:equipo_id>/permisos/<int:permiso_id>', methods=['DELETE'])
def quitar_permiso_equipo(equipo_id, permiso_id):
    ep = EquipoPermiso.query.filter_by(equipo_id=equipo_id, permiso_id=permiso_id).first()

    if not ep:
        return jsonify({"mensaje": "No estaba asignado"}), 404

    db.session.delete(ep)
    db.session.commit()

    return jsonify({"mensaje": "Permiso removido del equipo"}), 200

@bp.route('/consultores/<int:consultor_id>/permisos', methods=['GET'])
@permission_required("CONSULTORES_VER")
def permisos_por_consultor(consultor_id):
    perms = (
        db.session.query(Permiso)
        .join(ConsultorPermiso, ConsultorPermiso.permiso_id == Permiso.id)
        .filter(ConsultorPermiso.consultor_id == consultor_id)
        .order_by(Permiso.codigo.asc())
        .all()
    )
    return jsonify([p.to_dict() for p in perms]), 200

@bp.route('/consultores/<int:consultor_id>/permisos-efectivos', methods=['GET'])
def permisos_efectivos_consultor(consultor_id):
    return permisos_asignados(consultor_id)


@bp.route('/consultores/<int:consultor_id>/permisos', methods=['POST'])
def asignar_permiso_consultor(consultor_id):
    data = request.get_json() or {}
    permiso_id = data.get("permiso_id")

    if not permiso_id:
        return jsonify({"mensaje": "permiso_id requerido"}), 400

    if ConsultorPermiso.query.filter_by(consultor_id=consultor_id, permiso_id=permiso_id).first():
        return jsonify({"mensaje": "Permiso ya otorgado"}), 400

    cp = ConsultorPermiso(consultor_id=consultor_id, permiso_id=permiso_id)
    db.session.add(cp)
    db.session.commit()

    return jsonify({"mensaje": "Permiso asignado al consultor"}), 201

@bp.route('/consultores/<int:consultor_id>/permisos/<int:permiso_id>', methods=['DELETE'])
def quitar_permiso_consultor(consultor_id, permiso_id):
    cp = ConsultorPermiso.query.filter_by(consultor_id=consultor_id, permiso_id=permiso_id).first()

    if not cp:
        return jsonify({"mensaje": "No estaba asignado"}), 404

    db.session.delete(cp)
    db.session.commit()

    return jsonify({"mensaje": "Permiso removido del consultor"}), 200


# -------------------------------
#   OCUPACIONES
# -------------------------------

@bp.route("/ocupaciones", methods=["GET"])
def listar_ocupaciones():
    ocupaciones = Ocupacion.query.order_by(Ocupacion.codigo).all()
    return jsonify([o.to_dict() for o in ocupaciones]), 200


@bp.route("/ocupaciones", methods=["POST"])
def crear_ocupacion():
    data = request.get_json() or {}
    codigo = data.get("codigo", "").strip().upper()
    nombre = data.get("nombre")
    descripcion = data.get("descripcion")

    if not codigo or not nombre:
        return jsonify({"mensaje": "Código y nombre son obligatorios"}), 400

    if Ocupacion.query.filter_by(codigo=codigo).first():
        return jsonify({"mensaje": "La ocupación ya existe"}), 400

    o = Ocupacion(codigo=codigo, nombre=nombre, descripcion=descripcion)
    db.session.add(o)
    db.session.commit()
    return jsonify({"mensaje": "Ocupación creada", "ocupacion": o.to_dict()}), 201


@bp.route("/ocupaciones/<int:id>", methods=["PUT"])
def editar_ocupacion(id):
    o = Ocupacion.query.get_or_404(id)
    data = request.get_json() or {}

    o.nombre = data.get("nombre", o.nombre)
    o.descripcion = data.get("descripcion", o.descripcion)

    db.session.commit()
    return jsonify({"mensaje": "Ocupación actualizada", "ocupacion": o.to_dict()}), 200


@bp.route("/ocupaciones/<int:id>", methods=["DELETE"])
def eliminar_ocupacion(id):
    o = Ocupacion.query.get_or_404(id)
    db.session.delete(o)
    db.session.commit()
    return jsonify({"mensaje": "Ocupación eliminada"}), 200


# -------------------------------
#   TAREAS
# -------------------------------

@bp.route("/tareas", methods=["GET"])
def listar_tareas():
    tareas = Tarea.query.order_by(Tarea.codigo).all()
    return jsonify([t.to_dict() for t in tareas]), 200


@bp.route("/tareas", methods=["POST"])
def crear_tarea():
    data = request.get_json() or {}
    codigo = data.get("codigo", "").strip().upper()
    nombre = data.get("nombre")
    descripcion = data.get("descripcion")

    if not codigo or not nombre:
        return jsonify({"mensaje": "Código y nombre son obligatorios"}), 400

    if Tarea.query.filter_by(codigo=codigo).first():
        return jsonify({"mensaje": "La tarea ya existe"}), 400

    t = Tarea(codigo=codigo, nombre=nombre, descripcion=descripcion)
    db.session.add(t)
    db.session.commit()
    return jsonify({"mensaje": "Tarea creada", "tarea": t.to_dict()}), 201


@bp.route("/tareas/<int:id>", methods=["PUT"])
def editar_tarea(id):
    t = Tarea.query.get_or_404(id)
    data = request.get_json() or {}

    t.nombre = data.get("nombre", t.nombre)
    t.descripcion = data.get("descripcion", t.descripcion)

    db.session.commit()
    return jsonify({"mensaje": "Tarea actualizada", "tarea": t.to_dict()}), 200


@bp.route("/tareas/<int:id>", methods=["DELETE"])
def eliminar_tarea(id):
    t = Tarea.query.get_or_404(id)
    db.session.delete(t)
    db.session.commit()
    return jsonify({"mensaje": "Tarea eliminada"}), 200


# -------------------------------
#   ALIASES DE TAREAS
# -------------------------------

@bp.route("/tareas/<int:tarea_id>/alias", methods=["POST"])
def crear_alias_tarea(tarea_id):
    tarea = Tarea.query.get_or_404(tarea_id)

    data = request.get_json() or {}
    alias = data.get("alias", "").strip()

    if not alias:
        return jsonify({"mensaje": "Alias requerido"}), 400

    a = TareaAlias(alias=alias, tarea=tarea)
    db.session.add(a)
    db.session.commit()

    return jsonify({"mensaje": "Alias creado", "alias": a.to_dict()}), 201


@bp.route("/tareas/alias/<int:id>", methods=["DELETE"])
def eliminar_alias(id):
    alias = TareaAlias.query.get_or_404(id)
    db.session.delete(alias)
    db.session.commit()
    return jsonify({"mensaje": "Alias eliminado"}), 200


# -------------------------------
#   ASIGNAR / QUITAR TAREAS A UNA OCUPACIÓN
# -------------------------------

@bp.route("/ocupaciones/<int:ocupacion_id>/tareas", methods=["GET"])
def tareas_por_ocupacion(ocupacion_id):
    ocupacion = Ocupacion.query.get_or_404(ocupacion_id)
    return jsonify([t.to_dict_simple() for t in ocupacion.tareas]), 200


@bp.route("/ocupaciones/<int:ocupacion_id>/tareas", methods=["POST"])
def asignar_tarea_a_ocupacion(ocupacion_id):
    ocupacion = Ocupacion.query.get_or_404(ocupacion_id)
    data = request.get_json() or {}

    tarea_id = data.get("tarea_id")
    if not tarea_id:
        return jsonify({"mensaje": "tarea_id requerido"}), 400

    tarea = Tarea.query.get_or_404(tarea_id)

    if tarea in ocupacion.tareas:
        return jsonify({"mensaje": "La tarea ya está asignada"}), 400

    ocupacion.tareas.append(tarea)
    db.session.commit()

    return jsonify({"mensaje": "Tarea asignada", "ocupacion": ocupacion.to_dict()}), 201


@bp.route("/ocupaciones/<int:ocupacion_id>/tareas/<int:tarea_id>", methods=["DELETE"])
def quitar_tarea_de_ocupacion(ocupacion_id, tarea_id):
    ocupacion = Ocupacion.query.get_or_404(ocupacion_id)
    tarea = Tarea.query.get_or_404(tarea_id)

    if tarea not in ocupacion.tareas:
        return jsonify({"mensaje": "La tarea no estaba asignada"}), 404

    ocupacion.tareas.remove(tarea)
    db.session.commit()

    return jsonify({"mensaje": "Tarea removida", "ocupacion": ocupacion.to_dict()}), 200

@bp.route("/horarios-estadisticas", methods=["GET"])
@permission_required("GRAFICOS_VER")
def obtener_horarios():
    try:
        registros = Registro.query.all()

        # Normalizar ocupaciones antes de procesar
        registros = normalizar_ocupaciones(registros)

        data = []

        for r in registros:
            occ_text = r.ocupacion_azure or "00 - SIN CLASIFICAR"
            horas = float(r.horas_convertidas or 0)

            data.append({
                "ocupacion": occ_text,
                "tarea": r.tipo_tarea,
                "horas": horas
            })

        # Agrupación por ocupación
        agrupado = {}
        for d in data:
            agrupado.setdefault(d["ocupacion"], 0)
            agrupado[d["ocupacion"]] += d["horas"]

        # Resultado ordenado:
        resultado = [
            {"name": k, "value": round((h / sum(agrupado.values())) * 100, 2), "horas": h}
            for k, h in agrupado.items()
        ]

        resultado.sort(key=lambda x: x["value"], reverse=True)

        return jsonify({
            "ocupaciones": resultado
        })

    except Exception as e:
        print("❌ Error:", str(e))
        return jsonify({"error": str(e)}), 500


@bp.route('/permisos-asignados/<int:consultor_id>', methods=['GET'])
def permisos_asignados(consultor_id):

    consultor = (
        Consultor.query.options(
            joinedload(Consultor.rol_obj)
                .joinedload(Rol.permisos_asignados)
                .joinedload(RolPermiso.permiso),

            joinedload(Consultor.equipo_obj)
                .joinedload(Equipo.permisos_asignados)
                .joinedload(EquipoPermiso.permiso),

            joinedload(Consultor.permisos_especiales)
                .joinedload(ConsultorPermiso.permiso),

            joinedload(Consultor.modulos)
        )
        .filter_by(id=consultor_id)
        .first()
    )

    if not consultor:
        return jsonify({"error": "Consultor no encontrado"}), 404

    # ===============================
    # Recolección final de permisos
    # ===============================
    permisos_set = set()

    # Permisos del rol
    if consultor.rol_obj:
        for rp in consultor.rol_obj.permisos_asignados:
            permisos_set.add(rp.permiso.codigo)

    # Permisos del equipo
    if consultor.equipo_obj:
        for ep in consultor.equipo_obj.permisos_asignados:
            permisos_set.add(ep.permiso.codigo)

    # Permisos individuales
    for cp in consultor.permisos_especiales:
        permisos_set.add(cp.permiso.codigo)

    permisos_list = sorted(list(permisos_set))

    # ===============================
    #   RESPUESTA COMPLETA
    # ===============================
    return jsonify({
        "consultor": {
            "id": consultor.id,
            "usuario": consultor.usuario,
            "nombre": consultor.nombre,
            "rol": consultor.rol_obj.nombre if consultor.rol_obj else None,
            "equipo": consultor.equipo_obj.nombre if consultor.equipo_obj else None,
            "horario": consultor.horario_obj.rango if consultor.horario_obj else None,
            "modulos": [{"id": m.id, "nombre": m.nombre} for m in consultor.modulos],
        },
        "permisos": permisos_list,

        # Debug extendido (opcional)
        "detalle": {
            "rol_permisos": [rp.permiso.codigo for rp in consultor.rol_obj.permisos_asignados] if consultor.rol_obj else [],
            "equipo_permisos": [ep.permiso.codigo for ep in consultor.equipo_obj.permisos_asignados] if consultor.equipo_obj else [],
            "permisos_individuales": [cp.permiso.codigo for cp in consultor.permisos_especiales]
        }
    }), 200

# -------------------------------
#   HORARIOS ESTADÍSTICAS (NO CONFUNDIR CON /horarios)
# -------------------------------

@bp.route("/horas-ocupacion", methods=["GET"])
@permission_required("GRAFICOS_VER")
def horas_ocupacion():
    try:
        usuario = _get_usuario_from_request()
        rol_req = _get_rol_from_request()

        if not usuario:
            return jsonify({"error": "Usuario no enviado"}), 400

        usuario_norm = (usuario or "").strip().lower()

        consultor_login = (
            Consultor.query.options(
                joinedload(Consultor.rol_obj),
                joinedload(Consultor.equipo_obj),
            )
            .filter(func.lower(Consultor.usuario) == usuario_norm)
            .first()
        )
        if not consultor_login:
            return jsonify({"error": "Consultor no encontrado"}), 404

        scope, val = _scope_for_graficos(consultor_login, rol_req)

        C = aliased(Consultor)
        E = aliased(Equipo)

        # horas: usa total_horas si lo llenas, si no, usa tiempo_invertido
        horas_col = func.coalesce(Registro.total_horas, Registro.tiempo_invertido, 0)

        q = (
            db.session.query(
                func.coalesce(Ocupacion.nombre, "SIN OCUPACIÓN").label("ocupacion"),
                func.coalesce(func.sum(horas_col), 0).label("horas"),
            )
            .select_from(Registro)
            .outerjoin(C, func.lower(Registro.usuario_consultor) == func.lower(C.usuario))
            .outerjoin(E, C.equipo_id == E.id)
            .outerjoin(Ocupacion, Registro.ocupacion_id == Ocupacion.id)
        )

        if scope == "SELF":
            q = q.filter(func.lower(Registro.usuario_consultor) == usuario_norm)
        elif scope == "TEAM":
            if not int(val or 0):
                return jsonify({"error": "Consultor sin equipo asignado"}), 403
            q = q.filter(C.equipo_id == int(val))
        # ALL: sin filtro

        # (opcional) filtros por fecha
        desde = (request.args.get("desde") or "").strip()
        hasta = (request.args.get("hasta") or "").strip()
        if desde and hasta:
            q = q.filter(Registro.fecha.between(desde, hasta))
        elif desde:
            q = q.filter(Registro.fecha >= desde)
        elif hasta:
            q = q.filter(Registro.fecha <= hasta)

        q = q.group_by(func.coalesce(Ocupacion.nombre, "SIN OCUPACIÓN"))
        rows = q.order_by(func.sum(horas_col).desc()).all()

        out = [{"ocupacion": r.ocupacion, "horas": float(r.horas or 0)} for r in rows]
        return jsonify(out), 200

    except Exception as e:
        current_app.logger.exception("❌ Error en /horas-ocupacion")
        return jsonify({"error": str(e)}), 500

@bp.route('/equipos/<int:equipo_id>/permisos/codigo/<string:codigo>', methods=['DELETE'])
def quitar_permiso_equipo_por_codigo(equipo_id, codigo):
    p = Permiso.query.filter_by(codigo=codigo.upper()).first()
    if not p:
        return jsonify({"mensaje": "Permiso no encontrado"}), 404

    ep = EquipoPermiso.query.filter_by(equipo_id=equipo_id, permiso_id=p.id).first()
    if not ep:
        return jsonify({"mensaje": "Permiso no estaba asignado"}), 404

    db.session.delete(ep)
    db.session.commit()
    return jsonify({"mensaje": "Permiso removido"}), 200

@bp.route('/consultores/<int:consultor_id>/permisos/codigo/<string:codigo>', methods=['DELETE'])
def quitar_permiso_consultor_por_codigo(consultor_id, codigo):
    p = Permiso.query.filter_by(codigo=codigo.upper()).first()
    if not p:
        return jsonify({"mensaje": "Permiso no encontrado"}), 404

    cp = ConsultorPermiso.query.filter_by(consultor_id=consultor_id, permiso_id=p.id).first()
    if not cp:
        return jsonify({"mensaje": "Permiso no estaba asignado"}), 404

    db.session.delete(cp)
    db.session.commit()
    return jsonify({"mensaje": "Permiso removido"}), 200

# ========== ROLES ==========

@bp.route('/roles', methods=['POST'])
@permission_required("ROLES_ADMIN")
def crear_rol():
    data = request.get_json() or {}
    nombre = data.get("nombre", "").strip().upper()

    if not nombre:
        return jsonify({"mensaje": "Nombre requerido"}), 400

    if Rol.query.filter_by(nombre=nombre).first():
        return jsonify({"mensaje": "El rol ya existe"}), 400

    rol = Rol(nombre=nombre)
    db.session.add(rol)
    db.session.commit()

    return jsonify({"mensaje": "Rol creado", "rol": rol.to_dict()}), 201

@bp.route('/roles/<int:id>', methods=['PUT'])
@permission_required("ROLES_ADMIN")
def editar_rol(id):
    rol = Rol.query.get_or_404(id)
    data = request.get_json() or {}

    nuevo_nombre = data.get("nombre", "").strip().upper()

    if Rol.query.filter(Rol.id != id, Rol.nombre == nuevo_nombre).first():
        return jsonify({"mensaje": "Ese nombre ya lo tiene otro rol"}), 400

    rol.nombre = nuevo_nombre
    db.session.commit()

    return jsonify({"mensaje": "Rol actualizado", "rol": rol.to_dict()}), 200

@bp.route('/roles/<int:id>', methods=['DELETE'])
@permission_required("ROLES_ADMIN")
def eliminar_rol(id):
    rol = Rol.query.get_or_404(id)

    asignados = Consultor.query.filter_by(rol_id=id).count()
    if asignados > 0:
        return jsonify({"mensaje": "No se puede eliminar: tiene consultores asociados"}), 400

    db.session.delete(rol)
    db.session.commit()

    return jsonify({"mensaje": "Rol eliminado"}), 200

@bp.route('/consultores/<int:id>/rol', methods=['PUT'])
def asignar_rol_consultor(id):
    consultor = Consultor.query.get_or_404(id)
    data = request.get_json() or {}
    rol_id = data.get("rol_id")

    if not rol_id:
        return jsonify({"mensaje": "rol_id requerido"}), 400

    rol = Rol.query.get(rol_id)
    if not rol:
        return jsonify({"mensaje": "Rol no existe"}), 404

    consultor.rol_id = rol_id
    db.session.commit()

    return jsonify({"mensaje": "Rol asignado correctamente"}), 200

# ========== EQUIPOS ==========

@bp.route('/equipos', methods=['POST'])
@permission_required("EQUIPOS_ADMIN")
def crear_equipo():
    data = request.get_json() or {}
    nombre = data.get("nombre", "").strip().upper()

    if not nombre:
        return jsonify({"mensaje": "Nombre requerido"}), 400

    if Equipo.query.filter_by(nombre=nombre).first():
        return jsonify({"mensaje": "El equipo ya existe"}), 400

    eq = Equipo(nombre=nombre)
    db.session.add(eq)
    db.session.commit()

    return jsonify(eq.to_dict()), 201

@bp.route('/equipos/<int:id>', methods=['PUT'])
@permission_required("EQUIPOS_ADMIN")
def editar_equipo(id):
    equipo = Equipo.query.get_or_404(id)
    data = request.get_json() or {}

    nuevo = data.get("nombre", "").strip().upper()

    if Equipo.query.filter(Equipo.id != id, Equipo.nombre == nuevo).first():
        return jsonify({"mensaje": "Ya existe otro equipo con ese nombre"}), 400

    equipo.nombre = nuevo
    db.session.commit()

    return jsonify(equipo.to_dict()), 200

@bp.route('/equipos/<int:id>', methods=['DELETE'])
@permission_required("EQUIPOS_ADMIN")
def eliminar_equipo(id):
    equipo = Equipo.query.get_or_404(id)

    asignados = Consultor.query.filter_by(equipo_id=id).count()
    if asignados > 0:
        return jsonify({"mensaje": "No se puede eliminar: tiene consultores asignados"}), 400

    db.session.delete(equipo)
    db.session.commit()

    return jsonify({"mensaje": "Equipo eliminado"}), 200

@bp.route('/equipos/<int:equipo_id>/consultores', methods=['GET'])
def consultores_por_equipo(equipo_id):
    try:
        consultores = (
            Consultor.query
            .options(
                joinedload(Consultor.rol_obj),
                joinedload(Consultor.equipo_obj)
            )
            .filter(Consultor.equipo_id == equipo_id)
            .order_by(Consultor.nombre.asc())
            .all()
        )

        data = []
        for c in consultores:
            data.append({
                "id": c.id,
                "usuario": c.usuario,
                "nombre": c.nombre,
                "rol": c.rol_obj.nombre if c.rol_obj else None,
                "equipo": c.equipo_obj.nombre if c.equipo_obj else None,
            })

        return jsonify(data), 200

    except Exception as e:
        print("❌ Error consultores_por_equipo:", e)
        return jsonify({"error": "Error interno del servidor"}), 500


@bp.route('/consultores/<int:id>/equipo', methods=['PUT'])
def asignar_equipo_consultor(id):
    cons = Consultor.query.get_or_404(id)
    data = request.get_json() or {}
    equipo_id = data.get("equipo_id")

    if not equipo_id:
        return jsonify({"mensaje": "equipo_id requerido"}), 400

    # 🔒 VALIDACIÓN CRÍTICA
    if cons.equipo_id is not None:
        return jsonify({
            "mensaje": "El consultor ya está asignado a un equipo"
        }), 409

    eq = Equipo.query.get(equipo_id)
    if not eq:
        return jsonify({"mensaje": "Equipo no existe"}), 404

    cons.equipo_id = equipo_id
    db.session.commit()

    return jsonify({"mensaje": "Equipo asignado correctamente"}), 200


@bp.route('/consultores/<int:id>/equipo/remove', methods=['PUT'])
def remover_consultor_equipo(id):
    cons = Consultor.query.get_or_404(id)
    cons.equipo_id = None
    db.session.commit()
    return jsonify({"mensaje": "Consultor removido del equipo"}), 200

# ========== IMPORTAR REGISTROS DESDE EXCEL ==========
@bp.route('/registro/import-excel', methods=['POST'])
def importar_registro_excel():
    if 'file' not in request.files:
        return jsonify({"mensaje": "No se envió archivo"}), 400

    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({"mensaje": "Archivo vacío"}), 400

    try:
        import math
        import unicodedata

        def normalize_float(value):
            if value is None:
                return None
            if isinstance(value, float) and math.isnan(value):
                return None
            try:
                return float(value)
            except:
                return None

        def normalize_str(value):
            if value is None:
                return None
            v = str(value).strip()
            if v == "" or v.upper() in ["NA", "N/A", "NAN"]:
                return None
            return v

        # normaliza nombre de columna: quita tildes, mayus, espacios dobles
        def norm_col(s: str) -> str:
            s = str(s or "").strip()
            s = unicodedata.normalize("NFD", s)
            s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
            s = s.upper()
            s = " ".join(s.split())
            return s

        # 1) Leer Excel
        df = pd.read_excel(file, engine="openpyxl")
        df = df.where(pd.notnull(df), None)

        # 2) Normalizar headers (sin tocar data)
        original_cols = list(df.columns)
        norm_map = { norm_col(c): c for c in original_cols }  # NORMALIZADA -> ORIGINAL

        # 3) Mapa esperado (NORMALIZADO)
        expected = {
            "FECHA": "fecha",
            "MODULO": "modulo_nombre",
            "EQUIPO": "equipo",
            "CLIENTE": "cliente",
            "NRO CASO CLIENTE": "nro_caso_cliente",
            "NRO CASO INTERNO": "nro_caso_interno",
            "NRO CASO ESCALADO SAP": "nro_caso_escalado_sap",
            "TIPO TAREA AZURE": "tipo_tarea_raw",
            "CONSULTOR": "consultor",
            "HORA INICIO": "hora_inicio",
            "HORA FIN": "hora_fin",
            "TIEMPO INVERTIDO": "tiempo_invertido",
            "TIEMPO FACTURABLE": "tiempo_facturable",
            "ONCALL": "oncall",
            "DESBORDE": "desborde",
            "HORAS ADICIONALES": "horas_adicionales",
            "DESCRIPCION": "descripcion",  # soporta "Descripción" o "Descripcion"
            "DESCRIPCIÓN": "descripcion",

            # ✅ extras (si existen en el archivo)
            "CONSOLIDADO CON EL CLIENTE": "consolidado_cliente",
            "DIA": "dia",
            "MES": "mes",
            "AÑO": "anio",
            "ANO": "anio",
            "OCUPACION AZURE": "ocupacion_azure",
            "TAREA AZURE": "tarea_azure",
            "HORAS CONVERTIDAS": "horas_convertidas",
            "PROMEDIO": "promedio",
            "EXTEMPORANEO": "extemporaneo",
            "EXTEMPORÁNEO": "extemporaneo",
        }

        # 4) Renombrar usando equivalencias (solo las que existan)
        rename_real = {}
        for k_norm, new_name in expected.items():
            col_original = norm_map.get(norm_col(k_norm))
            if col_original:
                rename_real[col_original] = new_name

        df = df.rename(columns=rename_real)

        # 5) Parser tipo tarea
        def parse_tipo_tarea(valor):
            if not valor:
                return None, None
            valor = str(valor).strip()
            if '-' in valor:
                codigo, nombre = valor.split('-', 1)
                return codigo.strip(), nombre.strip()
            return valor.strip(), None

        # 6) Construcción de registros
        registros = []
        for _, row in df.iterrows():
            codigo_tarea, nombre_tarea = parse_tipo_tarea(row.get("tipo_tarea_raw"))

            registros.append(
                RegistroExcel(
                    fecha=norm_fecha(row.get("fecha")),
                    modulo_nombre=normalize_str(row.get("modulo_nombre")),
                    equipo=(normalize_str(row.get("equipo")) or "").upper() or None,
                    cliente=normalize_str(row.get("cliente")),

                    nro_caso_cliente=normalize_str(row.get("nro_caso_cliente")),
                    nro_caso_interno=normalize_str(row.get("nro_caso_interno")),
                    nro_caso_escalado_sap=normalize_str(row.get("nro_caso_escalado_sap")),

                    tipo_tarea_azure=normalize_str(codigo_tarea),
                    tipo_tarea_nombre=normalize_str(nombre_tarea),

                    consultor=(normalize_str(row.get("consultor")) or "").lower() or None,

                    hora_inicio=norm_hora(row.get("hora_inicio")),
                    hora_fin=norm_hora(row.get("hora_fin")),

                    tiempo_invertido=normalize_float(row.get("tiempo_invertido")),
                    tiempo_facturable=normalize_float(row.get("tiempo_facturable")),

                    oncall=normalize_str(row.get("oncall")),
                    desborde=normalize_str(row.get("desborde")),
                    horas_adicionales=normalize_str(row.get("horas_adicionales")),

                    descripcion=normalize_str(row.get("descripcion")),

                    # ✅ extras (si tu modelo RegistroExcel los tiene)
                    consolidado_cliente=normalize_str(row.get("consolidado_cliente")),
                    dia=normalize_str(row.get("dia")),
                    mes=normalize_str(row.get("mes")),
                    anio=normalize_str(row.get("anio")),
                    ocupacion_azure=normalize_str(row.get("ocupacion_azure")),
                    tarea_azure=normalize_str(row.get("tarea_azure")),
                    horas_convertidas=normalize_float(row.get("horas_convertidas")),
                    promedio=normalize_float(row.get("promedio")),
                    extemporaneo=normalize_str(row.get("extemporaneo")),
                )
            )

        db.session.bulk_save_objects(registros)
        db.session.commit()

        return jsonify({
            "mensaje": "Excel importado correctamente",
            "total_registros": len(registros)
        }), 200

    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        return jsonify({"mensaje": "Error importando Excel", "error": str(e)}), 500



@bp.route('/registros/importar-excel/preview', methods=['POST'])
def preview_import_excel():
    file = request.files.get('file')
    if not file:
        return jsonify({"error": "No se envió archivo"}), 400

    df = pd.read_excel(file, engine="openpyxl")
    df = df.where(pd.notnull(df), None)

    total = len(df)
    validos = df[df["Tipo Tarea Azure"].notna()].shape[0]
    errores = df[df["Tipo Tarea Azure"].isna()].to_dict(orient="records")

    return jsonify({
        "total": total,
        "validos": validos,
        "errores": errores[:20]
    })


@bp.route('/registros/importar-excel/commit', methods=['POST'])
def commit_import_excel():
    registros = request.json.get("registros", [])
    objs = [RegistroExcel(**r) for r in registros]

    db.session.bulk_save_objects(objs)
    db.session.commit()

    return jsonify({
        "mensaje": "Registros insertados",
        "total": len(objs)
    })

@bp.route('/registros/export', methods=['GET'])
def export_registros():
    try:
        usuario = _get_usuario_from_request()
        rol_req = _get_rol_from_request()

        if not usuario:
            return jsonify({'error': 'Usuario no enviado'}), 400

        usuario_norm = (usuario or "").strip().lower()

        consultor_login = (
            Consultor.query.options(
                joinedload(Consultor.rol_obj),
                joinedload(Consultor.equipo_obj),
            )
            .filter(func.lower(Consultor.usuario) == usuario_norm)
            .first()
        )
        if not consultor_login:
            return jsonify({'error': 'Consultor no encontrado'}), 404

        scope, val = scope_for(consultor_login, rol_req)

        C = aliased(Consultor)
        E = aliased(Equipo)

        q = (
            Registro.query
            .options(
                joinedload(Registro.consultor).joinedload(Consultor.equipo_obj),
                joinedload(Registro.tarea),
                joinedload(Registro.ocupacion),
                joinedload(Registro.proyecto),
                joinedload(Registro.fase_proyecto),
            )
            .outerjoin(C, func.lower(Registro.usuario_consultor) == func.lower(C.usuario))
            .outerjoin(E, C.equipo_id == E.id)
        )

        # Scope
        if scope == "SELF":
            q = q.filter(func.lower(Registro.usuario_consultor) == usuario_norm)

        elif scope == "TEAM":
            if not int(val or 0):
                return jsonify({'error': 'Consultor sin equipo asignado'}), 403
            q = q.filter(C.equipo_id == int(val))

        elif scope == "ROLE_POOL":
            if not int(val or 0):
                return jsonify({'error': 'Consultor sin rol asignado'}), 403
            q = q.filter(C.rol_id == int(val))

        # Filtros
        filtro_id = (request.args.get("id") or "").strip()
        filtro_fecha = (request.args.get("fecha") or "").strip()
        filtro_equipo = (request.args.get("equipo") or "").strip().upper()
        filtro_mes = (request.args.get("mes") or "").strip()
        filtro_anio = (request.args.get("anio") or "").strip()
        filtro_nro_caso = (request.args.get("nroCasoCliente") or "").strip()

        filtro_clientes = [v.strip() for v in _get_list_arg("cliente") if str(v).strip()]
        filtro_consultores = [v.strip() for v in _get_list_arg("consultor") if str(v).strip()]
        filtro_horas_adic = [str(v).strip().upper() for v in _get_list_arg("horasAdicionales") if str(v).strip()]
        filtro_tarea_ids = []
        filtro_ocupacion_ids = []

        for v in _get_list_arg("tarea_id"):
            try:
                filtro_tarea_ids.append(int(v))
            except Exception:
                pass

        for v in _get_list_arg("ocupacion_id"):
            try:
                filtro_ocupacion_ids.append(int(v))
            except Exception:
                pass

        filtro_clientes = list(dict.fromkeys(filtro_clientes))
        filtro_consultores = list(dict.fromkeys(filtro_consultores))
        filtro_horas_adic = list(dict.fromkeys(filtro_horas_adic))
        filtro_tarea_ids = list(dict.fromkeys(filtro_tarea_ids))
        filtro_ocupacion_ids = list(dict.fromkeys(filtro_ocupacion_ids))

        if filtro_id and filtro_id.isdigit():
            q = q.filter(Registro.id == int(filtro_id))

        if filtro_fecha:
            q = q.filter(Registro.fecha == filtro_fecha)

        if filtro_clientes:
            q = q.filter(Registro.cliente.in_(filtro_clientes))

        if filtro_consultores:
            q = q.filter(C.nombre.in_(filtro_consultores))

        if filtro_equipo:
            if scope == "TEAM":
                eq_login = (consultor_login.equipo_obj.nombre or "").strip().upper() if consultor_login.equipo_obj else ""
                if filtro_equipo != eq_login:
                    return jsonify({'error': 'No autorizado para consultar otro equipo'}), 403
            q = q.filter(func.upper(E.nombre) == filtro_equipo)

        if filtro_mes:
            try:
                q = q.filter(extract("month", cast(Registro.fecha, db.Date)) == int(filtro_mes))
            except Exception:
                pass

        if filtro_anio:
            try:
                q = q.filter(extract("year", cast(Registro.fecha, db.Date)) == int(filtro_anio))
            except Exception:
                pass

        if filtro_nro_caso:
            q = q.filter(Registro.nro_caso_cliente.ilike(f"%{filtro_nro_caso}%"))

        if filtro_horas_adic:
            q = q.filter(func.upper(Registro.horas_adicionales).in_(filtro_horas_adic))

        if filtro_tarea_ids:
            q = q.filter(Registro.tarea_id.in_(filtro_tarea_ids))

        if filtro_ocupacion_ids:
            q = q.filter(Registro.ocupacion_id.in_(filtro_ocupacion_ids))

        registros = q.order_by(Registro.fecha.desc(), Registro.id.desc()).all()

        data = []
        for r in registros:
            tarea = getattr(r, "tarea", None)
            ocup = getattr(r, "ocupacion", None)

            if tarea and getattr(tarea, "codigo", None) and getattr(tarea, "nombre", None):
                tipo_tarea_str = f"{tarea.codigo} - {tarea.nombre}"
            else:
                tipo_tarea_str = (r.tipo_tarea or "").strip() or None

            equipo_nombre = None
            if r.consultor and getattr(r.consultor, "equipo_obj", None):
                equipo_nombre = (r.consultor.equipo_obj.nombre or "").strip().upper()

            proyecto = getattr(r, "proyecto", None)
            fase_proyecto = getattr(r, "fase_proyecto", None)

            data.append({
                "id": r.id,
                "fecha": r.fecha,
                "modulo": r.modulo,
                "cliente": r.cliente,
                "equipo": equipo_nombre or (r.equipo or "").strip().upper() or "SIN EQUIPO",
                "nroCasoCliente": r.nro_caso_cliente,
                "nroCasoInterno": r.nro_caso_interno,
                "nroCasoEscaladoSap": r.nro_caso_escalado,

                "ocupacion_id": r.ocupacion_id,
                "ocupacion_codigo": ocup.codigo if ocup else None,
                "ocupacion_nombre": ocup.nombre if ocup else None,

                "tarea_id": r.tarea_id,
                "tipoTarea": tipo_tarea_str,
                "tarea": {
                    "id": tarea.id,
                    "codigo": getattr(tarea, "codigo", None),
                    "nombre": getattr(tarea, "nombre", None),
                } if tarea else None,

                "consultor": r.consultor.nombre if r.consultor else None,
                "usuario_consultor": (r.usuario_consultor or "").strip().lower(),

                "horaInicio": r.hora_inicio,
                "horaFin": r.hora_fin,
                "tiempoInvertido": r.tiempo_invertido,
                "tiempoFacturable": r.tiempo_facturable,
                "horasAdicionales": _calcular_horas_adicionales_por_horario(
                    r.hora_inicio,
                    r.hora_fin,
                    getattr(r, "horario_trabajo", None),
                    equipo_nombre or getattr(r, "equipo", None),
                    r.horas_adicionales,
                ),
                "horarioTrabajo": _normalizar_horario_trabajo_por_equipo(
                    getattr(r, "horario_trabajo", None),
                    equipo_nombre or getattr(r, "equipo", None),
                ),
                "horario_trabajo": _normalizar_horario_trabajo_por_equipo(
                    getattr(r, "horario_trabajo", None),
                    equipo_nombre or getattr(r, "equipo", None),
                ),
                "descripcion": r.descripcion,
                "totalHoras": r.total_horas,

                "bloqueado": bool(r.bloqueado),
                "oncall": r.oncall,
                "desborde": r.desborde,
                "actividadMalla": r.actividad_malla,

                "proyecto_id": r.proyecto_id,
                "fase_proyecto_id": r.fase_proyecto_id,

                "proyecto_codigo": proyecto.codigo if proyecto else None,
                "proyecto_nombre": proyecto.nombre if proyecto else None,
                "proyecto_fase": fase_proyecto.nombre if fase_proyecto else None,
            })

        return jsonify({
            "data": data,
            "total": len(data)
        }), 200

    except Exception as e:
        app.logger.exception("❌ Error en /registros/export")
        return jsonify({'error': str(e)}), 500

@bp.route("/horarios-permitidos", methods=["GET"])
def horarios_permitidos():
    usuario = (request.args.get("usuario") or "").strip().lower()

    
    if not usuario:
        return jsonify({
            "equipo": "",
            "dias": "",
            "horarios": []
        }), 200

    consultor = (
        Consultor.query
        .filter(func.lower(Consultor.usuario) == usuario)
        .first()
    )

    
    if not consultor:
        return jsonify({
            "equipo": "",
            "dias": "",
            "horarios": []
        }), 200

    equipo = (consultor.equipo_obj.nombre if consultor.equipo_obj else "").strip().upper()

    horarios_funcional = FUNCIONAL_HORARIOS_PERMITIDOS

    if equipo == "BASIS":
        horarios = [h.rango for h in Horario.query.order_by(Horario.rango).all()]
        dias = "DOMINGO_A_DOMINGO"
    else:
        horarios = horarios_funcional
        dias = "LUNES_A_VIERNES"

    return jsonify({
        "equipo": equipo,
        "dias": dias,
        "horarios": horarios
    }), 200

@bp.route('/registros/filtros', methods=['GET'])
def registros_filtros():
    try:
        usuario = _get_usuario_from_request()
        rol_req = _get_rol_from_request()

        if not usuario:
            return jsonify({'error': 'Usuario no enviado'}), 400

        usuario_norm = (usuario or "").strip().lower()

        consultor_login = (
            Consultor.query.options(
                joinedload(Consultor.rol_obj),
                joinedload(Consultor.equipo_obj),
            )
            .filter(func.lower(Consultor.usuario) == usuario_norm)
            .first()
        )
        if not consultor_login:
            return jsonify({'error': 'Consultor no encontrado'}), 404

        scope, val = scope_for(consultor_login, rol_req)

        C = aliased(Consultor)
        E = aliased(Equipo)

        base = (
            db.session.query(
                Registro.id,
                Registro.cliente,
                C.nombre.label("consultor"),
                E.nombre.label("equipo")
            )
            .select_from(Registro)
            .outerjoin(C, func.lower(Registro.usuario_consultor) == func.lower(C.usuario))
            .outerjoin(E, C.equipo_id == E.id)
        )

        if scope == "SELF":
            base = base.filter(func.lower(Registro.usuario_consultor) == usuario_norm)

        elif scope == "TEAM":
            if not int(val or 0):
                return jsonify({'error': 'Consultor sin equipo asignado'}), 403
            base = base.filter(C.equipo_id == int(val))

        elif scope == "ROLE_POOL":
            if not int(val or 0):
                return jsonify({'error': 'Consultor sin rol asignado'}), 403
            base = base.filter(C.rol_id == int(val))

        rows = base.all()

        consultores = sorted({str(r.consultor).strip() for r in rows if r.consultor})
        clientes = sorted({str(r.cliente).strip() for r in rows if r.cliente})
        equipos = sorted({str(r.equipo).strip().upper() for r in rows if r.equipo})

        return jsonify({
            "consultores": consultores,
            "clientes": clientes,
            "equipos": equipos,
            "total": len(rows)
        }), 200

    except Exception as e:
        app.logger.exception("❌ Error en /registros/filtros")
        return jsonify({'error': str(e)}), 500
    
@bp.route('/registros/conteos', methods=['GET'])
def registros_conteos():
    try:
        usuario = _get_usuario_from_request()
        rol_req = _get_rol_from_request()

        if not usuario:
            return jsonify({'error': 'Usuario no enviado'}), 400

        usuario_norm = (usuario or "").strip().lower()

        consultor_login = (
            Consultor.query.options(
                joinedload(Consultor.rol_obj),
                joinedload(Consultor.equipo_obj),
            )
            .filter(func.lower(Consultor.usuario) == usuario_norm)
            .first()
        )
        if not consultor_login:
            return jsonify({'error': 'Consultor no encontrado'}), 404

        scope, val = scope_for(consultor_login, rol_req)

        C = aliased(Consultor)
        E = aliased(Equipo)

        q = (
            db.session.query(
                E.nombre.label("equipo"),
                func.count(Registro.id).label("count")
            )
            .select_from(Registro)
            .outerjoin(C, func.lower(Registro.usuario_consultor) == func.lower(C.usuario))
            .outerjoin(E, C.equipo_id == E.id)
        )

        if scope == "SELF":
            q = q.filter(func.lower(Registro.usuario_consultor) == usuario_norm)

        elif scope == "TEAM":
            if not int(val or 0):
                return jsonify({'error': 'Consultor sin equipo asignado'}), 403
            q = q.filter(C.equipo_id == int(val))

        elif scope == "ROLE_POOL":
            if not int(val or 0):
                return jsonify({'error': 'Consultor sin rol asignado'}), 403
            q = q.filter(C.rol_id == int(val))

        rows = q.group_by(E.nombre).all()

        total_q = (
            db.session.query(func.count(Registro.id))
            .select_from(Registro)
            .outerjoin(C, func.lower(Registro.usuario_consultor) == func.lower(C.usuario))
        )

        if scope == "SELF":
            total_q = total_q.filter(func.lower(Registro.usuario_consultor) == usuario_norm)

        elif scope == "TEAM":
            total_q = total_q.filter(C.equipo_id == int(val))

        elif scope == "ROLE_POOL":
            total_q = total_q.filter(C.rol_id == int(val))

        total = int(total_q.scalar() or 0)

        equipos = [
            {
                "equipo": (r.equipo or "SIN EQUIPO").strip().upper(),
                "count": int(r.count or 0)
            }
            for r in rows
        ]
        equipos.sort(key=lambda x: x["equipo"])

        return jsonify({
            "total": total,
            "equipos": equipos
        }), 200

    except Exception as e:
        app.logger.exception("❌ Error en /registros/conteos")
        return jsonify({'error': str(e)}), 500
    


@bp.route("/resumen-calendario", methods=["GET"])
def resumen_calendario():
    try:
        usuario = _get_usuario_from_request()
        rol_req = _get_rol_from_request()

        if not usuario:
            return jsonify({"error": "Usuario no enviado"}), 400

        usuario_norm = (usuario or "").strip().lower()

        consultor_login = (
            Consultor.query
            .options(joinedload(Consultor.rol_obj), joinedload(Consultor.equipo_obj))
            .filter(func.lower(Consultor.usuario) == usuario_norm)
            .first()
        )
        if not consultor_login:
            return jsonify({"error": "Consultor no encontrado"}), 404

        scope, val = _scope_for_graficos(consultor_login, rol_req)

        # filtros de fecha opcionales
        desde = (request.args.get("desde") or "").strip()
        hasta = (request.args.get("hasta") or "").strip()

        # filtro opcional de equipo (ADMIN global), pero si es TEAM solo permite su equipo
        equipo_filter = (request.args.get("equipo") or "").strip().upper()
        if equipo_filter and scope == "TEAM":
            eq_login = (consultor_login.equipo_obj.nombre or "").strip().upper() if consultor_login.equipo_obj else ""
            if equipo_filter != eq_login:
                return jsonify({'error': 'No autorizado para consultar otro equipo'}), 403

        # base: traer usuario_consultor, nombre consultor, fecha, suma horas
        q = (
            db.session.query(
                func.lower(Registro.usuario_consultor).label("usuario_consultor"),
                Consultor.nombre.label("consultor"),
                Registro.fecha.label("fecha"),
                func.coalesce(func.sum(Registro.total_horas), 0).label("total_horas"),
            )
            .select_from(Registro)
            .join(Consultor, func.lower(Registro.usuario_consultor) == func.lower(Consultor.usuario))
            .outerjoin(Equipo, Consultor.equipo_id == Equipo.id)
        )

        # aplicar scope
        if scope == "SELF":
            q = q.filter(func.lower(Registro.usuario_consultor) == usuario_norm)
        elif scope == "TEAM":
            q = q.filter(Consultor.equipo_id == int(val))
        elif scope == "ROLE_POOL":
            q = q.filter(Consultor.rol_id == int(val))

        # aplicar equipo_filter si viene (solo realmente útil en ADMIN global)
        if equipo_filter:
            q = q.filter(func.upper(Equipo.nombre) == equipo_filter)

        # fecha filter
        if desde and hasta:
            q = q.filter(Registro.fecha.between(desde, hasta))
        elif desde:
            q = q.filter(Registro.fecha >= desde)
        elif hasta:
            q = q.filter(Registro.fecha <= hasta)

        # agrupar por consultor+fecha
        q = q.group_by(func.lower(Registro.usuario_consultor), Consultor.nombre, Registro.fecha)
        q = q.order_by(Consultor.nombre.asc(), Registro.fecha.asc())

        rows = q.all()

        # armar respuesta agrupada por consultor
        out = {}
        for r in rows:
            key = r.usuario_consultor or "na"
            if key not in out:
                out[key] = {
                    "consultor": r.consultor or r.usuario_consultor or "—",
                    "usuario_consultor": r.usuario_consultor,
                    "registros": []
                }
            out[key]["registros"].append({
                "fecha": r.fecha,
                "total_horas": float(r.total_horas or 0),
            })

        return jsonify(list(out.values())), 200

    except Exception as e:
        app.logger.exception("❌ Error en /resumen-calendario")
        return jsonify({"error": str(e)}), 500


# -------------------------------
#   REPORTES DE HORAS  (DIARIO)
# -------------------------------
@bp.route("/reporte/costos-cliente-dia", methods=["GET"])
def reporte_costos_cliente_dia():
    """
    Pivot diario:
      - filas: fecha
      - columnas: clientes (dinámicas)
      - valores: horas y costo

    Costo:
      Se calcula por consultor usando presupuesto vigente (vr_perfil / horas_base_mes),
      y se suma por (fecha, cliente) para no distorsionar cuando en un mismo día/cliente
      participaron varios consultores.

    Filtros opcionales:
      ?desde=YYYY-MM-DD&hasta=YYYY-MM-DD&equipo=BASIS&modulo=FI&cliente=HITSS&consultor=andres
    """
    try:
        desde = (request.args.get("desde") or "").strip()
        hasta = (request.args.get("hasta") or "").strip()
        equipo_filter = (request.args.get("equipo") or "").strip().upper()
        modulo_filter = (request.args.get("modulo") or "").strip().upper()
        cliente_filter = (request.args.get("cliente") or "").strip().upper()
        consultor_filter = (request.args.get("consultor") or "").strip().lower()

        # ----------------------------------------------------------
        # 1) Agregado base por (fecha, cliente, consultor)
        # ----------------------------------------------------------
        q = (
            db.session.query(
                Registro.fecha.label("fecha"),
                Registro.cliente.label("cliente"),
                Registro.modulo.label("modulo"),

                Consultor.id.label("consultor_id"),
                Consultor.nombre.label("consultor_nombre"),
                Consultor.usuario.label("consultor_usuario"),

                Equipo.nombre.label("equipo"),
                func.coalesce(
                    func.sum(
                        func.coalesce(Registro.tiempo_invertido, Registro.total_horas, 0)
                    ),
                    0
                ).label("horas"),
            )
            .select_from(Registro)
            .join(Consultor, func.lower(Registro.usuario_consultor) == func.lower(Consultor.usuario))
            .outerjoin(Equipo, Consultor.equipo_id == Equipo.id)
        )

        # ----------------------------------------------------------
        # 2) Filtros
        # ----------------------------------------------------------
        if desde and hasta:
            q = q.filter(Registro.fecha.between(desde, hasta))
        elif desde:
            q = q.filter(Registro.fecha >= desde)
        elif hasta:
            q = q.filter(Registro.fecha <= hasta)

        if equipo_filter:
            q = q.filter(func.upper(Equipo.nombre) == equipo_filter)

        if modulo_filter:
            q = q.filter(func.upper(Registro.modulo) == modulo_filter)

        if cliente_filter:
            q = q.filter(func.upper(Registro.cliente) == cliente_filter)

        if consultor_filter:
            q = q.filter(func.lower(Consultor.nombre).like(f"%{consultor_filter}%"))

        q = q.group_by(
            Registro.fecha,
            Registro.cliente,
            Registro.modulo,
            Consultor.id,
            Consultor.nombre,
            Consultor.usuario,
            Equipo.nombre,
        )

        raw = q.all()

        # ----------------------------------------------------------
        # 3) Columnas dinámicas (clientes) + pivot por fecha
        # ----------------------------------------------------------
        clientes_set = set()
        pivot = {}  # fecha -> obj fila

        # Vamos guardando consultores únicos por fecha
        # pivot[fecha]["consultoresSet"] = set()

        for r in raw:
            fecha = (r.fecha or "").strip()
            cliente = (r.cliente or "SIN CLIENTE").strip()
            clientes_set.add(cliente)

            if fecha not in pivot:
                pivot[fecha] = {
                    "key": fecha,
                    "fecha": fecha,
                    "clientesHoras": {},   # cliente -> horas
                    "clientesCosto": {},   # cliente -> costo
                    "totalHoras": 0.0,
                    "totalCosto": 0.0,
                    "consultoresSet": set(),  # 👈 aquí guardamos nombres únicos por fecha
                }

            # registrar consultor en el set por fecha
            nombre_cons = (r.consultor_nombre or "").strip()
            if nombre_cons:
                pivot[fecha]["consultoresSet"].add(nombre_cons)

        clientes = sorted(list(clientes_set))

        # ----------------------------------------------------------
        # 4) Presupuesto vigente por consultor (vr_perfil / horas_base_mes)
        #    (sin anio/mes porque tu modelo no lo tiene)
        # ----------------------------------------------------------
        presupuesto_map = {}  # consultor_id -> (vr, hb)

        if raw:
            cons_ids = sorted({int(r.consultor_id) for r in raw if r.consultor_id})

            pres_rows = (
                db.session.query(
                    ConsultorPresupuesto.consultor_id,
                    ConsultorPresupuesto.vr_perfil,
                    ConsultorPresupuesto.horas_base_mes,
                )
                .filter(ConsultorPresupuesto.consultor_id.in_(cons_ids))
                .filter(ConsultorPresupuesto.vigente == True)
                .all()
            )

            for pr in pres_rows:
                cid = int(pr.consultor_id)
                vr = float(pr.vr_perfil or 0)
                hb = float(pr.horas_base_mes or 0)
                presupuesto_map[cid] = (vr, hb)

        # ----------------------------------------------------------
        # 5) Agregar horas y costo por (fecha, cliente)
        #    sumando consultores si aplica
        # ----------------------------------------------------------
        temp = {}  # (fecha, cliente) -> {"horas": x, "costo": y}

        for r in raw:
            fecha = (r.fecha or "").strip()
            cliente = (r.cliente or "SIN CLIENTE").strip()
            horas = float(r.horas or 0.0)

            consultor_id = int(r.consultor_id) if r.consultor_id else 0
            vr, hb = presupuesto_map.get(consultor_id, (0.0, 0.0))
            valor_hora = round((vr / hb), 2) if hb > 0 else 0.0
            costo = round(horas * valor_hora, 2) if valor_hora > 0 else 0.0

            k = (fecha, cliente)
            if k not in temp:
                temp[k] = {"horas": 0.0, "costo": 0.0}

            temp[k]["horas"] += horas
            temp[k]["costo"] += costo

        # ----------------------------------------------------------
        # 6) Completar filas, asegurar columnas, totales
        # ----------------------------------------------------------
        totales_cliente_horas = {c: 0.0 for c in clientes}
        totales_cliente_costo = {c: 0.0 for c in clientes}
        total_general_horas = 0.0
        total_general_costo = 0.0

        data = []
        for fecha, obj in pivot.items():
            total_h = 0.0
            total_c = 0.0

            for c in clientes:
                v = temp.get((fecha, c), {"horas": 0.0, "costo": 0.0})
                h = round(float(v["horas"] or 0.0), 2)
                co = round(float(v["costo"] or 0.0), 2)

                obj["clientesHoras"][c] = h
                obj["clientesCosto"][c] = co

                totales_cliente_horas[c] += h
                totales_cliente_costo[c] += co

                total_h += h
                total_c += co

            obj["totalHoras"] = round(total_h, 2)
            obj["totalCosto"] = round(total_c, 2)

            # 👇 convertir set -> list y exponer count/list
            consultores_list = sorted(list(obj.get("consultoresSet", set())))
            obj["consultoresCount"] = len(consultores_list)
            obj["consultoresList"] = consultores_list
            obj.pop("consultoresSet", None)  # limpiar

            total_general_horas += obj["totalHoras"]
            total_general_costo += obj["totalCosto"]

            data.append(obj)

        data.sort(key=lambda x: x["fecha"], reverse=True)

        for c in clientes:
            totales_cliente_horas[c] = round(totales_cliente_horas[c], 2)
            totales_cliente_costo[c] = round(totales_cliente_costo[c], 2)

        total_general_horas = round(total_general_horas, 2)
        total_general_costo = round(total_general_costo, 2)

        return jsonify({
            "clientes": clientes,
            "rows": data,
            "totalesClienteHoras": totales_cliente_horas,
            "totalesClienteCosto": totales_cliente_costo,
            "totalGeneralHoras": total_general_horas,
            "totalGeneralCosto": total_general_costo,
        }), 200

    except Exception as e:
        current_app.logger.exception("❌ Error en /reporte/costos-cliente-dia")
        return jsonify({"error": str(e)}), 500

    
def _norm_name(s: str) -> str:
    s = (s or "").strip().upper()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("Á","A").replace("É","E").replace("Í","I").replace("Ó","O").replace("Ú","U").replace("Ñ","N")
    s = re.sub(r"[^A-Z0-9 ,.-]", "", s)
    return s


def _norm_doc(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"[^\d]", "", s)
    return s


def _parse_money_to_decimal(val) -> Decimal:
    if val is None:
        return Decimal("0.00")

    if isinstance(val, (int, float, Decimal)):
        try:
            return Decimal(str(val)).quantize(Decimal("0.01"))
        except InvalidOperation:
            return Decimal("0.00")

    s = str(val).strip()
    if not s:
        return Decimal("0.00")

    s = s.replace("$", "").strip()
    s = re.sub(r"[^\d,\.]", "", s)

    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        if "." in s and "," not in s:
            s = s.replace(".", "")
        if "," in s and "." not in s:
            s = s.replace(",", ".")

    try:
        return Decimal(s).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0.00")


def _col_idx(headers: dict, wanted: str):
    w = (wanted or "").strip().upper()
    if not w:
        return None
    if w in headers:
        return headers[w]
    for hk, i in headers.items():
        if w in hk:
            return i
    return None

@bp.route("/presupuestos/consultor", methods=["GET"])
def get_presupuestos_consultor():
    try:
        usuario = _get_usuario_from_request()
        rol_req = _get_rol_from_request()

        if not usuario:
            return jsonify({"error": "Usuario no enviado"}), 400

        usuario_norm = (usuario or "").strip().lower()

        consultor_login = (
            Consultor.query.options(
                joinedload(Consultor.rol_obj),
                joinedload(Consultor.equipo_obj),
            )
            .filter(func.lower(Consultor.usuario) == usuario_norm)
            .first()
        )

        if not consultor_login:
            return jsonify({"error": "Consultor no encontrado"}), 404

        rol_real = ""
        if getattr(consultor_login, "rol_obj", None) and getattr(consultor_login.rol_obj, "nombre", None):
            rol_real = consultor_login.rol_obj.nombre
        else:
            rol_real = getattr(consultor_login, "rol", "") or ""

        if not _is_admin_request(rol_real, consultor_login):
            return jsonify({"error": "No autorizado. Solo administradores."}), 403

        anio_raw = (request.args.get("anio") or "").strip()
        mes_raw = (request.args.get("mes") or "").strip()

        today = date.today()
        anio = int(anio_raw) if anio_raw.isdigit() else today.year
        mes = int(mes_raw) if mes_raw.isdigit() else today.month

        if mes < 1 or mes > 12:
            return jsonify({"error": "Mes inválido"}), 400

        scope, val = scope_for(consultor_login, rol_req)

        q_cons = Consultor.query

        if scope == "TEAM":
            if not int(val or 0):
                return jsonify({"error": "Consultor sin equipo asignado"}), 403
            q_cons = q_cons.filter(Consultor.equipo_id == int(val))

        elif scope == "ROLE_POOL":
            if not int(val or 0):
                return jsonify({"error": "Consultor sin rol asignado"}), 403
            q_cons = q_cons.filter(Consultor.rol_id == int(val))

        elif scope == "ALL":
            pass

        else:
            return jsonify({"error": "Scope no permitido"}), 403

        consultores = q_cons.order_by(Consultor.nombre.asc()).all()

        out = []

        for c in consultores:
            presupuesto = _presupuesto_consultor_mes(c.id, anio, mes)

            vr_perfil = Decimal("0.00")
            horas_base_mes = Decimal("0.00")
            valor_hora = Decimal("0.00")
            dias_habiles_mes = 0

            if presupuesto:
                vr_perfil = presupuesto["vr_perfil"]
                horas_base_mes = presupuesto["horas_base_mes"]
                valor_hora = presupuesto["valor_hora"]
                dias_habiles_mes = presupuesto["dias_habiles_mes"]
            else:
                meta_mes = _meta_horas_en_rango(*_month_bounds_local(anio, mes))
                horas_base_mes = meta_mes["horas"]
                dias_habiles_mes = meta_mes["dias_laborables"]

            out.append({
                "consultorId": c.id,
                "nombre": c.nombre,
                "usuario": c.usuario,
                "vrPerfil": float(vr_perfil),
                "diasHabilesMes": dias_habiles_mes,
                "horasBaseMes": float(horas_base_mes),
                "valorHora": float(valor_hora),
            })

        return jsonify(out), 200

    except Exception as e:
        app.logger.exception("❌ Error en /presupuestos/consultor")
        return jsonify({"error": str(e)}), 500

def _month_bounds_local(anio: int, mes: int):
    start = date(anio, mes, 1)
    if mes == 12:
        end = date(anio + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(anio, mes + 1, 1) - timedelta(days=1)
    return start, end


def _cap_is_standard_workday(d: date, co_holidays=None):
    co_holidays = co_holidays or set()

    # Lunes=0 ... Domingo=6
    if d.weekday() >= 5:
        return False

    if d in co_holidays:
        return False

    return True


def _meta_horas_en_rango(start_date: date, end_date: date):
    years = {start_date.year, end_date.year}
    co_holidays = _cap_colombia_holidays_for_years(years)

    total = Decimal("0.00")
    dias_laborables = 0
    dias_festivos = 0

    cur = start_date
    while cur <= end_date:
        if cur.weekday() < 5:
            if cur in co_holidays:
                dias_festivos += 1
            else:
                dias_laborables += 1
                total += Decimal(str(_cap_meta_hours_for_day(cur, co_holidays)))
        cur += timedelta(days=1)

    return {
        "horas": total.quantize(Decimal("0.01")),
        "dias_laborables": dias_laborables,
        "dias_festivos": dias_festivos,
    }


def _norm_doc(s: str) -> str:
    s = (s or "").strip()
    return re.sub(r"[^\d]", "", s)

    
@bp.route("/me", methods=["GET"])
@auth_required
def me():
    consultor = g.current_user
    permisos = sorted(list(obtener_permisos_finales(consultor)))

    return jsonify({
        "user": {
            "id": consultor.id,
            "usuario": consultor.usuario,
            "nombre": consultor.nombre,
            "rol": consultor.rol_obj.nombre.upper() if consultor.rol_obj else "CONSULTOR",
            "equipo": consultor.equipo_obj.nombre.upper() if consultor.equipo_obj else "SIN EQUIPO",
            "horario": consultor.horario_obj.rango if consultor.horario_obj else "N/D",
            "modulos": [{"id": m.id, "nombre": m.nombre} for m in consultor.modulos],
            "permisos": permisos
        }
    }), 200

## -------------------------------
##   PROYECTOS Y FASES

def _to_bool2(v, default=False):
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v == 1
    s = str(v).strip().lower()
    return s in ("1", "true", "si", "sí", "yes", "y")

def proyecto_fase_to_dict(f: ProyectoFase):
    return {
        "id": f.id,
        "nombre": f.nombre,
        "orden": int(f.orden or 0),
        "activo": bool(f.activo),
    }

def _money_to_json(v):
    if v is None:
        return None
    try:
        return float(v)
    except Exception:
        return None

def _date_to_json(v):
    return v.isoformat() if v else None

def _clean_int_list(values, field_name="ids"):
    """
    Normaliza una lista de ids enviada desde el frontend.
    Acepta strings o números.
    Elimina vacíos y duplicados conservando el orden.
    """
    if values is None:
        return []

    if not isinstance(values, list):
        raise ValueError(f"{field_name} debe ser una lista de ids")

    clean = []

    for value in values:
        if value in ("", None, "null", "None"):
            continue

        try:
            value_int = int(value)
        except Exception:
            raise ValueError(f"{field_name} contiene un id inválido: {value}")

        if value_int > 0:
            clean.append(value_int)

    return list(dict.fromkeys(clean))


def _validar_perfiles_ids(perfiles_ids):
    if not perfiles_ids:
        return False, "Debes seleccionar al menos 1 perfil", []

    perfiles_db = (
        Perfil.query
        .filter(Perfil.id.in_(perfiles_ids))
        .filter(Perfil.activo == True)
        .all()
    )

    found_ids = {int(p.id) for p in perfiles_db}
    missing = [pid for pid in perfiles_ids if int(pid) not in found_ids]

    if missing:
        return False, f"Perfiles no encontrados o inactivos: {missing}", []

    return True, None, perfiles_db


def _validar_modulos_ids(modulos_ids):
    if not modulos_ids:
        return False, "Debes seleccionar al menos 1 módulo", []

    modulos_db = (
        Modulo.query
        .filter(Modulo.id.in_(modulos_ids))
        .all()
    )

    found_ids = {int(m.id) for m in modulos_db}
    missing = [mid for mid in modulos_ids if int(mid) not in found_ids]

    if missing:
        return False, f"Módulos no encontrados: {missing}", []

    return True, None, modulos_db


def _validar_modulos_pertenecen_a_perfiles(perfiles_ids, modulos_ids):
    """
    Regla N:N:
    - Un perfil puede tener muchos módulos.
    - Un módulo puede pertenecer a muchos perfiles.
    La validación correcta es:
    cada módulo seleccionado debe pertenecer al menos a uno de los perfiles seleccionados.
    """
    relaciones = (
        ModuloPerfil.query
        .filter(ModuloPerfil.perfil_id.in_(perfiles_ids))
        .filter(ModuloPerfil.activo == True)
        .all()
    )

    modulos_permitidos = {int(r.modulo_id) for r in relaciones if r.modulo_id}

    modulos_invalidos = [
        int(mid)
        for mid in modulos_ids
        if int(mid) not in modulos_permitidos
    ]

    if modulos_invalidos:
        return False, (
            "Hay módulos que no pertenecen a los perfiles seleccionados: "
            f"{modulos_invalidos}"
        )

    return True, None


def proyecto_perfil_to_dict(x: ProyectoPerfil):
    perfil = getattr(x, "perfil", None)

    return {
        "id": x.id,
        "proyecto_id": x.proyecto_id,
        "perfil_id": x.perfil_id,
        "activo": bool(x.activo),
        "perfil": perfil_to_dict(perfil, include_modulos=True) if perfil else None,
    }


def proyecto_to_dict(p: Proyecto, include_modulos=True, include_fases=True, include_perfiles=True):
    fase = getattr(p, "fase", None)
    cli = getattr(p, "cliente", None)
    opp = getattr(p, "oportunidad", None)

    out = {
        "id": p.id,
        "codigo": p.codigo,
        "nombre": p.nombre,
        "activo": bool(p.activo),

        "cliente_id": p.cliente_id,
        "cliente": cli.to_dict() if cli else None,

        "fase_id": p.fase_id,
        "fase": {
            "id": fase.id,
            "nombre": fase.nombre,
            "orden": int(fase.orden or 0),
            "activo": bool(fase.activo),
        } if fase else None,

        "oportunidad_id": getattr(p, "oportunidad_id", None),
        "oportunidad": opp.to_dict() if opp else None,
        "tipo_negocio": getattr(p, "tipo_negocio", None),

        "codigo_ot_principal": getattr(p, "codigo_ot_principal", None),
        "fecha_inicio_ejecucion": _date_to_json(getattr(p, "fecha_inicio_ejecucion", None)),
        "fecha_fin_ejecucion": _date_to_json(getattr(p, "fecha_fin_ejecucion", None)),
        "fecha_inicio_facturacion": _date_to_json(getattr(p, "fecha_inicio_facturacion", None)),
        "fecha_fin_facturacion": _date_to_json(getattr(p, "fecha_fin_facturacion", None)),

        "moneda": getattr(p, "moneda", "COP"),
        "ingreso_total": _money_to_json(getattr(p, "ingreso_total", None)),
        "costo_objetivo_total": _money_to_json(getattr(p, "costo_objetivo_total", None)),
        "gasto_operativo_total": _money_to_json(getattr(p, "gasto_operativo_total", None)),
        "costo_administrativo_total": _money_to_json(getattr(p, "costo_administrativo_total", None)),
        "margen_objetivo_pct": _money_to_json(getattr(p, "margen_objetivo_pct", None)),
        "ebitda_objetivo": _money_to_json(getattr(p, "ebitda_objetivo", None)),
        "estado_financiero": getattr(p, "estado_financiero", "BORRADOR"),

        "alerta_umbral_1": _money_to_json(getattr(p, "alerta_umbral_1", None)),
        "alerta_umbral_2": _money_to_json(getattr(p, "alerta_umbral_2", None)),
        "alerta_umbral_3": _money_to_json(getattr(p, "alerta_umbral_3", None)),
    }

    if include_fases:
        fases = []
        fases_ids = []

        for pf in (getattr(p, "fases", None) or []):
            fx = getattr(pf, "fase", None)
            if not fx:
                continue

            if not bool(getattr(pf, "activo", True)):
                continue

            orden_final = int((pf.orden if pf.orden is not None else (fx.orden or 0)) or 0)

            fases.append({
                "id": fx.id,
                "fase_id": fx.id,
                "nombre": fx.nombre,
                "orden": orden_final,
                "activo": bool(pf.activo),
                "proyecto_fase_id": pf.id,
                "fase": {
                    "id": fx.id,
                    "nombre": fx.nombre,
                    "orden": int(fx.orden or 0),
                    "activo": bool(fx.activo),
                }
            })

            fases_ids.append(int(fx.id))

        fases.sort(key=lambda x: (x["orden"], x["nombre"]))
        out["fases"] = fases
        out["fases_ids"] = fases_ids

    if include_perfiles:
        perfiles = []
        perfiles_ids = []

        for pp in (getattr(p, "perfiles", None) or []):
            perfil = getattr(pp, "perfil", None)
            if not perfil:
                continue

            if not bool(getattr(pp, "activo", True)):
                continue

            perfiles.append(proyecto_perfil_to_dict(pp))
            perfiles_ids.append(int(pp.perfil_id))

        perfiles.sort(
            key=lambda x: (
                int(x.get("perfil", {}).get("orden", 0) if x.get("perfil") else 0),
                str(x.get("perfil", {}).get("nombre", "") if x.get("perfil") else "")
            )
        )

        out["perfiles"] = perfiles
        out["perfiles_ids"] = perfiles_ids

    if include_modulos:
        mods = []
        modulos_ids = []

        for pm in (getattr(p, "modulos", None) or []):
            modulo = getattr(pm, "modulo", None)
            if not modulo:
                continue

            if not bool(getattr(pm, "activo", True)):
                continue

            mods.append({
                "id": modulo.id,
                "modulo_id": modulo.id,
                "nombre": modulo.nombre,
                "activo": bool(pm.activo),
                "proyecto_modulo_id": pm.id,
                "modulo": {
                    "id": modulo.id,
                    "nombre": modulo.nombre,
                }
            })

            modulos_ids.append(int(modulo.id))

        mods.sort(key=lambda x: x["nombre"])
        out["modulos"] = mods
        out["modulos_ids"] = modulos_ids

    return out

@bp.route("/proyecto-fases", methods=["GET"])
@permission_required("PROYECTOS_VER")
def listar_proyecto_fases():
    fases = ProyectoFase.query.order_by(ProyectoFase.orden.asc(), ProyectoFase.nombre.asc()).all()
    return jsonify([proyecto_fase_to_dict(f) for f in fases]), 200

@bp.route("/proyecto-fases", methods=["POST"])
@permission_required("PROYECTOS_CREAR")
def crear_proyecto_fase():
    data = request.get_json(silent=True) or {}
    nombre = (data.get("nombre") or "").strip()
    orden = data.get("orden", 0)

    if not nombre:
        return jsonify({"mensaje": "nombre requerido"}), 400

    exists = ProyectoFase.query.filter(func.lower(ProyectoFase.nombre) == nombre.lower()).first()
    if exists:
        return jsonify({"mensaje": "La fase ya existe"}), 400

    f = ProyectoFase(
        nombre=nombre,
        orden=int(orden or 0),
        activo=_to_bool2(data.get("activo"), default=True),
    )
    db.session.add(f)
    db.session.commit()

    return jsonify({"mensaje": "Fase creada", "fase": proyecto_fase_to_dict(f)}), 201

@bp.route("/proyecto-fases/<int:id>", methods=["PUT"])
@permission_required("PROYECTOS_EDITAR")
def editar_proyecto_fase(id):
    f = ProyectoFase.query.get_or_404(id)
    data = request.get_json(silent=True) or {}

    nombre = data.get("nombre")
    if nombre is not None:
        nombre = str(nombre).strip()
        if not nombre:
            return jsonify({"mensaje": "nombre inválido"}), 400

        dupe = ProyectoFase.query.filter(
            ProyectoFase.id != id,
            func.lower(ProyectoFase.nombre) == nombre.lower()
        ).first()
        if dupe:
            return jsonify({"mensaje": "Ya existe otra fase con ese nombre"}), 400

        f.nombre = nombre

    if "orden" in data:
        f.orden = int(data.get("orden") or 0)

    if "activo" in data:
        f.activo = _to_bool2(data.get("activo"), default=True)

    db.session.commit()
    return jsonify({"mensaje": "Fase actualizada", "fase": proyecto_fase_to_dict(f)}), 200

@bp.route("/proyecto-fases/<int:id>", methods=["DELETE"])
@permission_required("PROYECTOS_ELIMINAR")
def eliminar_proyecto_fase(id):
    f = ProyectoFase.query.get_or_404(id)

    usados_main = Proyecto.query.filter(Proyecto.fase_id == id).count()
    usados_multi = 0
    try:
        usados_multi = ProyectoFaseProyecto.query.filter(ProyectoFaseProyecto.fase_id == id).count()
    except Exception:
        usados_multi = 0

    if (usados_main + usados_multi) > 0:
        return jsonify({"mensaje": "No se puede eliminar: hay proyectos usando esta fase"}), 400

    db.session.delete(f)
    db.session.commit()
    return jsonify({"mensaje": "Fase eliminada"}), 200

@bp.route("/proyectos", methods=["GET"])
@permission_required("PROYECTOS_VER")
def listar_proyectos():
    try:
        include_modulos = (request.args.get("include_modulos") or "0") == "1"
        include_fases = (request.args.get("include_fases") or "0") == "1"
        include_perfiles = (request.args.get("include_perfiles") or "0") == "1"

        solo_activos = (request.args.get("activos") or "").strip().lower()
        q = (request.args.get("q") or "").strip()

        opts = [
            joinedload(Proyecto.cliente),
            joinedload(Proyecto.oportunidad),
            joinedload(Proyecto.fase),
        ]

        if include_modulos:
            opts.append(
                joinedload(Proyecto.modulos).joinedload(ProyectoModulo.modulo)
            )

        if include_fases:
            opts.append(
                joinedload(Proyecto.fases).joinedload(ProyectoFaseProyecto.fase)
            )

        if include_perfiles:
            opts.append(
                joinedload(Proyecto.perfiles)
                .joinedload(ProyectoPerfil.perfil)
            )

        query = Proyecto.query.options(*opts)

        if solo_activos in ("1", "true", "si", "sí", "yes"):
            query = query.filter(Proyecto.activo == True)

        if q:
            like = f"%{q}%"
            query = query.filter(
                or_(
                    Proyecto.codigo.ilike(like),
                    Proyecto.nombre.ilike(like),
                    Proyecto.tipo_negocio.ilike(like),
                )
            )

        proyectos = (
            query
            .order_by(Proyecto.activo.desc(), Proyecto.codigo.asc())
            .all()
        )

        return jsonify([
            proyecto_to_dict(
                p,
                include_modulos=include_modulos,
                include_fases=include_fases,
                include_perfiles=include_perfiles,
            )
            for p in proyectos
        ]), 200

    except Exception as e:
        app.logger.exception("Error listando proyectos")
        return jsonify({
            "mensaje": f"Error listando proyectos: {str(e)}"
        }), 500

@bp.route("/proyectos/<int:id>", methods=["GET"])
@permission_required("PROYECTOS_VER")
def get_proyecto(id):
    opts = [joinedload(Proyecto.modulos).joinedload(ProyectoModulo.modulo)]
    if hasattr(Proyecto, "fase"):
        opts.append(joinedload(Proyecto.fase))
    if hasattr(Proyecto, "fases"):
        opts.append(joinedload(Proyecto.fases).joinedload(ProyectoFaseProyecto.fase))

    p = Proyecto.query.options(*opts).get_or_404(id)
    return jsonify(proyecto_to_dict(p, include_modulos=True, include_fases=True)), 200

@bp.route("/proyectos", methods=["POST"])
@permission_required("PROYECTOS_CREAR")
def crear_proyecto():
    data = request.get_json(silent=True) or {}

    try:
        nombre = (data.get("nombre") or "").strip()
        activo = _to_bool2(data.get("activo"), default=True)

        cliente_id = data.get("cliente_id", None)
        fase_id = data.get("fase_id")
        oportunidad_id = data.get("oportunidad_id", None)
        perfil_consultores = data.get("perfil_consultores")

        try:
            perfiles_ids = _clean_int_list(
                data.get("perfiles") or data.get("perfiles_ids") or [],
                "perfiles"
            )

            # Ahora los módulos pueden venir desde el frontend o calcularse
            # automáticamente desde los perfiles seleccionados.
            modulos_ids_payload = data.get("modulos") or data.get("modulos_ids") or []

            if modulos_ids_payload:
                modulos_ids = _clean_int_list(modulos_ids_payload, "modulos")
            else:
                modulos_ids = _get_modulos_ids_de_perfiles(perfiles_ids)

            fases_ids = _clean_int_list(
                data.get("fases") or data.get("fases_ids") or [],
                "fases"
            )
        except ValueError as e:
            return jsonify({"mensaje": str(e)}), 400

        if not nombre:
            return jsonify({"mensaje": "nombre requerido"}), 400

        if not perfiles_ids:
            return jsonify({
                "mensaje": "Debes seleccionar al menos un perfil"
            }), 400

        if not modulos_ids:
            return jsonify({
                "mensaje": "Los perfiles seleccionados no tienen módulos asociados"
            }), 400

        # ------------------------------------------------------------
        # oportunidad obligatoria
        # ------------------------------------------------------------
        if oportunidad_id in ("", "null", None):
            return jsonify({"mensaje": "oportunidad_id requerido"}), 400

        try:
            oportunidad_id = int(oportunidad_id)
        except Exception:
            return jsonify({"mensaje": "oportunidad_id inválido"}), 400

        opp = Oportunidad.query.get(oportunidad_id)

        if not opp:
            return jsonify({"mensaje": "La oportunidad no existe"}), 400

        prc = (opp.codigo_prc or "").strip().upper()

        if not prc:
            return jsonify({
                "mensaje": "La oportunidad no tiene código PRC"
            }), 400

        estado = _norm_key_for_match(opp.estado_oferta)
        resultado = _norm_key_for_match(opp.resultado_oferta)

        if estado != _norm_key_for_match("GANADA"):
            return jsonify({
                "mensaje": "Solo se pueden crear proyectos desde oportunidades GANADAS"
            }), 400

        resultados_permitidos = {
            _norm_key_for_match("PROYECTO"),
            _norm_key_for_match("BOLSA DE HORAS / CONTINUIDAD DE LA OPERACIÓN"),
        }

        if resultado not in resultados_permitidos:
            return jsonify({
                "mensaje": "La oportunidad ganada no es de tipo PROYECTO ni BOLSA DE HORAS / CONTINUIDAD DE LA OPERACIÓN"
            }), 400

        tipo_negocio = (
            "BOLSA_HORAS"
            if resultado == _norm_key_for_match("BOLSA DE HORAS / CONTINUIDAD DE LA OPERACIÓN")
            else "PROYECTO"
        )

        codigo = prc

        dupe = Proyecto.query.filter(
            func.lower(Proyecto.codigo) == codigo.lower()
        ).first()

        if dupe:
            return jsonify({
                "mensaje": "Ya existe un proyecto con ese código PRC"
            }), 400

        # ------------------------------------------------------------
        # cliente
        # ------------------------------------------------------------
        if cliente_id in ("", "null", None):
            cliente_id = None
        else:
            try:
                cliente_id = int(cliente_id)
            except Exception:
                return jsonify({"mensaje": "cliente_id inválido"}), 400

            if not Cliente.query.get(cliente_id):
                return jsonify({"mensaje": "Cliente no existe"}), 400

        # ------------------------------------------------------------
        # fase principal opcional
        # ------------------------------------------------------------
        if fase_id in ("", "null", None):
            fase_id = None
        else:
            try:
                fase_id = int(fase_id)
            except Exception:
                return jsonify({"mensaje": "fase_id inválido"}), 400

            if not ProyectoFase.query.get(fase_id):
                return jsonify({"mensaje": "Fase no existe"}), 400

        # ------------------------------------------------------------
        # perfiles obligatorios
        # ------------------------------------------------------------
        ok, msg, perfiles_db = _validar_perfiles_ids(perfiles_ids)

        if not ok:
            return jsonify({"mensaje": msg}), 400

        # ------------------------------------------------------------
        # módulos automáticos / enviados
        # ------------------------------------------------------------
        ok, msg, modulos_db = _validar_modulos_ids(modulos_ids)

        if not ok:
            return jsonify({"mensaje": msg}), 400

        # Cada módulo del proyecto debe pertenecer a mínimo uno de los perfiles.
        ok, msg = _validar_modulos_pertenecen_a_perfiles(
            perfiles_ids,
            modulos_ids
        )

        if not ok:
            return jsonify({"mensaje": msg}), 400

        # ------------------------------------------------------------
        # fases multi
        # ------------------------------------------------------------
        if fases_ids:
            fases_db = ProyectoFase.query.filter(
                ProyectoFase.id.in_(fases_ids)
            ).all()

            found_ids = {int(fx.id) for fx in fases_db}
            missing = [fid for fid in fases_ids if int(fid) not in found_ids]

            if missing:
                return jsonify({
                    "mensaje": f"Fases no encontradas: {missing}"
                }), 400

        # ------------------------------------------------------------
        # crear proyecto
        # ------------------------------------------------------------
        p = Proyecto(
            codigo=codigo,
            nombre=nombre,
            activo=activo,
            fase_id=fase_id,
            cliente_id=cliente_id,
            oportunidad_id=opp.id,
            tipo_negocio=tipo_negocio,
        )

        db.session.add(p)
        db.session.flush()

        # ------------------------------------------------------------
        # perfiles del proyecto
        # ------------------------------------------------------------
        for perfil_id in perfiles_ids:
            db.session.add(
                ProyectoPerfil(
                    proyecto_id=p.id,
                    perfil_id=int(perfil_id),
                    activo=True,
                )
            )

        # ------------------------------------------------------------
        # módulos del proyecto
        # Ahora salen automáticamente de los perfiles.
        # ------------------------------------------------------------
        for modulo_id in modulos_ids:
            db.session.add(
                ProyectoModulo(
                    proyecto_id=p.id,
                    modulo_id=int(modulo_id),
                    activo=True,
                )
            )

        # ------------------------------------------------------------
        # fases multi
        # ------------------------------------------------------------
        for fase_id_item in fases_ids:
            db.session.add(
                ProyectoFaseProyecto(
                    proyecto_id=p.id,
                    fase_id=int(fase_id_item),
                    activo=True,
                )
            )

        db.session.flush()

        # ------------------------------------------------------------
        # consultores asignados a cada perfil del proyecto
        # ------------------------------------------------------------
        _save_proyecto_perfil_consultores(
            p.id,
            perfil_consultores
        )

        db.session.commit()

        opts = [
            joinedload(Proyecto.cliente),
            joinedload(Proyecto.oportunidad),
            joinedload(Proyecto.modulos).joinedload(ProyectoModulo.modulo),
            joinedload(Proyecto.fase),
            joinedload(Proyecto.fases).joinedload(ProyectoFaseProyecto.fase),
            joinedload(Proyecto.perfiles).joinedload(ProyectoPerfil.perfil),
        ]

        p = Proyecto.query.options(*opts).get(p.id)

        return jsonify({
            "mensaje": "Proyecto creado",
            "proyecto": _proyecto_response_dict(
                p,
                include_modulos=True,
                include_fases=True,
                include_perfiles=True,
            )
        }), 201

    except ValueError as e:
        db.session.rollback()
        return jsonify({"mensaje": str(e)}), 400

    except IntegrityError as e:
        db.session.rollback()
        app.logger.exception("Error de integridad creando proyecto")
        return jsonify({
            "mensaje": f"No se pudo crear el proyecto por conflicto de datos: {str(e)}"
        }), 400

    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error creando proyecto")
        return jsonify({
            "mensaje": f"No se pudo crear el proyecto: {str(e)}"
        }), 500

@bp.route("/proyectos/<int:id>", methods=["PUT"])
@permission_required("PROYECTOS_EDITAR")
def editar_proyecto(id):
    opts = [
        joinedload(Proyecto.cliente),
        joinedload(Proyecto.oportunidad),
        joinedload(Proyecto.modulos).joinedload(ProyectoModulo.modulo),
        joinedload(Proyecto.fase),
        joinedload(Proyecto.fases).joinedload(ProyectoFaseProyecto.fase),
        joinedload(Proyecto.perfiles).joinedload(ProyectoPerfil.perfil),
    ]

    p = Proyecto.query.options(*opts).get_or_404(id)
    data = request.get_json(silent=True) or {}

    try:
        # ------------------------------------------------------------
        # nombre
        # ------------------------------------------------------------
        if "nombre" in data:
            nombre = (data.get("nombre") or "").strip()

            if not nombre:
                return jsonify({"mensaje": "nombre inválido"}), 400

            p.nombre = nombre

        # ------------------------------------------------------------
        # activo
        # ------------------------------------------------------------
        if "activo" in data:
            p.activo = _to_bool2(data.get("activo"), default=True)

        # ------------------------------------------------------------
        # cliente_id
        # ------------------------------------------------------------
        if "cliente_id" in data:
            cliente_id = data.get("cliente_id", None)

            if cliente_id in ("", "null", None):
                p.cliente_id = None
            else:
                try:
                    cliente_id = int(cliente_id)
                except Exception:
                    return jsonify({"mensaje": "cliente_id inválido"}), 400

                if not Cliente.query.get(cliente_id):
                    return jsonify({"mensaje": "Cliente no existe"}), 400

                p.cliente_id = cliente_id

        # ------------------------------------------------------------
        # fase_id principal
        # ------------------------------------------------------------
        if "fase_id" in data:
            fase_id = data.get("fase_id", None)

            if fase_id in ("", "null", None):
                p.fase_id = None
            else:
                try:
                    fase_id = int(fase_id)
                except Exception:
                    return jsonify({"mensaje": "fase_id inválido"}), 400

                if not ProyectoFase.query.get(fase_id):
                    return jsonify({"mensaje": "Fase no existe"}), 400

                p.fase_id = fase_id

        # ------------------------------------------------------------
        # oportunidad
        # ------------------------------------------------------------
        if "oportunidad_id" in data:
            oportunidad_id = data.get("oportunidad_id", None)

            if oportunidad_id in ("", "null", None):
                return jsonify({
                    "mensaje": "oportunidad_id requerido"
                }), 400

            try:
                oportunidad_id = int(oportunidad_id)
            except Exception:
                return jsonify({
                    "mensaje": "oportunidad_id inválido"
                }), 400

            opp = Oportunidad.query.get(oportunidad_id)

            if not opp:
                return jsonify({
                    "mensaje": "La oportunidad no existe"
                }), 400

            prc = (opp.codigo_prc or "").strip().upper()

            if not prc:
                return jsonify({
                    "mensaje": "La oportunidad no tiene código PRC válido"
                }), 400

            estado = _norm_key_for_match(opp.estado_oferta)
            resultado = _norm_key_for_match(opp.resultado_oferta)

            if estado != _norm_key_for_match("GANADA"):
                return jsonify({
                    "mensaje": "Solo se pueden asociar oportunidades GANADAS"
                }), 400

            resultados_permitidos = {
                _norm_key_for_match("PROYECTO"),
                _norm_key_for_match("BOLSA DE HORAS / CONTINUIDAD DE LA OPERACIÓN"),
            }

            if resultado not in resultados_permitidos:
                return jsonify({
                    "mensaje": "La oportunidad ganada no es de tipo PROYECTO ni BOLSA DE HORAS / CONTINUIDAD DE LA OPERACIÓN"
                }), 400

            dupe = Proyecto.query.filter(
                Proyecto.id != id,
                func.lower(Proyecto.codigo) == prc.lower()
            ).first()

            if dupe:
                return jsonify({
                    "mensaje": "Ya existe otro proyecto con ese código PRC"
                }), 400

            tipo_negocio = (
                "BOLSA_HORAS"
                if resultado == _norm_key_for_match("BOLSA DE HORAS / CONTINUIDAD DE LA OPERACIÓN")
                else "PROYECTO"
            )

            p.oportunidad_id = opp.id
            p.codigo = prc
            p.tipo_negocio = tipo_negocio

        # ------------------------------------------------------------
        # proteger código manual si ya existe oportunidad
        # ------------------------------------------------------------
        if "codigo" in data and p.oportunidad_id:
            opp_actual = Oportunidad.query.get(p.oportunidad_id)
            prc_actual = (
                (opp_actual.codigo_prc or "").strip().upper()
                if opp_actual else ""
            )

            if not prc_actual:
                return jsonify({
                    "mensaje": "La oportunidad asociada no tiene código PRC válido"
                }), 400

            p.codigo = prc_actual

        # ------------------------------------------------------------
        # perfiles
        # ------------------------------------------------------------
        perfiles_ids = None

        if "perfiles" in data or "perfiles_ids" in data:
            try:
                perfiles_ids = _clean_int_list(
                    data.get("perfiles") or data.get("perfiles_ids") or [],
                    "perfiles"
                )
            except ValueError as e:
                return jsonify({"mensaje": str(e)}), 400

            if not perfiles_ids:
                return jsonify({
                    "mensaje": "Debes seleccionar al menos un perfil"
                }), 400

            ok, msg, perfiles_db = _validar_perfiles_ids(perfiles_ids)

            if not ok:
                return jsonify({"mensaje": msg}), 400

            ProyectoPerfil.query.filter_by(
                proyecto_id=p.id
            ).delete(synchronize_session=False)

            for perfil_id in perfiles_ids:
                db.session.add(
                    ProyectoPerfil(
                        proyecto_id=p.id,
                        perfil_id=int(perfil_id),
                        activo=True,
                    )
                )

        # Si no llegaron perfiles en este PUT, se toman los actuales.
        if perfiles_ids is None:
            perfiles_ids = [
                int(x.perfil_id)
                for x in (
                    ProyectoPerfil.query
                    .filter_by(proyecto_id=p.id)
                    .filter(ProyectoPerfil.activo == True)
                    .all()
                )
                if x.perfil_id
            ]

        # ------------------------------------------------------------
        # módulos
        # Ahora pueden venir desde frontend o derivarse de perfiles.
        # ------------------------------------------------------------
        modulos_ids = None

        if "modulos" in data or "modulos_ids" in data:
            try:
                modulos_ids_payload = data.get("modulos") or data.get("modulos_ids") or []

                if modulos_ids_payload:
                    modulos_ids = _clean_int_list(
                        modulos_ids_payload,
                        "modulos"
                    )
                else:
                    modulos_ids = _get_modulos_ids_de_perfiles(perfiles_ids)
            except ValueError as e:
                return jsonify({"mensaje": str(e)}), 400

        elif "perfiles" in data or "perfiles_ids" in data:
            # Si cambiaron los perfiles, recalculamos módulos automáticamente.
            modulos_ids = _get_modulos_ids_de_perfiles(perfiles_ids)

        if modulos_ids is not None:
            if not modulos_ids:
                return jsonify({
                    "mensaje": "Los perfiles seleccionados no tienen módulos asociados"
                }), 400

            ok, msg, modulos_db = _validar_modulos_ids(modulos_ids)

            if not ok:
                return jsonify({"mensaje": msg}), 400

            ok, msg = _validar_modulos_pertenecen_a_perfiles(
                perfiles_ids,
                modulos_ids
            )

            if not ok:
                return jsonify({"mensaje": msg}), 400

            ProyectoModulo.query.filter_by(
                proyecto_id=p.id
            ).delete(synchronize_session=False)

            for modulo_id in modulos_ids:
                db.session.add(
                    ProyectoModulo(
                        proyecto_id=p.id,
                        modulo_id=int(modulo_id),
                        activo=True,
                    )
                )

        # ------------------------------------------------------------
        # fases multi
        # ------------------------------------------------------------
        if "fases" in data or "fases_ids" in data:
            try:
                fases_ids = _clean_int_list(
                    data.get("fases") or data.get("fases_ids") or [],
                    "fases"
                )
            except ValueError as e:
                return jsonify({"mensaje": str(e)}), 400

            if fases_ids:
                fases_db = ProyectoFase.query.filter(
                    ProyectoFase.id.in_(fases_ids)
                ).all()

                found_ids = {int(fx.id) for fx in fases_db}
                missing = [fid for fid in fases_ids if int(fid) not in found_ids]

                if missing:
                    return jsonify({
                        "mensaje": f"Fases no encontradas: {missing}"
                    }), 400

            ProyectoFaseProyecto.query.filter_by(
                proyecto_id=p.id
            ).delete(synchronize_session=False)

            for fase_id_item in fases_ids:
                db.session.add(
                    ProyectoFaseProyecto(
                        proyecto_id=p.id,
                        fase_id=int(fase_id_item),
                        activo=True,
                    )
                )

        db.session.flush()

        # ------------------------------------------------------------
        # consultores por perfil del proyecto
        # ------------------------------------------------------------
        if "perfil_consultores" in data:
            _save_proyecto_perfil_consultores(
                p.id,
                data.get("perfil_consultores")
            )

        db.session.commit()

        opts2 = [
            joinedload(Proyecto.cliente),
            joinedload(Proyecto.oportunidad),
            joinedload(Proyecto.modulos).joinedload(ProyectoModulo.modulo),
            joinedload(Proyecto.fase),
            joinedload(Proyecto.fases).joinedload(ProyectoFaseProyecto.fase),
            joinedload(Proyecto.perfiles).joinedload(ProyectoPerfil.perfil),
        ]

        p = Proyecto.query.options(*opts2).get(p.id)

        return jsonify({
            "mensaje": "Proyecto actualizado",
            "proyecto": _proyecto_response_dict(
                p,
                include_modulos=True,
                include_fases=True,
                include_perfiles=True,
            )
        }), 200

    except ValueError as e:
        db.session.rollback()
        return jsonify({"mensaje": str(e)}), 400

    except IntegrityError as e:
        db.session.rollback()
        app.logger.exception("Error de integridad actualizando proyecto")
        return jsonify({
            "mensaje": f"No se pudo actualizar el proyecto por conflicto de datos: {str(e)}"
        }), 400

    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error actualizando proyecto")
        return jsonify({
            "mensaje": f"No se pudo actualizar el proyecto: {str(e)}"
        }), 500
    
@bp.route("/proyectos/<int:id>/toggle-activo", methods=["PUT"])
@permission_required("PROYECTOS_EDITAR")
def toggle_activo_proyecto(id):
    p = Proyecto.query.get_or_404(id)
    p.activo = not bool(p.activo)
    db.session.commit()
    return jsonify({"mensaje": "Estado actualizado", "activo": bool(p.activo)}), 200

@bp.route("/proyectos/<int:id>", methods=["DELETE"])
@permission_required("PROYECTOS_ELIMINAR")
def eliminar_proyecto(id):
    p = Proyecto.query.get_or_404(id)
    try:
        ProyectoModulo.query.filter_by(proyecto_id=id).delete()
    except Exception:
        pass
    try:
        ProyectoFaseProyecto.query.filter_by(proyecto_id=id).delete()
    except Exception:
        pass
    db.session.delete(p)
    db.session.commit()
    return jsonify({"mensaje": "Proyecto eliminado"}), 200

@bp.route("/proyectos/activos-por-modulo", methods=["GET"])
@permission_required("PROYECTOS_VER")
def proyectos_activos_por_modulo():
    modulo = (request.args.get("modulo") or "").strip().upper()
    if not modulo:
        return jsonify({"mensaje": "modulo requerido"}), 400

    opts = [
        joinedload(Proyecto.modulos).joinedload(ProyectoModulo.modulo),
    ]
    if hasattr(Proyecto, "fase"):
        opts.append(joinedload(Proyecto.fase))
    if hasattr(Proyecto, "fases"):
        opts.append(joinedload(Proyecto.fases).joinedload(ProyectoFaseProyecto.fase))

    query = (
        Proyecto.query
        .join(ProyectoModulo, ProyectoModulo.proyecto_id == Proyecto.id)
        .join(Modulo, Modulo.id == ProyectoModulo.modulo_id)
        .options(*opts)
        .filter(Proyecto.activo == True)
        .filter(ProyectoModulo.activo == True)
        .filter(func.upper(Modulo.nombre) == modulo)
        .order_by(Proyecto.codigo.asc())
    )

    proyectos = query.all()
    return jsonify([{
        "id": p.id,
        "codigo": p.codigo,
        "nombre": p.nombre,
        "fase_id": p.fase_id,
        "fase": p.fase.nombre if getattr(p, "fase", None) else None,
        "fases": [{"id": x["id"], "nombre": x["nombre"], "orden": x["orden"], "activo": x["activo"]} for x in (proyecto_to_dict(p, include_modulos=False, include_fases=True).get("fases") or [])],
        "fases_ids": proyecto_to_dict(p, include_modulos=False, include_fases=True).get("fases_ids") or [],
    } for p in proyectos]), 200

def pm_to_dict(x: ProyectoMapeo):
    return {
        "id": x.id,
        "proyecto_id": x.proyecto_id,
        "valor_origen": x.valor_origen,
        "tipo_match": x.tipo_match,
        "activo": bool(x.activo),
    }

@bp.route("/proyecto-mapeos", methods=["GET"])
@permission_required("PROYECTOS_VER")
def listar_proyecto_mapeos():
    proyecto_id = request.args.get("proyecto_id")
    q = ProyectoMapeo.query
    if proyecto_id:
        try:
            pid = int(proyecto_id)
            q = q.filter(ProyectoMapeo.proyecto_id == pid)
        except:
            pass
    rows = q.order_by(ProyectoMapeo.proyecto_id.asc(), ProyectoMapeo.valor_origen.asc()).all()
    return jsonify([pm_to_dict(x) for x in rows]), 200

## -------------------------------
##  Proyectos Mapeos (para agrupar valores en reportes)
@bp.route("/proyecto-mapeos", methods=["POST"])
@permission_required("PROYECTOS_EDITAR")
def crear_proyecto_mapeo():
    data = request.get_json(silent=True) or {}
    proyecto_id = data.get("proyecto_id")
    valor_origen = (data.get("valor_origen") or "").strip().upper()
    tipo_match = (data.get("tipo_match") or "EXACT").strip().upper()

    if not proyecto_id:
        return jsonify({"mensaje": "proyecto_id requerido"}), 400
    if not valor_origen:
        return jsonify({"mensaje": "valor_origen requerido"}), 400
    if tipo_match not in {"EXACT", "CONTAINS", "REGEX"}:
        return jsonify({"mensaje": "tipo_match inválido"}), 400

    try:
        proyecto_id = int(proyecto_id)
    except:
        return jsonify({"mensaje": "proyecto_id inválido"}), 400

    if not Proyecto.query.get(proyecto_id):
        return jsonify({"mensaje": "Proyecto no existe"}), 400

    x = ProyectoMapeo(
        proyecto_id=proyecto_id,
        valor_origen=valor_origen,
        tipo_match=tipo_match,
        activo=_to_bool2(data.get("activo"), default=True)
    )
    db.session.add(x)

    try:
        db.session.commit()
        return jsonify({"mensaje": "Mapeo creado", "mapeo": pm_to_dict(x)}), 201
    except IntegrityError:
        db.session.rollback()
        return jsonify({"mensaje": "Ya existe ese valor_origen para ese proyecto"}), 400


@bp.route("/proyecto-mapeos/<int:id>", methods=["PUT"])
@permission_required("PROYECTOS_EDITAR")
def editar_proyecto_mapeo(id):
    x = ProyectoMapeo.query.get_or_404(id)
    data = request.get_json(silent=True) or {}

    if "valor_origen" in data:
        valor_origen = (data.get("valor_origen") or "").strip().upper()
        if not valor_origen:
            return jsonify({"mensaje": "valor_origen requerido"}), 400
        x.valor_origen = valor_origen

    if "tipo_match" in data:
        tipo_match = (data.get("tipo_match") or "").strip().upper()
        if tipo_match not in {"EXACT", "CONTAINS", "REGEX"}:
            return jsonify({"mensaje": "tipo_match inválido"}), 400
        x.tipo_match = tipo_match

    if "activo" in data:
        x.activo = _to_bool2(data.get("activo"), default=True)

    try:
        db.session.commit()
        return jsonify({"mensaje": "Mapeo actualizado", "mapeo": pm_to_dict(x)}), 200
    except IntegrityError:
        db.session.rollback()
        return jsonify({"mensaje": "Duplicado: ya existe ese valor_origen para el proyecto"}), 400


@bp.route("/proyecto-mapeos/<int:id>", methods=["DELETE"])
@permission_required("PROYECTOS_EDITAR")
def eliminar_proyecto_mapeo(id):
    x = ProyectoMapeo.query.get_or_404(id)
    db.session.delete(x)
    db.session.commit()
    return jsonify({"mensaje": "Mapeo eliminado"}), 200

@bp.route("/proyectos/<int:proyecto_id>/mapeos", methods=["GET"])
@permission_required("PROYECTOS_VER")
def listar_mapeos_por_proyecto(proyecto_id):
    rows = (
        ProyectoMapeo.query
        .filter(ProyectoMapeo.proyecto_id == proyecto_id)
        .order_by(ProyectoMapeo.valor_origen.asc())
        .all()
    )
    return jsonify([pm_to_dict(x) for x in rows]), 200

@bp.route("/proyectos/<int:proyecto_id>/mapeos", methods=["POST"])
@permission_required("PROYECTOS_EDITAR")
def crear_mapeo_por_proyecto(proyecto_id):
    data = request.get_json(silent=True) or {}
    valor_origen = (data.get("valor_origen") or "").strip().upper()
    tipo_match = (data.get("tipo_match") or "EXACT").strip().upper()

    if not valor_origen:
        return jsonify({"mensaje": "valor_origen requerido"}), 400
    if tipo_match not in {"EXACT", "CONTAINS", "REGEX"}:
        return jsonify({"mensaje": "tipo_match inválido"}), 400

    proyecto = Proyecto.query.get(proyecto_id)
    if not proyecto:
        return jsonify({"mensaje": "Proyecto no existe"}), 404

    x = ProyectoMapeo(
        proyecto_id=proyecto_id,
        valor_origen=valor_origen,
        tipo_match=tipo_match,
        activo=_to_bool2(data.get("activo"), default=True)
    )
    db.session.add(x)

    try:
        db.session.commit()
        return jsonify({"mensaje": "Mapeo creado", "mapeo": pm_to_dict(x)}), 201
    except IntegrityError:
        db.session.rollback()
        return jsonify({"mensaje": "Ya existe ese valor_origen para ese proyecto"}), 400
    
def _cost_parse_date(v):
    if v in (None, "", "null", "None"):
        return None
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(str(v).strip()[:10], "%Y-%m-%d").date()
    except Exception:
        return None

def _cost_parse_decimal(v):
    if v in (None, "", "null", "None"):
        return None

    if isinstance(v, Decimal):
        return v

    s = str(v).strip()
    s = s.replace("$", "").replace(" ", "")

    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")

    try:
        return Decimal(s)
    except Exception:
        return None

def _presupuesto_mensual_to_dict(x: ProyectoPresupuestoMensual):
    return {
        "id": x.id,
        "proyecto_id": x.proyecto_id,
        "anio": x.anio,
        "mes": x.mes,
        "ingreso_planeado": _money_to_json(x.ingreso_planeado),
        "costo_planeado": _money_to_json(x.costo_planeado),
        "gasto_operativo_planeado": _money_to_json(x.gasto_operativo_planeado),
        "costo_administrativo_planeado": _money_to_json(x.costo_administrativo_planeado),
        "ebitda_planeado": _money_to_json(x.ebitda_planeado),
        "margen_planeado_pct": _money_to_json(x.margen_planeado_pct),
        "activo": bool(x.activo),
    }

def _perfil_plan_to_dict(x: ProyectoPerfilPlan):
    return {
        "id": x.id,
        "proyecto_id": x.proyecto_id,
        "anio": x.anio,
        "mes": x.mes,

        "perfil_id": x.perfil_id,
        "perfil": _perfil_to_dict(x.perfil) if x.perfil else None,

        "modulo_id": x.modulo_id,
        "modulo": {
            "id": x.modulo.id,
            "nombre": x.modulo.nombre,
        } if x.modulo else None,

        "consultor_id": x.consultor_id,
        "consultor": {
            "id": x.consultor.id,
            "nombre": x.consultor.nombre,
            "usuario": x.consultor.usuario,
        } if x.consultor else None,

        "horas_estimadas": _money_to_json(x.horas_estimadas),

        "fte_estimado": _money_to_json(x.fte_estimado),
        "valor_hora_ingreso": _money_to_json(x.fte_estimado),

        "valor_hora_planeado": _money_to_json(x.valor_hora_planeado),
        "costo_estimado": _money_to_json(x.costo_estimado),
        "ingreso_estimado": _money_to_json(x.ingreso_estimado),
        "observacion": x.observacion,
        "orden": int(x.orden or 0),
        "activo": bool(x.activo),
    }

def _costo_adicional_to_dict(x: ProyectoCostoAdicional):
    return {
        "id": x.id,
        "proyecto_id": x.proyecto_id,
        "anio": x.anio,
        "mes": x.mes,
        "tipo_costo": x.tipo_costo,
        "categoria": x.categoria,
        "descripcion": x.descripcion,
        "valor": _money_to_json(x.valor),
        "activo": bool(x.activo),
    }

def _ym_from_registro_fecha(fecha_str):
    s = str(fecha_str or "").strip()
    if not s:
        return None, None

    s10 = s[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            d = datetime.strptime(s10, fmt).date()
            return d.year, d.month
        except Exception:
            continue

    if len(s) >= 7 and s[4] == "-":
        try:
            return int(s[:4]), int(s[5:7])
        except Exception:
            return None, None

    return None, None

def _valor_hora_consultor(consultor_id, anio, mes):
    if not consultor_id or not anio or not mes:
        return Decimal("0")

    row = (
        ConsultorPresupuesto.query
        .filter(ConsultorPresupuesto.consultor_id == consultor_id)
        .filter(ConsultorPresupuesto.anio == anio)
        .filter(ConsultorPresupuesto.mes == mes)
        .filter(ConsultorPresupuesto.vigente == True)
        .order_by(ConsultorPresupuesto.id.desc())
        .first()
    )

    if not row:
        row = (
            ConsultorPresupuesto.query
            .filter(ConsultorPresupuesto.consultor_id == consultor_id)
            .filter(ConsultorPresupuesto.anio == anio)
            .filter(ConsultorPresupuesto.mes == mes)
            .order_by(ConsultorPresupuesto.id.desc())
            .first()
        )

    if not row:
        return Decimal("0")

    vr = Decimal(str(row.vr_perfil or 0))
    hb = Decimal(str(row.horas_base_mes or 0))

    if hb <= 0:
        return Decimal("0")

    return vr / hb
    
@bp.route("/oportunidades/elegibles-proyecto", methods=["GET"])
@permission_required("OPORTUNIDADES_VER")
def listar_oportunidades_elegibles_proyecto():
    try:
        query = Oportunidad.query

        # Solo oportunidades con PRC
        query = query.filter(Oportunidad.codigo_prc.isnot(None))
        query = query.filter(func.trim(Oportunidad.codigo_prc) != "")

        # Orden seguro para MySQL/MariaDB
        query = query.order_by(
            Oportunidad.fecha_cierre_oportunidad.desc(),
            Oportunidad.id.desc()
        )

        rows = query.all()

        resultados_permitidos = {
            _norm_key_for_match("PROYECTO"),
            _norm_key_for_match("BOLSA DE HORAS / CONTINUIDAD DE LA OPERACIÓN"),
        }

        data = []
        for o in rows:
            estado = _norm_key_for_match(o.estado_oferta)
            resultado = _norm_key_for_match(o.resultado_oferta)
            prc = (o.codigo_prc or "").strip().upper()

            if estado != _norm_key_for_match("GANADA"):
                continue

            if resultado not in resultados_permitidos:
                continue

            tipo_negocio = (
                "BOLSA_HORAS"
                if resultado == _norm_key_for_match("BOLSA DE HORAS / CONTINUIDAD DE LA OPERACIÓN")
                else "PROYECTO"
            )

            data.append({
                "id": o.id,
                "codigo_prc": prc,
                "nombre_cliente": o.nombre_cliente,
                "servicio": o.servicio,
                "fecha_cierre_oportunidad": o.fecha_cierre_oportunidad.isoformat() if o.fecha_cierre_oportunidad else None,
                "pm_asignado_claro": o.pm_asignado_claro,
                "pm_asignado_hitss": o.pm_asignado_hitss,
                "estado_oferta": o.estado_oferta,
                "resultado_oferta": o.resultado_oferta,
                "tipo_negocio": tipo_negocio,
            })

        return jsonify(data), 200

    except Exception:
        return jsonify({
            "mensaje": "Error interno en /oportunidades/elegibles-proyecto",
            "trace": traceback.format_exc()
        }), 500

## ----------
## Costos Endpoint 
def _cost_parse_date(v):
    if v in (None, "", "null", "None"):
        return None
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(str(v).strip()[:10], "%Y-%m-%d").date()
    except Exception:
        return None

def _cost_parse_decimal(v):
    if v in (None, "", "null", "None"):
        return None

    if isinstance(v, Decimal):
        return v

    s = str(v).strip()
    s = s.replace("$", "").replace(" ", "")

    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")

    try:
        return Decimal(s)
    except Exception:
        return None

def _perfil_to_dict(x: Perfil, permitido_en_modulos=True, include_modulos=False, modulos_permitidos_ids=None):
    modulos_permitidos_ids = (
        {int(m) for m in modulos_permitidos_ids}
        if modulos_permitidos_ids
        else None
    )

    out = {
        "id": x.id,
        "codigo": x.codigo,
        "nombre": x.nombre,
        "descripcion": getattr(x, "descripcion", None),
        "activo": bool(x.activo),
        "orden": int(x.orden or 0),
        "permitido_en_modulos": bool(permitido_en_modulos),
    }

    if include_modulos:
        modulos = []

        for mp in (getattr(x, "modulos", None) or []):
            modulo = getattr(mp, "modulo", None)

            if not modulo:
                continue

            if not bool(getattr(mp, "activo", True)):
                continue

            if modulos_permitidos_ids is not None and int(modulo.id) not in modulos_permitidos_ids:
                continue

            modulos.append({
                "id": modulo.id,
                "modulo_id": modulo.id,
                "nombre": modulo.nombre,
                "activo": bool(getattr(mp, "activo", True)),
                "modulo_perfil_id": mp.id,
            })

        modulos.sort(key=lambda m: str(m["nombre"] or "").upper())
        out["modulos"] = modulos

    return out

def _presupuesto_mensual_to_dict(x: ProyectoPresupuestoMensual):
    return {
        "id": x.id,
        "proyecto_id": x.proyecto_id,
        "anio": x.anio,
        "mes": x.mes,
        "ingreso_planeado": _money_to_json(x.ingreso_planeado),
        "costo_planeado": _money_to_json(x.costo_planeado),
        "gasto_operativo_planeado": _money_to_json(x.gasto_operativo_planeado),
        "costo_administrativo_planeado": _money_to_json(x.costo_administrativo_planeado),
        "ebitda_planeado": _money_to_json(x.ebitda_planeado),
        "margen_planeado_pct": _money_to_json(x.margen_planeado_pct),
        "activo": bool(x.activo),
    }

def _costo_adicional_to_dict(x: ProyectoCostoAdicional):
    return {
        "id": x.id,
        "proyecto_id": x.proyecto_id,
        "anio": x.anio,
        "mes": x.mes,
        "tipo_costo": x.tipo_costo,
        "categoria": x.categoria,
        "descripcion": x.descripcion,
        "valor": _money_to_json(x.valor),
        "activo": bool(x.activo),
    }

def _ym_from_registro_fecha(fecha_str):
    s = str(fecha_str or "").strip()
    if not s:
        return None, None

    s10 = s[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            d = datetime.strptime(s10, fmt).date()
            return d.year, d.month
        except Exception:
            continue

    if len(s) >= 7 and s[4] == "-":
        try:
            return int(s[:4]), int(s[5:7])
        except Exception:
            return None, None

    return None, None

def _valor_hora_consultor(consultor_id, anio, mes):
    if not consultor_id or not anio or not mes:
        return Decimal("0")

    row = (
        ConsultorPresupuesto.query
        .filter(ConsultorPresupuesto.consultor_id == consultor_id)
        .filter(ConsultorPresupuesto.anio == anio)
        .filter(ConsultorPresupuesto.mes == mes)
        .filter(ConsultorPresupuesto.vigente == True)
        .order_by(ConsultorPresupuesto.id.desc())
        .first()
    )

    if not row:
        row = (
            ConsultorPresupuesto.query
            .filter(ConsultorPresupuesto.consultor_id == consultor_id)
            .filter(ConsultorPresupuesto.anio == anio)
            .filter(ConsultorPresupuesto.mes == mes)
            .order_by(ConsultorPresupuesto.id.desc())
            .first()
        )

    if not row:
        return Decimal("0")

    vr = Decimal(str(row.vr_perfil or 0))
    hb = Decimal(str(row.horas_base_mes or 0))

    if hb <= 0:
        return Decimal("0")

    return vr / hb

@bp.route("/proyectos/<int:proyecto_id>/costos", methods=["GET"])
@permission_required("PROYECTOS_VER")
def get_proyecto_costos(proyecto_id):
    p = (
        Proyecto.query.options(
            joinedload(Proyecto.cliente),
            joinedload(Proyecto.oportunidad),
            joinedload(Proyecto.modulos).joinedload(ProyectoModulo.modulo),
            joinedload(Proyecto.perfiles)
                .joinedload(ProyectoPerfil.perfil)
                .joinedload(Perfil.modulos)
                .joinedload(ModuloPerfil.modulo),
            joinedload(Proyecto.fase),
            joinedload(Proyecto.fases).joinedload(ProyectoFaseProyecto.fase),
            joinedload(Proyecto.presupuestos_mensuales),
            joinedload(Proyecto.perfiles_plan).joinedload(ProyectoPerfilPlan.perfil),
            joinedload(Proyecto.perfiles_plan).joinedload(ProyectoPerfilPlan.modulo),
            joinedload(Proyecto.costos_adicionales),
        )
        .get_or_404(proyecto_id)
    )

    perfiles_proyecto = []
    perfiles_proyecto_ids = set()

    for pp in (getattr(p, "perfiles", None) or []):
        perfil = getattr(pp, "perfil", None)

        if not perfil:
            continue

        if not bool(getattr(pp, "activo", True)):
            continue

        perfiles_proyecto.append(perfil)
        perfiles_proyecto_ids.add(int(perfil.id))

    modulos_proyecto = []
    modulos_proyecto_ids = set()

    for pm in (getattr(p, "modulos", None) or []):
        modulo = getattr(pm, "modulo", None)

        if not modulo:
            continue

        if not bool(getattr(pm, "activo", True)):
            continue

        modulos_proyecto.append(modulo)
        modulos_proyecto_ids.add(int(modulo.id))

    modulos_planeacion_map = {}

    for perfil in perfiles_proyecto:
        for mp in (getattr(perfil, "modulos", None) or []):
            modulo = getattr(mp, "modulo", None)

            if not modulo:
                continue

            if not bool(getattr(mp, "activo", True)):
                continue

            if modulos_proyecto_ids and int(modulo.id) not in modulos_proyecto_ids:
                continue

            modulos_planeacion_map[int(modulo.id)] = modulo

    modulos_planeacion = sorted(
        modulos_planeacion_map.values(),
        key=lambda m: str(m.nombre or "").upper()
    )

    consultor_join_cond = db.or_(
        func.lower(func.trim(Registro.usuario_consultor)) == func.lower(func.trim(Consultor.usuario)),
        func.lower(func.trim(Registro.usuario_consultor)) == func.lower(func.trim(Consultor.nombre)),
    )

    rows_filtros = (
        _apply_project_filter_shared(
            db.session.query(
                func.coalesce(Equipo.nombre, Registro.equipo).label("equipo"),
                Registro.modulo.label("modulo"),
                Registro.usuario_consultor.label("usuario_consultor"),
                Consultor.usuario.label("consultor_usuario"),
                Consultor.nombre.label("consultor_nombre"),
            )
            .select_from(Registro)
            .outerjoin(Consultor, consultor_join_cond)
            .outerjoin(Equipo, Consultor.equipo_id == Equipo.id),
            proyecto_id
        )
        .all()
    )

    equipos_map = {}
    modulos_map = {}
    consultores_map = {}

    for r in rows_filtros:
        equipo = (r.equipo or "").strip().upper()
        modulo = (r.modulo or "").strip().upper()

        usuario = (
            r.consultor_usuario
            or r.usuario_consultor
            or ""
        ).strip().lower()

        nombre = (
            r.consultor_nombre
            or r.usuario_consultor
            or usuario
            or ""
        ).strip()

        if equipo:
            equipos_map[equipo] = {
                "id": equipo,
                "nombre": equipo,
            }

        if modulo:
            modulos_map[modulo] = {
                "id": modulo,
                "nombre": modulo,
            }

        if usuario:
            if usuario not in consultores_map:
                consultores_map[usuario] = {
                    "id": usuario,
                    "usuario": usuario,
                    "nombre": nombre,
                    "equipos": [],
                    "modulos": [],
                }

            if equipo and equipo not in consultores_map[usuario]["equipos"]:
                consultores_map[usuario]["equipos"].append(equipo)

            if modulo and modulo not in consultores_map[usuario]["modulos"]:
                consultores_map[usuario]["modulos"].append(modulo)

    return jsonify({
        "proyecto": proyecto_to_dict(
            p,
            include_modulos=True,
            include_fases=True,
            include_perfiles=True,
        ),
        "catalogos": {
            "perfiles": [
                _perfil_to_dict(
                    perfil,
                    include_modulos=True,
                    modulos_permitidos_ids=modulos_proyecto_ids,
                )
                for perfil in sorted(
                    perfiles_proyecto,
                    key=lambda x: (int(x.orden or 0), str(x.nombre or "").upper())
                )
            ],
            "equipos": sorted(
                equipos_map.values(),
                key=lambda x: x["nombre"].upper()
            ),
            "modulos": sorted(
                modulos_map.values(),
                key=lambda x: x["nombre"].upper()
            ),
            "consultores": sorted(
                consultores_map.values(),
                key=lambda x: x["nombre"].upper()
            ),
            "modulos_planeacion": [
                {
                    "id": m.id,
                    "nombre": m.nombre,
                }
                for m in modulos_planeacion
            ],
        },
        "presupuesto_mensual": [
            _presupuesto_mensual_to_dict(x)
            for x in sorted(
                p.presupuestos_mensuales,
                key=lambda r: (r.anio or 0, r.mes or 0)
            )
        ],
        "perfil_plan": [
            _perfil_plan_to_dict(x)
            for x in sorted(
                p.perfiles_plan,
                key=lambda r: (r.anio or 0, r.mes or 0, r.orden or 0, r.id or 0)
            )
        ],
        "costos_adicionales": [
            _costo_adicional_to_dict(x)
            for x in sorted(
                p.costos_adicionales,
                key=lambda r: (r.anio or 0, r.mes or 0, r.id or 0)
            )
        ],
    }), 200

@bp.route("/proyectos/<int:proyecto_id>/costos/cabecera", methods=["PUT"])
@permission_required("PROYECTOS_EDITAR")
def update_proyecto_costos_cabecera(proyecto_id):
    p = Proyecto.query.get_or_404(proyecto_id)
    data = request.get_json(silent=True) or {}

    if "oportunidad_id" in data:
        opp_id = data.get("oportunidad_id")
        if opp_id in (None, "", "null"):
            p.oportunidad_id = None
        else:
            try:
                opp_id = int(opp_id)
            except Exception:
                return jsonify({"mensaje": "oportunidad_id inválido"}), 400

            if not Oportunidad.query.get(opp_id):
                return jsonify({"mensaje": "La oportunidad no existe"}), 400

            p.oportunidad_id = opp_id

    if "codigo_ot_principal" in data:
        p.codigo_ot_principal = (data.get("codigo_ot_principal") or "").strip().upper() or None

    if "fecha_inicio_ejecucion" in data:
        p.fecha_inicio_ejecucion = _cost_parse_date(data.get("fecha_inicio_ejecucion"))
    if "fecha_fin_ejecucion" in data:
        p.fecha_fin_ejecucion = _cost_parse_date(data.get("fecha_fin_ejecucion"))
    if "fecha_inicio_facturacion" in data:
        p.fecha_inicio_facturacion = _cost_parse_date(data.get("fecha_inicio_facturacion"))
    if "fecha_fin_facturacion" in data:
        p.fecha_fin_facturacion = _cost_parse_date(data.get("fecha_fin_facturacion"))

    if p.fecha_inicio_ejecucion and p.fecha_fin_ejecucion and p.fecha_fin_ejecucion < p.fecha_inicio_ejecucion:
        return jsonify({"mensaje": "La fecha fin de ejecución no puede ser menor a la fecha inicio"}), 400

    if p.fecha_inicio_facturacion and p.fecha_fin_facturacion and p.fecha_fin_facturacion < p.fecha_inicio_facturacion:
        return jsonify({"mensaje": "La fecha fin de facturación no puede ser menor a la fecha inicio"}), 400

    if "moneda" in data:
        moneda = (data.get("moneda") or "COP").strip().upper()
        if moneda not in {"COP", "USD"}:
            return jsonify({"mensaje": "moneda inválida"}), 400
        p.moneda = moneda

    if "ingreso_total" in data:
        p.ingreso_total = _cost_parse_decimal(data.get("ingreso_total"))
    if "costo_objetivo_total" in data:
        p.costo_objetivo_total = _cost_parse_decimal(data.get("costo_objetivo_total"))
    if "gasto_operativo_total" in data:
        p.gasto_operativo_total = _cost_parse_decimal(data.get("gasto_operativo_total"))
    if "costo_administrativo_total" in data:
        p.costo_administrativo_total = _cost_parse_decimal(data.get("costo_administrativo_total"))
    if "margen_objetivo_pct" in data:
        p.margen_objetivo_pct = _cost_parse_decimal(data.get("margen_objetivo_pct"))
    if "ebitda_objetivo" in data:
        p.ebitda_objetivo = _cost_parse_decimal(data.get("ebitda_objetivo"))

    if "estado_financiero" in data:
        estado = (data.get("estado_financiero") or "").strip().upper()
        if estado not in {"BORRADOR", "CONFIGURADO", "ACTIVO", "PAUSADO", "CERRADO"}:
            return jsonify({"mensaje": "estado_financiero inválido"}), 400
        p.estado_financiero = estado

    if "alerta_umbral_1" in data:
        p.alerta_umbral_1 = _cost_parse_decimal(data.get("alerta_umbral_1")) or Decimal("70.00")
    if "alerta_umbral_2" in data:
        p.alerta_umbral_2 = _cost_parse_decimal(data.get("alerta_umbral_2")) or Decimal("85.00")
    if "alerta_umbral_3" in data:
        p.alerta_umbral_3 = _cost_parse_decimal(data.get("alerta_umbral_3")) or Decimal("95.00")

    db.session.commit()

    return jsonify({
        "mensaje": "Cabecera financiera actualizada",
        "proyecto": proyecto_to_dict(p, include_modulos=True, include_fases=True)
    }), 200

@bp.route("/proyectos/<int:proyecto_id>/costos/presupuesto-mensual", methods=["POST"])
@permission_required("PROYECTOS_EDITAR")
def save_proyecto_presupuesto_mensual(proyecto_id):
    Proyecto.query.get_or_404(proyecto_id)
    data = request.get_json(silent=True) or {}
    rows = data.get("rows") or []

    # validar duplicados antes de borrar
    seen = set()
    for row in rows:
        anio = int(row.get("anio") or 0)
        mes = int(row.get("mes") or 0)
        if anio <= 0 or mes < 1 or mes > 12:
            return jsonify({"mensaje": f"Periodo inválido en presupuesto mensual: {anio}-{mes}"}), 400

        key = (anio, mes)
        if key in seen:
            return jsonify({"mensaje": f"Periodo duplicado en presupuesto mensual: {anio}-{mes}"}), 400
        seen.add(key)

    ProyectoPresupuestoMensual.query.filter_by(proyecto_id=proyecto_id).delete()

    for row in rows:
        db.session.add(ProyectoPresupuestoMensual(
            proyecto_id=proyecto_id,
            anio=int(row.get("anio")),
            mes=int(row.get("mes")),
            ingreso_planeado=_cost_parse_decimal(row.get("ingreso_planeado")),
            costo_planeado=_cost_parse_decimal(row.get("costo_planeado")),
            gasto_operativo_planeado=_cost_parse_decimal(row.get("gasto_operativo_planeado")),
            costo_administrativo_planeado=_cost_parse_decimal(row.get("costo_administrativo_planeado")),
            ebitda_planeado=_cost_parse_decimal(row.get("ebitda_planeado")),
            margen_planeado_pct=_cost_parse_decimal(row.get("margen_planeado_pct")),
            activo=_to_bool2(row.get("activo"), default=True),
        ))

    db.session.commit()

    rows_db = (
        ProyectoPresupuestoMensual.query
        .filter_by(proyecto_id=proyecto_id)
        .order_by(ProyectoPresupuestoMensual.anio.asc(), ProyectoPresupuestoMensual.mes.asc())
        .all()
    )

    return jsonify({
        "mensaje": "Presupuesto mensual guardado",
        "rows": [_presupuesto_mensual_to_dict(x) for x in rows_db]
    }), 200

@bp.route("/proyectos/<int:proyecto_id>/costos/perfil-plan", methods=["POST"])
@permission_required("PROYECTOS_EDITAR")
def save_proyecto_perfil_plan(proyecto_id):
    Proyecto.query.get_or_404(proyecto_id)

    data = request.get_json(silent=True) or {}
    rows = data.get("rows") or []

    try:
        if not isinstance(rows, list):
            return jsonify({
                "mensaje": "El payload de planeación por perfil debe enviar 'rows' como lista."
            }), 400

        # ============================================================
        # 1) PERFILES Y MÓDULOS PERMITIDOS EN EL PROYECTO
        # ============================================================
        proyecto_perfil_ids = {
            int(x.perfil_id)
            for x in (
                ProyectoPerfil.query
                .filter_by(proyecto_id=proyecto_id)
                .filter(ProyectoPerfil.activo == True)
                .all()
            )
            if x.perfil_id
        }

        proyecto_modulo_ids = {
            int(x.modulo_id)
            for x in (
                ProyectoModulo.query
                .filter_by(proyecto_id=proyecto_id)
                .filter(ProyectoModulo.activo == True)
                .all()
            )
            if x.modulo_id
        }

        if rows and not proyecto_perfil_ids:
            return jsonify({
                "mensaje": (
                    "El proyecto no tiene perfiles asociados. "
                    "Primero agrega perfiles al proyecto antes de planear costos."
                )
            }), 400

        if rows and not proyecto_modulo_ids:
            return jsonify({
                "mensaje": (
                    "El proyecto no tiene módulos asociados. "
                    "Primero agrega módulos al proyecto antes de planear costos."
                )
            }), 400

        # ============================================================
        # 2) VALIDAR FILAS ANTES DE BORRAR / GUARDAR
        # ============================================================
        seen = set()
        normalized_rows = []

        for idx, row in enumerate(rows):
            if not isinstance(row, dict):
                return jsonify({
                    "mensaje": f"Fila inválida en planeación por perfil (índice {idx})."
                }), 400

            try:
                anio = int(row.get("anio") or 0)
                mes = int(row.get("mes") or 0)
                perfil_id = int(row.get("perfil_id") or 0)
                modulo_id = int(row.get("modulo_id") or 0)
            except Exception:
                return jsonify({
                    "mensaje": (
                        f"Fila inválida en planeación por perfil (índice {idx}). "
                        "Perfil y módulo deben ser IDs numéricos. "
                        f"Valores recibidos: perfil_id={row.get('perfil_id')}, "
                        f"modulo_id={row.get('modulo_id')}"
                    )
                }), 400

            if anio <= 0 or mes < 1 or mes > 12 or perfil_id <= 0 or modulo_id <= 0:
                return jsonify({
                    "mensaje": f"Fila inválida en planeación por perfil (índice {idx})"
                }), 400

            # --------------------------------------------------------
            # Validar que el perfil esté asociado directamente al proyecto
            # --------------------------------------------------------
            if perfil_id not in proyecto_perfil_ids:
                return jsonify({
                    "mensaje": (
                        f"El perfil {perfil_id} no está asociado al proyecto. "
                        "Primero debes agregarlo en la configuración del proyecto."
                    )
                }), 400

            # --------------------------------------------------------
            # Validar que el módulo esté asociado directamente al proyecto
            # --------------------------------------------------------
            if modulo_id not in proyecto_modulo_ids:
                return jsonify({
                    "mensaje": (
                        f"El módulo {modulo_id} no está asociado al proyecto. "
                        "Primero debes agregarlo en la configuración del proyecto."
                    )
                }), 400

            perfil = Perfil.query.get(perfil_id)
            if not perfil:
                return jsonify({
                    "mensaje": f"Perfil no existe: {perfil_id}"
                }), 400

            modulo = Modulo.query.get(modulo_id)
            if not modulo:
                return jsonify({
                    "mensaje": f"Módulo no existe: {modulo_id}"
                }), 400

            # --------------------------------------------------------
            # Validar que el módulo pertenezca al perfil
            # --------------------------------------------------------
            relacion_perfil_modulo = (
                ModuloPerfil.query
                .filter(ModuloPerfil.perfil_id == perfil_id)
                .filter(ModuloPerfil.modulo_id == modulo_id)
                .filter(ModuloPerfil.activo == True)
                .first()
            )

            if not relacion_perfil_modulo:
                return jsonify({
                    "mensaje": (
                        f"El módulo '{modulo.nombre}' no pertenece al perfil "
                        f"'{perfil.nombre}'. Selecciona un módulo permitido "
                        "para ese perfil."
                    )
                }), 400

            consultor_id_raw = row.get("consultor_id")
            consultor_id = None

            if consultor_id_raw not in ("", "null", "None", None):
                try:
                    consultor_id = int(consultor_id_raw)
                except Exception:
                    return jsonify({
                        "mensaje": f"Consultor inválido en planeación por perfil (índice {idx})"
                    }), 400

                consultor = Consultor.query.get(consultor_id)
                if not consultor:
                    return jsonify({
                        "mensaje": f"Consultor no existe: {consultor_id}"
                    }), 400

            key = (anio, mes, perfil_id, modulo_id, consultor_id)

            if key in seen:
                return jsonify({
                    "mensaje": (
                        "Fila duplicada para el mismo periodo/perfil/módulo/consultor: "
                        f"{anio}-{mes}, perfil {perfil_id}, módulo {modulo_id}, "
                        f"consultor {consultor_id or 'SIN CONSULTOR'}"
                    )
                }), 400

            seen.add(key)

            # --------------------------------------------------------
            # Recalcular costo e ingreso en backend
            # --------------------------------------------------------
            valor_hora_ingreso_raw = (
                row.get("valor_hora_ingreso")
                if row.get("valor_hora_ingreso") not in ("", None, "null", "None")
                else row.get("fte_estimado")
            )

            horas_estimadas = _cost_parse_decimal(row.get("horas_estimadas")) or Decimal("0.00")
            valor_hora_ingreso_dec = _cost_parse_decimal(valor_hora_ingreso_raw) or Decimal("0.00")
            valor_hora_planeado_dec = _cost_parse_decimal(row.get("valor_hora_planeado")) or Decimal("0.00")

            costo_estimado = (horas_estimadas * valor_hora_planeado_dec).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP
            )

            ingreso_estimado = (horas_estimadas * valor_hora_ingreso_dec).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP
            )

            try:
                orden = int(row.get("orden") or idx)
            except Exception:
                orden = idx

            normalized_rows.append({
                "anio": anio,
                "mes": mes,
                "perfil_id": perfil_id,
                "modulo_id": modulo_id,
                "consultor_id": consultor_id,
                "horas_estimadas": horas_estimadas,
                "fte_estimado": valor_hora_ingreso_dec,
                "valor_hora_planeado": valor_hora_planeado_dec,
                "costo_estimado": costo_estimado,
                "ingreso_estimado": ingreso_estimado,
                "observacion": (row.get("observacion") or "").strip() or None,
                "orden": orden,
                "activo": _to_bool2(row.get("activo"), default=True),
            })

        # ============================================================
        # 3) BORRAR PLANEACIÓN ACTUAL DEL PROYECTO
        #    IMPORTANTE:
        #    Ya NO se autoasocian módulos al proyecto desde esta ruta.
        # ============================================================
        ProyectoPerfilPlan.query.filter_by(
            proyecto_id=proyecto_id
        ).delete(synchronize_session=False)

        # ============================================================
        # 4) INSERTAR NUEVA PLANEACIÓN
        # ============================================================
        for row in normalized_rows:
            db.session.add(ProyectoPerfilPlan(
                proyecto_id=proyecto_id,
                anio=row["anio"],
                mes=row["mes"],
                perfil_id=row["perfil_id"],
                modulo_id=row["modulo_id"],
                consultor_id=row["consultor_id"],

                horas_estimadas=row["horas_estimadas"],
                fte_estimado=row["fte_estimado"],
                valor_hora_planeado=row["valor_hora_planeado"],
                costo_estimado=row["costo_estimado"],
                ingreso_estimado=row["ingreso_estimado"],

                observacion=row["observacion"],
                orden=row["orden"],
                activo=row["activo"],
            ))

        db.session.commit()

        # ============================================================
        # 5) RETORNAR FILAS GUARDADAS
        # ============================================================
        rows_db = (
            ProyectoPerfilPlan.query
            .options(
                joinedload(ProyectoPerfilPlan.perfil),
                joinedload(ProyectoPerfilPlan.modulo),
                joinedload(ProyectoPerfilPlan.consultor),
            )
            .filter_by(proyecto_id=proyecto_id)
            .order_by(
                ProyectoPerfilPlan.anio.asc(),
                ProyectoPerfilPlan.mes.asc(),
                ProyectoPerfilPlan.orden.asc(),
                ProyectoPerfilPlan.id.asc(),
            )
            .all()
        )

        return jsonify({
            "mensaje": "Planeación por perfil guardada",
            "rows": [_perfil_plan_to_dict(x) for x in rows_db]
        }), 200

    except IntegrityError as e:
        db.session.rollback()
        app.logger.exception("Error de integridad guardando planeación por perfil")

        return jsonify({
            "mensaje": "No se pudo guardar la planeación por perfil por un conflicto de datos duplicados o llaves foráneas.",
            "detalle": str(e)
        }), 400

    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error guardando planeación por perfil")

        return jsonify({
            "mensaje": f"No se pudo guardar la planeación por perfil: {str(e)}"
        }), 500

@bp.route("/proyectos/<int:proyecto_id>/costos/costos-adicionales", methods=["POST"])
@permission_required("PROYECTOS_EDITAR")
def save_proyecto_costos_adicionales(proyecto_id):
    Proyecto.query.get_or_404(proyecto_id)
    data = request.get_json(silent=True) or {}
    rows = data.get("rows") or []

    for idx, row in enumerate(rows):
        anio = int(row.get("anio") or 0)
        mes = int(row.get("mes") or 0)

        if anio <= 0 or mes < 1 or mes > 12:
            return jsonify({"mensaje": f"Fila inválida en costos adicionales (índice {idx})"}), 400

    ProyectoCostoAdicional.query.filter_by(proyecto_id=proyecto_id).delete()

    for row in rows:
        tipo_costo = (row.get("tipo_costo") or "").strip().upper()
        if tipo_costo not in {"OPERATIVO", "ADMINISTRATIVO", "OTRO"}:
            tipo_costo = "OTRO"

        db.session.add(ProyectoCostoAdicional(
            proyecto_id=proyecto_id,
            anio=int(row.get("anio")),
            mes=int(row.get("mes")),
            tipo_costo=tipo_costo,
            categoria=(row.get("categoria") or "").strip() or None,
            descripcion=(row.get("descripcion") or "").strip() or None,
            valor=_cost_parse_decimal(row.get("valor")) or Decimal("0"),
            activo=_to_bool2(row.get("activo"), default=True),
        ))

    db.session.commit()

    rows_db = (
        ProyectoCostoAdicional.query
        .filter_by(proyecto_id=proyecto_id)
        .order_by(ProyectoCostoAdicional.anio.asc(), ProyectoCostoAdicional.mes.asc(), ProyectoCostoAdicional.id.asc())
        .all()
    )

    return jsonify({
        "mensaje": "Costos adicionales guardados",
        "rows": [_costo_adicional_to_dict(x) for x in rows_db]
    }), 200

@bp.route("/proyectos/<int:proyecto_id>/costos/resumen", methods=["GET"])
@permission_required("PROYECTOS_VER")
def get_proyecto_costos_resumen(proyecto_id):
    def _dec(v):
        try:
            if v is None or v == "":
                return Decimal("0.00")
            if isinstance(v, Decimal):
                return v
            return Decimal(str(v))
        except Exception:
            return Decimal("0.00")

    def _money(v):
        return float(_dec(v).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))

    def _fecha_to_parts(v):
        if v is None:
            return None, None

        if isinstance(v, datetime):
            return v.year, v.month

        if isinstance(v, date):
            return v.year, v.month

        s = str(v).strip()
        if not s:
            return None, None

        if len(s) >= 7 and s[4] == "-":
            try:
                return int(s[:4]), int(s[5:7])
            except Exception:
                return None, None

        try:
            dt = datetime.fromisoformat(s[:19])
            return dt.year, dt.month
        except Exception:
            return None, None

    def _empty_period(anio, mes):
        return {
            "anio": int(anio),
            "mes": int(mes),
            "ingreso_planeado": Decimal("0.00"),
            "costo_planeado": Decimal("0.00"),
            "gasto_operativo_planeado": Decimal("0.00"),
            "costo_administrativo_planeado": Decimal("0.00"),
            "ebitda_planeado": Decimal("0.00"),
            "margen_planeado_pct": Decimal("0.00"),
            "horas_planeadas": Decimal("0.00"),
            "costo_adicional": Decimal("0.00"),
            "horas_reales": Decimal("0.00"),
            "costo_real": Decimal("0.00"),
        }

    filtro_equipos = [
        str(v).strip().upper()
        for v in request.args.getlist("equipo")
        if str(v).strip()
    ]

    filtro_modulos = [
        str(v).strip().upper()
        for v in request.args.getlist("modulo")
        if str(v).strip()
    ]

    filtro_consultores = [
        str(v).strip().lower()
        for v in request.args.getlist("consultor")
        if str(v).strip()
    ]

    hay_filtros = bool(filtro_equipos or filtro_modulos or filtro_consultores)

    p = (
        Proyecto.query.options(
            joinedload(Proyecto.presupuestos_mensuales),
            joinedload(Proyecto.perfiles_plan).joinedload(ProyectoPerfilPlan.modulo),
            joinedload(Proyecto.costos_adicionales),
        )
        .get_or_404(proyecto_id)
    )

    plan_mensual = {}

    # ---------------------------------------------------------
    # 1) Presupuesto mensual base
    # ---------------------------------------------------------
    for x in (p.presupuestos_mensuales or []):
        key = f"{int(x.anio):04d}-{int(x.mes):02d}"

        plan_mensual[key] = {
            "anio": int(x.anio),
            "mes": int(x.mes),
            "ingreso_planeado": _dec(x.ingreso_planeado),
            "costo_planeado": _dec(x.costo_planeado),
            "gasto_operativo_planeado": _dec(x.gasto_operativo_planeado),
            "costo_administrativo_planeado": _dec(x.costo_administrativo_planeado),
            "ebitda_planeado": _dec(x.ebitda_planeado),
            "margen_planeado_pct": _dec(x.margen_planeado_pct),
            "horas_planeadas": Decimal("0.00"),
            "costo_adicional": Decimal("0.00"),
            "horas_reales": Decimal("0.00"),
            "costo_real": Decimal("0.00"),
        }

    # ---------------------------------------------------------
    # 2) Horas planeadas desde planeación por perfil
    #    Nota:
    #    - Si filtras por módulo, se cruza contra módulo.
    #    - Si filtras solo por equipo/consultor, no se suma planeado,
    #      porque la planeación ya no tiene consultor/equipo asignado.
    # ---------------------------------------------------------
    for x in (p.perfiles_plan or []):
        if not bool(getattr(x, "activo", True)):
            continue

        modulo_nombre = ""
        if getattr(x, "modulo", None):
            modulo_nombre = (x.modulo.nombre or "").strip().upper()

        if filtro_modulos and modulo_nombre not in filtro_modulos:
            continue

        if (filtro_equipos or filtro_consultores) and not filtro_modulos:
            continue

        key = f"{int(x.anio):04d}-{int(x.mes):02d}"

        if key not in plan_mensual:
            plan_mensual[key] = _empty_period(x.anio, x.mes)

        plan_mensual[key]["horas_planeadas"] += _dec(x.horas_estimadas)

    # ---------------------------------------------------------
    # 3) Costos adicionales
    # ---------------------------------------------------------
    for x in (p.costos_adicionales or []):
        if not bool(getattr(x, "activo", True)):
            continue

        key = f"{int(x.anio):04d}-{int(x.mes):02d}"

        if key not in plan_mensual:
            plan_mensual[key] = _empty_period(x.anio, x.mes)

        plan_mensual[key]["costo_adicional"] += _dec(x.valor)

    # ---------------------------------------------------------
    # 4) Registros reales del proyecto
    #    Equipo real tomado desde Consultor.equipo_id -> Equipo.nombre
    #    con fallback a Registro.equipo.
    # ---------------------------------------------------------
    consultor_join_cond = db.or_(
        func.lower(func.trim(Registro.usuario_consultor)) == func.lower(func.trim(Consultor.usuario)),
        func.lower(func.trim(Registro.usuario_consultor)) == func.lower(func.trim(Consultor.nombre)),
    )

    rows_reg_query = _apply_project_filter_shared(
        db.session.query(
            Registro.fecha.label("fecha"),
            Registro.usuario_consultor.label("usuario_consultor"),
            Consultor.id.label("consultor_id"),
            Consultor.usuario.label("consultor_usuario"),
            Consultor.nombre.label("consultor_nombre"),
            Equipo.nombre.label("equipo_nombre"),
            Registro.modulo.label("modulo_nombre"),
            func.coalesce(
                func.sum(
                    func.coalesce(Registro.tiempo_invertido, Registro.total_horas, 0)
                ),
                0
            ).label("horas"),
        )
        .select_from(Registro)
        .outerjoin(Consultor, consultor_join_cond)
        .outerjoin(Equipo, Consultor.equipo_id == Equipo.id),
        proyecto_id
    )

    if filtro_equipos:
        rows_reg_query = rows_reg_query.filter(
            func.upper(
                func.coalesce(Equipo.nombre, Registro.equipo, "")
            ).in_(filtro_equipos)
        )

    if filtro_modulos:
        rows_reg_query = rows_reg_query.filter(
            func.upper(func.coalesce(Registro.modulo, "")).in_(filtro_modulos)
        )

    if filtro_consultores:
        rows_reg_query = rows_reg_query.filter(
            func.lower(
                func.coalesce(Consultor.usuario, Registro.usuario_consultor, "")
            ).in_(filtro_consultores)
        )

    rows_reg_query = rows_reg_query.group_by(
        Registro.fecha,
        Registro.usuario_consultor,
        Consultor.id,
        Consultor.usuario,
        Consultor.nombre,
        Equipo.nombre,
        Registro.modulo,
    )

    rows_reg = rows_reg_query.all()

    detalle_consultores_mes = {}
    periodos_reales_filtrados = set()

    for r in rows_reg:
        anio, mes = _fecha_to_parts(r.fecha)
        if not anio or not mes:
            continue

        periodo = f"{anio:04d}-{mes:02d}"
        periodos_reales_filtrados.add(periodo)

        horas = _dec(r.horas)

        if horas <= 0:
            continue

        consultor_id = r.consultor_id
        consultor_nombre = r.consultor_nombre or r.usuario_consultor or "SIN NOMBRE"
        usuario_consultor = (
            r.consultor_usuario
            or r.usuario_consultor
            or ""
        ).strip().lower()

        valor_hora = Decimal("0.00")

        if consultor_id:
            presupuesto = None

            try:
                presupuesto = _presupuesto_consultor_mes(consultor_id, anio, mes)
            except Exception:
                presupuesto = None

            if presupuesto:
                valor_hora = _dec(presupuesto.get("valor_hora"))

        costo_real_reg = (horas * valor_hora).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP
        )

        app.logger.info("COSTO_REAL_PROYECTO_DEBUG", {
            "periodo": periodo,
            "consultor_id": consultor_id,
            "consultor": consultor_nombre,
            "usuario_consultor": usuario_consultor,
            "equipo": r.equipo_nombre,
            "modulo": r.modulo_nombre,
            "horas": str(horas),
            "valor_hora": str(valor_hora),
            "costo_real_reg": str(costo_real_reg),
        })

        if periodo not in plan_mensual:
            plan_mensual[periodo] = _empty_period(anio, mes)

        plan_mensual[periodo]["horas_reales"] += horas
        plan_mensual[periodo]["costo_real"] += costo_real_reg

        detail_key = f"{periodo}||{consultor_id or usuario_consultor or 'SIN_USUARIO'}"

        if detail_key not in detalle_consultores_mes:
            detalle_consultores_mes[detail_key] = {
                "periodo": periodo,
                "consultor_id": consultor_id,
                "consultor": consultor_nombre,
                "usuario_consultor": usuario_consultor or None,
                "horas_reales": Decimal("0.00"),
                "valor_hora": valor_hora,
                "costo_real": Decimal("0.00"),
            }

        detalle_consultores_mes[detail_key]["horas_reales"] += horas
        detalle_consultores_mes[detail_key]["costo_real"] += costo_real_reg

        if detalle_consultores_mes[detail_key]["valor_hora"] <= 0 and valor_hora > 0:
            detalle_consultores_mes[detail_key]["valor_hora"] = valor_hora

    # ---------------------------------------------------------
    # 5) Si hay filtros, dejar solo períodos relacionados
    # ---------------------------------------------------------
    if hay_filtros:
        plan_mensual = {
            key: value
            for key, value in plan_mensual.items()
            if key in periodos_reales_filtrados
            or _dec(value.get("horas_planeadas")) > 0
            or _dec(value.get("horas_reales")) > 0
        }

    # ---------------------------------------------------------
    # 6) Salida mensual
    # ---------------------------------------------------------
    meses_out = []

    total_ingreso_planeado = Decimal("0.00")
    total_costo_planeado = Decimal("0.00")
    total_costo_real = Decimal("0.00")

    for key in sorted(plan_mensual.keys()):
        item = plan_mensual[key]

        ingreso_planeado = _dec(item["ingreso_planeado"])
        costo_planeado = _dec(item["costo_planeado"])
        costo_adicional = _dec(item["costo_adicional"])
        costo_planeado_total = costo_planeado + costo_adicional
        costo_real = _dec(item["costo_real"])
        variacion_costo = costo_planeado_total - costo_real

        pct_uso = None
        if costo_planeado_total > 0:
            pct_uso = float(
                ((costo_real / costo_planeado_total) * Decimal("100")).quantize(
                    Decimal("0.01"),
                    rounding=ROUND_HALF_UP
                )
            )

        meses_out.append({
            "periodo": key,
            "anio": item["anio"],
            "mes": item["mes"],
            "ingreso_planeado": _money(ingreso_planeado),
            "costo_planeado": _money(costo_planeado),
            "costo_adicional": _money(costo_adicional),
            "costo_planeado_total": _money(costo_planeado_total),
            "horas_planeadas": _money(item["horas_planeadas"]),
            "horas_reales": _money(item["horas_reales"]),
            "costo_real": _money(costo_real),
            "variacion_costo": _money(variacion_costo),
            "pct_uso": pct_uso,
        })

        total_ingreso_planeado += ingreso_planeado
        total_costo_planeado += costo_planeado_total
        total_costo_real += costo_real

    # ---------------------------------------------------------
    # 7) Cards
    # ---------------------------------------------------------
    if hay_filtros:
        ingreso_total = total_ingreso_planeado
        costo_objetivo_total = total_costo_planeado
    else:
        ingreso_total = _dec(getattr(p, "ingreso_total", None))
        costo_objetivo_total = _dec(getattr(p, "costo_objetivo_total", None))

        if ingreso_total <= 0:
            ingreso_total = total_ingreso_planeado

        if costo_objetivo_total <= 0:
            costo_objetivo_total = total_costo_planeado

    # ---------------------------------------------------------
    # 8) Detalle por consultor y mes
    # ---------------------------------------------------------
    detalle_out = []

    for _, x in sorted(
        detalle_consultores_mes.items(),
        key=lambda kv: (kv[1]["periodo"], (kv[1]["consultor"] or "").upper())
    ):
        detalle_out.append({
            "periodo": x["periodo"],
            "consultor_id": x["consultor_id"],
            "consultor": x["consultor"],
            "usuario_consultor": x["usuario_consultor"],
            "horas_reales": _money(x["horas_reales"]),
            "valor_hora": _money(x["valor_hora"]),
            "costo_real": _money(x["costo_real"]),
        })

    return jsonify({
        "cards": {
            "ingreso_total": _money(ingreso_total),
            "costo_objetivo_total": _money(costo_objetivo_total),
            "costo_planeado_acumulado": _money(total_costo_planeado),
            "costo_real_acumulado": _money(total_costo_real),
            "margen_planeado": _money(ingreso_total - total_costo_planeado),
            "margen_real": _money(ingreso_total - total_costo_real),
        },
        "meses": meses_out,
        "detalle_consultores_mes": detalle_out,
    }), 200

@bp.route("/proyectos/<int:proyecto_id>/costos/graficas", methods=["GET"])
@permission_required("PROYECTOS_VER")
def get_proyecto_costos_graficas(proyecto_id):
    def _dec(v):
        try:
            if v is None or v == "":
                return Decimal("0.00")
            if isinstance(v, Decimal):
                return v
            return Decimal(str(v))
        except Exception:
            return Decimal("0.00")

    def _money(v):
        return float(_dec(v).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))

    def _up(v):
        return str(v or "").strip().upper()

    def _low(v):
        return str(v or "").strip().lower()

    def _mod_key(v):
        s = str(v or "").strip().upper()
        s = unicodedata.normalize("NFD", s)
        s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
        s = re.sub(r"\s+", " ", s)
        return s

    def _fecha_to_parts(v):
        if v is None:
            return None, None, None

        if isinstance(v, datetime):
            return v.year, v.month, v.date()

        if isinstance(v, date):
            return v.year, v.month, v

        s = str(v).strip()
        if not s:
            return None, None, None

        try:
            dt = datetime.fromisoformat(s[:19])
            return dt.year, dt.month, dt.date()
        except Exception:
            pass

        try:
            dt = datetime.strptime(s[:10], "%Y-%m-%d")
            return dt.year, dt.month, dt.date()
        except Exception:
            return None, None, None

    def _period_key(anio, mes):
        return f"{int(anio):04d}-{int(mes):02d}"

    def _period_tuple(anio, mes):
        return (int(anio), int(mes))

    def _ensure_bucket(perfil_nombre):
        if perfil_nombre not in horas_por_perfil:
            horas_por_perfil[perfil_nombre] = {
                "perfil": perfil_nombre,
                "estimadas": Decimal("0.00"),
                "reales": Decimal("0.00"),
            }

        if perfil_nombre not in costos_por_perfil:
            costos_por_perfil[perfil_nombre] = {
                "perfil": perfil_nombre,
                "estimado": Decimal("0.00"),
                "real": Decimal("0.00"),
            }

        if perfil_nombre not in acumulado_horas_por_perfil:
            acumulado_horas_por_perfil[perfil_nombre] = {
                "perfil": perfil_nombre,
                "estimadas": Decimal("0.00"),
                "reales": Decimal("0.00"),
            }

    # ---------------------------------------------------------
    # Filtros globales
    # ---------------------------------------------------------
    filtro_equipos = [
        str(v).strip().upper()
        for v in request.args.getlist("equipo")
        if str(v).strip()
    ]

    filtro_modulos = [
        _mod_key(v)
        for v in request.args.getlist("modulo")
        if str(v).strip()
    ]

    filtro_consultores = [
        str(v).strip().lower()
        for v in request.args.getlist("consultor")
        if str(v).strip()
    ]

    # ---------------------------------------------------------
    # Proyecto
    # ---------------------------------------------------------
    p = (
        Proyecto.query.options(
            joinedload(Proyecto.perfiles_plan).joinedload(ProyectoPerfilPlan.perfil),
            joinedload(Proyecto.perfiles_plan).joinedload(ProyectoPerfilPlan.modulo),
            joinedload(Proyecto.presupuestos_mensuales),
            joinedload(Proyecto.costos_adicionales),
            joinedload(Proyecto.mapeos),
        )
        .get_or_404(proyecto_id)
    )

    # ---------------------------------------------------------
    # Rango de meses
    # ---------------------------------------------------------
    anio_desde = request.args.get("anio_desde", type=int)
    mes_desde = request.args.get("mes_desde", type=int)
    anio_hasta = request.args.get("anio_hasta", type=int)
    mes_hasta = request.args.get("mes_hasta", type=int)

    anio_single = request.args.get("anio", type=int)
    mes_single = request.args.get("mes", type=int)

    if anio_desde and mes_desde and anio_hasta and mes_hasta:
        if mes_desde < 1 or mes_desde > 12 or mes_hasta < 1 or mes_hasta > 12:
            return jsonify({
                "mensaje": "Rango de meses inválido. El mes debe estar entre 1 y 12."
            }), 400

        periodo_desde = _period_tuple(anio_desde, mes_desde)
        periodo_hasta = _period_tuple(anio_hasta, mes_hasta)

        if periodo_desde > periodo_hasta:
            periodo_desde, periodo_hasta = periodo_hasta, periodo_desde

    elif anio_single and mes_single:
        if mes_single < 1 or mes_single > 12:
            return jsonify({
                "mensaje": "Periodo inválido. El mes debe estar entre 1 y 12."
            }), 400

        periodo_desde = _period_tuple(anio_single, mes_single)
        periodo_hasta = _period_tuple(anio_single, mes_single)

    else:
        hoy = datetime.now(ZoneInfo("America/Bogota")).date()
        periodo_desde = _period_tuple(hoy.year, hoy.month)
        periodo_hasta = _period_tuple(hoy.year, hoy.month)

    anio_desde, mes_desde = periodo_desde
    anio_hasta, mes_hasta = periodo_hasta

    def _period_in_range(anio_periodo, mes_periodo):
        if not anio_periodo or not mes_periodo:
            return False

        current = _period_tuple(anio_periodo, mes_periodo)
        return periodo_desde <= current <= periodo_hasta

    # ---------------------------------------------------------
    # Perfil vigente del consultor como fallback
    # ---------------------------------------------------------
    consultor_perfiles_rows = (
        ConsultorPerfil.query
        .options(joinedload(ConsultorPerfil.perfil))
        .filter(ConsultorPerfil.activo == True)
        .order_by(
            ConsultorPerfil.consultor_id.asc(),
            ConsultorPerfil.fecha_inicio.desc(),
            ConsultorPerfil.id.desc(),
        )
        .all()
    )

    perfiles_por_consultor = defaultdict(list)

    for cp in consultor_perfiles_rows:
        perfiles_por_consultor[cp.consultor_id].append(cp)

    def _perfil_vigente(consultor_id, fecha_ref):
        if not consultor_id or not fecha_ref:
            return None

        rows_cp = perfiles_por_consultor.get(consultor_id, [])

        for cp in rows_cp:
            fi = cp.fecha_inicio
            ff = cp.fecha_fin

            if fi and fecha_ref < fi:
                continue

            if ff and fecha_ref > ff:
                continue

            if cp.perfil:
                return cp.perfil

        return None

    # ---------------------------------------------------------
    # Estructuras de salida
    # ---------------------------------------------------------
    horas_por_perfil = {}
    costos_por_perfil = {}
    acumulado_horas_por_perfil = {}

    # ---------------------------------------------------------
    # Mapa de planeación por periodo + módulo
    # Sirve para distribuir el real.
    # ---------------------------------------------------------
    planeacion_periodo_modulo = defaultdict(list)
    planeacion_modulo_global = defaultdict(list)

    for row in (p.perfiles_plan or []):
        if not bool(getattr(row, "activo", True)):
            continue

        if not row.anio or not row.mes:
            continue

        modulo_nombre = _mod_key(
            getattr(row.modulo, "nombre", None)
            if getattr(row, "modulo", None)
            else None
        )

        perfil_nombre = (
            getattr(getattr(row, "perfil", None), "nombre", None)
            or "SIN PERFIL"
        )

        if not modulo_nombre or not perfil_nombre:
            continue

        horas_est = _dec(row.horas_estimadas)
        costo_est = _dec(row.costo_estimado)

        item = {
            "perfil": perfil_nombre,
            "modulo": modulo_nombre,
            "horas_estimadas": horas_est,
            "costo_estimado": costo_est,
        }

        planeacion_periodo_modulo[
            (int(row.anio), int(row.mes), modulo_nombre)
        ].append(item)

        planeacion_modulo_global[modulo_nombre].append(item)

    def _distribuir_real_por_planeacion(anio_reg, mes_reg, modulo_reg, horas_real, costo_real):
        modulo_key = _mod_key(modulo_reg)

        if not modulo_key:
            return False

        targets = planeacion_periodo_modulo.get(
            (int(anio_reg), int(mes_reg), modulo_key),
            []
        )

        # Si no hay planeación exacta para ese mes, usa la planeación global del módulo.
        if not targets:
            targets = planeacion_modulo_global.get(modulo_key, [])

        if not targets:
            return False

        total_horas_peso = sum((_dec(t["horas_estimadas"]) for t in targets), Decimal("0.00"))
        total_costo_peso = sum((_dec(t["costo_estimado"]) for t in targets), Decimal("0.00"))

        usar_peso_horas = total_horas_peso > 0
        usar_peso_costos = total_costo_peso > 0

        total_targets = Decimal(str(len(targets)))

        for t in targets:
            perfil_nombre = t["perfil"]
            _ensure_bucket(perfil_nombre)

            if usar_peso_horas:
                factor_horas = _dec(t["horas_estimadas"]) / total_horas_peso
            else:
                factor_horas = Decimal("1.00") / total_targets

            if usar_peso_costos:
                factor_costos = _dec(t["costo_estimado"]) / total_costo_peso
            else:
                factor_costos = factor_horas

            horas_asignadas = (horas_real * factor_horas).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP
            )

            costo_asignado = (costo_real * factor_costos).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP
            )

            horas_por_perfil[perfil_nombre]["reales"] += horas_asignadas
            costos_por_perfil[perfil_nombre]["real"] += costo_asignado
            acumulado_horas_por_perfil[perfil_nombre]["reales"] += horas_asignadas

        return True

    # ---------------------------------------------------------
    # PLANEADO por perfil dentro del rango
    # ---------------------------------------------------------
    for row in (p.perfiles_plan or []):
        if not bool(getattr(row, "activo", True)):
            continue

        if not row.anio or not row.mes:
            continue

        if not _period_in_range(row.anio, row.mes):
            continue

        row_modulo = _mod_key(
            getattr(row.modulo, "nombre", None)
            if getattr(row, "modulo", None)
            else None
        )

        if filtro_modulos and row_modulo not in filtro_modulos:
            continue

        perfil_nombre = (
            getattr(getattr(row, "perfil", None), "nombre", None)
            or "SIN PERFIL"
        )

        _ensure_bucket(perfil_nombre)

        horas_est = _dec(row.horas_estimadas)
        costo_est = _dec(row.costo_estimado)

        horas_por_perfil[perfil_nombre]["estimadas"] += horas_est
        costos_por_perfil[perfil_nombre]["estimado"] += costo_est
        acumulado_horas_por_perfil[perfil_nombre]["estimadas"] += horas_est

    # ---------------------------------------------------------
    # REAL por perfil desde registros del proyecto
    # ---------------------------------------------------------
    consultor_join_cond = db.or_(
        func.lower(func.trim(Registro.usuario_consultor)) == func.lower(func.trim(Consultor.usuario)),
        func.lower(func.trim(Registro.usuario_consultor)) == func.lower(func.trim(Consultor.nombre)),
    )

    rows_reg = (
        _apply_project_filter_shared(
            db.session.query(Registro, Consultor, Equipo)
            .select_from(Registro)
            .outerjoin(Consultor, consultor_join_cond)
            .outerjoin(Equipo, Consultor.equipo_id == Equipo.id),
            proyecto_id
        )
        .all()
    )

    for reg, consultor, equipo in rows_reg:
        ry, rm, fecha_ref = _fecha_to_parts(reg.fecha)

        if not ry or not rm:
            continue

        if not _period_in_range(ry, rm):
            continue

        modulo_reg = _mod_key(reg.modulo)
        equipo_reg = _up(getattr(equipo, "nombre", None) or reg.equipo)
        usuario_reg = _low(getattr(consultor, "usuario", None) or reg.usuario_consultor)

        if filtro_modulos and modulo_reg not in filtro_modulos:
            continue

        if filtro_equipos and equipo_reg not in filtro_equipos:
            continue

        if filtro_consultores and usuario_reg not in filtro_consultores:
            continue

        horas_real = _dec(
            reg.total_horas
            if reg.total_horas is not None
            else (
                reg.tiempo_invertido
                if reg.tiempo_invertido is not None
                else 0
            )
        )

        if horas_real <= 0:
            continue

        valor_hora = Decimal("0.00")

        if consultor and getattr(consultor, "id", None):
            try:
                presupuesto = _presupuesto_consultor_mes(
                    consultor.id,
                    int(ry),
                    int(rm)
                )

                if presupuesto:
                    valor_hora = _dec(presupuesto.get("valor_hora"))
            except Exception:
                valor_hora = Decimal("0.00")

        costo_real = (horas_real * valor_hora).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP
        )

        # 1) Primero intenta distribuir por planeación del módulo.
        asignado_por_planeacion = _distribuir_real_por_planeacion(
            ry,
            rm,
            modulo_reg,
            horas_real,
            costo_real
        )

        if asignado_por_planeacion:
            continue

        # 2) Fallback: perfil vigente del consultor.
        perfil_obj = _perfil_vigente(getattr(consultor, "id", None), fecha_ref)
        perfil_nombre = getattr(perfil_obj, "nombre", None)

        # 3) Último fallback: no lo ocultamos, pero lo marcamos mejor.
        if not perfil_nombre:
            perfil_nombre = f"MÓDULO SIN PLANEACIÓN: {modulo_reg or 'SIN MÓDULO'}"

        _ensure_bucket(perfil_nombre)

        horas_por_perfil[perfil_nombre]["reales"] += horas_real
        costos_por_perfil[perfil_nombre]["real"] += costo_real
        acumulado_horas_por_perfil[perfil_nombre]["reales"] += horas_real

    # ---------------------------------------------------------
    # Serialización
    # ---------------------------------------------------------
    horas_out = [
        {
            "perfil": x["perfil"],
            "estimadas": _money(x["estimadas"]),
            "reales": _money(x["reales"]),
        }
        for x in sorted(
            horas_por_perfil.values(),
            key=lambda r: max(r["estimadas"], r["reales"]),
            reverse=True
        )
    ]

    costos_out = [
        {
            "perfil": x["perfil"],
            "estimado": _money(x["estimado"]),
            "real": _money(x["real"]),
        }
        for x in sorted(
            costos_por_perfil.values(),
            key=lambda r: max(r["estimado"], r["real"]),
            reverse=True
        )
    ]

    acumulado_out = [
        {
            "perfil": x["perfil"],
            "estimadas": _money(x["estimadas"]),
            "reales": _money(x["reales"]),
        }
        for x in sorted(
            acumulado_horas_por_perfil.values(),
            key=lambda r: max(r["estimadas"], r["reales"]),
            reverse=True
        )
    ]

    return jsonify({
        "anio_desde": int(anio_desde),
        "mes_desde": int(mes_desde),
        "anio_hasta": int(anio_hasta),
        "mes_hasta": int(mes_hasta),
        "periodo_desde": _period_key(anio_desde, mes_desde),
        "periodo_hasta": _period_key(anio_hasta, mes_hasta),
        "periodo": f"{_period_key(anio_desde, mes_desde)} a {_period_key(anio_hasta, mes_hasta)}",

        "horas_por_perfil": horas_out,
        "costos_por_perfil": costos_out,
        "acumulado_horas_por_perfil": acumulado_out,
    }), 200

## -------------------------------
## Modulos (para categorizar proyectos y reportes)
@bp.route('/modulos', methods=['POST'])
@permission_required("MODULOS_CREAR")
def crear_modulo():
    data = request.get_json(silent=True) or {}
    nombre = (data.get("nombre") or "").strip()

    if not nombre:
        return jsonify({"mensaje": "nombre requerido"}), 400

    dupe = Modulo.query.filter(func.lower(Modulo.nombre) == nombre.lower()).first()
    if dupe:
        return jsonify({"mensaje": "El módulo ya existe"}), 400

    m = Modulo(nombre=nombre)
    db.session.add(m)
    db.session.commit()

    return jsonify({"mensaje": "Módulo creado", "modulo": {"id": m.id, "nombre": m.nombre}}), 201


@bp.route('/modulos/<int:id>', methods=['PUT'])
@permission_required("MODULOS_EDITAR")
def editar_modulo(id):
    m = Modulo.query.get_or_404(id)
    data = request.get_json(silent=True) or {}
    nombre = (data.get("nombre") or "").strip()

    if not nombre:
        return jsonify({"mensaje": "nombre requerido"}), 400

    dupe = (
        Modulo.query
        .filter(Modulo.id != id)
        .filter(func.lower(Modulo.nombre) == nombre.lower())
        .first()
    )
    if dupe:
        return jsonify({"mensaje": "Ya existe otro módulo con ese nombre"}), 400

    m.nombre = nombre
    db.session.commit()

    return jsonify({"mensaje": "Módulo actualizado", "modulo": {"id": m.id, "nombre": m.nombre}}), 200


@bp.route('/modulos/<int:id>', methods=['DELETE'])
@permission_required("MODULOS_ELIMINAR")
def eliminar_modulo(id):
    m = Modulo.query.get_or_404(id)

    # Opcional: bloquear si está asignado a consultores
    # if m.consultores and len(m.consultores) > 0:
    #     return jsonify({"mensaje": "No se puede eliminar: hay consultores con este módulo"}), 400

    db.session.delete(m)
    db.session.commit()

    return jsonify({"mensaje": "Módulo eliminado"}), 200


##Ruta para graficos de proyectos 

def _registro_fecha_expr():
    fecha_txt = func.left(func.cast(Registro.fecha, db.String), 10)
    return func.coalesce(
        func.str_to_date(fecha_txt, "%Y-%m-%d"),
        func.str_to_date(fecha_txt, "%d/%m/%Y"),
        func.str_to_date(fecha_txt, "%d-%m-%Y"),
    )

def _safe_float_report(v):
    try:
        return float(v or 0)
    except Exception:
        return 0.0

def _get_modulos_ids_de_perfiles(perfiles_ids):
    perfiles_ids = [int(x) for x in (perfiles_ids or []) if x]

    if not perfiles_ids:
        return []

    rows = (
        ModuloPerfil.query
        .filter(ModuloPerfil.perfil_id.in_(perfiles_ids))
        .filter(ModuloPerfil.activo == True)
        .all()
    )

    modulos_ids = sorted({
        int(r.modulo_id)
        for r in rows
        if r.modulo_id
    })

    return modulos_ids


def _proyecto_perfil_consultores_to_map(proyecto_id):
    rows = (
        ProyectoPerfilConsultor.query
        .options(
            joinedload(ProyectoPerfilConsultor.perfil),
            joinedload(ProyectoPerfilConsultor.consultor),
        )
        .filter(ProyectoPerfilConsultor.proyecto_id == proyecto_id)
        .filter(ProyectoPerfilConsultor.activo == True)
        .all()
    )

    mapa = {}
    detalle = []

    for row in rows:
        pid = str(row.perfil_id)
        cid = int(row.consultor_id)

        if pid not in mapa:
            mapa[pid] = []

        if cid not in mapa[pid]:
            mapa[pid].append(cid)

        detalle.append({
            "id": row.id,
            "proyecto_id": row.proyecto_id,
            "perfil_id": row.perfil_id,
            "consultor_id": row.consultor_id,
            "activo": bool(row.activo),
            "perfil": {
                "id": row.perfil.id,
                "nombre": row.perfil.nombre,
                "codigo": row.perfil.codigo,
            } if row.perfil else None,
            "consultor": {
                "id": row.consultor.id,
                "nombre": row.consultor.nombre,
                "usuario": row.consultor.usuario,
            } if row.consultor else None,
        })

    return mapa, detalle


def _proyecto_response_dict(p, include_modulos=True, include_fases=True, include_perfiles=True):
    out = proyecto_to_dict(
        p,
        include_modulos=include_modulos,
        include_fases=include_fases,
        include_perfiles=include_perfiles,
    )

    try:
        mapa, detalle = _proyecto_perfil_consultores_to_map(p.id)
        out["perfil_consultores"] = mapa
        out["perfil_consultores_detalle"] = detalle
    except Exception:
        out["perfil_consultores"] = {}
        out["perfil_consultores_detalle"] = []

    return out


def _validar_consultor_pertenece_a_modulos_perfil(consultor, perfil_id):
    relaciones = (
        ModuloPerfil.query
        .filter(ModuloPerfil.perfil_id == int(perfil_id))
        .filter(ModuloPerfil.activo == True)
        .all()
    )

    modulos_perfil_ids = {
        int(r.modulo_id)
        for r in relaciones
        if r.modulo_id
    }

    if not modulos_perfil_ids:
        return False, "El perfil no tiene módulos asociados"

    modulos_consultor_ids = {
        int(m.id)
        for m in (getattr(consultor, "modulos", None) or [])
        if getattr(m, "id", None)
    }

    if not modulos_consultor_ids:
        return False, (
            f"El consultor {consultor.nombre or consultor.usuario} "
            "no tiene módulos asociados"
        )

    if not modulos_consultor_ids.intersection(modulos_perfil_ids):
        return False, (
            f"El consultor {consultor.nombre or consultor.usuario} "
            f"no pertenece a los módulos del perfil {perfil_id}"
        )

    return True, None


def _save_proyecto_perfil_consultores(proyecto_id, perfil_consultores):
    """
    Espera payload:
    {
      "1": [3, 4],
      "2": [5]
    }

    Donde:
    - key = perfil_id
    - value = lista de consultor_id
    """
    if perfil_consultores is None:
        return

    if not isinstance(perfil_consultores, dict):
        raise ValueError("perfil_consultores debe ser un objeto")

    perfiles_proyecto_ids = {
        int(x.perfil_id)
        for x in (
            ProyectoPerfil.query
            .filter(ProyectoPerfil.proyecto_id == proyecto_id)
            .filter(ProyectoPerfil.activo == True)
            .all()
        )
        if x.perfil_id
    }

    if not perfiles_proyecto_ids:
        raise ValueError("El proyecto no tiene perfiles asociados")

    # Validar todo antes de borrar
    normalized = []
    seen = set()

    for perfil_id_raw, consultores_ids in perfil_consultores.items():
        try:
            perfil_id = int(perfil_id_raw)
        except Exception:
            raise ValueError(f"perfil_id inválido: {perfil_id_raw}")

        if perfil_id not in perfiles_proyecto_ids:
            raise ValueError(
                f"El perfil {perfil_id} no está asociado al proyecto"
            )

        perfil = Perfil.query.get(perfil_id)
        if not perfil:
            raise ValueError(f"Perfil no existe: {perfil_id}")

        if not isinstance(consultores_ids, list):
            raise ValueError(
                f"Los consultores del perfil {perfil_id} deben venir como lista"
            )

        for consultor_id_raw in consultores_ids:
            try:
                consultor_id = int(consultor_id_raw)
            except Exception:
                raise ValueError(f"consultor_id inválido: {consultor_id_raw}")

            key = (perfil_id, consultor_id)

            if key in seen:
                continue

            consultor = (
                Consultor.query
                .options(joinedload(Consultor.modulos))
                .get(consultor_id)
            )

            if not consultor:
                raise ValueError(f"Consultor no existe: {consultor_id}")

            if getattr(consultor, "activo", True) is False:
                raise ValueError(
                    f"El consultor {consultor.nombre or consultor.usuario} está inactivo"
                )

            ok, msg = _validar_consultor_pertenece_a_modulos_perfil(
                consultor,
                perfil_id
            )

            if not ok:
                raise ValueError(msg)

            seen.add(key)

            normalized.append({
                "perfil_id": perfil_id,
                "consultor_id": consultor_id,
            })

    ProyectoPerfilConsultor.query.filter_by(
        proyecto_id=proyecto_id
    ).delete(synchronize_session=False)

    for row in normalized:
        db.session.add(
            ProyectoPerfilConsultor(
                proyecto_id=proyecto_id,
                perfil_id=row["perfil_id"],
                consultor_id=row["consultor_id"],
                activo=True,
            )
        )

@bp.route("/reporte/horas-consultor-cliente-detalle", methods=["GET"])
@permission_required("GRAFICOS_VER")
def reporte_horas_consultor_cliente_detalle():
    try:
        usuario = _get_usuario_from_request()
        rol_req = _get_rol_from_request()

        if not usuario:
            return jsonify({"error": "Usuario no enviado"}), 400

        usuario_norm = (usuario or "").strip().lower()

        consultor_login = (
            Consultor.query.options(
                joinedload(Consultor.rol_obj),
                joinedload(Consultor.equipo_obj),
            )
            .filter(func.lower(Consultor.usuario) == usuario_norm)
            .first()
        )
        if not consultor_login:
            return jsonify({"error": "Consultor no encontrado"}), 404

        scope, val = _scope_for_graficos(consultor_login, rol_req)

        desde = (request.args.get("desde") or "").strip()
        hasta = (request.args.get("hasta") or "").strip()
        equipo_filter = (request.args.get("equipo") or "").strip().upper()
        cliente_filter = (request.args.get("cliente") or "").strip()
        consultor_filter = (request.args.get("consultor") or "").strip()
        modulo_filter = (request.args.get("modulo") or "").strip().upper()

        max_rows = request.args.get("max_rows", type=int) or 5000
        max_rows = min(max(max_rows, 1), 20000)

        C = aliased(Consultor)
        E = aliased(Equipo)

        fecha_expr = _registro_fecha_expr()

        q = (
            Registro.query
            .options(
                joinedload(Registro.consultor).joinedload(Consultor.equipo_obj),
                joinedload(Registro.tarea),
                joinedload(Registro.ocupacion),
                joinedload(Registro.proyecto),
                joinedload(Registro.fase_proyecto),
            )
            .outerjoin(C, func.lower(Registro.usuario_consultor) == func.lower(C.usuario))
            .outerjoin(E, C.equipo_id == E.id)
        )

        # -------------------------
        # Scope
        # -------------------------
        if scope == "SELF":
            q = q.filter(func.lower(Registro.usuario_consultor) == usuario_norm)

        elif scope == "TEAM":
            if not int(val or 0):
                return jsonify({"error": "Consultor sin equipo asignado"}), 403
            q = q.filter(C.equipo_id == int(val))

        # ALL -> sin filtro

        # -------------------------
        # Filtros
        # -------------------------
        if desde:
            q = q.filter(fecha_expr >= desde)

        if hasta:
            q = q.filter(fecha_expr <= hasta)

        if equipo_filter:
            if scope == "TEAM":
                eq_login = (
                    (consultor_login.equipo_obj.nombre or "").strip().upper()
                    if consultor_login.equipo_obj else ""
                )
                if equipo_filter != eq_login:
                    return jsonify({"error": "No autorizado para consultar otro equipo"}), 403

            q = q.filter(func.upper(E.nombre) == equipo_filter)

        if cliente_filter:
            q = q.filter(Registro.cliente.ilike(f"%{cliente_filter}%"))

        if consultor_filter:
            q = q.filter(C.nombre.ilike(f"%{consultor_filter}%"))

        if modulo_filter:
            q = q.filter(func.upper(Registro.modulo) == modulo_filter)

        registros = (
            q.order_by(fecha_expr.desc(), C.nombre.asc(), Registro.id.desc())
             .limit(max_rows)
             .all()
        )

        # -------------------------
        # Presupuestos vigentes
        # -------------------------
        consultor_ids = sorted({
            r.consultor.id for r in registros
            if getattr(r, "consultor", None) and getattr(r.consultor, "id", None)
        })

        presupuesto_map = {}
        if consultor_ids:
            pres_rows = (
                ConsultorPresupuesto.query
                .filter(ConsultorPresupuesto.consultor_id.in_(consultor_ids))
                .filter(ConsultorPresupuesto.vigente == True)
                .all()
            )
            for p in pres_rows:
                presupuesto_map[int(p.consultor_id)] = {
                    "horas_base_mes": _safe_float_report(p.horas_base_mes),
                    "vr_perfil": _safe_float_report(p.vr_perfil),
                }

        # -------------------------
        # Acumuladores
        # -------------------------
        clientes_set = set()
        filtros_equipos = set()
        filtros_clientes = set()
        filtros_consultores = set()
        filtros_modulos = set()

        resumen_map = {}
        totales_cliente = defaultdict(float)
        total_general = 0.0

        graf_cliente = defaultdict(float)
        graf_consultor = defaultdict(float)
        graf_equipo = defaultdict(float)
        graf_fecha = defaultdict(float)

        registros_out = []

        for r in registros:
            consultor_obj = getattr(r, "consultor", None)
            tarea = getattr(r, "tarea", None)
            ocup = getattr(r, "ocupacion", None)

            consultor_id = getattr(consultor_obj, "id", None)
            consultor_nombre = (
                getattr(consultor_obj, "nombre", None)
                or (r.usuario_consultor or "SIN CONSULTOR")
            ).strip()

            equipo_nombre = (
                (consultor_obj.equipo_obj.nombre if consultor_obj and consultor_obj.equipo_obj else None)
                or (r.equipo or "SIN EQUIPO")
            )
            equipo_nombre = str(equipo_nombre).strip().upper()

            cliente_nombre = str(r.cliente or "SIN CLIENTE").strip()
            modulo_nombre = str(r.modulo or "SIN MODULO").strip().upper()

            horas = _safe_float_report(
                r.tiempo_invertido if r.tiempo_invertido is not None else r.total_horas
            )

            fecha_str = _safe_fecha_iso(r.fecha)

            if tarea and getattr(tarea, "codigo", None) and getattr(tarea, "nombre", None):
                tipo_tarea_str = f"{tarea.codigo} - {tarea.nombre}"
            else:
                tipo_tarea_str = (getattr(r, "tipo_tarea", None) or "").strip() or None

            clientes_set.add(cliente_nombre)
            filtros_equipos.add(equipo_nombre)
            filtros_clientes.add(cliente_nombre)
            filtros_consultores.add(consultor_nombre)
            filtros_modulos.add(modulo_nombre)

            key = consultor_id or f"user::{(r.usuario_consultor or '').strip().lower()}"

            if key not in resumen_map:
                resumen_map[key] = {
                    "consultorId": consultor_id,
                    "consultor": consultor_nombre,
                    "equipo": equipo_nombre,
                    "presupuestoHoras": 0.0,
                    "totalHoras": 0.0,
                    "diferenciaHoras": 0.0,
                    "porcentajeUso": None,
                    "clientes": defaultdict(float),
                }

            resumen_map[key]["clientes"][cliente_nombre] += horas
            resumen_map[key]["totalHoras"] += horas

            totales_cliente[cliente_nombre] += horas
            total_general += horas

            graf_cliente[cliente_nombre] += horas
            graf_consultor[consultor_nombre] += horas
            graf_equipo[equipo_nombre] += horas
            if fecha_str:
                graf_fecha[fecha_str] += horas

            registros_out.append({
                "id": r.id,
                "fecha": fecha_str,
                "cliente": cliente_nombre,
                "consultor": consultor_nombre,
                "consultorId": consultor_id,
                "usuario_consultor": (r.usuario_consultor or "").strip().lower(),
                "equipo": equipo_nombre,
                "modulo": modulo_nombre,
                "nroCasoCliente": r.nro_caso_cliente,
                "nroCasoInterno": r.nro_caso_interno,
                "nroCasoEscaladoSap": r.nro_caso_escalado,
                "tipoTarea": tipo_tarea_str,
                "ocupacion": ocup.nombre if ocup else None,
                "horaInicio": r.hora_inicio,
                "horaFin": r.hora_fin,
                "tiempoInvertido": _safe_float_report(r.tiempo_invertido),
                "tiempoFacturable": _safe_float_report(r.tiempo_facturable),
                "totalHoras": round(horas, 2),
                "descripcion": r.descripcion,
                "proyecto": r.proyecto.nombre if getattr(r, "proyecto", None) else None,
                "faseProyecto": r.fase_proyecto.nombre if getattr(r, "fase_proyecto", None) else None,
            })

        clientes = sorted(clientes_set)

        rows_out = []
        for _, item in resumen_map.items():
            presupuesto_horas = 0.0
            if item["consultorId"] and item["consultorId"] in presupuesto_map:
                presupuesto_horas = _safe_float_report(
                    presupuesto_map[item["consultorId"]]["horas_base_mes"]
                )

            total_horas = round(_safe_float_report(item["totalHoras"]), 2)
            diferencia_horas = round(presupuesto_horas - total_horas, 2)
            porcentaje = round((total_horas / presupuesto_horas) * 100, 2) if presupuesto_horas > 0 else None

            clientes_dict = {}
            for c in clientes:
                clientes_dict[c] = round(_safe_float_report(item["clientes"].get(c, 0)), 2)

            rows_out.append({
                "consultorId": item["consultorId"],
                "consultor": item["consultor"],
                "equipo": item["equipo"],
                "presupuestoHoras": round(presupuesto_horas, 2),
                "totalHoras": total_horas,
                "diferenciaHoras": diferencia_horas,
                "porcentajeUso": porcentaje,
                "clientes": clientes_dict,
            })

        rows_out.sort(key=lambda x: ((x["equipo"] or ""), (x["consultor"] or "")))

        totales_cliente_out = {
            c: round(_safe_float_report(v), 2)
            for c, v in totales_cliente.items()
        }

        graficos = {
            "porCliente": [
                {"name": k, "horas": round(v, 2)}
                for k, v in sorted(graf_cliente.items(), key=lambda x: x[1], reverse=True)
            ],
            "porConsultor": [
                {"name": k, "horas": round(v, 2)}
                for k, v in sorted(graf_consultor.items(), key=lambda x: x[1], reverse=True)
            ],
            "porEquipo": [
                {"name": k, "horas": round(v, 2)}
                for k, v in sorted(graf_equipo.items(), key=lambda x: x[1], reverse=True)
            ],
            "porFecha": [
                {"fecha": k, "horas": round(v, 2)}
                for k, v in sorted(graf_fecha.items(), key=lambda x: x[0])
            ],
        }

        filtros = {
            "equipos": sorted(filtros_equipos),
            "clientes": sorted(filtros_clientes),
            "consultores": sorted(filtros_consultores),
            "modulos": sorted(filtros_modulos),
        }

        return jsonify({
            "clientes": clientes,
            "rows": rows_out,
            "registros": registros_out,
            "totalesCliente": totales_cliente_out,
            "totalGeneral": round(total_general, 2),
            "graficos": graficos,
            "filtros": filtros,
        }), 200

    except Exception as e:
        app.logger.exception("❌ Error en /reporte/horas-consultor-cliente-detalle")
        return jsonify({"error": str(e)}), 500
    

def _normalize_iso_range(desde, hasta):
    d = (desde or "").strip()
    h = (hasta or "").strip()

    if d and h and d > h:
        return h, d

    return d, h

def _dashboard_costos_parse_periodo_request():
    modo = (request.args.get("modo") or "").strip().lower()

    # caso 1: mes / año
    mes = (request.args.get("mes") or "").strip()
    anio = (request.args.get("anio") or "").strip()

    # caso 2: rango de meses
    mes_desde = (request.args.get("mes_desde") or "").strip()
    anio_desde = (request.args.get("anio_desde") or "").strip()
    mes_hasta = (request.args.get("mes_hasta") or "").strip()
    anio_hasta = (request.args.get("anio_hasta") or "").strip()

    # caso 3: rango de fechas
    desde = (request.args.get("desde") or "").strip()
    hasta = (request.args.get("hasta") or "").strip()

    hoy = date.today()

    # ---------------------------
    # Mes / año
    # ---------------------------
    if modo == "mes" or (mes and anio and not desde and not hasta):
        try:
            y = int(anio)
            m = int(mes)
            if m < 1 or m > 12:
                raise ValueError("Mes inválido")
        except Exception:
            raise ValueError("Parámetros de mes/año inválidos")

        d = date(y, m, 1)
        if m == 12:
            h = date(y + 1, 1, 1) - timedelta(days=1)
        else:
            h = date(y, m + 1, 1) - timedelta(days=1)

        return d, h, "mes"

    # ---------------------------
    # Rango de meses
    # ---------------------------
    if modo == "rango_meses" or (mes_desde and anio_desde and mes_hasta and anio_hasta):
        try:
            md = int(mes_desde)
            ad = int(anio_desde)
            mh = int(mes_hasta)
            ah = int(anio_hasta)

            if md < 1 or md > 12 or mh < 1 or mh > 12:
                raise ValueError("Mes inválido")
        except Exception:
            raise ValueError("Parámetros de rango de meses inválidos")

        d = date(ad, md, 1)

        if mh == 12:
            h = date(ah + 1, 1, 1) - timedelta(days=1)
        else:
            h = date(ah, mh + 1, 1) - timedelta(days=1)

        if h < d:
            raise ValueError("El mes final no puede ser menor al mes inicial")

        return d, h, "rango_meses"

    # ---------------------------
    # Rango de fechas
    # ---------------------------
    if modo == "rango_fechas" or (desde or hasta):
        if not desde or not hasta:
            raise ValueError("Debes enviar desde y hasta")

        desde, hasta = _normalize_iso_range(desde, hasta)

        try:
            d = datetime.strptime(desde, "%Y-%m-%d").date()
            h = datetime.strptime(hasta, "%Y-%m-%d").date()
        except Exception:
            raise ValueError("Formato de fechas inválido, usa YYYY-MM-DD")

        return d, h, "rango_fechas"

    # default: mes actual
    d = date(hoy.year, hoy.month, 1)
    if hoy.month == 12:
        h = date(hoy.year + 1, 1, 1) - timedelta(days=1)
    else:
        h = date(hoy.year, hoy.month + 1, 1) - timedelta(days=1)

    return d, h, "mes"

# ==========================================
# CAPACIDAD SEMANAL / MENSUAL
# ==========================================

def _cap_norm(value):
    txt = str(value or "").strip().upper()
    txt = unicodedata.normalize("NFD", txt)
    return "".join(ch for ch in txt if unicodedata.category(ch) != "Mn")


def _cap_parse_iso_date(value):
    s = str(value or "").strip()
    if not s:
        return None

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except Exception:
            continue

    return None


def _cap_team_kind(equipo_nombre):
    eq = _cap_norm(equipo_nombre)
    if "BASIS" in eq:
        return "BASIS"
    if "FUNCIONAL" in eq:
        return "FUNCIONAL"
    return "OTRO"


def _cap_month_bounds(anio, mes):
    start = date(anio, mes, 1)
    if mes == 12:
        end = date(anio + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(anio, mes + 1, 1) - timedelta(days=1)
    return start, end


def _cap_weeks_for_month(anio, mes):
    """
    Semanas calendario lunes-domingo, recortadas al mes.
    """
    month_start, month_end = _cap_month_bounds(anio, mes)
    cursor = month_start - timedelta(days=month_start.weekday())  # lunes
    out = []
    idx = 1

    while cursor <= month_end:
        week_start = cursor
        week_end = cursor + timedelta(days=6)

        real_start = max(week_start, month_start)
        real_end = min(week_end, month_end)

        out.append({
            "index": idx,
            "start": real_start,
            "end": real_end,
        })

        idx += 1
        cursor = week_end + timedelta(days=1)

    return out



def _cap_is_holiday(fecha_obj, co_holidays):
    return fecha_obj in co_holidays


def _cap_is_standard_workday(fecha_obj, co_holidays):
    """
    Regla única para todos:
    - lunes a viernes
    - sin festivos de Colombia
    """
    if fecha_obj.weekday() > 4:
        return False

    if _cap_is_holiday(fecha_obj, co_holidays):
        return False

    return True


def _cap_meta_hours_for_day(d: date, co_holidays=None):
    co_holidays = co_holidays or set()

    if not _cap_is_standard_workday(d, co_holidays):
        return Decimal("0.00")

    return Decimal("8.00") if d.weekday() == 0 else Decimal("9.00")



def _cap_work_days_text():
    return "Lunes 8 h / martes a viernes 9 h (sin festivos CO)"


def _cap_parse_month_year_from_request():
    tz = ZoneInfo("America/Bogota")
    now = datetime.now(tz)

    mes_raw = (request.args.get("mes") or "").strip()
    anio_raw = (request.args.get("anio") or "").strip()

    mes = now.month
    anio = now.year

    if mes_raw and re.match(r"^\d{4}-\d{2}$", mes_raw):
        y, m = mes_raw.split("-")
        anio = int(y)
        mes = int(m)
        return anio, mes

    if mes_raw.isdigit():
        mes = int(mes_raw)

    if anio_raw.isdigit():
        anio = int(anio_raw)

    if mes < 1 or mes > 12:
        raise ValueError("Mes inválido")

    return anio, mes


@bp.route("/resumen-capacidad-semanal", methods=["GET"])
def resumen_capacidad_semanal():
    try:
        usuario = _get_usuario_from_request()
        rol_req = _get_rol_from_request()

        if not usuario:
            return jsonify({"error": "Usuario no enviado"}), 400

        usuario_norm = (usuario or "").strip().lower()

        consultor_login = (
            Consultor.query.options(
                joinedload(Consultor.rol_obj),
                joinedload(Consultor.equipo_obj),
                joinedload(Consultor.horario_obj),
            )
            .filter(func.lower(Consultor.usuario) == usuario_norm)
            .first()
        )

        if not consultor_login:
            return jsonify({"error": "Consultor no encontrado"}), 404

        scope, val = scope_for(consultor_login, rol_req)

        # Solo ADMIN global y ADMIN_EQUIPO
        if scope in {"SELF", "ROLE_POOL"}:
            return jsonify({"error": "No autorizado para ver capacidad semanal"}), 403

        anio, mes = _cap_parse_month_year_from_request()
        month_start, month_end = _cap_month_bounds(anio, mes)
        semanas_mes = _cap_weeks_for_month(anio, mes)
        co_holidays = _cap_colombia_holidays_for_years([anio])

        equipo_filter = (request.args.get("equipo") or "").strip().upper()
        consultor_filter = (request.args.get("consultor") or "").strip().lower()

        if scope == "TEAM":
            eq_login = (
                (consultor_login.equipo_obj.nombre or "").strip().upper()
                if consultor_login.equipo_obj else ""
            )

            if not eq_login:
                return jsonify({"error": "Consultor sin equipo asignado"}), 403

            if equipo_filter and equipo_filter != eq_login:
                return jsonify({"error": "No autorizado para consultar otro equipo"}), 403

            equipo_filter = eq_login

        q = (
            db.session.query(
                func.lower(Registro.usuario_consultor).label("usuario_consultor"),
                Consultor.id.label("consultor_id"),
                Consultor.nombre.label("consultor"),
                Equipo.nombre.label("equipo"),
                Registro.fecha.label("fecha"),
                func.coalesce(func.sum(Registro.total_horas), 0).label("total_horas"),
            )
            .select_from(Registro)
            .join(
                Consultor,
                func.lower(Registro.usuario_consultor) == func.lower(Consultor.usuario)
            )
            .outerjoin(Equipo, Consultor.equipo_id == Equipo.id)
        )

        if scope == "TEAM":
            q = q.filter(Consultor.equipo_id == int(val))
        elif scope == "ALL":
            pass
        else:
            return jsonify({"error": "Scope no permitido"}), 403

        q = q.filter(
            Registro.fecha.between(month_start.isoformat(), month_end.isoformat())
        )

        if equipo_filter:
            q = q.filter(func.upper(Equipo.nombre) == equipo_filter)

        if consultor_filter:
            q = q.filter(func.lower(Consultor.nombre).like(f"%{consultor_filter}%"))

        q = q.group_by(
            func.lower(Registro.usuario_consultor),
            Consultor.id,
            Consultor.nombre,
            Equipo.nombre,
            Registro.fecha,
        ).order_by(
            Consultor.nombre.asc(),
            Registro.fecha.asc(),
        )

        raw = q.all()

        grouped = {}

        for r in raw:
            consultor_id = int(r.consultor_id) if r.consultor_id else None
            key = consultor_id or (r.usuario_consultor or "na")

            fecha_obj = _cap_parse_iso_date(r.fecha)
            if not fecha_obj:
                continue

            if key not in grouped:
                grouped[key] = {
                    "consultorId": consultor_id,
                    "consultor": r.consultor or r.usuario_consultor or "—",
                    "equipo": (r.equipo or "SIN EQUIPO").strip().upper(),
                    "diasHoras": defaultdict(float),
                }

            horas = float(r.total_horas or 0)
            grouped[key]["diasHoras"][fecha_obj.isoformat()] += horas

        rows_out = []

        for _, item in grouped.items():
            equipo_kind = _cap_team_kind(item["equipo"])

            meta_mes = 0.0
            dias_laborables_mes = 0
            dias_festivos_mes = 0

            cursor = month_start
            while cursor <= month_end:
                if cursor.weekday() <= 4 and _cap_is_holiday(cursor, co_holidays):
                    dias_festivos_mes += 1

                meta_dia = _cap_meta_hours_for_day(cursor, co_holidays)
                if meta_dia > 0:
                    meta_mes += meta_dia
                    dias_laborables_mes += 1

                cursor += timedelta(days=1)

            meta_mes = round(meta_mes, 2)

            horas_mes = round(
                sum(float(v or 0) for v in item["diasHoras"].values()),
                2
            )

            semanas_out = []

            for wk in semanas_mes:
                horas_semana = 0.0
                meta_semana = 0.0
                dias_out = []

                cursor = wk["start"]
                while cursor <= wk["end"]:
                    fecha_key = cursor.isoformat()
                    horas_dia = round(float(item["diasHoras"].get(fecha_key, 0)), 2)
                    meta_dia = round(_cap_meta_hours_for_day(cursor, co_holidays), 2)

                    es_laborable = meta_dia > 0

                    if es_laborable:
                        meta_semana += meta_dia

                    if horas_dia > 0:
                        horas_semana += horas_dia

                    # Mostrar:
                    # - siempre días laborables
                    # - fines de semana / festivos con horas solo si NO es BASIS
                    mostrar_dia = es_laborable or (horas_dia > 0 and equipo_kind != "BASIS")

                    if mostrar_dia:
                        dias_out.append({
                            "fecha": fecha_key,
                            "horas": horas_dia,
                            "metaDia": meta_dia,
                        })

                    cursor += timedelta(days=1)

                horas_semana = round(horas_semana, 2)
                meta_semana = round(meta_semana, 2)

                porcentaje_semanal = (
                    round((horas_semana / meta_semana) * 100, 2)
                    if meta_semana > 0 else 0.0
                )
                diferencia_semana = round(meta_semana - horas_semana, 2)

                semanas_out.append({
                    "label": f"Semana {wk['index']}",
                    "inicio": wk["start"].isoformat(),
                    "fin": wk["end"].isoformat(),
                    "metaSemanal": meta_semana,
                    "horasSemana": horas_semana,
                    "porcentajeSemanal": porcentaje_semanal,
                    "diferenciaSemana": diferencia_semana,
                    "dias": dias_out,
                })

            porcentaje_mes = (
                round((horas_mes / meta_mes) * 100, 2)
                if meta_mes > 0 else 0.0
            )

            rows_out.append({
                "consultorId": item["consultorId"],
                "consultor": item["consultor"],
                "equipo": item["equipo"],
                "metaMes": meta_mes,
                "metaDiaObjetivo": 9.0,
                "horasMes": horas_mes,
                "porcentajeMes": porcentaje_mes,
                "diasTrabajoTexto": _cap_work_days_text(),
                "diasLaborablesMes": dias_laborables_mes,
                "diasFestivosMes": dias_festivos_mes,
                "semanas": semanas_out,
            })

        rows_out.sort(key=lambda x: ((x["equipo"] or ""), (x["consultor"] or "")))

        return jsonify({
            "mes": mes,
            "anio": anio,
            "desde": month_start.isoformat(),
            "hasta": month_end.isoformat(),
            "rows": rows_out,
        }), 200

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        app.logger.exception("❌ Error en /resumen-capacidad-semanal")
        return jsonify({"error": str(e)}), 500


@bp.route("/capacidad-semanal-ocupaciones", methods=["GET"])
def capacidad_semanal_ocupaciones():
    try:
        usuario = _get_usuario_from_request()
        rol_req = _get_rol_from_request()

        if not usuario:
            return jsonify({"error": "Usuario no enviado"}), 400

        usuario_norm = (usuario or "").strip().lower()

        consultor_login = (
            Consultor.query.options(
                joinedload(Consultor.rol_obj),
                joinedload(Consultor.equipo_obj),
                joinedload(Consultor.horario_obj),
            )
            .filter(func.lower(Consultor.usuario) == usuario_norm)
            .first()
        )

        if not consultor_login:
            return jsonify({"error": "Consultor no encontrado"}), 404

        scope, val = scope_for(consultor_login, rol_req)

        if scope in {"SELF", "ROLE_POOL"}:
            return jsonify({"error": "No autorizado para ver ocupaciones de capacidad semanal"}), 403

        anio, mes = _cap_parse_month_year_from_request()
        month_start, month_end = _cap_month_bounds(anio, mes)

        equipo_filter = (request.args.get("equipo") or "").strip().upper()
        consultor_filter = (request.args.get("consultor") or "").strip().lower()

        if scope == "TEAM":
            eq_login = (
                (consultor_login.equipo_obj.nombre or "").strip().upper()
                if consultor_login.equipo_obj else ""
            )

            if not eq_login:
                return jsonify({"error": "Consultor sin equipo asignado"}), 403

            if equipo_filter and equipo_filter != eq_login:
                return jsonify({"error": "No autorizado para consultar otro equipo"}), 403

            equipo_filter = eq_login

        q = (
            db.session.query(
                Consultor.id.label("consultor_id"),
                Consultor.nombre.label("consultor"),
                Equipo.nombre.label("equipo"),
                Ocupacion.codigo.label("ocupacion_codigo"),
                Ocupacion.nombre.label("ocupacion_nombre"),
                func.coalesce(
                    func.sum(
                        func.coalesce(Registro.tiempo_invertido, Registro.total_horas, 0)
                    ),
                    0
                ).label("horas")
            )
            .select_from(Registro)
            .join(
                Consultor,
                func.lower(Registro.usuario_consultor) == func.lower(Consultor.usuario)
            )
            .outerjoin(Equipo, Consultor.equipo_id == Equipo.id)
            .outerjoin(Ocupacion, Registro.ocupacion_id == Ocupacion.id)
            .filter(Registro.fecha >= month_start.isoformat())
            .filter(Registro.fecha <= month_end.isoformat())
        )

        if scope == "TEAM":
            q = q.filter(func.upper(Equipo.nombre) == equipo_filter)
        elif equipo_filter:
            q = q.filter(func.upper(Equipo.nombre) == equipo_filter)

        if consultor_filter:
            q = q.filter(func.lower(Consultor.nombre) == consultor_filter)

        q = q.group_by(
            Consultor.id,
            Consultor.nombre,
            Equipo.nombre,
            Ocupacion.codigo,
            Ocupacion.nombre,
        )

        rows = q.all()

        data = []
        for r in rows:
            data.append({
                "consultorId": r.consultor_id,
                "consultor": r.consultor,
                "equipo": r.equipo,
                "ocupacion_codigo": r.ocupacion_codigo or "",
                "ocupacion_nombre": r.ocupacion_nombre or "SIN OCUPACIÓN",
                "horas": round(float(r.horas or 0), 2),
            })

        data.sort(key=lambda x: (
            (x["equipo"] or ""),
            (x["consultor"] or ""),
            -float(x["horas"] or 0),
            (x["ocupacion_nombre"] or ""),
        ))

        return jsonify({
            "mes": mes,
            "anio": anio,
            "desde": month_start.isoformat(),
            "hasta": month_end.isoformat(),
            "rows": data,
        }), 200

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        app.logger.exception("❌ Error en /capacidad-semanal-ocupaciones")
        return jsonify({"error": str(e)}), 500


@bp.route("/modulos/<int:modulo_id>/perfiles", methods=["GET"])
@permission_required("MODULOS_VER")
def get_modulo_perfiles(modulo_id):
    modulo = Modulo.query.get_or_404(modulo_id)

    rows = (
        ModuloPerfil.query.options(
            joinedload(ModuloPerfil.modulo),
            joinedload(ModuloPerfil.perfil),
        )
        .filter(ModuloPerfil.modulo_id == modulo_id)
        .order_by(ModuloPerfil.id.asc())
        .all()
    )

    return jsonify({
        "modulo": {
            "id": modulo.id,
            "nombre": modulo.nombre,
        },
        "perfiles": [modulo_perfil_to_dict(x) for x in rows]
    }), 200

## Consultor Perfil 

@bp.route("/consultores/<int:consultor_id>/perfiles", methods=["GET"])
@permission_required("CONSULTORES_VER")
def get_consultor_perfiles(consultor_id):
    consultor = Consultor.query.get_or_404(consultor_id)

    rows = (
        ConsultorPerfil.query.options(
            joinedload(ConsultorPerfil.consultor),
            joinedload(ConsultorPerfil.perfil),
        )
        .filter(ConsultorPerfil.consultor_id == consultor_id)
        .order_by(ConsultorPerfil.id.asc())
        .all()
    )

    return jsonify({
        "consultor": {
            "id": consultor.id,
            "nombre": consultor.nombre,
            "usuario": consultor.usuario,
        },
        "perfiles": [consultor_perfil_to_dict(x) for x in rows]
    }), 200

@bp.route("/consultores/<int:consultor_id>/perfiles", methods=["PUT"])
@permission_required("CONSULTORES_EDITAR")
def save_consultor_perfiles(consultor_id):
    consultor = Consultor.query.get_or_404(consultor_id)

    data = request.get_json(silent=True) or {}
    rows = data.get("rows") or []

    if not isinstance(rows, list):
        return jsonify({"mensaje": "rows debe ser una lista"}), 400

    seen = set()
    rows_clean = []

    try:
        for idx, row in enumerate(rows):
            if not isinstance(row, dict):
                return jsonify({"mensaje": f"Fila inválida en posición {idx}"}), 400

            perfil_id = row.get("perfil_id")
            try:
                perfil_id = int(perfil_id)
            except Exception:
                return jsonify({"mensaje": f"perfil_id inválido en fila {idx}"}), 400

            if perfil_id in seen:
                return jsonify({
                    "mensaje": f"Perfil duplicado en la misma carga: {perfil_id}"
                }), 400
            seen.add(perfil_id)

            perfil = Perfil.query.get(perfil_id)
            if not perfil:
                return jsonify({"mensaje": f"Perfil no existe: {perfil_id}"}), 400

            fecha_inicio = _parse_date_safe(row.get("fecha_inicio"))
            fecha_fin = _parse_date_safe(row.get("fecha_fin"))

            if row.get("fecha_inicio") not in (None, "", "null", "None") and not fecha_inicio:
                return jsonify({
                    "mensaje": f"fecha_inicio inválida para el perfil {perfil.nombre}"
                }), 400

            if row.get("fecha_fin") not in (None, "", "null", "None") and not fecha_fin:
                return jsonify({
                    "mensaje": f"fecha_fin inválida para el perfil {perfil.nombre}"
                }), 400

            if fecha_inicio and fecha_fin and fecha_fin < fecha_inicio:
                return jsonify({
                    "mensaje": f"La fecha fin no puede ser menor que la fecha inicio para el perfil {perfil.nombre}"
                }), 400

            rows_clean.append({
                "perfil_id": perfil_id,
                "activo": _to_bool2(row.get("activo"), default=True),
                "fecha_inicio": fecha_inicio,
                "fecha_fin": fecha_fin,
            })

        # borrar relaciones actuales del consultor
        ConsultorPerfil.query.filter_by(
            consultor_id=consultor_id
        ).delete(synchronize_session=False)

        # insertar nuevas relaciones
        for row in rows_clean:
            db.session.add(
                ConsultorPerfil(
                    consultor_id=consultor_id,
                    perfil_id=row["perfil_id"],
                    activo=row["activo"],
                    fecha_inicio=row["fecha_inicio"],
                    fecha_fin=row["fecha_fin"],
                )
            )

        db.session.commit()

        rows_db = (
            ConsultorPerfil.query.options(
                joinedload(ConsultorPerfil.consultor),
                joinedload(ConsultorPerfil.perfil),
            )
            .filter_by(consultor_id=consultor_id)
            .order_by(ConsultorPerfil.id.asc())
            .all()
        )

        return jsonify({
            "mensaje": "Perfiles del consultor actualizados",
            "consultor": {
                "id": consultor.id,
                "nombre": consultor.nombre,
                "usuario": consultor.usuario,
            },
            "perfiles": [consultor_perfil_to_dict(x) for x in rows_db]
        }), 200

    except IntegrityError as e:
        db.session.rollback()
        return jsonify({
            "mensaje": "Error de integridad al guardar perfiles del consultor",
            "detalle": str(e)
        }), 400

    except Exception as e:
        db.session.rollback()
        app.logger.exception("❌ Error en save_consultor_perfiles")
        return jsonify({
            "mensaje": "Error interno al guardar perfiles del consultor",
            "detalle": str(e)
        }), 500

@bp.route("/perfiles/<int:perfil_id>/consultores", methods=["GET"])
@permission_required("PERFILES_VER")
def get_perfil_consultores(perfil_id):
    perfil = Perfil.query.get_or_404(perfil_id)

    rows = (
        ConsultorPerfil.query.options(
            joinedload(ConsultorPerfil.consultor),
            joinedload(ConsultorPerfil.perfil),
        )
        .filter(ConsultorPerfil.perfil_id == perfil_id)
        .order_by(ConsultorPerfil.id.asc())
        .all()
    )

    return jsonify({
        "perfil": perfil_to_dict(perfil),
        "consultores": [consultor_perfil_to_dict(x) for x in rows]
    }), 200

# =========================================================
# HELPERS COSTO / PRESUPUESTO
# =========================================================

def _to_date_safe(value):
    if isinstance(value, date):
        return value
    s = str(value or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except Exception:
            continue
    return None

def _daterange(start_date: date, end_date: date):
    cur = start_date
    while cur <= end_date:
        yield cur
        cur += timedelta(days=1)

def _month_bounds_local(anio: int, mes: int):
    start = date(anio, mes, 1)
    if mes == 12:
        end = date(anio + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(anio, mes + 1, 1) - timedelta(days=1)
    return start, end

def _iter_months_between(start_date: date, end_date: date):
    cur = date(start_date.year, start_date.month, 1)
    stop = date(end_date.year, end_date.month, 1)
    while cur <= stop:
        yield cur.year, cur.month
        if cur.month == 12:
            cur = date(cur.year + 1, 1, 1)
        else:
            cur = date(cur.year, cur.month + 1, 1)

def _cap_is_standard_workday(d: date, co_holidays=None):
    co_holidays = co_holidays or set()
    return d.weekday() < 5 and d not in co_holidays

def _cap_meta_hours_for_day(d: date, co_holidays=None):
    """
    Regla única:
    - lunes: 8h
    - martes a viernes: 9h
    - fines de semana / festivos: 0h
    """
    co_holidays = co_holidays or set()

    if not _cap_is_standard_workday(d, co_holidays):
        return 0.0

    return 8.0 if d.weekday() == 0 else 9.0

def _cap_colombia_holidays_for_years(years):
    import holidays
    out = set()
    for y in set(int(x) for x in years):
        for h in holidays.CO(years=[y]).keys():
            out.add(h)
    return out

def _presupuesto_consultor_mes(consultor_id: int, anio: int, mes: int):
    """
    Regla:
    1. Busca presupuesto exacto del periodo.
    2. Si no existe, busca el último presupuesto anterior o igual al periodo.
    3. Si tampoco existe, usa el último presupuesto disponible del consultor.
    4. horas_base_mes SIEMPRE se recalcula con el mes solicitado.
    """

    row = (
        ConsultorPresupuesto.query
        .filter(ConsultorPresupuesto.consultor_id == consultor_id)
        .filter(ConsultorPresupuesto.anio == anio)
        .filter(ConsultorPresupuesto.mes == mes)
        .order_by(ConsultorPresupuesto.vigente.desc(), ConsultorPresupuesto.id.desc())
        .first()
    )

    if not row:
        row = (
            ConsultorPresupuesto.query
            .filter(ConsultorPresupuesto.consultor_id == consultor_id)
            .filter(
                or_(
                    ConsultorPresupuesto.anio < anio,
                    and_(
                        ConsultorPresupuesto.anio == anio,
                        ConsultorPresupuesto.mes <= mes
                    )
                )
            )
            .order_by(
                ConsultorPresupuesto.anio.desc(),
                ConsultorPresupuesto.mes.desc(),
                ConsultorPresupuesto.id.desc()
            )
            .first()
        )

    if not row:
        row = (
            ConsultorPresupuesto.query
            .filter(ConsultorPresupuesto.consultor_id == consultor_id)
            .order_by(
                ConsultorPresupuesto.vigente.desc(),
                ConsultorPresupuesto.anio.desc(),
                ConsultorPresupuesto.mes.desc(),
                ConsultorPresupuesto.id.desc()
            )
            .first()
        )

    if not row:
        return None

    meta_mes = _meta_horas_en_rango(*_month_bounds_local(anio, mes))
    horas_base_mes = meta_mes["horas"]
    dias_habiles_mes = meta_mes["dias_laborables"]

    vr = Decimal(str(row.vr_perfil or 0)).quantize(Decimal("0.01"))
    valor_hora = Decimal("0.00")

    if horas_base_mes > 0:
        valor_hora = (vr / horas_base_mes).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP
        )

    return {
        "row": row,
        "vr_perfil": vr,
        "horas_base_mes": horas_base_mes,
        "valor_hora": valor_hora,
        "dias_habiles_mes": dias_habiles_mes,
    }

def _cost_parse_periodo_request():
    """
    Soporta:
      - ?mes=4&anio=2026
      - ?desde=2026-04-01&hasta=2026-04-30
    """
    desde_raw = (request.args.get("desde") or "").strip()
    hasta_raw = (request.args.get("hasta") or "").strip()

    if desde_raw or hasta_raw:
        desde = _to_date_safe(desde_raw) if desde_raw else None
        hasta = _to_date_safe(hasta_raw) if hasta_raw else None

        if not desde and not hasta:
            raise ValueError("Rango inválido")

        if desde and not hasta:
            hasta = desde
        if hasta and not desde:
            desde = hasta

        if hasta < desde:
            raise ValueError("Hasta no puede ser menor que desde")

        return desde, hasta, "rango"

    mes_raw = (request.args.get("mes") or "").strip()
    anio_raw = (request.args.get("anio") or "").strip()

    today = date.today()
    mes = int(mes_raw) if mes_raw.isdigit() else today.month
    anio = int(anio_raw) if anio_raw.isdigit() else today.year

    if mes < 1 or mes > 12:
        raise ValueError("Mes inválido")

    desde, hasta = _month_bounds_local(anio, mes)
    return desde, hasta, "mes"

# =========================================================
# IMPORTADOR DE PRESUPUESTO POR PERIODO
# =========================================================

def _norm_name(s: str) -> str:
    s = (s or "").strip().upper()
    s = re.sub(r"\s+", " ", s)
    s = (
        s.replace("Á", "A")
         .replace("É", "E")
         .replace("Í", "I")
         .replace("Ó", "O")
         .replace("Ú", "U")
         .replace("Ñ", "N")
    )
    s = re.sub(r"[^A-Z0-9 ,.-]", "", s)
    return s

def _norm_doc(s: str) -> str:
    s = (s or "").strip()
    return re.sub(r"[^\d]", "", s)

def _parse_money_to_decimal(val) -> Decimal:
    if val is None:
        return Decimal("0.00")

    if isinstance(val, (int, float, Decimal)):
        try:
            return Decimal(str(val)).quantize(Decimal("0.01"))
        except InvalidOperation:
            return Decimal("0.00")

    s = str(val).strip()
    if not s:
        return Decimal("0.00")

    s = s.replace("$", "").strip()
    s = re.sub(r"[^\d,\.]", "", s)

    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        if "." in s and "," not in s:
            s = s.replace(".", "")
        if "," in s and "." not in s:
            s = s.replace(",", ".")

    try:
        return Decimal(s).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0.00")

def _col_idx(headers: dict, wanted: str):
    w = (wanted or "").strip().upper()
    if not w:
        return None
    if w in headers:
        return headers[w]
    for hk, i in headers.items():
        if w in hk:
            return i
    return None

@bp.route("/presupuestos/consultor/import-excel", methods=["POST"])
def import_presupuesto_consultor_excel():
    try:
        f = request.files.get("file")
        if not f:
            return jsonify({"error": "Falta archivo (file)"}), 400

        anio = request.form.get("anio")
        mes = request.form.get("mes")

        try:
            anio = int(str(anio).strip()) if anio else None
        except Exception:
            anio = None

        try:
            mes = int(str(mes).strip()) if mes else None
        except Exception:
            mes = None

        now = datetime.now()
        if not anio:
            anio = now.year
        if not mes or mes < 1 or mes > 12:
            mes = now.month

        meta_mes_info = _meta_horas_en_rango(*_month_bounds_local(anio, mes))
        horas_base = meta_mes_info["horas"]

        sheet_name = (request.form.get("sheet") or "").strip() or None

        wb = load_workbook(f, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.worksheets[0]

        headers = {}
        for idx, cell in enumerate(ws[1], start=1):
            key = (str(cell.value).strip() if cell.value is not None else "")
            if key:
                headers[key.upper()] = idx

        meses_nombre = {
            1: "ENERO",
            2: "FEBRERO",
            3: "MARZO",
            4: "ABRIL",
            5: "MAYO",
            6: "JUNIO",
            7: "JULIO",
            8: "AGOSTO",
            9: "SEPTIEMBRE",
            10: "OCTUBRE",
            11: "NOVIEMBRE",
            12: "DICIEMBRE",
        }

        mes_nombre = meses_nombre.get(mes, "")
        posibles_col_total = [
            f"TOTAL {mes_nombre}",
            f"TOTAL_{mes_nombre}",
            f"TOTAL-{mes_nombre}",
            "TOTAL",
            "VR PERFIL",
        ]

        ci_nombre = _col_idx(headers, "NOMBRE COLABORADOR")
        ci_cedula = _col_idx(headers, "CEDULA")

        ci_total = None
        total_header_detectado = None
        for cand in posibles_col_total:
            ci_total = _col_idx(headers, cand)
            if ci_total is not None:
                total_header_detectado = cand
                break

        if ci_nombre is None or ci_cedula is None or ci_total is None:
            return jsonify({
                "error": "No encontré las columnas requeridas",
                "detalle": {
                    "requeridas": [
                        "NOMBRE COLABORADOR",
                        "CEDULA",
                        f"TOTAL {mes_nombre}"
                    ],
                    "headers_encontrados": list(headers.keys())[:100]
                }
            }), 400

        consultores = Consultor.query.all()

        by_name = {}
        by_doc = {}

        for c in consultores:
          nombre_norm = _norm_name(c.nombre) if getattr(c, "nombre", None) else ""
          if nombre_norm:
              by_name[nombre_norm] = c

          cedula_norm = _norm_doc(str(getattr(c, "cedula", "") or ""))
          if cedula_norm:
              by_doc[cedula_norm] = c

        updated = 0
        created = 0
        not_found = []
        invalid_rows = []
        seen = set()

        for r in range(2, ws.max_row + 1):
            raw_name = ws.cell(row=r, column=ci_nombre).value
            raw_doc = ws.cell(row=r, column=ci_cedula).value
            raw_total = ws.cell(row=r, column=ci_total).value

            if raw_name is None and raw_doc is None and raw_total is None:
                continue

            vr = _parse_money_to_decimal(raw_total)
            if vr <= 0:
                invalid_rows.append({
                    "row": r,
                    "nombre": str(raw_name or ""),
                    "cedula": str(raw_doc or ""),
                    "valor": str(raw_total or "")
                })
                continue

            c = None

            doc_norm = _norm_doc(str(raw_doc or ""))
            if doc_norm:
                c = by_doc.get(doc_norm)

            if not c:
                name_norm = _norm_name(str(raw_name or ""))
                if name_norm:
                    c = by_name.get(name_norm)

            if not c:
                not_found.append({
                    "row": r,
                    "nombre": str(raw_name or ""),
                    "cedula": str(raw_doc or ""),
                    "valor": str(raw_total or "")
                })
                continue

            key = f"{c.id}"
            if key in seen:
                continue
            seen.add(key)

            existente = (
                ConsultorPresupuesto.query
                .filter_by(
                    consultor_id=c.id,
                    anio=anio,
                    mes=mes,
                    vigente=True
                )
                .order_by(ConsultorPresupuesto.id.desc())
                .first()
            )

            if existente:
                existente.vr_perfil = vr
                existente.horas_base_mes = horas_base
                updated += 1
            else:
                db.session.query(ConsultorPresupuesto).filter_by(
                    consultor_id=c.id,
                    anio=anio,
                    mes=mes,
                    vigente=True
                ).update({"vigente": False})

                db.session.add(ConsultorPresupuesto(
                    consultor_id=c.id,
                    anio=anio,
                    mes=mes,
                    vr_perfil=vr,
                    horas_base_mes=horas_base,
                    vigente=True
                ))
                created += 1

        db.session.commit()

        return jsonify({
            "ok": True,
            "anio": anio,
            "mes": mes,
            "columnaValorDetectada": total_header_detectado,
            "horasBaseCalculadas": float(horas_base),
            "diasLaborablesMes": meta_mes_info["dias_laborables"],
            "diasFestivosMes": meta_mes_info["dias_festivos"],
            "created": created,
            "updated": updated,
            "notFoundCount": len(not_found),
            "invalidCount": len(invalid_rows),
            "notFound": not_found[:50],
            "invalidRows": invalid_rows[:50]
        }), 200

    except Exception as e:
        db.session.rollback()
        app.logger.exception("❌ Error en /presupuestos/consultor/import-excel")
        return jsonify({"error": str(e)}), 500
    
# =========================================================
# RESUMEN DE COSTO POR CONSULTOR
# =========================================================

@bp.route("/resumen-costo-consultor", methods=["GET"])
def resumen_costo_consultor():
    try:
        usuario = _get_usuario_from_request()
        rol_req = _get_rol_from_request()

        if not usuario:
            return jsonify({"error": "Usuario no enviado"}), 400

        usuario_norm = (usuario or "").strip().lower()

        consultor_login = (
            Consultor.query.options(
                joinedload(Consultor.rol_obj),
                joinedload(Consultor.equipo_obj),
            )
            .filter(func.lower(Consultor.usuario) == usuario_norm)
            .first()
        )

        if not consultor_login:
            return jsonify({"error": "Consultor no encontrado"}), 404

        rol_real = ""
        if getattr(consultor_login, "rol_obj", None) and getattr(consultor_login.rol_obj, "nombre", None):
            rol_real = consultor_login.rol_obj.nombre
        else:
            rol_real = getattr(consultor_login, "rol", "") or ""

        if not _is_admin_request(rol_real, consultor_login):
            return jsonify({"error": "No autorizado. Solo administradores."}), 403

        scope, val = scope_for(consultor_login, rol_req)
        desde, hasta, modo = _cost_parse_periodo_request()

        equipo_filter = (request.args.get("equipo") or "").strip().upper()
        consultor_filter = (request.args.get("consultor") or "").strip().lower()

        ocupacion_ids = [
            int(x) for x in request.args.getlist("ocupacion_id")
            if str(x).strip().isdigit()
        ]

        q = (
            db.session.query(
                Consultor.id.label("consultor_id"),
                Consultor.nombre.label("consultor"),
                Consultor.usuario.label("usuario_consultor"),
                Equipo.nombre.label("equipo"),
                func.substr(func.cast(Registro.fecha, db.String), 1, 7).label("periodo"),
                func.coalesce(
                    func.sum(
                        func.coalesce(Registro.tiempo_invertido, Registro.total_horas, 0)
                    ),
                    0
                ).label("horas"),
            )
            .select_from(Registro)
            .join(
                Consultor,
                func.lower(Registro.usuario_consultor) == func.lower(Consultor.usuario)
            )
            .outerjoin(Equipo, Consultor.equipo_id == Equipo.id)
            .filter(Registro.fecha >= desde.isoformat())
            .filter(Registro.fecha <= hasta.isoformat())
        )

        if ocupacion_ids:
            q = q.filter(Registro.ocupacion_id.in_(ocupacion_ids))

        if scope == "TEAM":
            if not int(val or 0):
                return jsonify({"error": "Consultor sin equipo asignado"}), 403

            q = q.filter(Consultor.equipo_id == int(val))

            eq_login = (
                (consultor_login.equipo_obj.nombre or "").strip().upper()
                if consultor_login.equipo_obj else ""
            )

            if not eq_login:
                return jsonify({"error": "Consultor sin equipo asignado"}), 403

            if equipo_filter and equipo_filter != eq_login:
                return jsonify({"error": "No autorizado para consultar otro equipo"}), 403

            equipo_filter = eq_login

        elif scope == "ALL":
            pass

        elif scope == "ROLE_POOL":
            if not int(val or 0):
                return jsonify({"error": "Consultor sin rol asignado"}), 403
            q = q.filter(Consultor.rol_id == int(val))

        else:
            return jsonify({"error": "Scope no permitido"}), 403

        if equipo_filter:
            q = q.filter(func.upper(Equipo.nombre) == equipo_filter)

        if consultor_filter:
            q = q.filter(func.lower(Consultor.nombre).like(f"%{consultor_filter}%"))

        q = q.group_by(
            Consultor.id,
            Consultor.nombre,
            Consultor.usuario,
            Equipo.nombre,
            func.substr(func.cast(Registro.fecha, db.String), 1, 7),
        ).order_by(Consultor.nombre.asc())

        raw = q.all()

        rows_map = {}
        total_horas_general = Decimal("0.00")
        total_meta_general = Decimal("0.00")
        total_costo_general = Decimal("0.00")

        for item in raw:
            cid = int(item.consultor_id)
            nombre = item.consultor or "—"
            usuario_cons = item.usuario_consultor or ""
            equipo = item.equipo or "SIN EQUIPO"
            periodo = str(item.periodo or "").strip()

            try:
                anio = int(periodo[:4])
                mes = int(periodo[5:7])
            except Exception:
                continue

            horas_reg_mes = Decimal(str(item.horas or 0)).quantize(Decimal("0.01"))

            month_start, month_end = _month_bounds_local(anio, mes)
            tramo_inicio = max(desde, month_start)
            tramo_fin = min(hasta, month_end)

            meta_tramo_info = _meta_horas_en_rango(tramo_inicio, tramo_fin)
            meta_tramo = meta_tramo_info["horas"]

            presupuesto = _presupuesto_consultor_mes(cid, anio, mes)

            vr_perfil = Decimal("0.00")
            horas_base_mes = Decimal("0.00")
            valor_hora_mes = Decimal("0.00")
            dias_habiles_mes = _meta_horas_en_rango(*_month_bounds_local(anio, mes))["dias_laborables"]

            if presupuesto:
                vr_perfil = presupuesto["vr_perfil"]
                horas_base_mes = presupuesto["horas_base_mes"]
                valor_hora_mes = presupuesto["valor_hora"]

            costo_mes = (horas_reg_mes * valor_hora_mes).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP
            )

            if cid not in rows_map:
                rows_map[cid] = {
                    "consultorId": cid,
                    "consultor": nombre,
                    "usuarioConsultor": usuario_cons,
                    "equipo": equipo,
                    "horasPeriodo": Decimal("0.00"),
                    "metaHorasPeriodo": Decimal("0.00"),
                    "costoPeriodo": Decimal("0.00"),
                    "presupuestos": [],
                }

            rows_map[cid]["horasPeriodo"] += horas_reg_mes
            rows_map[cid]["metaHorasPeriodo"] += meta_tramo
            rows_map[cid]["costoPeriodo"] += costo_mes
            rows_map[cid]["presupuestos"].append({
                "anio": anio,
                "mes": mes,
                "vrPerfil": float(vr_perfil),
                "diasHabilesMes": dias_habiles_mes,
                "horasBaseMes": float(horas_base_mes),
                "valorHoraMes": float(valor_hora_mes),
                "horasRegistradasMesEnFiltro": float(horas_reg_mes),
                "metaHorasMesEnFiltro": float(meta_tramo),
                "costoMesEnFiltro": float(costo_mes),
                "diasLaborablesMesEnFiltro": meta_tramo_info["dias_laborables"],
                "diasFestivosMesEnFiltro": meta_tramo_info["dias_festivos"],
            })

            total_horas_general += horas_reg_mes
            total_meta_general += meta_tramo
            total_costo_general += costo_mes

        rows_out = []

        for _, item in rows_map.items():
            horas_periodo = item["horasPeriodo"].quantize(Decimal("0.01"))
            meta_periodo = item["metaHorasPeriodo"].quantize(Decimal("0.01"))
            costo_periodo = item["costoPeriodo"].quantize(Decimal("0.01"))

            valor_hora_promedio = Decimal("0.00")
            if horas_periodo > 0:
                valor_hora_promedio = (costo_periodo / horas_periodo).quantize(
                    Decimal("0.01"),
                    rounding=ROUND_HALF_UP
                )

            porcentaje = Decimal("0.00")
            if meta_periodo > 0:
                porcentaje = ((horas_periodo / meta_periodo) * Decimal("100")).quantize(
                    Decimal("0.01"),
                    rounding=ROUND_HALF_UP
                )

            item["presupuestos"].sort(key=lambda x: (x["anio"], x["mes"]))

            rows_out.append({
                "consultorId": item["consultorId"],
                "consultor": item["consultor"],
                "usuarioConsultor": item["usuarioConsultor"],
                "equipo": item["equipo"],
                "horasPeriodo": float(horas_periodo),
                "metaHorasPeriodo": float(meta_periodo),
                "diferenciaHoras": float((meta_periodo - horas_periodo).quantize(Decimal("0.01"))),
                "porcentajeUsoPeriodo": float(porcentaje),
                "costoPeriodo": float(costo_periodo),
                "valorHoraPromedio": float(valor_hora_promedio),
                "presupuestos": item["presupuestos"],
            })

        rows_out.sort(key=lambda x: (-x["costoPeriodo"], x["consultor"]))

        porcentaje_general = Decimal("0.00")
        if total_meta_general > 0:
            porcentaje_general = ((total_horas_general / total_meta_general) * Decimal("100")).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP
            )

        return jsonify({
            "modo": modo,
            "desde": desde.isoformat(),
            "hasta": hasta.isoformat(),
            "totalConsultores": len(rows_out),
            "totalHorasPeriodo": float(total_horas_general.quantize(Decimal("0.01"))),
            "totalMetaPeriodo": float(total_meta_general.quantize(Decimal("0.01"))),
            "totalCostoPeriodo": float(total_costo_general.quantize(Decimal("0.01"))),
            "porcentajeGeneral": float(porcentaje_general),
            "rows": rows_out,
        }), 200

    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    except Exception as e:
        app.logger.exception("❌ Error en /resumen-costo-consultor")
        return jsonify({"error": str(e)}), 500

# =========================================================
# DASHBOARD COSTOS RESUMEN
# Agrupa por cliente + ocupación y calcula costo real
# usando valor hora del consultor según el periodo.
# =========================================================

@bp.route("/dashboard/costos-resumen", methods=["GET"])
@permission_required("GRAFICOS_VER")
def dashboard_costos_resumen():
    try:
        usuario = _get_usuario_from_request()
        rol_req = _rol_from_request()

        if not usuario:
            return jsonify({"error": "Usuario no enviado"}), 400

        usuario_norm = (usuario or "").strip().lower()

        consultor_login = (
            Consultor.query.options(
                joinedload(Consultor.rol_obj),
                joinedload(Consultor.equipo_obj),
            )
            .filter(func.lower(Consultor.usuario) == usuario_norm)
            .first()
        )

        if not consultor_login:
            return jsonify({"error": "Consultor no encontrado"}), 404

        scope, val = _scope_for_graficos(consultor_login, rol_req)
        desde, hasta, modo = _dashboard_costos_parse_periodo_request()

        C = aliased(Consultor)
        E = aliased(Equipo)
        O = aliased(Ocupacion)

        fecha_expr = _registro_fecha_expr()

        def _get_multi(name):
            vals = request.args.getlist(name)
            if vals:
                return [str(v).strip() for v in vals if str(v).strip()]
            single = (request.args.get(name) or "").strip()
            return [single] if single else []

        def _dec(v):
            if isinstance(v, Decimal):
                return v
            if v is None:
                return Decimal("0.00")
            try:
                return Decimal(str(v)).quantize(Decimal("0.01"))
            except Exception:
                return Decimal("0.00")

        def _money_num(v):
            if v is None:
                return Decimal("0.00")

            s = str(v).strip()
            if not s:
                return Decimal("0.00")

            s = (
                s.replace("\u00A0", " ")
                 .replace(" ", "")
                 .replace("COP", "")
                 .replace("$", "")
                 .replace("%", "")
            )

            if "," in s and "." in s:
                if s.rfind(",") > s.rfind("."):
                    s = s.replace(".", "").replace(",", ".")
                else:
                    s = s.replace(",", "")
            elif "," in s and "." not in s:
                s = s.replace(",", ".")
            else:
                s = s.replace(",", "")

            try:
                return Decimal(s).quantize(Decimal("0.01"))
            except Exception:
                return Decimal("0.00")

        def _client_norm(v):
            return " ".join(str(v or "").replace("\u00A0", " ").strip().upper().split())

        def _sql_client_norm(col):
            c = func.upper(func.trim(func.replace(func.coalesce(col, ""), "\u00A0", " ")))
            c = func.replace(c, "  ", " ")
            c = func.replace(c, "  ", " ")
            return c

        clientes_filter = [_client_norm(x) for x in _get_multi("cliente")]
        consultores_filter = _get_multi("consultor")
        modulos_filter = [x.upper() for x in _get_multi("modulo")]
        estados_ot_filter = _get_multi("estado_ot")
        servicios_filter = _get_multi("servicio")

        ocupacion_ids = [
            int(x) for x in request.args.getlist("ocupacion_id")
            if str(x).strip().isdigit()
        ]

        equipo_filter = (request.args.get("equipo") or "").strip().upper()
        filtro_proyecto_id = (request.args.get("proyecto_id") or "").strip()
        cliente_vinculado = (request.args.get("cliente_vinculado") or "").strip()

        # -------------------------------------------------
        # 1) RESUMEN OPERATIVO (Registro)
        # -------------------------------------------------
        q = (
            db.session.query(
                Registro.id.label("registro_id"),
                fecha_expr.label("fecha_real"),
                Registro.cliente.label("cliente"),
                Registro.usuario_consultor.label("usuario_consultor"),
                Registro.total_horas.label("total_horas"),
                Registro.tiempo_invertido.label("tiempo_invertido"),
                Registro.modulo.label("modulo"),
                Registro.proyecto_id.label("proyecto_id"),
                C.id.label("consultor_id"),
                C.nombre.label("consultor"),
                E.nombre.label("equipo"),
                O.codigo.label("ocupacion_codigo"),
                O.nombre.label("ocupacion_nombre"),
            )
            .select_from(Registro)
            .outerjoin(C, func.lower(Registro.usuario_consultor) == func.lower(C.usuario))
            .outerjoin(E, C.equipo_id == E.id)
            .outerjoin(O, Registro.ocupacion_id == O.id)
            .filter(fecha_expr >= desde)
            .filter(fecha_expr <= hasta)
        )

        if scope == "SELF":
            q = q.filter(func.lower(Registro.usuario_consultor) == usuario_norm)

        elif scope == "TEAM":
            if not int(val or 0):
                return jsonify({"error": "Consultor sin equipo asignado"}), 403
            q = q.filter(C.equipo_id == int(val))

        if clientes_filter:
            q = q.filter(_sql_client_norm(Registro.cliente).in_(clientes_filter))

        if consultores_filter:
            q = q.filter(C.nombre.in_(consultores_filter))

        if modulos_filter:
            q = q.filter(func.upper(func.coalesce(Registro.modulo, "")).in_(modulos_filter))

        if equipo_filter:
            q = q.filter(func.upper(func.coalesce(E.nombre, "")) == equipo_filter)

        if ocupacion_ids:
            q = q.filter(Registro.ocupacion_id.in_(ocupacion_ids))

        if filtro_proyecto_id.isdigit():
            q = _apply_project_filter_shared(q, int(filtro_proyecto_id))

        registros = q.all()

        resumen_map = {}
        graf_cliente_operativo = defaultdict(lambda: Decimal("0.00"))
        graf_ocupacion = defaultdict(lambda: Decimal("0.00"))
        graf_mensual_operativo = defaultdict(lambda: {
            "horas": Decimal("0.00"),
            "costo": Decimal("0.00"),
        })

        total_horas = Decimal("0.00")
        total_costo = Decimal("0.00")

        clientes_set = set()
        ocupaciones_set = set()
        consultores_set = set()

        costo_operativo_por_cliente = defaultdict(lambda: Decimal("0.00"))
        horas_operativas_por_cliente = defaultdict(lambda: Decimal("0.00"))

        for r in registros:
            fecha_reg = r.fecha_real
            if isinstance(fecha_reg, datetime):
                fecha_reg = fecha_reg.date()

            if not isinstance(fecha_reg, date):
                continue

            consultor_id = int(r.consultor_id or 0)

            cliente = (r.cliente or "SIN CLIENTE").strip() or "SIN CLIENTE"
            cliente_norm = _client_norm(cliente)

            ocupacion = (
                f"{(r.ocupacion_codigo or '').strip()} - {(r.ocupacion_nombre or '').strip()}".strip(" -")
                if (r.ocupacion_codigo or r.ocupacion_nombre)
                else "SIN OCUPACIÓN"
            )
            consultor_nombre = (r.consultor or "SIN NOMBRE").strip() or "SIN NOMBRE"
            equipo = (r.equipo or "SIN EQUIPO").strip() or "SIN EQUIPO"

            horas = _dec(r.tiempo_invertido if r.tiempo_invertido is not None else r.total_horas)

            valor_hora = Decimal("0.00")
            if consultor_id:
                presupuesto = _presupuesto_consultor_mes(
                    consultor_id,
                    int(fecha_reg.year),
                    int(fecha_reg.month),
                )
                if isinstance(presupuesto, dict):
                    valor_hora = _dec(presupuesto.get("valor_hora"))
                else:
                    valor_hora = _dec(presupuesto)

            costo = (horas * valor_hora).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP
            )

            key = f"{cliente}||{ocupacion}||{equipo}"

            if key not in resumen_map:
                resumen_map[key] = {
                    "cliente": cliente,
                    "clienteNorm": cliente_norm,
                    "ocupacion": ocupacion,
                    "equipo": equipo,
                    "horas": Decimal("0.00"),
                    "costoTotal": Decimal("0.00"),
                    "consultoresSet": set(),
                    "registrosCount": 0,
                    "detallePeriodos": defaultdict(lambda: {
                        "horas": Decimal("0.00"),
                        "costo": Decimal("0.00"),
                    }),
                    "detalleConsultores": defaultdict(lambda: {
                        "horas": Decimal("0.00"),
                        "costo": Decimal("0.00"),
                        "registrosCount": 0,
                    }),
                }

            bucket = resumen_map[key]
            bucket["horas"] = _dec(bucket["horas"]) + horas
            bucket["costoTotal"] = _dec(bucket["costoTotal"]) + costo
            bucket["consultoresSet"].add(consultor_nombre)
            bucket["registrosCount"] += 1

            periodo = f"{fecha_reg.year:04d}-{fecha_reg.month:02d}"
            bucket["detallePeriodos"][periodo]["horas"] = _dec(bucket["detallePeriodos"][periodo]["horas"]) + horas
            bucket["detallePeriodos"][periodo]["costo"] = _dec(bucket["detallePeriodos"][periodo]["costo"]) + costo

            bucket["detalleConsultores"][consultor_nombre]["horas"] = _dec(bucket["detalleConsultores"][consultor_nombre]["horas"]) + horas
            bucket["detalleConsultores"][consultor_nombre]["costo"] = _dec(bucket["detalleConsultores"][consultor_nombre]["costo"]) + costo
            bucket["detalleConsultores"][consultor_nombre]["registrosCount"] += 1

            graf_cliente_operativo[cliente] = _dec(graf_cliente_operativo[cliente]) + costo
            graf_ocupacion[ocupacion] = _dec(graf_ocupacion[ocupacion]) + costo

            graf_mensual_operativo[periodo]["horas"] = _dec(graf_mensual_operativo[periodo]["horas"]) + horas
            graf_mensual_operativo[periodo]["costo"] = _dec(graf_mensual_operativo[periodo]["costo"]) + costo

            costo_operativo_por_cliente[cliente_norm] = _dec(costo_operativo_por_cliente[cliente_norm]) + costo
            horas_operativas_por_cliente[cliente_norm] = _dec(horas_operativas_por_cliente[cliente_norm]) + horas

            total_horas = _dec(total_horas) + horas
            total_costo = _dec(total_costo) + costo

            clientes_set.add(cliente)
            ocupaciones_set.add(ocupacion)
            consultores_set.add(consultor_nombre)

        rows_out = []
        for _, item in resumen_map.items():
            horas_row = _dec(item["horas"])
            costo_row = _dec(item["costoTotal"])

            valor_hora_promedio = Decimal("0.00")
            if horas_row > 0:
                valor_hora_promedio = (costo_row / horas_row).quantize(
                    Decimal("0.01"),
                    rounding=ROUND_HALF_UP
                )

            detalle_periodos = [
                {
                    "periodo": periodo,
                    "horas": float(_dec(vals["horas"])),
                    "costo": float(_dec(vals["costo"])),
                }
                for periodo, vals in sorted(item["detallePeriodos"].items())
            ]

            detalle_consultores = [
                {
                    "consultor": nombre,
                    "horas": float(_dec(vals["horas"])),
                    "costo": float(_dec(vals["costo"])),
                    "registrosCount": int(vals["registrosCount"]),
                }
                for nombre, vals in sorted(item["detalleConsultores"].items(), key=lambda x: x[0])
            ]

            consultores_list = sorted(list(item["consultoresSet"]))

            rows_out.append({
                "cliente": item["cliente"],
                "clienteNorm": item["clienteNorm"],
                "ocupacion": item["ocupacion"],
                "equipo": item["equipo"],
                "horas": float(horas_row),
                "costoTotal": float(costo_row),
                "valorHoraPromedio": float(valor_hora_promedio),
                "consultoresCount": len(consultores_list),
                "consultores": consultores_list,
                "registrosCount": int(item["registrosCount"]),
                "detallePeriodos": detalle_periodos,
                "detalleConsultores": detalle_consultores,
            })

        rows_out.sort(key=lambda x: (-x["costoTotal"], x["cliente"], x["ocupacion"], x["equipo"]))

        # -----------------------------------------
        # 2) OPORTUNIDADES GANADAS / OT
        #    Valor comercial = OTC + MRC
        # -----------------------------------------
        qo = db.session.query(
            Oportunidad.id.label("id"),
            Oportunidad.nombre_cliente.label("nombre_cliente"),
            Oportunidad.servicio.label("servicio"),
            Oportunidad.codigo_prc.label("codigo_prc"),
            Oportunidad.fecha_creacion.label("fecha_creacion"),
            Oportunidad.estado_oferta.label("estado_oferta"),
            Oportunidad.resultado_oferta.label("resultado_oferta"),
            Oportunidad.estado_ot.label("estado_ot"),
            Oportunidad.otc.label("otc"),
            Oportunidad.mrc.label("mrc"),
        ).filter(
            or_(
                _sql_norm_estado(Oportunidad.estado_oferta) == _norm_key_for_match("GANADA"),
                _sql_norm_estado(Oportunidad.estado_oferta) == _norm_key_for_match("OT"),
                _sql_norm_estado(Oportunidad.resultado_oferta) == _norm_key_for_match("OT"),
            )
        )

        if clientes_filter:
            qo = qo.filter(_sql_client_norm(Oportunidad.nombre_cliente).in_(clientes_filter))

        if estados_ot_filter:
            qo = qo.filter(Oportunidad.estado_ot.in_(estados_ot_filter))

        if servicios_filter:
            qo = qo.filter(Oportunidad.servicio.in_(servicios_filter))

        if filtro_proyecto_id:
            try:
                qo = qo.join(
                    Proyecto,
                    Proyecto.oportunidad_id == Oportunidad.id
                ).filter(Proyecto.id == int(filtro_proyecto_id))
            except Exception:
                pass

        opp_rows_db = qo.order_by(
            Oportunidad.id.desc()
        ).all()

        opp_rows_out = []
        valor_comercial_por_cliente = defaultdict(lambda: Decimal("0.00"))
        oportunidades_por_cliente = defaultdict(int)
        valor_por_resultado = defaultdict(lambda: {
            "valor": Decimal("0.00"),
            "oportunidades": 0,
        })

        for op in opp_rows_db:
            fecha_op = op.fecha_creacion
            fecha_iso = None

            if isinstance(fecha_op, datetime):
                fecha_iso = fecha_op.date().isoformat()
            elif isinstance(fecha_op, date):
                fecha_iso = fecha_op.isoformat()
            else:
                try:
                    fecha_iso = str(fecha_op)[:10]
                except Exception:
                    fecha_iso = None

            otc = _money_num(op.otc)
            mrc = _money_num(op.mrc)

            valor_oportunidad = (otc + mrc).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP
            )

            mrc_normalizado = (mrc + (otc / Decimal("12"))).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP
            )

            if valor_oportunidad <= Decimal("0.00"):
                continue

            cliente_opp = (op.nombre_cliente or "SIN CLIENTE").strip() or "SIN CLIENTE"
            cliente_opp_norm = _client_norm(cliente_opp)
            prc = (op.codigo_prc or "SIN PRC").strip() or "SIN PRC"
            resultado = (op.resultado_oferta or "SIN RESULTADO").strip() or "SIN RESULTADO"

            valor_comercial_por_cliente[cliente_opp_norm] = _dec(valor_comercial_por_cliente[cliente_opp_norm]) + valor_oportunidad
            oportunidades_por_cliente[cliente_opp_norm] += 1

            valor_por_resultado[resultado]["valor"] = _dec(valor_por_resultado[resultado]["valor"]) + valor_oportunidad
            valor_por_resultado[resultado]["oportunidades"] += 1

            opp_rows_out.append({
                "id": op.id,
                "cliente": cliente_opp,
                "clienteNorm": cliente_opp_norm,
                "servicio": (op.servicio or "-").strip() or "-",
                "codigo_prc": prc,
                "fecha_creacion": fecha_iso,
                "estado_oferta": op.estado_oferta or "GANADA",
                "resultado_oferta": resultado,
                "estado_ot": op.estado_ot or "-",
                "otc": float(otc),
                "mrc": float(mrc),
                "valorOportunidad": float(valor_oportunidad),
                "mrcNormalizado": float(mrc_normalizado),
            })

        # -----------------------------------------
        # 3) MÁRGENES POR CLIENTE
        #    Ingreso = OTC + MRC
        #    Costo = suma del resumen operativo
        # -----------------------------------------
        clientes_margin_keys = sorted(
            set(list(valor_comercial_por_cliente.keys()) + list(costo_operativo_por_cliente.keys()))
        )

        margenes_por_cliente = []
        for cliente_norm in clientes_margin_keys:
            ingreso = _dec(valor_comercial_por_cliente.get(cliente_norm))
            costo = _dec(costo_operativo_por_cliente.get(cliente_norm))
            horas = _dec(horas_operativas_por_cliente.get(cliente_norm))
            margen = (ingreso - costo).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

            margen_pct = None
            if ingreso > 0:
                margen_pct = float(
                    ((margen / ingreso) * Decimal("100")).quantize(
                        Decimal("0.01"),
                        rounding=ROUND_HALF_UP
                    )
                )

            display_name = next(
                (
                    r["cliente"] for r in rows_out
                    if r.get("clienteNorm") == cliente_norm
                ),
                next(
                    (
                        o["cliente"] for o in opp_rows_out
                        if o.get("clienteNorm") == cliente_norm
                    ),
                    cliente_norm
                )
            )

            margenes_por_cliente.append({
                "cliente": display_name,
                "clienteNorm": cliente_norm,
                "ingreso": float(ingreso),
                "costo": float(costo),
                "horas": float(horas),
                "margen": float(margen),
                "margenPct": margen_pct,
                "oportunidadesCount": int(oportunidades_por_cliente.get(cliente_norm, 0)),
            })

        margenes_por_cliente.sort(key=lambda x: (-x["ingreso"], x["cliente"]))

        # -----------------------------------------
        # 4) RESUMEN FINANCIERO GLOBAL
        # -----------------------------------------
        ingreso_total = sum((_dec(v) for v in valor_comercial_por_cliente.values()), Decimal("0.00"))
        costo_total = _dec(total_costo)
        margen_total = (ingreso_total - costo_total).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        margen_pct_total = None
        if ingreso_total > 0:
            margen_pct_total = float(
                ((margen_total / ingreso_total) * Decimal("100")).quantize(
                    Decimal("0.01"),
                    rounding=ROUND_HALF_UP
                )
            )

        resumen_financiero = {
            "ingreso": float(ingreso_total),
            "costo": float(costo_total),
            "margen": float(margen_total),
            "margenPct": margen_pct_total,
        }

        # -----------------------------------------
        # 5) VÍNCULO DIRECTO POR CLIENTE
        #    usar ?cliente_vinculado=JGB
        # -----------------------------------------
        vinculo_cliente = None
        if cliente_vinculado:
            cliente_vinculado_norm = _client_norm(cliente_vinculado)
            resumen_rows_cliente = [
                r for r in rows_out
                if r.get("clienteNorm") == cliente_vinculado_norm
            ]
            oportunidades_cliente = [
                o for o in opp_rows_out
                if o.get("clienteNorm") == cliente_vinculado_norm
            ]

            ingreso_cliente = _dec(valor_comercial_por_cliente.get(cliente_vinculado_norm))
            costo_cliente = _dec(costo_operativo_por_cliente.get(cliente_vinculado_norm))
            margen_cliente = (ingreso_cliente - costo_cliente).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP
            )

            margen_pct_cliente = None
            if ingreso_cliente > 0:
                margen_pct_cliente = float(
                    ((margen_cliente / ingreso_cliente) * Decimal("100")).quantize(
                        Decimal("0.01"),
                        rounding=ROUND_HALF_UP
                    )
                )

            vinculo_cliente = {
                "cliente": cliente_vinculado,
                "clienteNorm": cliente_vinculado_norm,
                "ingreso": float(ingreso_cliente),
                "costo": float(costo_cliente),
                "margen": float(margen_cliente),
                "margenPct": margen_pct_cliente,
                "resumenRowsCount": len(resumen_rows_cliente),
                "oportunidadesCount": len(oportunidades_cliente),
                "resumenRows": resumen_rows_cliente,
                "oportunidadesRows": oportunidades_cliente,
            }

        return jsonify({
            "modo": modo,
            "desde": desde.isoformat(),
            "hasta": hasta.isoformat(),
            "totalHoras": float(_dec(total_horas)),
            "totalCosto": float(_dec(total_costo)),
            "totalClientes": len(clientes_set),
            "totalOcupaciones": len(ocupaciones_set),
            "totalConsultores": len(consultores_set),

            "resumenFinanciero": resumen_financiero,
            "margenesPorCliente": margenes_por_cliente,
            "vinculoCliente": vinculo_cliente,

            "rows": rows_out,

            "graficos": {
                "porCliente": [
                    {
                        "name": item["cliente"],
                        "costo": item["ingreso"],
                        "margen": item["margen"],
                        "margenPct": item["margenPct"],
                        "oportunidadesCount": item["oportunidadesCount"],
                    }
                    for item in margenes_por_cliente
                ],

                "porClienteOperativo": [
                    {"name": k, "costo": float(_dec(v))}
                    for k, v in sorted(graf_cliente_operativo.items(), key=lambda x: x[1], reverse=True)
                ],

                "porOcupacion": [
                    {"name": k, "costo": float(_dec(v))}
                    for k, v in sorted(graf_ocupacion.items(), key=lambda x: x[1], reverse=True)
                ],

                "porMes": [
                    {
                        "periodo": periodo,
                        "horas": float(_dec(vals["horas"])),
                        "costo": float(_dec(vals["costo"])),
                    }
                    for periodo, vals in sorted(graf_mensual_operativo.items(), key=lambda x: x[0])
                ],

                "oportunidadesPorResultado": [
                    {
                        "name": resultado,
                        "costo": float(_dec(vals["valor"])),
                        "oportunidades": int(vals["oportunidades"]),
                    }
                    for resultado, vals in sorted(
                        valor_por_resultado.items(),
                        key=lambda x: x[1]["valor"],
                        reverse=True
                    )
                ],
            },

            "oportunidadesGanadas": {
                "rows": opp_rows_out,
                "resumenPorCliente": [
                    {
                        "cliente": item["cliente"],
                        "clienteNorm": item["clienteNorm"],
                        "valorTotal": item["ingreso"],
                        "oportunidadesCount": item["oportunidadesCount"],
                    }
                    for item in margenes_por_cliente
                    if item["ingreso"] > 0
                ],
            },

            "filtrosAplicados": {
                "equipo": equipo_filter,
                "clientes": clientes_filter,
                "consultores": consultores_filter,
                "modulos": modulos_filter,
                "ocupacion_ids": ocupacion_ids,
                "estado_ot": estados_ot_filter,
                "servicios": servicios_filter,
                "proyecto_id": int(filtro_proyecto_id) if filtro_proyecto_id.isdigit() else None,
                "cliente_vinculado": cliente_vinculado or None,
            },
        }), 200

    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    except Exception as e:
        app.logger.exception("❌ Error en /dashboard/costos-resumen")
        return jsonify({"error": str(e)}), 500
    
@bp.route("/dashboard/costos-filtros", methods=["GET"])
@permission_required("GRAFICOS_VER")
def dashboard_costos_filtros():
    try:
        usuario = _get_usuario_from_request()
        rol_req = _rol_from_request()

        if not usuario:
            return jsonify({"error": "Usuario no enviado"}), 400

        usuario_norm = (usuario or "").strip().lower()

        consultor_login = (
            Consultor.query.options(
                joinedload(Consultor.rol_obj),
                joinedload(Consultor.equipo_obj),
            )
            .filter(func.lower(Consultor.usuario) == usuario_norm)
            .first()
        )

        if not consultor_login:
            return jsonify({"error": "Consultor no encontrado"}), 404

        # mismo alcance que el resumen
        scope, val = _scope_for_graficos(consultor_login, rol_req)

        # mismo parser de período que usa /dashboard/costos-resumen
        desde, hasta, modo = _dashboard_costos_parse_periodo_request()

        C = aliased(Consultor)
        E = aliased(Equipo)
        fecha_expr = _registro_fecha_expr()

        def _get_multi(name):
            vals = request.args.getlist(name)
            if vals:
                return [str(v).strip() for v in vals if str(v).strip()]
            single = (request.args.get(name) or "").strip()
            return [single] if single else []

        def _norm(v):
            return " ".join(str(v or "").strip().upper().split())

        # -----------------------------
        # filtros recibidos
        # -----------------------------
        equipo_filter = _norm(request.args.get("equipo"))
        clientes_filter = [_norm(v) for v in _get_multi("cliente")]
        consultores_filter = [_norm(v) for v in _get_multi("consultor")]
        modulos_filter = [_norm(v) for v in _get_multi("modulo")]
        estados_ot_filter = [_norm(v) for v in _get_multi("estado_ot")]
        servicios_filter = [_norm(v) for v in _get_multi("servicio")]

        ocupacion_ids = []
        for v in request.args.getlist("ocupacion_id"):
            try:
                ocupacion_ids.append(int(v))
            except Exception:
                pass

        filtro_proyecto_id = (request.args.get("proyecto_id") or "").strip()

        # =========================================================
        # 1) CATÁLOGOS OPERATIVOS DESDE REGISTRO
        #    cliente / consultor / modulo / equipo
        # =========================================================
        q = (
            db.session.query(
                Registro.cliente.label("cliente"),
                C.nombre.label("consultor"),
                Registro.modulo.label("modulo"),
                E.nombre.label("equipo"),
            )
            .select_from(Registro)
            .outerjoin(C, func.lower(Registro.usuario_consultor) == func.lower(C.usuario))
            .outerjoin(E, C.equipo_id == E.id)
            .filter(fecha_expr >= desde, fecha_expr <= hasta)
        )

        # alcance
        if scope == "SELF":
            q = q.filter(func.lower(Registro.usuario_consultor) == usuario_norm)

        elif scope == "TEAM":
            if not int(val or 0):
                return jsonify({"error": "Consultor sin equipo asignado"}), 403
            q = q.filter(C.equipo_id == int(val))

        elif scope == "ALL":
            pass

        # filtros contextuales
        if equipo_filter:
            if scope in ("TEAM", "SELF"):
                eq_login = ""
                if consultor_login.equipo_obj:
                    eq_login = _norm(consultor_login.equipo_obj.nombre)

                if equipo_filter != eq_login:
                    return jsonify({"error": "No autorizado para consultar otro equipo"}), 403

            q = q.filter(func.upper(E.nombre) == equipo_filter)

        if clientes_filter:
            q = q.filter(func.upper(Registro.cliente).in_(clientes_filter))

        if consultores_filter:
            q = q.filter(func.upper(C.nombre).in_(consultores_filter))

        if modulos_filter:
            q = q.filter(func.upper(Registro.modulo).in_(modulos_filter))

        if ocupacion_ids:
            q = q.filter(Registro.ocupacion_id.in_(ocupacion_ids))

        if filtro_proyecto_id:
            try:
                q = q.filter(Registro.proyecto_id == int(filtro_proyecto_id))
            except Exception:
                return jsonify({"error": "proyecto_id inválido"}), 400

        rows = q.all()

        clientes = sorted({
            str(r.cliente).strip()
            for r in rows
            if r.cliente and str(r.cliente).strip()
        })

        consultores = sorted({
            str(r.consultor).strip()
            for r in rows
            if r.consultor and str(r.consultor).strip()
        })

        modulos = sorted({
            _norm(r.modulo)
            for r in rows
            if r.modulo and str(r.modulo).strip()
        })

        equipos = sorted({
            _norm(r.equipo)
            for r in rows
            if r.equipo and str(r.equipo).strip()
        })

        # =========================================================
        # 2) CATÁLOGOS DE OPORTUNIDADES
        #    estado_ot / servicio
        # =========================================================
        oq = (
            db.session.query(
                Oportunidad.estado_ot.label("estado_ot"),
                Oportunidad.servicio.label("servicio"),
                Oportunidad.nombre_cliente.label("cliente"),
            )
            .select_from(Oportunidad)
        )

        # filtrar por período de creación de oportunidad
        oq = oq.filter(
            Oportunidad.fecha_creacion.isnot(None),
            Oportunidad.fecha_creacion >= desde,
            Oportunidad.fecha_creacion <= hasta,
        )

        if clientes_filter:
            oq = oq.filter(func.upper(Oportunidad.nombre_cliente).in_(clientes_filter))

        if estados_ot_filter:
            oq = oq.filter(func.upper(Oportunidad.estado_ot).in_(estados_ot_filter))

        if servicios_filter:
            oq = oq.filter(func.upper(Oportunidad.servicio).in_(servicios_filter))

        if filtro_proyecto_id:
            try:
                proyecto_id_int = int(filtro_proyecto_id)
                oq = oq.join(
                    Proyecto,
                    Proyecto.oportunidad_id == Oportunidad.id
                ).filter(Proyecto.id == proyecto_id_int)
            except Exception:
                return jsonify({"error": "proyecto_id inválido"}), 400

        opp_rows = oq.all()

        estados_ot = sorted({
            str(r.estado_ot).strip()
            for r in opp_rows
            if r.estado_ot and str(r.estado_ot).strip()
        })

        servicios = sorted({
            str(r.servicio).strip()
            for r in opp_rows
            if r.servicio and str(r.servicio).strip()
        })

        return jsonify({
            "modo": modo,
            "desde": desde.isoformat(),
            "hasta": hasta.isoformat(),

            "clientes": clientes,
            "consultores": consultores,
            "modulos": modulos,
            "equipos": equipos,
            "estados_ot": estados_ot,
            "servicios": servicios,

            "filtrosAplicados": {
                "equipo": equipo_filter or None,
                "clientes": clientes_filter,
                "consultores": consultores_filter,
                "modulos": modulos_filter,
                "ocupacion_ids": ocupacion_ids,
                "estado_ot": estados_ot_filter,
                "servicios": servicios_filter,
                "proyecto_id": int(filtro_proyecto_id) if filtro_proyecto_id.isdigit() else None,
            }
        }), 200

    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    except Exception as e:
        app.logger.exception("❌ Error en /dashboard/costos-filtros")
        return jsonify({"error": str(e)}), 500
    

@bp.route('/dashboard/proyectos-horas', methods=['GET'])
@permission_required("GRAFICOS_VER")
def obtener_proyectos_horas_dashboard():
    try:
        usuario = _get_usuario_from_request()
        rol_req = _get_rol_from_request()

        if not usuario:
            return jsonify({"error": "Usuario no enviado"}), 400

        usuario_norm = (usuario or "").strip().lower()

        consultor_login = (
            Consultor.query.options(
                joinedload(Consultor.rol_obj),
                joinedload(Consultor.equipo_obj),
            )
            .filter(func.lower(Consultor.usuario) == usuario_norm)
            .first()
        )

        if not consultor_login:
            return jsonify({"error": "Consultor no encontrado"}), 404

        scope, val = _scope_for_graficos(consultor_login, rol_req)

        C = aliased(Consultor)
        E = aliased(Equipo)

        q = (
            Registro.query
            .options(
                joinedload(Registro.consultor).joinedload(Consultor.equipo_obj),
                joinedload(Registro.tarea),
                joinedload(Registro.ocupacion),
                joinedload(Registro.proyecto),
                joinedload(Registro.fase_proyecto),
            )
            .outerjoin(C, func.lower(Registro.usuario_consultor) == func.lower(C.usuario))
            .outerjoin(E, C.equipo_id == E.id)
        )

        if scope == "SELF":
            q = q.filter(func.lower(Registro.usuario_consultor) == usuario_norm)

        elif scope == "TEAM":
            if not int(val or 0):
                return jsonify({"error": "Consultor sin equipo asignado"}), 403
            q = q.filter(C.equipo_id == int(val))

        elif scope == "ALL":
            pass

        def _get_list_param(name, upper=False):
            vals = request.args.getlist(name)
            if not vals:
                vals = request.args.getlist(f"{name}[]")

            one = request.args.get(name)
            if one and not vals:
                vals = [one]

            out = []
            for v in vals:
                s = str(v or "").strip()
                if not s:
                    continue
                out.append(s.upper() if upper else s)

            seen = set()
            cleaned = []
            for v in out:
                if v not in seen:
                    seen.add(v)
                    cleaned.append(v)
            return cleaned

        def _get_int_list_param(name):
            vals = request.args.getlist(name)
            if not vals:
                vals = request.args.getlist(f"{name}[]")

            one = request.args.get(name)
            if one and not vals:
                vals = [one]

            out = []
            for v in vals:
                try:
                    out.append(int(v))
                except Exception:
                    pass

            seen = set()
            cleaned = []
            for v in out:
                if v not in seen:
                    seen.add(v)
                    cleaned.append(v)
            return cleaned

        equipo_filter = (request.args.get("equipo") or "").strip().upper()

        if equipo_filter:
            if scope in ("TEAM", "SELF"):
                eq_login = ""
                if consultor_login.equipo_obj:
                    eq_login = (consultor_login.equipo_obj.nombre or "").strip().upper()

                if equipo_filter != eq_login:
                    return jsonify({"error": "No autorizado para consultar otro equipo"}), 403

            q = q.filter(func.upper(E.nombre) == equipo_filter)

        filtro_mes = (request.args.get("mes") or "").strip()
        filtro_desde = (request.args.get("desde") or "").strip()
        filtro_hasta = (request.args.get("hasta") or "").strip()
        filtro_modulo = (request.args.get("modulo") or "").strip()
        filtro_cliente = (request.args.get("cliente") or "").strip()
        filtro_consultor = (request.args.get("consultor") or "").strip()
        filtro_proyecto_ids = request.args.getlist("proyecto_id")
        if not filtro_proyecto_ids:
            filtro_proyecto_ids = request.args.getlist("proyecto_id[]")

        filtro_tarea_ids = _get_int_list_param("tarea_id")
        filtro_ocupacion_ids = _get_int_list_param("ocupacion_id")

        if filtro_mes:
            partes = filtro_mes.split("-")
            if len(partes) != 2:
                return jsonify({"error": "mes inválido, usa YYYY-MM"}), 400

            try:
                y = int(partes[0])
                m = int(partes[1])
            except ValueError:
                return jsonify({"error": "mes inválido, usa YYYY-MM"}), 400

            if m < 1 or m > 12:
                return jsonify({"error": "mes inválido, usa YYYY-MM"}), 400

            prefijo_mes = f"{y:04d}-{m:02d}"

            q = q.filter(
                func.substr(func.cast(Registro.fecha, db.String), 1, 7) == prefijo_mes
            )
        else:
            if filtro_desde:
                q = q.filter(func.cast(Registro.fecha, db.String) >= filtro_desde)
            if filtro_hasta:
                q = q.filter(func.cast(Registro.fecha, db.String) <= filtro_hasta)

        if filtro_modulo:
            q = q.filter(func.upper(Registro.modulo) == filtro_modulo.upper())

        if filtro_cliente:
            q = q.filter(Registro.cliente.ilike(f"%{filtro_cliente}%"))

        if filtro_consultor:
            q = q.filter(C.nombre.ilike(f"%{filtro_consultor}%"))

        if filtro_tarea_ids:
            q = q.filter(Registro.tarea_id.in_(filtro_tarea_ids))

        if filtro_ocupacion_ids:
            q = q.filter(Registro.ocupacion_id.in_(filtro_ocupacion_ids))

        if filtro_proyecto_ids:
            try:
                filtros_proyecto = []
                for pid in filtro_proyecto_ids:
                    subq = Registro.query.with_entities(Registro.id)
                    subq = _apply_project_filter_shared(subq, int(pid))
                    filtros_proyecto.append(Registro.id.in_(subq))

                q = q.filter(or_(*filtros_proyecto))
            except Exception:
                return jsonify({"error": "proyecto_id inválido"}), 400

        q = q.order_by(Registro.fecha.desc(), Registro.id.desc())

        tiene_filtro_temporal = bool(
            filtro_mes or filtro_desde or filtro_hasta or filtro_proyecto_ids
        )

        if tiene_filtro_temporal:
            registros = q.all()
            truncated = False
            max_rows = None
        else:
            max_rows = request.args.get("max_rows", type=int) or 5000
            max_rows = min(max(max_rows, 1), 10000)

            registros = q.limit(max_rows + 1).all()
            truncated = len(registros) > max_rows

            if truncated:
                registros = registros[:max_rows]

        data = []

        for r in registros:
            tarea = getattr(r, "tarea", None)
            ocup = getattr(r, "ocupacion", None)
            proyecto = getattr(r, "proyecto", None)
            fase_proyecto = getattr(r, "fase_proyecto", None)

            horas = _safe_float_report(
                r.tiempo_invertido if r.tiempo_invertido is not None else r.total_horas
            )

            if tarea and getattr(tarea, "codigo", None) and getattr(tarea, "nombre", None):
                tipo_tarea_str = f"{tarea.codigo} - {tarea.nombre}"
            else:
                tipo_tarea_str = (getattr(r, "tipo_tarea", "") or "").strip() or None

            equipo_nombre = None
            if r.consultor and getattr(r.consultor, "equipo_obj", None):
                equipo_nombre = (r.consultor.equipo_obj.nombre or "").strip().upper()

            equipo_raw = getattr(r, "equipo", None)

            data.append({
                "id": r.id,
                "fecha": _safe_fecha_iso(r.fecha),
                "modulo": r.modulo,
                "cliente": r.cliente,
                "equipo": equipo_nombre or (str(equipo_raw).strip().upper() if equipo_raw else "SIN EQUIPO"),

                "nroCasoCliente": r.nro_caso_cliente,
                "nroCasoInterno": r.nro_caso_interno,
                "nroCasoEscaladoSap": r.nro_caso_escalado,

                "ocupacion_id": r.ocupacion_id,
                "ocupacion_codigo": ocup.codigo if ocup else None,
                "ocupacion_nombre": ocup.nombre if ocup else None,

                "tarea_id": r.tarea_id,
                "tipoTarea": tipo_tarea_str,
                "tarea": {
                    "id": tarea.id,
                    "codigo": getattr(tarea, "codigo", None),
                    "nombre": getattr(tarea, "nombre", None),
                } if tarea else None,

                "consultor": r.consultor.nombre if r.consultor else None,
                "usuario_consultor": (r.usuario_consultor or "").strip().lower(),

                "horaInicio": r.hora_inicio,
                "horaFin": r.hora_fin,
                "tiempoInvertido": round(horas, 2),
                "tiempoFacturable": _safe_float_report(r.tiempo_facturable),
                "horasAdicionales": _calcular_horas_adicionales_por_horario(
                    r.hora_inicio,
                    r.hora_fin,
                    getattr(r, "horario_trabajo", None),
                    equipo_nombre or getattr(r, "equipo", None),
                    r.horas_adicionales,
                ),
                "horarioTrabajo": _normalizar_horario_trabajo_por_equipo(
                    getattr(r, "horario_trabajo", None),
                    equipo_nombre or getattr(r, "equipo", None),
                ),
                "horario_trabajo": _normalizar_horario_trabajo_por_equipo(
                    getattr(r, "horario_trabajo", None),
                    equipo_nombre or getattr(r, "equipo", None),
                ),
                "descripcion": r.descripcion,
                "totalHoras": round(horas, 2),

                "bloqueado": bool(r.bloqueado),
                "oncall": r.oncall,
                "desborde": r.desborde,
                "actividadMalla": r.actividad_malla,

                "proyecto_id": r.proyecto_id,
                "fase_proyecto_id": r.fase_proyecto_id,
                "proyecto": {
                    "id": proyecto.id,
                    "codigo": proyecto.codigo,
                    "nombre": proyecto.nombre,
                    "activo": bool(getattr(proyecto, "activo", True)),
                } if proyecto else None,
                "fase_proyecto": {
                    "id": fase_proyecto.id,
                    "nombre": fase_proyecto.nombre,
                } if fase_proyecto else None,
                "proyecto_codigo": proyecto.codigo if proyecto else None,
                "proyecto_nombre": proyecto.nombre if proyecto else None,
                "proyecto_fase": fase_proyecto.nombre if fase_proyecto else None,
            })

        return jsonify({
            "data": data,
            "truncated": truncated,
            "max_rows": max_rows,
        }), 200

    except Exception as e:
        err = traceback.format_exc()
        app.logger.error(f"❌ Error en /dashboard/proyectos-horas: {e}\n{err}")
        return jsonify({
            "error": "Error interno del servidor",
            "detalle": str(e)
        }), 500

@bp.route('/proyectos/dashboard', methods=['GET'])
@auth_required
def dashboard_proyectos():
    try:
        # =========================
        # PARAMETROS
        # =========================
        desde = request.args.get("desde")
        hasta = request.args.get("hasta")

        equipo = request.args.getlist("equipo")
        consultor = request.args.getlist("consultor")
        cliente = request.args.getlist("cliente")
        modulo = request.args.getlist("modulo")
        proyecto = request.args.getlist("proyecto")

        page = int(request.args.get("page", 1))
        page_size = int(request.args.get("page_size", 500))

        # =========================
        # QUERY BASE
        # =========================
        query = Registro.query

        # 👉 SOLO registros que tengan proyecto
        query = query.filter(Registro.proyecto_id.isnot(None))

        # =========================
        # FILTRO FECHA 🔥
        # =========================
        if desde:
            query = query.filter(Registro.fecha >= desde)

        if hasta:
            query = query.filter(Registro.fecha <= hasta)

        # =========================
        # FILTROS DINÁMICOS
        # =========================
        if equipo:
            query = query.filter(Registro.equipo.in_(equipo))

        if consultor:
            query = query.filter(Registro.usuario_consultor.in_(consultor))

        if cliente:
            query = query.filter(Registro.cliente.in_(cliente))

        if modulo:
            query = query.filter(Registro.modulo.in_(modulo))

        if proyecto:
            query = query.filter(Registro.proyecto_id.in_(proyecto))

        # =========================
        # PAGINACIÓN
        # =========================
        total = query.count()

        registros = (
            query
            .order_by(Registro.fecha.asc())  # importante para gráficas
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )

        return jsonify({
            "data": [registro_to_dict(r) for r in registros],
            "total": total,
            "page": page,
            "page_size": page_size
        }), 200

    except Exception as e:
        return jsonify({"mensaje": str(e)}), 500


## Cambiar password
@bp.route('/cambiar-password', methods=['POST'])
def cambiar_password():
    data = request.get_json(silent=True) or {}

    usuario = (data.get("usuario") or "").strip().lower()
    password_actual = data.get("passwordActual") or ""
    nueva_password = data.get("nuevaPassword") or ""
    confirmar_password = data.get("confirmarPassword") or ""

    if not usuario or not password_actual or not nueva_password or not confirmar_password:
        return jsonify({
            "mensaje": "Usuario, contraseña actual, nueva contraseña y confirmación son obligatorios"
        }), 400

    if nueva_password != confirmar_password:
        return jsonify({
            "mensaje": "La nueva contraseña y la confirmación no coinciden"
        }), 400

    if len(nueva_password) < 6:
        return jsonify({
            "mensaje": "La nueva contraseña debe tener mínimo 6 caracteres"
        }), 400

    consultor = Consultor.query.filter(
        func.lower(Consultor.usuario) == usuario
    ).first()

    if not consultor:
        return jsonify({"mensaje": "Usuario no encontrado"}), 404

    if not bool(getattr(consultor, "activo", True)):
        return jsonify({
            "mensaje": "Usuario inactivo. Contacte al administrador."
        }), 403

    stored = consultor.password or ""

    if stored.startswith("$2"):
        ok_pass = bcrypt.checkpw(
            password_actual.encode("utf-8"),
            stored.encode("utf-8")
        )
    else:
        ok_pass = stored == password_actual

    if not ok_pass:
        return jsonify({
            "mensaje": "La contraseña actual no es correcta"
        }), 401

    hashed = bcrypt.hashpw(
        nueva_password.encode("utf-8"),
        bcrypt.gensalt()
    ).decode("utf-8")

    consultor.password = hashed

    try:
        db.session.commit()
        return jsonify({
            "mensaje": "Contraseña actualizada correctamente"
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "mensaje": f"No se pudo actualizar la contraseña: {str(e)}"
        }), 500

# ============================================================
# PERFILES + MÓDULOS
# ============================================================

def _get_modulos_ids_from_payload(data):
    modulos_ids = data.get("modulos") or data.get("modulos_ids") or []

    if not isinstance(modulos_ids, list):
        raise ValueError("modulos debe ser una lista de ids")

    clean = []

    for mid in modulos_ids:
        try:
            mid_int = int(mid)
        except Exception:
            raise ValueError(f"modulo_id inválido: {mid}")

        if mid_int > 0:
            clean.append(mid_int)

    return list(dict.fromkeys(clean))


def _sync_perfil_modulos(perfil_id, modulos_ids):
    ModuloPerfil.query.filter_by(
        perfil_id=perfil_id
    ).delete(synchronize_session=False)

    for mid in modulos_ids:
        db.session.add(
            ModuloPerfil(
                perfil_id=perfil_id,
                modulo_id=int(mid),
                activo=True,
            )
        )


@bp.route("/perfiles", methods=["GET"])
@permission_required("PERFILES_VER")
def listar_perfiles_catalogo():
    try:
        include_modulos = (request.args.get("include_modulos") or "0") == "1"
        q = (request.args.get("q") or "").strip()
        activos = (request.args.get("activos") or "").strip().lower()

        query = Perfil.query

        if include_modulos:
            query = query.options(
                selectinload(Perfil.modulos).joinedload(ModuloPerfil.modulo)
            )

        if q:
            like = f"%{q}%"
            query = query.filter(
                or_(
                    Perfil.codigo.ilike(like),
                    Perfil.nombre.ilike(like),
                    Perfil.descripcion.ilike(like),
                )
            )

        if activos in ("1", "true", "si", "sí", "yes"):
            query = query.filter(Perfil.activo == True)

        rows = query.order_by(
            Perfil.activo.desc(),
            Perfil.orden.asc(),
            Perfil.nombre.asc(),
        ).all()

        return jsonify([
            perfil_to_dict(x, include_modulos=include_modulos)
            for x in rows
        ]), 200

    except Exception as e:
        app.logger.exception("Error listando perfiles")
        return jsonify({"mensaje": f"Error listando perfiles: {str(e)}"}), 500


@bp.route("/perfiles/<int:perfil_id>", methods=["GET"])
@permission_required("PERFILES_VER")
def get_perfil_catalogo(perfil_id):
    perfil = (
        Perfil.query
        .options(selectinload(Perfil.modulos).joinedload(ModuloPerfil.modulo))
        .get_or_404(perfil_id)
    )

    return jsonify(perfil_to_dict(perfil, include_modulos=True)), 200


@bp.route("/perfiles", methods=["POST"])
@permission_required("PERFILES_CREAR")
def crear_perfil_catalogo():
    data = request.get_json(silent=True) or {}

    try:
        nombre = (data.get("nombre") or "").strip()
        codigo = (data.get("codigo") or _perfil_build_code(nombre)).strip().upper()
        descripcion = (data.get("descripcion") or "").strip() or None
        orden = int(data.get("orden") or 0)
        activo = _to_bool2(data.get("activo"), default=True)
        modulos_ids = _get_modulos_ids_from_payload(data)

        if not nombre:
            return jsonify({"mensaje": "nombre requerido"}), 400

        if not codigo:
            return jsonify({"mensaje": "codigo requerido"}), 400

        if not modulos_ids:
            return jsonify({
                "mensaje": "Debes asignar al menos un módulo al perfil"
            }), 400

        dupe_codigo = Perfil.query.filter(
            func.lower(Perfil.codigo) == codigo.lower()
        ).first()

        if dupe_codigo:
            return jsonify({"mensaje": "Ya existe un perfil con ese código"}), 400

        dupe_nombre = Perfil.query.filter(
            func.lower(Perfil.nombre) == nombre.lower()
        ).first()

        if dupe_nombre:
            return jsonify({"mensaje": "Ya existe un perfil con ese nombre"}), 400

        mods = Modulo.query.filter(Modulo.id.in_(modulos_ids)).all()
        found_ids = {int(m.id) for m in mods}
        missing = [mid for mid in modulos_ids if mid not in found_ids]

        if missing:
            return jsonify({"mensaje": f"Módulos no encontrados: {missing}"}), 400

        perfil = Perfil(
            codigo=codigo,
            nombre=nombre,
            descripcion=descripcion,
            orden=orden,
            activo=activo,
        )

        db.session.add(perfil)
        db.session.flush()

        _sync_perfil_modulos(perfil.id, modulos_ids)

        db.session.commit()

        perfil_db = (
            Perfil.query
            .options(selectinload(Perfil.modulos).joinedload(ModuloPerfil.modulo))
            .get(perfil.id)
        )

        return jsonify({
            "mensaje": "Perfil creado",
            "perfil": perfil_to_dict(perfil_db, include_modulos=True)
        }), 201

    except ValueError as e:
        db.session.rollback()
        return jsonify({"mensaje": str(e)}), 400

    except IntegrityError as e:
        db.session.rollback()
        app.logger.exception("Error de integridad creando perfil")
        return jsonify({"mensaje": f"No se pudo crear el perfil: {str(e)}"}), 400

    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error creando perfil")
        return jsonify({"mensaje": f"No se pudo crear el perfil: {str(e)}"}), 500


@bp.route("/perfiles/<int:perfil_id>", methods=["PUT"])
@permission_required("PERFILES_EDITAR")
def editar_perfil_catalogo(perfil_id):
    perfil = Perfil.query.get_or_404(perfil_id)
    data = request.get_json(silent=True) or {}

    try:
        if "nombre" in data:
            nombre = (data.get("nombre") or "").strip()
            if not nombre:
                return jsonify({"mensaje": "nombre inválido"}), 400

            dupe = Perfil.query.filter(
                func.lower(Perfil.nombre) == nombre.lower(),
                Perfil.id != perfil_id
            ).first()

            if dupe:
                return jsonify({"mensaje": "Ya existe otro perfil con ese nombre"}), 400

            perfil.nombre = nombre

        if "codigo" in data:
            codigo = (data.get("codigo") or "").strip().upper()
            if not codigo:
                return jsonify({"mensaje": "codigo inválido"}), 400

            dupe = Perfil.query.filter(
                func.lower(Perfil.codigo) == codigo.lower(),
                Perfil.id != perfil_id
            ).first()

            if dupe:
                return jsonify({"mensaje": "Ya existe otro perfil con ese código"}), 400

            perfil.codigo = codigo

        if "descripcion" in data:
            perfil.descripcion = (data.get("descripcion") or "").strip() or None

        if "orden" in data:
            perfil.orden = int(data.get("orden") or 0)

        if "activo" in data:
            perfil.activo = _to_bool2(data.get("activo"), default=True)

        if "modulos" in data or "modulos_ids" in data:
            modulos_ids = _get_modulos_ids_from_payload(data)

            if not modulos_ids:
                return jsonify({
                    "mensaje": "Debes asignar al menos un módulo al perfil"
                }), 400

            mods = Modulo.query.filter(Modulo.id.in_(modulos_ids)).all()
            found_ids = {int(m.id) for m in mods}
            missing = [mid for mid in modulos_ids if mid not in found_ids]

            if missing:
                return jsonify({"mensaje": f"Módulos no encontrados: {missing}"}), 400

            _sync_perfil_modulos(perfil.id, modulos_ids)

        db.session.commit()

        perfil_db = (
            Perfil.query
            .options(selectinload(Perfil.modulos).joinedload(ModuloPerfil.modulo))
            .get(perfil.id)
        )

        return jsonify({
            "mensaje": "Perfil actualizado",
            "perfil": perfil_to_dict(perfil_db, include_modulos=True)
        }), 200

    except ValueError as e:
        db.session.rollback()
        return jsonify({"mensaje": str(e)}), 400

    except IntegrityError as e:
        db.session.rollback()
        app.logger.exception("Error de integridad actualizando perfil")
        return jsonify({"mensaje": f"No se pudo actualizar el perfil: {str(e)}"}), 400

    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error actualizando perfil")
        return jsonify({"mensaje": f"No se pudo actualizar el perfil: {str(e)}"}), 500


@bp.route("/perfiles/<int:perfil_id>", methods=["DELETE"])
@permission_required("PERFILES_ELIMINAR")
def desactivar_perfil_catalogo(perfil_id):
    perfil = Perfil.query.get_or_404(perfil_id)

    try:
        perfil.activo = False
        db.session.commit()

        return jsonify({
            "mensaje": "Perfil desactivado",
            "perfil": perfil_to_dict(perfil, include_modulos=False)
        }), 200

    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error desactivando perfil")
        return jsonify({"mensaje": f"No se pudo desactivar el perfil: {str(e)}"}), 500

### Herencias oportunidades 

@bp.route("/oportunidades/<int:id>/marcar-principal", methods=["PUT"])
@permission_required("OPORTUNIDADES_EDITAR")
def marcar_oportunidad_principal(id):
    try:
        oportunidad = Oportunidad.query.get_or_404(id)

        cliente_key = _norm_key_for_match(oportunidad.nombre_cliente)

        if not cliente_key:
            return jsonify({
                "mensaje": "La oportunidad no tiene nombre de cliente válido"
            }), 400

        ultimo_consecutivo = (
            db.session.query(func.max(Oportunidad.consecutivo_principal))
            .filter(
                or_(
                    Oportunidad.cliente_grupo_key == cliente_key,
                    _sql_norm_estado(Oportunidad.nombre_cliente) == cliente_key
                )
            )
            .scalar()
        )

        consecutivo = int(ultimo_consecutivo or 0) + 1

        oportunidad.tipo_oportunidad = "PRINCIPAL"
        oportunidad.oportunidad_padre_id = None
        oportunidad.cliente_grupo_key = cliente_key
        oportunidad.consecutivo_principal = consecutivo
        oportunidad.consecutivo_sub = None
        oportunidad.codigo_control = str(consecutivo)

        db.session.commit()

        return jsonify({
            "mensaje": "Oportunidad marcada como principal",
            "oportunidad": oportunidad.to_dict()
        }), 200

    except Exception as e:
        db.session.rollback()
        app.logger.exception(f"Error marcando oportunidad como principal id={id}")

        return jsonify({
            "mensaje": f"Error marcando oportunidad como principal: {str(e)}",
            "trace": traceback.format_exc()
        }), 500
    
@bp.route("/oportunidades/<int:id>/asignar-principal", methods=["PUT"])
@permission_required("OPORTUNIDADES_EDITAR")
def asignar_oportunidad_a_principal(id):
    try:
        data = request.get_json() or {}
        padre_id = data.get("oportunidad_padre_id")

        if not padre_id:
            return jsonify({"mensaje": "Debe enviar oportunidad_padre_id"}), 400

        oportunidad = Oportunidad.query.get_or_404(id)
        principal = Oportunidad.query.get_or_404(int(padre_id))

        if oportunidad.id == principal.id:
            return jsonify({"mensaje": "Una oportunidad no puede asignarse a sí misma"}), 400

        if _norm_key_for_match(oportunidad.nombre_cliente) != _norm_key_for_match(principal.nombre_cliente):
            return jsonify({
                "mensaje": "Solo se pueden asociar oportunidades del mismo cliente"
            }), 400

        if _norm_key_for_match(principal.tipo_oportunidad) != "PRINCIPAL":
            return jsonify({
                "mensaje": "La oportunidad seleccionada no está marcada como principal"
            }), 400

        cliente_key = principal.cliente_grupo_key or _norm_key_for_match(principal.nombre_cliente)

        if not principal.consecutivo_principal:
            ultimo_principal = (
                db.session.query(func.max(Oportunidad.consecutivo_principal))
                .filter(
                    or_(
                        Oportunidad.cliente_grupo_key == cliente_key,
                        _sql_norm_estado(Oportunidad.nombre_cliente) == cliente_key
                    )
                )
                .scalar()
            )

            principal.consecutivo_principal = int(ultimo_principal or 0) + 1
            principal.codigo_control = str(principal.consecutivo_principal)
            principal.cliente_grupo_key = cliente_key

        ultimo_sub = (
            db.session.query(func.max(Oportunidad.consecutivo_sub))
            .filter(Oportunidad.oportunidad_padre_id == principal.id)
            .scalar()
        )

        consecutivo_sub = int(ultimo_sub or 0) + 1

        oportunidad.tipo_oportunidad = "SUBOPORTUNIDAD"
        oportunidad.oportunidad_padre_id = principal.id
        oportunidad.cliente_grupo_key = cliente_key
        oportunidad.consecutivo_principal = principal.consecutivo_principal
        oportunidad.consecutivo_sub = consecutivo_sub
        oportunidad.codigo_control = f"{principal.codigo_control or principal.consecutivo_principal}.{consecutivo_sub}"

        db.session.commit()

        return jsonify({
            "mensaje": "Oportunidad asignada a principal",
            "oportunidad": oportunidad.to_dict()
        }), 200

    except Exception as e:
        db.session.rollback()
        app.logger.exception(f"Error asignando oportunidad id={id} a principal")

        return jsonify({
            "mensaje": f"Error asignando oportunidad a principal: {str(e)}",
            "trace": traceback.format_exc()
        }), 500
    
@bp.route("/oportunidades/<int:id>/quitar-principal", methods=["PUT"])
@permission_required("OPORTUNIDADES_EDITAR")
def quitar_oportunidad_de_principal(id):
    try:
        oportunidad = Oportunidad.query.get_or_404(id)

        oportunidad.tipo_oportunidad = "SUBOPORTUNIDAD"
        oportunidad.oportunidad_padre_id = None
        oportunidad.consecutivo_sub = None

        db.session.commit()

        return jsonify({
            "mensaje": "Oportunidad retirada de la principal",
            "oportunidad": oportunidad.to_dict()
        }), 200

    except Exception:
        db.session.rollback()
        return jsonify({
            "mensaje": "Error retirando oportunidad de la principal",
            "trace": traceback.format_exc()
        }), 500
    
@bp.route("/oportunidades/principales", methods=["GET"])
@permission_required("OPORTUNIDADES_VER")
def listar_oportunidades_principales():
    try:
        cliente = (request.args.get("cliente") or "").strip()

        query = Oportunidad.query.filter(
            func.upper(Oportunidad.tipo_oportunidad) == "PRINCIPAL"
        )

        if cliente:
            query = query.filter(
                func.upper(func.trim(Oportunidad.nombre_cliente)) ==
                cliente.strip().upper()
            )

        rows = query.order_by(
            Oportunidad.nombre_cliente.asc(),
            Oportunidad.consecutivo_principal.asc(),
            Oportunidad.id.asc()
        ).all()

        return jsonify([o.to_dict() for o in rows]), 200

    except Exception:
        return jsonify({
            "mensaje": "Error consultando oportunidades principales",
            "trace": traceback.format_exc()
        }), 500
    
POST_PRC_CLEAR_FIELDS = [
    "codigo_prc",
    "fecha_firma_aos",
    "pm_asignado_claro",
    "pm_asignado_hitss",
    "descripcion_ot",
    "num_enlace",
    "num_incidente",
    "num_ot",
    "estado_ot",
    "proyeccion_ingreso",
    "fecha_compromiso",
    "fecha_cierre",
    "estado_proyecto",
    "anio_creacion_ot",
    "fecha_acta_cierre_ot",
    "seguimiento_ot",
    "tipo_servicio",
    "semestre_ejecucion",
    "publicacion_sharepoint",
]


@bp.route("/oportunidades/<int:id>/copiar-como-principal", methods=["POST"])
@permission_required("OPORTUNIDADES_EDITAR")
def copiar_oportunidad_como_principal(id):
    try:
        origen = Oportunidad.query.get_or_404(id)

        cliente_key = _norm_key_for_match(origen.nombre_cliente)

        if not cliente_key:
            return jsonify({
                "mensaje": "La oportunidad origen no tiene nombre de cliente válido"
            }), 400

        ultimo_principal = (
            db.session.query(func.max(Oportunidad.consecutivo_principal))
            .filter(_sql_norm_estado(Oportunidad.tipo_oportunidad) == "PRINCIPAL")
            .scalar()
        )

        consecutivo_principal = int(ultimo_principal or 0) + 1

        data = {}

        for column in Oportunidad.__table__.columns:
            field = column.name

            if field in ["id", "created_at", "updated_at"]:
                continue

            data[field] = getattr(origen, field, None)

        # Limpia bloque OT/Proyecto desde CODIGO PRC hacia la derecha
        for field in POST_PRC_CLEAR_FIELDS:
            if field in data:
                data[field] = None

        data["tipo_oportunidad"] = "PRINCIPAL"
        data["oportunidad_padre_id"] = None
        data["cliente_grupo_key"] = cliente_key
        data["consecutivo_principal"] = consecutivo_principal
        data["consecutivo_sub"] = None
        data["codigo_control"] = str(consecutivo_principal)

        principal = Oportunidad(**data)

        db.session.add(principal)
        db.session.flush()

        # La oportunidad original queda asignada a la principal creada
        origen.tipo_oportunidad = "SUBOPORTUNIDAD"
        origen.oportunidad_padre_id = principal.id
        origen.cliente_grupo_key = cliente_key
        origen.consecutivo_principal = consecutivo_principal
        origen.consecutivo_sub = 1
        origen.codigo_control = f"{principal.codigo_control}.1"

        db.session.commit()

        return jsonify({
            "mensaje": "Oportunidad principal creada desde copia",
            "principal": principal.to_dict(),
            "suboportunidad": origen.to_dict(),
        }), 201

    except Exception as e:
        db.session.rollback()
        app.logger.exception(f"Error copiando oportunidad como principal id={id}")

        return jsonify({
            "mensaje": f"Error copiando oportunidad como principal: {str(e)}",
            "trace": traceback.format_exc()
        }), 500

# ============================================================
# BASE DE REGISTRO DE INFORMACION COE SAP FUNCIONAL
# ============================================================

COE_SAP_FUNCIONAL_ALIASES = {
    "numero": [
        "NUMERO",
        "NÚMERO",
        "NO",
        "NO.",
        "NRO",
        "NRO.",
        "N°",
        "#",
    ],
    "id_interaccion": [
        "ID DE INTERACCION",
        "ID DE INTERACCIÓN",
        "ID INTERACCION",
        "ID INTERACCIÓN",
    ],
    "compania": [
        "COMPANIA",
        "COMPAÑIA",
        "COMPAÑÍA",
    ],
    "fecha_entrega": [
        "FECHA DE ENTREGA",
        "FECHA ENTREGA",
    ],
    "fecha_resolucion": [
        "FECHA DE RESOLUCION",
        "FECHA DE RESOLUCIÓN",
        "FECHA RESOLUCION",
        "FECHA RESOLUCIÓN",
    ],
    "fecha_cierre": [
        "FECHA DE CIERRE",
        "FECHA CIERRE",
    ],
    "estado": [
        "ESTADO",
    ],
    "titulo": [
        "TITULO",
        "TÍTULO",
    ],
    "asignado_a": [
        "ASIGNADO A",
        "ASIGNADO",
    ],
    "nombre_completo_contacto": [
        "NOMBRE COMPLETO CONTACTO",
        "CONTACTO",
        "NOMBRE CONTACTO",
    ],
    "incumplimiento_sla": [
        "INCUMPLIMIENTO DE SLA",
        "INCUMPLIMIENTO SLA",
    ],
    "alerta": [
        "ALERTA",
    ],
    "estado_alerta_ans": [
        "ESTADO DE ALERTA ANS",
        "ESTADO ALERTA ANS",
    ],
    "impacto": [
        "IMPACTO",
    ],
    "urgencia": [
        "URGENCIA",
    ],
    "prioridad": [
        "PRIORIDAD",
    ],
    "accion_actualizacion": [
        "ACCION DE ACTUALIZACION",
        "ACCIÓN DE ACTUALIZACIÓN",
        "ACCION ACTUALIZACION",
        "ACCIÓN ACTUALIZACIÓN",
    ],
    "canal_resolucion": [
        "CANAL DE RESOLUCION",
        "CANAL DE RESOLUCIÓN",
        "CANAL RESOLUCION",
        "CANAL RESOLUCIÓN",
    ],
    "clr_txt_servicio": [
        "CLR TXT SERVICIO",
    ],
    "clr_txt_client_type": [
        "CLR TXT CLIENT TYPE",
    ],
}

def _coe_norm_col(value):
    s = str(value or "")

    # Quitar BOM y caracteres invisibles comunes
    s = s.replace("\ufeff", "")
    s = s.replace("\u200b", "")
    s = s.replace("\u200c", "")
    s = s.replace("\u200d", "")
    s = s.replace("\u00A0", " ")

    s = s.strip().upper()

    # Quitar tildes
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")

    # Normalizar símbolos comunes
    s = s.replace("º", "")
    s = s.replace("°", "")
    s = s.replace(".", " ")
    s = s.replace(":", " ")
    s = s.replace("-", " ")
    s = s.replace("_", " ")
    s = s.replace("/", " ")

    # Normalizar espacios
    s = re.sub(r"\s+", " ", s).strip()

    return s


def _coe_parse_bool(value):
    if value is None:
        return None

    if isinstance(value, bool):
        return value

    s = str(value).strip().lower()

    if s in ("", "nan", "none", "null"):
        return None

    if s in ("true", "1", "si", "sí", "s", "yes", "y", "x"):
        return True

    if s in ("false", "0", "no", "n"):
        return False

    return None


def _coe_parse_datetime(value):
    if value is None:
        return None

    if hasattr(value, "to_pydatetime"):
        try:
            return value.to_pydatetime()
        except Exception:
            pass

    if isinstance(value, datetime):
        return value

    s = str(value).strip()

    if s == "" or s.lower() in ("nan", "none", "null"):
        return None

    try:
        parsed = pd.to_datetime(s, errors="coerce", dayfirst=True)

        if pd.isna(parsed):
            return None

        return parsed.to_pydatetime()
    except Exception:
        return None


def _coe_parse_str(value):
    if value is None:
        return None

    s = str(value).replace("\u00A0", " ").strip()

    if s == "" or s.lower() in ("nan", "none", "null"):
        return None

    return s


def _coe_format_datetime(value):
    if value is None:
        return None

    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            pass

    s = str(value).strip()

    if s == "" or s.lower() in ("nan", "none", "null"):
        return None

    try:
        parsed = pd.to_datetime(s, errors="coerce", dayfirst=True)

        if pd.isna(parsed):
            return s

        return parsed.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return s


def _coe_format_bool(value):
    if value is None:
        return None

    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        return bool(value)

    s = str(value).strip().lower()

    if s in ("true", "1", "si", "sí", "s", "yes", "y", "x"):
        return True

    if s in ("false", "0", "no", "n"):
        return False

    return None


def _coe_find_column(columnas_normalizadas, aliases):
    for alias in aliases:
        alias_norm = _coe_norm_col(alias)

        if alias_norm in columnas_normalizadas:
            return columnas_normalizadas[alias_norm]

    return None


def _coe_read_csv_from_bytes(contenido):
    intentos = [
        {"sep": None, "encoding": "utf-8-sig"},
        {"sep": ";", "encoding": "utf-8-sig"},
        {"sep": ",", "encoding": "utf-8-sig"},
        {"sep": None, "encoding": "latin1"},
        {"sep": ";", "encoding": "latin1"},
        {"sep": ",", "encoding": "latin1"},
        {"sep": None, "encoding": "cp1252"},
        {"sep": ";", "encoding": "cp1252"},
        {"sep": ",", "encoding": "cp1252"},
    ]

    ultimo_error = None

    for intento in intentos:
        try:
            return pd.read_csv(
                BytesIO(contenido),
                dtype=str,
                sep=intento["sep"],
                engine="python",
                encoding=intento["encoding"],
            )
        except Exception as e:
            ultimo_error = e

    raise ultimo_error


def _leer_archivo_coe_sap_funcional(file):
    filename = (file.filename or "").lower().strip()

    contenido = file.read()

    if not contenido:
        raise ValueError("El archivo está vacío.")

    if filename.endswith(".csv"):
        df = _coe_read_csv_from_bytes(contenido)
    else:
        try:
            df = pd.read_excel(
                BytesIO(contenido),
                dtype=str,
                engine="openpyxl"
            )
        except Exception:
            df = pd.read_excel(
                BytesIO(contenido),
                dtype=str
            )

    df = df.where(pd.notnull(df), None)

    columnas_originales = list(df.columns)

    columnas_normalizadas = {
        _coe_norm_col(c): c
        for c in columnas_originales
    }

    app.logger.info(
        "COE SAP Funcional - columnas originales recibidas: %s",
        columnas_originales
    )

    app.logger.info(
        "COE SAP Funcional - columnas normalizadas recibidas: %s",
        list(columnas_normalizadas.keys())
    )

    columnas_encontradas = {}

    for campo, aliases in COE_SAP_FUNCIONAL_ALIASES.items():
        col = _coe_find_column(columnas_normalizadas, aliases)

        if col:
            columnas_encontradas[campo] = col

    obligatorias = {
        "numero": "NUMERO",
    }

    faltantes = []

    for campo, nombre_visible in obligatorias.items():
        if campo not in columnas_encontradas:
            faltantes.append(nombre_visible)

    if faltantes:
        raise ValueError(
            "Faltan columnas obligatorias: "
            + ", ".join(faltantes)
            + ". Columnas recibidas normalizadas: "
            + ", ".join(list(columnas_normalizadas.keys()))
        )

    registros = []

    for _, row in df.iterrows():
        obj = {}

        for campo, col_original in columnas_encontradas.items():
            raw = row.get(col_original)

            if campo in ("fecha_entrega", "fecha_resolucion", "fecha_cierre"):
                obj[campo] = _coe_parse_datetime(raw)

            elif campo in ("incumplimiento_sla", "alerta"):
                obj[campo] = _coe_parse_bool(raw)

            else:
                obj[campo] = _coe_parse_str(raw)

        if not obj.get("numero"):
            continue

        registros.append(obj)

    return registros


def coe_sap_funcional_to_dict(r):
    return {
        "id": r.id,
        "numero": r.numero,
        "idInteraccion": r.id_interaccion,
        "compania": r.compania,
        "fechaEntrega": _coe_format_datetime(r.fecha_entrega),
        "fechaResolucion": _coe_format_datetime(r.fecha_resolucion),
        "fechaCierre": _coe_format_datetime(r.fecha_cierre),
        "estado": r.estado,
        "titulo": r.titulo,
        "asignadoA": r.asignado_a,
        "nombreCompletoContacto": r.nombre_completo_contacto,
        "incumplimientoSla": _coe_format_bool(r.incumplimiento_sla),
        "alerta": _coe_format_bool(r.alerta),
        "estadoAlertaAns": r.estado_alerta_ans,
        "impacto": r.impacto,
        "urgencia": r.urgencia,
        "prioridad": r.prioridad,
        "accionActualizacion": r.accion_actualizacion,
        "canalResolucion": r.canal_resolucion,
        "clrTxtServicio": r.clr_txt_servicio,
        "clrTxtClientType": r.clr_txt_client_type,
        "origenCargue": r.origen_cargue,
        "fechaCargue": _coe_format_datetime(r.fecha_cargue),
        "usuarioCargue": r.usuario_cargue,
    }


@bp.route("/coe-sap-funcional", methods=["GET"])
@permission_required("BASE_REGISTRO_VER")
def listar_coe_sap_funcional():
    try:
        page = max(int(request.args.get("page", 1)), 1)
        page_size = min(max(int(request.args.get("page_size", 50)), 1), 1000)

        q = (request.args.get("q") or "").strip()
        estado = (request.args.get("estado") or "").strip()
        prioridad = (request.args.get("prioridad") or "").strip()
        compania = (request.args.get("compania") or "").strip()
        asignado_a = (request.args.get("asignado_a") or "").strip()
        impacto = (request.args.get("impacto") or "").strip()
        urgencia = (request.args.get("urgencia") or "").strip()
        estado_alerta_ans = (request.args.get("estado_alerta_ans") or "").strip()
        canal_resolucion = (request.args.get("canal_resolucion") or "").strip()
        clr_txt_servicio = (request.args.get("clr_txt_servicio") or "").strip()
        clr_txt_client_type = (request.args.get("clr_txt_client_type") or "").strip()
        fecha_desde = (request.args.get("fecha_desde") or "").strip()
        fecha_hasta = (request.args.get("fecha_hasta") or "").strip()

        qry = BaseRegistroInfoCoeSapFuncional.query

        # Búsqueda general
        if q:
            like = f"%{q}%"

            qry = qry.filter(or_(
                BaseRegistroInfoCoeSapFuncional.numero.ilike(like),
                BaseRegistroInfoCoeSapFuncional.id_interaccion.ilike(like),
                BaseRegistroInfoCoeSapFuncional.compania.ilike(like),
                BaseRegistroInfoCoeSapFuncional.fecha_entrega.cast(db.String).ilike(like),
                BaseRegistroInfoCoeSapFuncional.fecha_resolucion.cast(db.String).ilike(like),
                BaseRegistroInfoCoeSapFuncional.fecha_cierre.cast(db.String).ilike(like),
                BaseRegistroInfoCoeSapFuncional.estado.ilike(like),
                BaseRegistroInfoCoeSapFuncional.titulo.ilike(like),
                BaseRegistroInfoCoeSapFuncional.asignado_a.ilike(like),
                BaseRegistroInfoCoeSapFuncional.nombre_completo_contacto.ilike(like),
                BaseRegistroInfoCoeSapFuncional.estado_alerta_ans.ilike(like),
                BaseRegistroInfoCoeSapFuncional.impacto.ilike(like),
                BaseRegistroInfoCoeSapFuncional.urgencia.ilike(like),
                BaseRegistroInfoCoeSapFuncional.prioridad.ilike(like),
                BaseRegistroInfoCoeSapFuncional.accion_actualizacion.ilike(like),
                BaseRegistroInfoCoeSapFuncional.canal_resolucion.ilike(like),
                BaseRegistroInfoCoeSapFuncional.clr_txt_servicio.ilike(like),
                BaseRegistroInfoCoeSapFuncional.clr_txt_client_type.ilike(like),
            ))

        # Filtros específicos
        if estado:
            qry = qry.filter(
                BaseRegistroInfoCoeSapFuncional.estado.ilike(f"%{estado}%")
            )

        if prioridad:
            qry = qry.filter(
                BaseRegistroInfoCoeSapFuncional.prioridad.ilike(f"%{prioridad}%")
            )

        if compania:
            qry = qry.filter(
                BaseRegistroInfoCoeSapFuncional.compania.ilike(f"%{compania}%")
            )

        if asignado_a:
            qry = qry.filter(
                BaseRegistroInfoCoeSapFuncional.asignado_a.ilike(f"%{asignado_a}%")
            )

        if impacto:
            qry = qry.filter(
                BaseRegistroInfoCoeSapFuncional.impacto.ilike(f"%{impacto}%")
            )

        if urgencia:
            qry = qry.filter(
                BaseRegistroInfoCoeSapFuncional.urgencia.ilike(f"%{urgencia}%")
            )

        if estado_alerta_ans:
            qry = qry.filter(
                BaseRegistroInfoCoeSapFuncional.estado_alerta_ans.ilike(
                    f"%{estado_alerta_ans}%"
                )
            )

        if canal_resolucion:
            qry = qry.filter(
                BaseRegistroInfoCoeSapFuncional.canal_resolucion.ilike(
                    f"%{canal_resolucion}%"
                )
            )

        if clr_txt_servicio:
            qry = qry.filter(
                BaseRegistroInfoCoeSapFuncional.clr_txt_servicio.ilike(
                    f"%{clr_txt_servicio}%"
                )
            )

        if clr_txt_client_type:
            qry = qry.filter(
                BaseRegistroInfoCoeSapFuncional.clr_txt_client_type.ilike(
                    f"%{clr_txt_client_type}%"
                )
            )

        # Filtro por fecha de entrega
        if fecha_desde:
            try:
                desde_dt = datetime.strptime(fecha_desde[:10], "%Y-%m-%d")
                qry = qry.filter(
                    BaseRegistroInfoCoeSapFuncional.fecha_entrega >= desde_dt
                )
            except Exception:
                pass

        if fecha_hasta:
            try:
                hasta_dt = datetime.strptime(fecha_hasta[:10], "%Y-%m-%d") + timedelta(days=1)
                qry = qry.filter(
                    BaseRegistroInfoCoeSapFuncional.fecha_entrega < hasta_dt
                )
            except Exception:
                pass

        total = qry.count()

        rows = (
            qry.order_by(
                BaseRegistroInfoCoeSapFuncional.id.desc()
            )
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )

        return jsonify({
            "data": [coe_sap_funcional_to_dict(r) for r in rows],
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": math.ceil(total / page_size) if page_size else 1,
        }), 200

    except Exception as e:
        app.logger.exception("Error listando COE SAP Funcional")
        return jsonify({
            "mensaje": "Error interno",
            "error": str(e),
            "trace": traceback.format_exc(),
        }), 500


@bp.route("/coe-sap-funcional/import-principal", methods=["POST"])
@permission_required("BASE_REGISTRO_IMPORTAR")
def importar_coe_sap_funcional_principal():
    file = request.files.get("file")

    if not file:
        return jsonify({"mensaje": "Archivo no recibido"}), 400

    try:
        registros = _leer_archivo_coe_sap_funcional(file)

        if not registros:
            return jsonify({
                "mensaje": "El archivo no contiene registros válidos"
            }), 400

        usuario_cargue = ""

        try:
            usuario_cargue = g.current_user.usuario if g.current_user else ""
        except Exception:
            usuario_cargue = ""

        registros_por_numero = {}

        for reg in registros:
            numero = str(reg.get("numero") or "").strip()

            if not numero:
                continue

            reg["numero"] = numero
            registros_por_numero[numero] = reg

        registros_limpios = list(registros_por_numero.values())

        if not registros_limpios:
            return jsonify({
                "mensaje": "El archivo no contiene registros con Número válido"
            }), 400

        try:
            db.session.execute(text("TRUNCATE TABLE base_registro_info_coe_sap_funcional"))
            db.session.commit()
        except Exception:
            db.session.rollback()
            db.session.execute(text("DELETE FROM base_registro_info_coe_sap_funcional"))
            db.session.execute(text("ALTER TABLE base_registro_info_coe_sap_funcional AUTO_INCREMENT = 1"))
            db.session.commit()

        objetos = []

        for reg in registros_limpios:
            reg["origen_cargue"] = "PRINCIPAL"
            reg["usuario_cargue"] = usuario_cargue
            reg["fecha_cargue"] = datetime.utcnow()

            objetos.append(BaseRegistroInfoCoeSapFuncional(**reg))

        db.session.bulk_save_objects(objetos)
        db.session.commit()

        return jsonify({
            "mensaje": "Carga principal COE SAP Funcional realizada correctamente",
            "total_recibidos": len(registros),
            "duplicados_archivo": len(registros) - len(registros_limpios),
            "insertados": len(objetos),
        }), 200

    except ValueError as e:
        db.session.rollback()
        return jsonify({
            "mensaje": "Archivo inválido",
            "error": str(e),
        }), 400

    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error importando carga principal COE SAP Funcional")
        return jsonify({
            "mensaje": "Error importando carga principal",
            "error": str(e),
            "trace": traceback.format_exc(),
        }), 500


@bp.route("/coe-sap-funcional/import-adicional", methods=["POST"])
@permission_required("BASE_REGISTRO_IMPORTAR")
def importar_coe_sap_funcional_adicional():
    file = request.files.get("file")

    if not file:
        return jsonify({"mensaje": "Archivo no recibido"}), 400

    try:
        registros = _leer_archivo_coe_sap_funcional(file)

        if not registros:
            return jsonify({
                "mensaje": "El archivo no contiene registros válidos"
            }), 400

        usuario_cargue = ""

        try:
            usuario_cargue = g.current_user.usuario if g.current_user else ""
        except Exception:
            usuario_cargue = ""

        registros_por_numero = {}

        for reg in registros:
            numero = str(reg.get("numero") or "").strip()

            if not numero:
                continue

            reg["numero"] = numero
            registros_por_numero[numero] = reg

        registros_limpios = list(registros_por_numero.values())

        if not registros_limpios:
            return jsonify({
                "mensaje": "El archivo no contiene registros con Número válido"
            }), 400

        insertados = 0
        actualizados = 0

        for reg in registros_limpios:
            numero = reg.get("numero")

            existente = BaseRegistroInfoCoeSapFuncional.query.filter_by(
                numero=numero
            ).first()

            if existente:
                for k, v in reg.items():
                    if hasattr(existente, k):
                        setattr(existente, k, v)

                existente.origen_cargue = "ADICIONAL"
                existente.usuario_cargue = usuario_cargue
                existente.fecha_cargue = datetime.utcnow()

                actualizados += 1

            else:
                reg["origen_cargue"] = "ADICIONAL"
                reg["usuario_cargue"] = usuario_cargue
                reg["fecha_cargue"] = datetime.utcnow()

                db.session.add(BaseRegistroInfoCoeSapFuncional(**reg))
                insertados += 1

        db.session.commit()

        return jsonify({
            "mensaje": "Carga adicional COE SAP Funcional procesada correctamente",
            "total_recibidos": len(registros),
            "duplicados_archivo": len(registros) - len(registros_limpios),
            "insertados": insertados,
            "actualizados": actualizados,
        }), 200

    except ValueError as e:
        db.session.rollback()
        return jsonify({
            "mensaje": "Archivo inválido",
            "error": str(e),
        }), 400

    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error importando carga adicional COE SAP Funcional")
        return jsonify({
            "mensaje": "Error importando carga adicional",
            "error": str(e),
            "trace": traceback.format_exc(),
        }), 500


@bp.route("/coe-sap-funcional/filters", methods=["GET"])
@permission_required("BASE_REGISTRO_VER")
def filtros_coe_sap_funcional():
    try:
        base = BaseRegistroInfoCoeSapFuncional.query

        def distinct_col(col):
            rows = (
                base.with_entities(col)
                .filter(col.isnot(None))
                .filter(func.trim(col) != "")
                .distinct()
                .order_by(col.asc())
                .all()
            )

            return [r[0] for r in rows if r[0]]

        return jsonify({
            "estado": distinct_col(BaseRegistroInfoCoeSapFuncional.estado),
            "prioridad": distinct_col(BaseRegistroInfoCoeSapFuncional.prioridad),
            "asignado_a": distinct_col(BaseRegistroInfoCoeSapFuncional.asignado_a),
            "compania": distinct_col(BaseRegistroInfoCoeSapFuncional.compania),
            "impacto": distinct_col(BaseRegistroInfoCoeSapFuncional.impacto),
            "urgencia": distinct_col(BaseRegistroInfoCoeSapFuncional.urgencia),
            "estado_alerta_ans": distinct_col(BaseRegistroInfoCoeSapFuncional.estado_alerta_ans),
            "canal_resolucion": distinct_col(BaseRegistroInfoCoeSapFuncional.canal_resolucion),
            "clr_txt_servicio": distinct_col(BaseRegistroInfoCoeSapFuncional.clr_txt_servicio),
            "clr_txt_client_type": distinct_col(BaseRegistroInfoCoeSapFuncional.clr_txt_client_type),
        }), 200

    except Exception as e:
        app.logger.exception("Error generando filtros COE SAP Funcional")
        return jsonify({
            "mensaje": "Error interno",
            "error": str(e),
            "trace": traceback.format_exc(),
        }), 500

# ============================================================
# CALIFICACION COE SAP FUNCIONAL
# ============================================================

CALIFICACION_FUNCIONALES_ESTIMADAS = [
    "fi", "mm", "sd", "co", "ps", "slcm", "crm", "crm2",
    "pca", "fm", "pp", "pm", "hcm", "ssff", "fiori", "wf"
]

CALIFICACION_EJECUTADAS = [
    "fi", "mm", "sd", "co", "ps", "pca", "fm", "hcm",
    "ssff", "fiori", "wf", "abap", "basis"
]


def _calificacion_decimal(value):
    try:
        if value is None or value == "":
            return 0
        return float(value)
    except Exception:
        return 0


def _calificacion_fecha(value):
    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    try:
        parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
        if pd.isna(parsed):
            return None
        return parsed.to_pydatetime()
    except Exception:
        return None


def _calificacion_fecha_str(value):
    if value is None:
        return None

    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return str(value)

    return str(value)


def _calificacion_diff_dias(fecha_fin, fecha_ini):
    if not fecha_fin or not fecha_ini:
        return None

    try:
        return round((fecha_fin - fecha_ini).total_seconds() / 86400, 2)
    except Exception:
        return None


def _calificacion_networkdays(fecha_inicio, fecha_fin):
    if not fecha_inicio or not fecha_fin:
        return None

    try:
        inicio = fecha_inicio.date()
        fin = fecha_fin.date()

        if fin < inicio:
            return 0

        dias = 0
        actual = inicio

        while actual <= fin:
            if actual.weekday() < 5:
                dias += 1
            actual += timedelta(days=1)

        return dias
    except Exception:
        return None


def _calificacion_estado_consolidado(estado):
    s = str(estado or "").strip().upper()

    if not s:
        return ""

    if s in ("CERRADO", "CERRADA", "RESUELTO", "RESUELTA", "SOLUCIONADO", "SOLUCIONADA"):
        return "CERRADO"

    if s in ("CANCELADO", "CANCELADA", "ANULADO", "ANULADA"):
        return "CANCELADO"

    if s in ("PENDIENTE", "ABIERTO", "ABIERTA", "ASIGNADO", "ASIGNADA", "EN PROCESO", "EN CURSO"):
        return "ABIERTO"

    if "ESPERA" in s:
        return "EN ESPERA"

    return estado


def _calificacion_responsable_estado(estado):
    consolidado = _calificacion_estado_consolidado(estado)

    if consolidado in ("CERRADO", "CANCELADO"):
        return "CLARO"

    if consolidado in ("ABIERTO", "EN ESPERA"):
        return "CONSULTOR"

    return ""


def _calificacion_recalcular(campos):
    fecha_asignacion = campos.get("fecha_asignacion")
    fecha_respuesta = campos.get("fecha_respuesta")
    fecha_resolucion = campos.get("fecha_resolucion")
    fecha_finalizacion_cierre = campos.get("fecha_finalizacion_cierre")
    fecha_estimacion = campos.get("fecha_estimacion")
    fecha_aprobacion_estimacion = campos.get("fecha_aprobacion_estimacion")

    if campos.get("numero"):
        campos["sistema"] = str(campos.get("numero") or "")[:2]

    if fecha_asignacion:
        campos["dia_creacion"] = fecha_asignacion.day
        campos["mes_creacion"] = fecha_asignacion.month
        campos["anio_creacion"] = fecha_asignacion.year
    else:
        campos["dia_creacion"] = None
        campos["mes_creacion"] = None
        campos["anio_creacion"] = None

    if fecha_finalizacion_cierre:
        campos["dia_cierre"] = fecha_finalizacion_cierre.day
        campos["mes_cierre"] = fecha_finalizacion_cierre.month
        campos["anio_cierre"] = fecha_finalizacion_cierre.year
    else:
        campos["dia_cierre"] = None
        campos["mes_cierre"] = None
        campos["anio_cierre"] = None

    campos["tiempo_respuesta"] = _calificacion_diff_dias(fecha_respuesta, fecha_asignacion)
    campos["tiempo_resolucion"] = _calificacion_diff_dias(fecha_resolucion, fecha_asignacion)
    campos["tiempo_finalizacion_cierre"] = _calificacion_diff_dias(fecha_finalizacion_cierre, fecha_asignacion)

    fecha_fin_estimacion = fecha_estimacion or datetime.utcnow()

    if fecha_asignacion:
        campos["dias_entrega_estimacion"] = _calificacion_networkdays(
            fecha_asignacion + timedelta(days=1),
            fecha_fin_estimacion
        )
    else:
        campos["dias_entrega_estimacion"] = None

    if fecha_estimacion:
        campos["mes_estimacion"] = fecha_estimacion.month
        campos["anio_estimacion"] = fecha_estimacion.year
    else:
        campos["mes_estimacion"] = None
        campos["anio_estimacion"] = None

    if fecha_aprobacion_estimacion:
        campos["mes_aprobado_estimacion"] = fecha_aprobacion_estimacion.month
        campos["anio_aprobado_estimacion"] = fecha_aprobacion_estimacion.year
    else:
        campos["mes_aprobado_estimacion"] = None
        campos["anio_aprobado_estimacion"] = None

    total_funcionales = 0

    for modulo in CALIFICACION_FUNCIONALES_ESTIMADAS:
        total_funcionales += _calificacion_decimal(campos.get(f"horas_estimadas_{modulo}"))

    campos["total_horas_funcionales"] = total_funcionales

    total_estimadas = (
        total_funcionales
        + _calificacion_decimal(campos.get("horas_estimadas_abap"))
        + _calificacion_decimal(campos.get("horas_estimadas_basis"))
    )

    campos["total_horas_estimadas"] = total_estimadas

    campos["total_horas_estimadas2"] = (
        total_estimadas
        + _calificacion_decimal(campos.get("horas_estimadas_pmo"))
    )

    estado = campos.get("estado")
    campos["estado_consolidado"] = _calificacion_estado_consolidado(estado)
    campos["responsable_estado"] = _calificacion_responsable_estado(estado)

    if not campos.get("estado_herramienta_gestion"):
        campos["estado_herramienta_gestion"] = estado

    return campos


def _calificacion_to_dict(r):
    return {
        "id": r.id,
        "baseRegistroId": r.base_registro_id,

        "numero": r.numero,
        "sistema": r.sistema,
        "casoSm": r.caso_sm,

        "documentacion": r.documentacion,
        "casoTransporte": r.caso_transporte,
        "controlHoras": r.control_horas,
        "errorSap": r.error_sap,
        "notaOssSap": r.nota_oss_sap,

        "tipoContrato": r.tipo_contrato,
        "sociedad": r.sociedad,
        "asunto": r.asunto,
        "observaciones": r.observaciones,
        "nombreSolicitante": r.nombre_solicitante,

        "impacto": r.impacto,
        "urgencia": r.urgencia,
        "prioridad": r.prioridad,

        "tipoSolicitud": r.tipo_solicitud,
        "modulo": r.modulo,
        "categoria": r.categoria,
        "subcategoria": r.subcategoria,
        "articulo": r.articulo,

        "estado": r.estado,
        "estadoHerramientaGestion": r.estado_herramienta_gestion,
        "responsableEstado": r.responsable_estado,
        "estadoConsolidado": r.estado_consolidado,

        "asignadoA": r.asignado_a,
        "apoyo1": r.apoyo_1,
        "apoyo2": r.apoyo_2,
        "apoyo3": r.apoyo_3,

        "requiereAbap": r.requiere_abap,
        "asignacionAbap": r.asignacion_abap,

        "fechaAsignacion": _calificacion_fecha_str(r.fecha_asignacion),
        "diaCreacion": r.dia_creacion,
        "mesCreacion": r.mes_creacion,
        "anioCreacion": r.anio_creacion,

        "horaUltimaActualizacion": _calificacion_fecha_str(r.hora_ultima_actualizacion),
        "fechaRespuesta": _calificacion_fecha_str(r.fecha_respuesta),
        "fechaResolucion": _calificacion_fecha_str(r.fecha_resolucion),
        "fechaFinalizacionCierre": _calificacion_fecha_str(r.fecha_finalizacion_cierre),

        "diaCierre": r.dia_cierre,
        "mesCierre": r.mes_cierre,
        "anioCierre": r.anio_cierre,

        "tiempoRespuesta": float(r.tiempo_respuesta or 0) if r.tiempo_respuesta is not None else None,
        "tiempoResolucion": float(r.tiempo_resolucion or 0) if r.tiempo_resolucion is not None else None,
        "tiempoFinalizacionCierre": float(r.tiempo_finalizacion_cierre or 0) if r.tiempo_finalizacion_cierre is not None else None,

        "fechaCompromiso": _calificacion_fecha_str(r.fecha_compromiso),
        "liderClaro": r.lider_claro,
        "tipoIngreso": r.tipo_ingreso,

        "fechaEstimacion": _calificacion_fecha_str(r.fecha_estimacion),
        "diasEntregaEstimacion": r.dias_entrega_estimacion,
        "mesEstimacion": r.mes_estimacion,
        "anioEstimacion": r.anio_estimacion,

        "fechaAprobacionEstimacion": _calificacion_fecha_str(r.fecha_aprobacion_estimacion),
        "mesAprobadoEstimacion": r.mes_aprobado_estimacion,
        "anioAprobadoEstimacion": r.anio_aprobado_estimacion,
        "estadoEstimacion": r.estado_estimacion,

        "horasEstimadasFi": float(r.horas_estimadas_fi or 0),
        "horasEstimadasMm": float(r.horas_estimadas_mm or 0),
        "horasEstimadasSd": float(r.horas_estimadas_sd or 0),
        "horasEstimadasCo": float(r.horas_estimadas_co or 0),
        "horasEstimadasPs": float(r.horas_estimadas_ps or 0),
        "horasEstimadasSlcm": float(r.horas_estimadas_slcm or 0),
        "horasEstimadasCrm": float(r.horas_estimadas_crm or 0),
        "horasEstimadasCrm2": float(r.horas_estimadas_crm2 or 0),
        "horasEstimadasPca": float(r.horas_estimadas_pca or 0),
        "horasEstimadasFm": float(r.horas_estimadas_fm or 0),
        "horasEstimadasPp": float(r.horas_estimadas_pp or 0),
        "horasEstimadasPm": float(r.horas_estimadas_pm or 0),
        "horasEstimadasHcm": float(r.horas_estimadas_hcm or 0),
        "horasEstimadasSsff": float(r.horas_estimadas_ssff or 0),
        "horasEstimadasFiori": float(r.horas_estimadas_fiori or 0),
        "horasEstimadasWf": float(r.horas_estimadas_wf or 0),

        "horasEstimadasAbap": float(r.horas_estimadas_abap or 0),
        "horasEstimadasBasis": float(r.horas_estimadas_basis or 0),
        "horasEstimadasPmo": float(r.horas_estimadas_pmo or 0),

        "totalHorasFuncionales": float(r.total_horas_funcionales or 0),
        "totalHorasEstimadas": float(r.total_horas_estimadas or 0),
        "totalHorasEstimadas2": float(r.total_horas_estimadas2 or 0),

        "horasEjecutadasFi": float(r.horas_ejecutadas_fi or 0),
        "horasEjecutadasMm": float(r.horas_ejecutadas_mm or 0),
        "horasEjecutadasSd": float(r.horas_ejecutadas_sd or 0),
        "horasEjecutadasCo": float(r.horas_ejecutadas_co or 0),
        "horasEjecutadasPs": float(r.horas_ejecutadas_ps or 0),
        "horasEjecutadasPca": float(r.horas_ejecutadas_pca or 0),
        "horasEjecutadasFm": float(r.horas_ejecutadas_fm or 0),
        "horasEjecutadasHcm": float(r.horas_ejecutadas_hcm or 0),
        "horasEjecutadasSsff": float(r.horas_ejecutadas_ssff or 0),
        "horasEjecutadasFiori": float(r.horas_ejecutadas_fiori or 0),
        "horasEjecutadasWf": float(r.horas_ejecutadas_wf or 0),
        "horasEjecutadasAbap": float(r.horas_ejecutadas_abap or 0),
        "horasEjecutadasBasis": float(r.horas_ejecutadas_basis or 0),

        "horasGarantia": float(r.horas_garantia or 0),
        "horasProyectoAbap": float(r.horas_proyecto_abap or 0),

        "creadoPor": r.creado_por,
        "actualizadoPor": r.actualizado_por,
        "createdAt": _calificacion_fecha_str(r.created_at),
        "updatedAt": _calificacion_fecha_str(r.updated_at),
    }


def _calificacion_hora_to_dict(h):
    return {
        "id": h.id,
        "calificacionId": h.calificacion_id,
        "numero": h.numero,
        "tipo": h.tipo,
        "modulo": h.modulo,
        "horas": float(h.horas or 0),
        "observacion": h.observacion,
        "origen": getattr(h, "origen", None),
        "excelFila": getattr(h, "excel_fila", None),
        "usuarioRegistro": h.usuario_registro,
        "createdAt": _calificacion_fecha_str(h.created_at),
    }


def _calificacion_usuario_actual():
    try:
        return g.current_user.usuario if g.current_user else ""
    except Exception:
        return ""


def _calificacion_campos_desde_base(base):
    campos = {
        "base_registro_id": base.id,

        "numero": base.numero,
        "sistema": str(base.numero or "")[:2],
        "caso_sm": base.id_interaccion,

        "sociedad": base.compania,
        "asunto": base.titulo,
        "observaciones": base.accion_actualizacion,
        "nombre_solicitante": base.nombre_completo_contacto,

        "impacto": base.impacto,
        "urgencia": base.urgencia,
        "prioridad": base.prioridad,

        "tipo_solicitud": base.clr_txt_client_type,
        "articulo": base.clr_txt_servicio,

        "estado": base.estado,
        "estado_herramienta_gestion": base.estado,
        "asignado_a": base.asignado_a,

        "fecha_asignacion": base.fecha_entrega,
        "hora_ultima_actualizacion": base.fecha_cargue,
        "fecha_resolucion": base.fecha_resolucion,
        "fecha_finalizacion_cierre": base.fecha_cierre,

        "tipo_contrato": "BOLSA DE HORAS",
    }

    return _calificacion_recalcular(campos)

# ============================================================
# IMPORTACION EXCEL HISTORICO - CALIFICACION COE SAP FUNCIONAL
# ============================================================

CALIFICACION_EXCEL_ALIASES = {
    "numero": [
        "ID",
        "NUMERO",
        "NÚMERO",
    ],
    "caso_sm": [
        "N CASO SM",
        "N° CASO SM",
        "NO CASO SM",
        "NRO CASO SM",
        "CASO SM",
    ],
    "documentacion": [
        "DOCUMENTACION",
        "DOCUMENTACIÓN",
    ],
    "caso_transporte": [
        "CASO TRANSPORTE",
    ],
    "control_horas": [
        "CONTROL HORAS",
    ],
    "error_sap": [
        "N ERROR SAP",
        "N° ERROR SAP",
        "NO ERROR SAP",
        "NUMERO ERROR SAP",
    ],
    "nota_oss_sap": [
        "N NOTA OSS SAP",
        "N° NOTA OSS SAP",
        "NO NOTA OSS SAP",
        "NUMERO NOTA OSS SAP",
    ],
    "tipo_contrato": [
        "TIPO CONTRATO",
    ],
    "sociedad": [
        "SOCIEDAD",
        "COMPAÑIA",
        "COMPAÑÍA",
        "COMPANIA",
    ],
    "asunto": [
        "ASUNTO",
        "TITULO",
        "TÍTULO",
    ],
    "observaciones": [
        "OBSERVACIONES",
    ],
    "nombre_solicitante": [
        "NOMBRE DEL SOLICITANTE",
        "SOLICITANTE",
        "NOMBRE COMPLETO CONTACTO",
    ],
    "impacto": [
        "IMPACTO",
    ],
    "urgencia": [
        "URGENCIA",
    ],
    "prioridad": [
        "PRIORIDAD",
    ],
    "tipo_solicitud": [
        "TIPO DE SOLICITUD",
        "TIPO SOLICITUD",
    ],
    "modulo": [
        "MODULO",
        "MÓDULO",
    ],
    "categoria": [
        "CATEGORIA",
        "CATEGORÍA",
    ],
    "subcategoria": [
        "SUBCATEGORIA",
        "SUBCATEGORÍA",
    ],
    "articulo": [
        "ARTICULO",
        "ARTÍCULO",
    ],
    "estado": [
        "ESTADO",
    ],
    "estado_herramienta_gestion": [
        "ESTADO CASOS EN HERRAMIENTAS DE GESTION",
        "ESTADO CASOS EN HERRAMIENTAS DE GESTIÓN",
    ],
    "responsable_estado": [
        "RESPONSABLE ESTADO",
    ],
    "estado_consolidado": [
        "ESTADO CONSOLIDADO",
    ],
    "asignado_a": [
        "ASIGNADO A",
    ],
    "apoyo_1": [
        "APOYO 1",
    ],
    "apoyo_2": [
        "APOYO 2",
    ],
    "apoyo_3": [
        "APOYO 3",
    ],
    "requiere_abap": [
        "REQUIERE ABAP",
    ],
    "asignacion_abap": [
        "ASIGNACION ABAP",
        "ASIGNACIÓN ABAP",
    ],
    "fecha_asignacion": [
        "FECHA DE ASIGNACION",
        "FECHA DE ASIGNACIÓN",
    ],
    "hora_ultima_actualizacion": [
        "HORA DE LA ULTIMA ACTUALIZACION",
        "HORA DE LA ÚLTIMA ACTUALIZACIÓN",
    ],
    "fecha_respuesta": [
        "FECHA DE RESPUESTA",
    ],
    "fecha_resolucion": [
        "FECHA DE RESOLUCION",
        "FECHA DE RESOLUCIÓN",
    ],
    "fecha_finalizacion_cierre": [
        "FECHA DE FINALIZACION CIERRE",
        "FECHA DE FINALIZACIÓN CIERRE",
        "FECHA DE FINALIZACION / CIERRE",
        "FECHA DE FINALIZACIÓN / CIERRE",
    ],
    "fecha_compromiso": [
        "FECHA COMPROMISO",
    ],
    "lider_claro": [
        "LIDER CLARO",
        "LÍDER CLARO",
    ],
    "tipo_ingreso": [
        "TIPO DE INGRESO",
        "TIPO INGRESO",
    ],
    "fecha_estimacion": [
        "FECHA ESTIMACION",
        "FECHA ESTIMACIÓN",
    ],
    "fecha_aprobacion_estimacion": [
        "FECHA APROBACION ESTIMACION",
        "FECHA APROBACIÓN ESTIMACIÓN",
    ],
    "estado_estimacion": [
        "ESTADO ESTIMACION",
        "ESTADO ESTIMACIÓN",
    ],
}


CALIFICACION_EXCEL_HORAS = [
    ("ESTIMADA", "FI", "horas_estimadas_fi", ["HORAS ESTIMADAS FI"]),
    ("ESTIMADA", "MM", "horas_estimadas_mm", ["HORAS ESTIMADAS MM"]),
    ("ESTIMADA", "SD", "horas_estimadas_sd", ["HORAS ESTIMADAS SD"]),
    ("ESTIMADA", "CO", "horas_estimadas_co", ["HORAS ESTIMADAS CO"]),
    ("ESTIMADA", "PS", "horas_estimadas_ps", ["HORAS ESTIMADAS PS"]),
    ("ESTIMADA", "SLCM", "horas_estimadas_slcm", ["HORAS ESTIMADAS SLCM"]),
    ("ESTIMADA", "CRM", "horas_estimadas_crm", ["HORAS ESTIMADAS CRM"]),
    ("ESTIMADA", "CRM2", "horas_estimadas_crm2", ["HORAS ESTIMADAS CRM2"]),
    ("ESTIMADA", "PCA", "horas_estimadas_pca", ["HORAS ESTIMADAS PCA", "HORAS ESTIMADAS PCA OPEX RENTABILIDAD"]),
    ("ESTIMADA", "FM", "horas_estimadas_fm", ["HORAS ESTIMADAS FM"]),
    ("ESTIMADA", "PP", "horas_estimadas_pp", ["HORAS ESTIMADAS PP"]),
    ("ESTIMADA", "PM", "horas_estimadas_pm", ["HORAS ESTIMADAS PM"]),
    ("ESTIMADA", "HCM", "horas_estimadas_hcm", ["HORAS ESTIMADAS HCM"]),
    ("ESTIMADA", "SSFF", "horas_estimadas_ssff", ["HORAS ESTIMADAS SSFF"]),
    ("ESTIMADA", "FIORI", "horas_estimadas_fiori", ["HORAS ESTIMADAS FIORI"]),
    ("ESTIMADA", "WF", "horas_estimadas_wf", ["HORAS ESTIMADAS WF"]),
    ("ESTIMADA", "ABAP", "horas_estimadas_abap", ["HORAS ESTIMADAS ABAP"]),
    ("ESTIMADA", "BASIS", "horas_estimadas_basis", ["HORAS ESTIMADAS BASIS"]),
    ("ESTIMADA", "PMO", "horas_estimadas_pmo", ["HORAS ESTIMADAS PMO"]),

    ("EJECUTADA", "FI", "horas_ejecutadas_fi", ["HORAS EJECUTADAS FI"]),
    ("EJECUTADA", "MM", "horas_ejecutadas_mm", ["HORAS EJECUTADAS MM"]),
    ("EJECUTADA", "SD", "horas_ejecutadas_sd", ["HORAS EJECUTADAS SD"]),
    ("EJECUTADA", "CO", "horas_ejecutadas_co", ["HORAS EJECUTADAS CO"]),
    ("EJECUTADA", "PS", "horas_ejecutadas_ps", ["HORAS EJECUTADAS PS"]),
    ("EJECUTADA", "PCA", "horas_ejecutadas_pca", ["HORAS EJECUTADAS PCA", "HORAS EJECUTADAS PCA OPEX RENTABILIDAD"]),
    ("EJECUTADA", "FM", "horas_ejecutadas_fm", ["HORAS EJECUTADAS FM"]),
    ("EJECUTADA", "HCM", "horas_ejecutadas_hcm", ["HORAS EJECUTADAS HCM"]),
    ("EJECUTADA", "SSFF", "horas_ejecutadas_ssff", ["HORAS EJECUTADAS SSFF"]),
    ("EJECUTADA", "FIORI", "horas_ejecutadas_fiori", ["HORAS EJECUTADAS FIORI"]),
    ("EJECUTADA", "WF", "horas_ejecutadas_wf", ["HORAS EJECUTADAS WF"]),
    ("EJECUTADA", "ABAP", "horas_ejecutadas_abap", ["HORAS EJECUTADAS ABAP"]),
    ("EJECUTADA", "BASIS", "horas_ejecutadas_basis", ["HORAS EJECUTADAS BASIS"]),

    ("GARANTIA", "GARANTIA", "horas_garantia", ["HORAS GARANTIA", "HORAS GARANTÍA"]),
    ("PROYECTO_ABAP", "ABAP", "horas_proyecto_abap", ["HORAS PROYECTO ABAP"]),
]


CALIFICACION_EXCEL_FECHAS = {
    "fecha_asignacion",
    "hora_ultima_actualizacion",
    "fecha_respuesta",
    "fecha_resolucion",
    "fecha_finalizacion_cierre",
    "fecha_compromiso",
    "fecha_estimacion",
    "fecha_aprobacion_estimacion",
}


CALIFICACION_EXCEL_MANUALES = [
    "caso_sm",
    "documentacion",
    "caso_transporte",
    "control_horas",
    "error_sap",
    "nota_oss_sap",
    "tipo_contrato",
    "tipo_solicitud",
    "modulo",
    "categoria",
    "subcategoria",
    "articulo",
    "estado_herramienta_gestion",
    "responsable_estado",
    "estado_consolidado",
    "apoyo_1",
    "apoyo_2",
    "apoyo_3",
    "requiere_abap",
    "asignacion_abap",
    "fecha_respuesta",
    "fecha_compromiso",
    "lider_claro",
    "tipo_ingreso",
    "fecha_estimacion",
    "fecha_aprobacion_estimacion",
    "estado_estimacion",
]


CALIFICACION_EXCEL_SOLO_EXCEL = [
    "sociedad",
    "asunto",
    "observaciones",
    "nombre_solicitante",
    "impacto",
    "urgencia",
    "prioridad",
    "estado",
    "asignado_a",
    "fecha_asignacion",
    "hora_ultima_actualizacion",
    "fecha_resolucion",
    "fecha_finalizacion_cierre",
]


def _calificacion_norm_col(value):
    s = str(value or "")

    s = s.replace("\ufeff", "")
    s = s.replace("\u200b", "")
    s = s.replace("\u200c", "")
    s = s.replace("\u200d", "")
    s = s.replace("\u00A0", " ")

    s = s.strip().upper()

    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")

    s = s.replace("º", "")
    s = s.replace("°", "")
    s = s.replace(".", " ")
    s = s.replace(":", " ")
    s = s.replace("-", " ")
    s = s.replace("_", " ")
    s = s.replace("/", " ")
    s = s.replace("(", " ")
    s = s.replace(")", " ")

    s = re.sub(r"\s+", " ", s).strip()

    return s


def _calificacion_get_excel(row, aliases):
    for alias in aliases:
        key = _calificacion_norm_col(alias)

        if key in row:
            return row.get(key)

    return None


def _calificacion_value_present(value):
    if value is None:
        return False

    s = str(value).strip()

    if s == "" or s.lower() in ("nan", "none", "null"):
        return False

    return True


def _calificacion_excel_str(value):
    if not _calificacion_value_present(value):
        return None

    return str(value).replace("\u00A0", " ").strip()


def _calificacion_excel_fecha(value):
    if not _calificacion_value_present(value):
        return None

    return _calificacion_fecha(value)


def _calificacion_excel_horas(value):
    if not _calificacion_value_present(value):
        return 0

    try:
        # Excel puede traer duración como timedelta.
        if isinstance(value, timedelta):
            return round(value.total_seconds() / 3600, 2)

        # Excel puede traer hora como time.
        if hasattr(value, "hour") and hasattr(value, "minute") and not isinstance(value, datetime):
            return round(value.hour + (value.minute / 60) + (value.second / 3600), 2)

        # Si viene datetime real, normalmente no debería ser una hora.
        # Solo se toma la parte horaria si el año es 1899/1900.
        if isinstance(value, datetime):
            if value.year in (1899, 1900):
                return round(value.hour + (value.minute / 60) + (value.second / 3600), 2)
            return 0

        s = str(value).strip()

        if s == "" or s.lower() in ("nan", "none", "null"):
            return 0

        s = s.replace(",", ".")

        # Formato tipo "1 day, 11:00:00"
        m_day = re.search(r"(\d+)\s+day[s]?,\s+(\d{1,2}):(\d{2})(?::(\d{2}))?", s, re.IGNORECASE)
        if m_day:
            dias = int(m_day.group(1))
            horas = int(m_day.group(2))
            minutos = int(m_day.group(3))
            segundos = int(m_day.group(4) or 0)
            return round((dias * 24) + horas + (minutos / 60) + (segundos / 3600), 2)

        # Formato tipo "18:30:00" o "18:30"
        m_hora = re.match(r"^(\d{1,4}):(\d{2})(?::(\d{2}))?$", s)
        if m_hora:
            horas = int(m_hora.group(1))
            minutos = int(m_hora.group(2))
            segundos = int(m_hora.group(3) or 0)
            return round(horas + (minutos / 60) + (segundos / 3600), 2)

        return round(float(s), 2)

    except Exception:
        return 0


def _calificacion_excel_detectar_header(filas):
    for idx, fila in enumerate(filas[:50]):
        normalizados = [_calificacion_norm_col(v) for v in fila if v is not None]

        if "ID" in normalizados and ("SOCIEDAD" in normalizados or "ASUNTO" in normalizados):
            return idx

    raise ValueError("No se encontró la fila de encabezados. Debe existir una columna ID y columnas como SOCIEDAD o ASUNTO.")


def _leer_excel_historico_calificacion(file):
    filename = (file.filename or "").lower().strip()
    contenido = file.read()

    if not contenido:
        raise ValueError("El archivo está vacío.")

    if filename.endswith(".csv"):
        df = _coe_read_csv_from_bytes(contenido)
        df = df.where(pd.notnull(df), None)

        columnas = [_calificacion_norm_col(c) for c in list(df.columns)]
        rows = []

        for index, row in df.iterrows():
            item = {
                columnas[i]: row.iloc[i]
                for i in range(len(columnas))
                if columnas[i]
            }

            item["_excel_fila"] = int(index) + 2
            rows.append(item)

        return rows

    wb = load_workbook(
        BytesIO(contenido),
        data_only=True,
        read_only=True
    )

    ws = wb["BASE"] if "BASE" in wb.sheetnames else wb.active

    filas = list(ws.iter_rows(values_only=True))

    if not filas:
        raise ValueError("El archivo no tiene información.")

    header_idx = _calificacion_excel_detectar_header(filas)
    headers_raw = filas[header_idx]

    headers = [_calificacion_norm_col(h) for h in headers_raw]

    rows = []

    for excel_idx, fila in enumerate(filas[header_idx + 1:], start=header_idx + 2):
        item = {}

        tiene_datos = False

        for i, value in enumerate(fila):
            if i >= len(headers):
                continue

            header = headers[i]

            if not header:
                continue

            item[header] = value

            if _calificacion_value_present(value):
                tiene_datos = True

        if not tiene_datos:
            continue

        item["_excel_fila"] = excel_idx
        rows.append(item)

    return rows


def _calificacion_extraer_campos_excel(row):
    campos = {}

    for campo, aliases in CALIFICACION_EXCEL_ALIASES.items():
        value = _calificacion_get_excel(row, aliases)

        if not _calificacion_value_present(value):
            continue

        if campo in CALIFICACION_EXCEL_FECHAS:
            campos[campo] = _calificacion_excel_fecha(value)
        else:
            campos[campo] = _calificacion_excel_str(value)

    return campos


def _calificacion_horas_desde_row_excel(row):
    horas = []

    for tipo, modulo, campo_modelo, aliases in CALIFICACION_EXCEL_HORAS:
        value = _calificacion_get_excel(row, aliases)
        cantidad = _calificacion_excel_horas(value)

        if cantidad and cantidad > 0:
            horas.append({
                "tipo": tipo,
                "modulo": modulo,
                "campo_modelo": campo_modelo,
                "horas": cantidad,
                "excel_fila": row.get("_excel_fila"),
            })

    return horas


def _calificacion_merge_ultimo_valor(rows, campo):
    aliases = CALIFICACION_EXCEL_ALIASES.get(campo, [])

    ultimo = None

    for row in rows:
        value = _calificacion_get_excel(row, aliases)

        if _calificacion_value_present(value):
            ultimo = value

    if not _calificacion_value_present(ultimo):
        return None

    if campo in CALIFICACION_EXCEL_FECHAS:
        return _calificacion_excel_fecha(ultimo)

    return _calificacion_excel_str(ultimo)


def _calificacion_comparar_base_excel(base, campos_excel):
    diferencias = []

    comparaciones = [
        ("sociedad", getattr(base, "compania", None), campos_excel.get("sociedad")),
        ("asunto", getattr(base, "titulo", None), campos_excel.get("asunto")),
        ("observaciones", getattr(base, "accion_actualizacion", None), campos_excel.get("observaciones")),
        ("nombre_solicitante", getattr(base, "nombre_completo_contacto", None), campos_excel.get("nombre_solicitante")),
        ("impacto", getattr(base, "impacto", None), campos_excel.get("impacto")),
        ("urgencia", getattr(base, "urgencia", None), campos_excel.get("urgencia")),
        ("prioridad", getattr(base, "prioridad", None), campos_excel.get("prioridad")),
        ("estado", getattr(base, "estado", None), campos_excel.get("estado")),
        ("asignado_a", getattr(base, "asignado_a", None), campos_excel.get("asignado_a")),
    ]

    for campo, valor_base, valor_excel in comparaciones:
        if not _calificacion_value_present(valor_base) or not _calificacion_value_present(valor_excel):
            continue

        if _calificacion_norm_col(valor_base) != _calificacion_norm_col(valor_excel):
            diferencias.append({
                "campo": campo,
                "base": str(valor_base),
                "excel": str(valor_excel),
            })

    return diferencias

@bp.route("/coe-sap-funcional/calificacion/generar", methods=["POST"])
@permission_required("BASE_REGISTRO_IMPORTAR")
def generar_calificacion_coe_sap_funcional():
    try:
        usuario = _calificacion_usuario_actual()

        bases = BaseRegistroInfoCoeSapFuncional.query.all()

        creados = 0
        actualizados = 0

        for base in bases:
            if not base.numero:
                continue

            campos = _calificacion_campos_desde_base(base)

            existente = CoeSapFuncionalCalificacion.query.filter_by(
                numero=base.numero
            ).first()

            if existente:
                # Se actualizan solo campos automáticos.
                # Los manuales se conservan.
                campos_automaticos = [
                    "base_registro_id",
                    "sistema",
                    "caso_sm",
                    "sociedad",
                    "asunto",
                    "observaciones",
                    "nombre_solicitante",
                    "impacto",
                    "urgencia",
                    "prioridad",
                    "tipo_solicitud",
                    "articulo",
                    "estado",
                    "estado_herramienta_gestion",
                    "responsable_estado",
                    "estado_consolidado",
                    "asignado_a",
                    "fecha_asignacion",
                    "dia_creacion",
                    "mes_creacion",
                    "anio_creacion",
                    "hora_ultima_actualizacion",
                    "fecha_resolucion",
                    "fecha_finalizacion_cierre",
                    "dia_cierre",
                    "mes_cierre",
                    "anio_cierre",
                    "tiempo_resolucion",
                    "tiempo_finalizacion_cierre",
                    "dias_entrega_estimacion",
                    "mes_estimacion",
                    "anio_estimacion",
                    "mes_aprobado_estimacion",
                    "anio_aprobado_estimacion",
                    "total_horas_funcionales",
                    "total_horas_estimadas",
                    "total_horas_estimadas2",
                ]

                for campo in campos_automaticos:
                    if campo in campos and hasattr(existente, campo):
                        setattr(existente, campo, campos[campo])

                existente.actualizado_por = usuario
                existente.updated_at = datetime.utcnow()
                actualizados += 1

            else:
                campos["creado_por"] = usuario
                campos["actualizado_por"] = usuario
                nuevo = CoeSapFuncionalCalificacion(**campos)
                db.session.add(nuevo)
                creados += 1

        db.session.commit()

        return jsonify({
            "mensaje": "Calificación generada correctamente desde la base COE SAP Funcional",
            "base_registros": len(bases),
            "creados": creados,
            "actualizados": actualizados,
        }), 200

    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error generando calificación COE SAP Funcional")
        return jsonify({
            "mensaje": "Error generando calificación",
            "error": str(e),
            "trace": traceback.format_exc(),
        }), 500


@bp.route("/coe-sap-funcional/calificacion", methods=["GET"])
@permission_required("BASE_REGISTRO_VER")
def listar_calificacion_coe_sap_funcional():
    try:
        page = max(int(request.args.get("page", 1)), 1)
        page_size = min(max(int(request.args.get("page_size", 50)), 1), 1000)

        q = (request.args.get("q") or "").strip()
        estado = (request.args.get("estado") or "").strip()
        sociedad = (request.args.get("sociedad") or "").strip()
        asignado_a = (request.args.get("asignado_a") or "").strip()
        sistema = (request.args.get("sistema") or "").strip()
        modulo = (request.args.get("modulo") or "").strip()
        estado_consolidado = (request.args.get("estado_consolidado") or "").strip()

        qry = CoeSapFuncionalCalificacion.query

        if q:
            like = f"%{q}%"
            qry = qry.filter(or_(
                CoeSapFuncionalCalificacion.numero.ilike(like),
                CoeSapFuncionalCalificacion.sistema.ilike(like),
                CoeSapFuncionalCalificacion.caso_sm.ilike(like),
                CoeSapFuncionalCalificacion.sociedad.ilike(like),
                CoeSapFuncionalCalificacion.asunto.ilike(like),
                CoeSapFuncionalCalificacion.observaciones.ilike(like),
                CoeSapFuncionalCalificacion.nombre_solicitante.ilike(like),
                CoeSapFuncionalCalificacion.asignado_a.ilike(like),
                CoeSapFuncionalCalificacion.estado.ilike(like),
                CoeSapFuncionalCalificacion.estado_consolidado.ilike(like),
                CoeSapFuncionalCalificacion.modulo.ilike(like),
                CoeSapFuncionalCalificacion.categoria.ilike(like),
                CoeSapFuncionalCalificacion.subcategoria.ilike(like),
                CoeSapFuncionalCalificacion.articulo.ilike(like),
            ))

        if estado:
            qry = qry.filter(CoeSapFuncionalCalificacion.estado.ilike(f"%{estado}%"))

        if sociedad:
            qry = qry.filter(CoeSapFuncionalCalificacion.sociedad.ilike(f"%{sociedad}%"))

        if asignado_a:
            qry = qry.filter(CoeSapFuncionalCalificacion.asignado_a.ilike(f"%{asignado_a}%"))

        if sistema:
            qry = qry.filter(CoeSapFuncionalCalificacion.sistema.ilike(f"%{sistema}%"))

        if modulo:
            qry = qry.filter(CoeSapFuncionalCalificacion.modulo.ilike(f"%{modulo}%"))

        if estado_consolidado:
            qry = qry.filter(
                CoeSapFuncionalCalificacion.estado_consolidado.ilike(f"%{estado_consolidado}%")
            )

        total = qry.count()

        rows = (
            qry.order_by(CoeSapFuncionalCalificacion.id.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )

        return jsonify({
            "data": [_calificacion_to_dict(r) for r in rows],
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": math.ceil(total / page_size) if page_size else 1,
        }), 200

    except Exception as e:
        app.logger.exception("Error listando calificación COE SAP Funcional")
        return jsonify({
            "mensaje": "Error interno",
            "error": str(e),
            "trace": traceback.format_exc(),
        }), 500


@bp.route("/coe-sap-funcional/calificacion/<int:calificacion_id>", methods=["PATCH"])
@permission_required("BASE_REGISTRO_IMPORTAR")
def actualizar_calificacion_coe_sap_funcional(calificacion_id):
    try:
        row = CoeSapFuncionalCalificacion.query.get(calificacion_id)

        if not row:
            return jsonify({"mensaje": "Registro de calificación no encontrado"}), 404

        data = request.get_json(silent=True) or {}
        usuario = _calificacion_usuario_actual()

        campos_editables = {
            "documentacion": "documentacion",
            "casoTransporte": "caso_transporte",
            "controlHoras": "control_horas",
            "errorSap": "error_sap",
            "notaOssSap": "nota_oss_sap",
            "tipoContrato": "tipo_contrato",
            "tipoSolicitud": "tipo_solicitud",
            "modulo": "modulo",
            "categoria": "categoria",
            "subcategoria": "subcategoria",
            "articulo": "articulo",
            "apoyo1": "apoyo_1",
            "apoyo2": "apoyo_2",
            "apoyo3": "apoyo_3",
            "requiereAbap": "requiere_abap",
            "asignacionAbap": "asignacion_abap",
            "fechaRespuesta": "fecha_respuesta",
            "fechaCompromiso": "fecha_compromiso",
            "liderClaro": "lider_claro",
            "tipoIngreso": "tipo_ingreso",
            "fechaEstimacion": "fecha_estimacion",
            "fechaAprobacionEstimacion": "fecha_aprobacion_estimacion",
            "estadoEstimacion": "estado_estimacion",

            "horasEstimadasFi": "horas_estimadas_fi",
            "horasEstimadasMm": "horas_estimadas_mm",
            "horasEstimadasSd": "horas_estimadas_sd",
            "horasEstimadasCo": "horas_estimadas_co",
            "horasEstimadasPs": "horas_estimadas_ps",
            "horasEstimadasSlcm": "horas_estimadas_slcm",
            "horasEstimadasCrm": "horas_estimadas_crm",
            "horasEstimadasCrm2": "horas_estimadas_crm2",
            "horasEstimadasPca": "horas_estimadas_pca",
            "horasEstimadasFm": "horas_estimadas_fm",
            "horasEstimadasPp": "horas_estimadas_pp",
            "horasEstimadasPm": "horas_estimadas_pm",
            "horasEstimadasHcm": "horas_estimadas_hcm",
            "horasEstimadasSsff": "horas_estimadas_ssff",
            "horasEstimadasFiori": "horas_estimadas_fiori",
            "horasEstimadasWf": "horas_estimadas_wf",
            "horasEstimadasAbap": "horas_estimadas_abap",
            "horasEstimadasBasis": "horas_estimadas_basis",
            "horasEstimadasPmo": "horas_estimadas_pmo",
        }

        campos_fecha = {
            "fechaRespuesta",
            "fechaCompromiso",
            "fechaEstimacion",
            "fechaAprobacionEstimacion",
        }

        campos_decimal = {
            "horasEstimadasFi",
            "horasEstimadasMm",
            "horasEstimadasSd",
            "horasEstimadasCo",
            "horasEstimadasPs",
            "horasEstimadasSlcm",
            "horasEstimadasCrm",
            "horasEstimadasCrm2",
            "horasEstimadasPca",
            "horasEstimadasFm",
            "horasEstimadasPp",
            "horasEstimadasPm",
            "horasEstimadasHcm",
            "horasEstimadasSsff",
            "horasEstimadasFiori",
            "horasEstimadasWf",
            "horasEstimadasAbap",
            "horasEstimadasBasis",
            "horasEstimadasPmo",
        }

        for json_key, model_key in campos_editables.items():
            if json_key not in data:
                continue

            value = data.get(json_key)

            if json_key in campos_fecha:
                value = _calificacion_fecha(value)

            if json_key in campos_decimal:
                value = _calificacion_decimal(value)

            setattr(row, model_key, value)

        campos = {
            c.name: getattr(row, c.name)
            for c in CoeSapFuncionalCalificacion.__table__.columns
        }

        campos = _calificacion_recalcular(campos)

        for k, v in campos.items():
            if hasattr(row, k):
                setattr(row, k, v)

        row.actualizado_por = usuario
        row.updated_at = datetime.utcnow()

        db.session.commit()

        return jsonify({
            "mensaje": "Calificación actualizada correctamente",
            "data": _calificacion_to_dict(row),
        }), 200

    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error actualizando calificación COE SAP Funcional")
        return jsonify({
            "mensaje": "Error actualizando calificación",
            "error": str(e),
            "trace": traceback.format_exc(),
        }), 500


@bp.route("/coe-sap-funcional/calificacion/<int:calificacion_id>/horas", methods=["GET"])
@permission_required("BASE_REGISTRO_VER")
def listar_horas_calificacion_coe_sap_funcional(calificacion_id):
    try:
        row = CoeSapFuncionalCalificacion.query.get(calificacion_id)

        if not row:
            return jsonify({"mensaje": "Registro de calificación no encontrado"}), 404

        horas = (
            CoeSapFuncionalCalificacionHora.query
            .filter_by(calificacion_id=calificacion_id)
            .order_by(CoeSapFuncionalCalificacionHora.id.desc())
            .all()
        )

        return jsonify({
            "data": [_calificacion_hora_to_dict(h) for h in horas],
            "total": len(horas),
        }), 200

    except Exception as e:
        app.logger.exception("Error listando horas calificación COE SAP Funcional")
        return jsonify({
            "mensaje": "Error interno",
            "error": str(e),
            "trace": traceback.format_exc(),
        }), 500


@bp.route("/coe-sap-funcional/calificacion/<int:calificacion_id>/horas", methods=["POST"])
@permission_required("BASE_REGISTRO_IMPORTAR")
def agregar_horas_calificacion_coe_sap_funcional(calificacion_id):
    try:
        row = CoeSapFuncionalCalificacion.query.get(calificacion_id)

        if not row:
            return jsonify({"mensaje": "Registro de calificación no encontrado"}), 404

        data = request.get_json(silent=True) or {}

        tipo = str(data.get("tipo") or "").strip().upper()
        modulo = str(data.get("modulo") or "").strip().upper()
        horas = _calificacion_decimal(data.get("horas"))
        observacion = data.get("observacion")

        if tipo not in ("ESTIMADA", "EJECUTADA", "GARANTIA", "PROYECTO_ABAP"):
            return jsonify({
                "mensaje": "Tipo inválido. Usa ESTIMADA, EJECUTADA, GARANTIA o PROYECTO_ABAP"
            }), 400

        if not modulo:
            return jsonify({"mensaje": "Debes seleccionar un módulo"}), 400

        if horas <= 0:
            return jsonify({"mensaje": "Las horas deben ser mayores a cero"}), 400

        usuario = _calificacion_usuario_actual()

        nueva_hora = CoeSapFuncionalCalificacionHora(
            calificacion_id=row.id,
            numero=row.numero,
            tipo=tipo,
            modulo=modulo,
            horas=horas,
            observacion=observacion,
            usuario_registro=usuario,
            created_at=datetime.utcnow(),
        )

        db.session.add(nueva_hora)

        modulo_key = modulo.lower()

        if tipo == "ESTIMADA":
            campo = f"horas_estimadas_{modulo_key}"
            if hasattr(row, campo):
                setattr(row, campo, _calificacion_decimal(getattr(row, campo)) + horas)

        elif tipo == "EJECUTADA":
            campo = f"horas_ejecutadas_{modulo_key}"
            if hasattr(row, campo):
                setattr(row, campo, _calificacion_decimal(getattr(row, campo)) + horas)

        elif tipo == "GARANTIA":
            row.horas_garantia = _calificacion_decimal(row.horas_garantia) + horas

        elif tipo == "PROYECTO_ABAP":
            row.horas_proyecto_abap = _calificacion_decimal(row.horas_proyecto_abap) + horas

        campos = {
            c.name: getattr(row, c.name)
            for c in CoeSapFuncionalCalificacion.__table__.columns
        }

        campos = _calificacion_recalcular(campos)

        for k, v in campos.items():
            if hasattr(row, k):
                setattr(row, k, v)

        row.actualizado_por = usuario
        row.updated_at = datetime.utcnow()

        db.session.commit()

        return jsonify({
            "mensaje": "Horas agregadas correctamente",
            "hora": _calificacion_hora_to_dict(nueva_hora),
            "data": _calificacion_to_dict(row),
        }), 200

    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error agregando horas calificación COE SAP Funcional")
        return jsonify({
            "mensaje": "Error agregando horas",
            "error": str(e),
            "trace": traceback.format_exc(),
        }), 500
    
@bp.route("/coe-sap-funcional/calificacion/import-excel", methods=["POST"])
@permission_required("BASE_REGISTRO_IMPORTAR")
def importar_excel_historico_calificacion_coe_sap_funcional():
    file = request.files.get("file")

    if not file:
        return jsonify({"mensaje": "Archivo no recibido"}), 400

    try:
        usuario = _calificacion_usuario_actual()

        rows_excel = _leer_excel_historico_calificacion(file)

        if not rows_excel:
            return jsonify({
                "mensaje": "El Excel no contiene registros válidos"
            }), 400

        grupos = {}

        for row in rows_excel:
            numero_raw = _calificacion_get_excel(
                row,
                CALIFICACION_EXCEL_ALIASES["numero"]
            )

            numero = _calificacion_excel_str(numero_raw)

            if not numero:
                continue

            if numero not in grupos:
                grupos[numero] = []

            grupos[numero].append(row)

        if not grupos:
            return jsonify({
                "mensaje": "No se encontraron registros con ID válido en el Excel"
            }), 400

        creados = 0
        actualizados = 0
        creados_solo_excel = 0
        no_encontrados_en_base = 0
        duplicados_excel = 0
        horas_movimientos = 0

        diferencias_detectadas = []

        for numero, filas_caso in grupos.items():
            if len(filas_caso) > 1:
                duplicados_excel += len(filas_caso) - 1

            base = BaseRegistroInfoCoeSapFuncional.query.filter_by(
                numero=numero
            ).first()

            campos_excel_primer_row = _calificacion_extraer_campos_excel(filas_caso[0])

            if base:
                diferencias = _calificacion_comparar_base_excel(
                    base,
                    campos_excel_primer_row
                )

                if diferencias and len(diferencias_detectadas) < 100:
                    diferencias_detectadas.append({
                        "numero": numero,
                        "diferencias": diferencias,
                    })

            else:
                no_encontrados_en_base += 1

            row_calificacion = CoeSapFuncionalCalificacion.query.filter_by(
                numero=numero
            ).first()

            if row_calificacion:
                actualizados += 1
            else:
                if base:
                    campos_nuevo = _calificacion_campos_desde_base(base)
                else:
                    campos_nuevo = {
                        "numero": numero,
                        "sistema": str(numero or "")[:2],
                        "tipo_contrato": "BOLSA DE HORAS",
                    }

                    creados_solo_excel += 1

                campos_nuevo["creado_por"] = usuario
                campos_nuevo["actualizado_por"] = usuario

                row_calificacion = CoeSapFuncionalCalificacion(**campos_nuevo)
                db.session.add(row_calificacion)
                db.session.flush()

                creados += 1

            # Si existe base principal, actualizamos los campos oficiales desde la base.
            # Si no existe base, se toman del Excel.
            if base:
                campos_base = _calificacion_campos_desde_base(base)

                campos_automaticos = [
                    "base_registro_id",
                    "sistema",
                    "caso_sm",
                    "sociedad",
                    "asunto",
                    "observaciones",
                    "nombre_solicitante",
                    "impacto",
                    "urgencia",
                    "prioridad",
                    "tipo_solicitud",
                    "articulo",
                    "estado",
                    "estado_herramienta_gestion",
                    "asignado_a",
                    "fecha_asignacion",
                    "hora_ultima_actualizacion",
                    "fecha_resolucion",
                    "fecha_finalizacion_cierre",
                ]

                for campo in campos_automaticos:
                    if campo in campos_base and hasattr(row_calificacion, campo):
                        setattr(row_calificacion, campo, campos_base[campo])

            else:
                for campo in CALIFICACION_EXCEL_SOLO_EXCEL:
                    value = _calificacion_merge_ultimo_valor(filas_caso, campo)

                    if value is not None and hasattr(row_calificacion, campo):
                        setattr(row_calificacion, campo, value)

            # Campos manuales desde Excel histórico.
            for campo in CALIFICACION_EXCEL_MANUALES:
                value = _calificacion_merge_ultimo_valor(filas_caso, campo)

                if value is not None and hasattr(row_calificacion, campo):
                    setattr(row_calificacion, campo, value)

            # Limpiar horas importadas anteriormente desde Excel para que el proceso sea repetible.
            try:
                CoeSapFuncionalCalificacionHora.query.filter_by(
                    calificacion_id=row_calificacion.id,
                    origen="EXCEL"
                ).delete()
            except Exception:
                # Por si la columna origen aún no existe en alguna BD de pruebas.
                pass

            # Inicializar horas en cero antes de sumar lo importado.
            for _, _, campo_modelo, _ in CALIFICACION_EXCEL_HORAS:
                if hasattr(row_calificacion, campo_modelo):
                    setattr(row_calificacion, campo_modelo, 0)

            # Sumar horas de todas las filas del caso.
            for row_excel in filas_caso:
                horas_row = _calificacion_horas_desde_row_excel(row_excel)

                for hora_item in horas_row:
                    campo_modelo = hora_item["campo_modelo"]
                    cantidad = hora_item["horas"]

                    if hasattr(row_calificacion, campo_modelo):
                        actual = _calificacion_decimal(
                            getattr(row_calificacion, campo_modelo)
                        )

                        setattr(row_calificacion, campo_modelo, actual + cantidad)

                    movimiento = CoeSapFuncionalCalificacionHora(
                        calificacion_id=row_calificacion.id,
                        numero=numero,
                        tipo=hora_item["tipo"],
                        modulo=hora_item["modulo"],
                        horas=cantidad,
                        observacion="Importado desde Excel histórico",
                        usuario_registro=usuario,
                        created_at=datetime.utcnow(),
                    )

                    if hasattr(movimiento, "origen"):
                        movimiento.origen = "EXCEL"

                    if hasattr(movimiento, "excel_fila"):
                        movimiento.excel_fila = hora_item.get("excel_fila")

                    db.session.add(movimiento)
                    horas_movimientos += 1

            campos_actuales = {
                c.name: getattr(row_calificacion, c.name)
                for c in CoeSapFuncionalCalificacion.__table__.columns
            }

            campos_actuales = _calificacion_recalcular(campos_actuales)

            for k, v in campos_actuales.items():
                if hasattr(row_calificacion, k):
                    setattr(row_calificacion, k, v)

            row_calificacion.actualizado_por = usuario
            row_calificacion.updated_at = datetime.utcnow()

        db.session.commit()

        return jsonify({
            "mensaje": "Excel histórico de calificación procesado correctamente",
            "filas_excel": len(rows_excel),
            "casos_unicos_excel": len(grupos),
            "creados": creados,
            "actualizados": actualizados,
            "creados_solo_excel": creados_solo_excel,
            "no_encontrados_en_base": no_encontrados_en_base,
            "duplicados_excel": duplicados_excel,
            "horas_movimientos": horas_movimientos,
            "diferencias_muestra": diferencias_detectadas,
        }), 200

    except ValueError as e:
        db.session.rollback()

        return jsonify({
            "mensaje": "Archivo inválido",
            "error": str(e),
        }), 400

    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error importando Excel histórico de calificación COE SAP Funcional")

        return jsonify({
            "mensaje": "Error importando Excel histórico de calificación",
            "error": str(e),
            "trace": traceback.format_exc(),
        }), 500